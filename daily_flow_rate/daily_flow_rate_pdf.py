"""Recognize scanned daily mean flow tables arranged by day and month.

The supported form has one day column followed by twelve month columns.  The
daily body has no horizontal rules, so row order is reconstructed from real
ink anchors instead of fitting a fixed pitch.  Each selected row anchor must
be supported by several columns and is checked against the printed 1-31 day
sequence before the 12 month cells are recognized.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import traceback
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import cv2
import numpy as np
import pypdfium2 as pdfium
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Cached models must remain usable when the machine has no model-source probe.
os.environ.setdefault("PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK", "True")

from paddleocr import TextRecognition
from scipy.interpolate import PchipInterpolator

from flood_hydro_factor_pdf import log, safe_filename, verify_device
from reconstruct_grid_table import prepare_cell, result_value


MONTH_NAMES = [
    "一月",
    "二月",
    "三月",
    "四月",
    "五月",
    "六月",
    "七月",
    "八月",
    "九月",
    "十月",
    "十一月",
    "十二月",
]
COMMON_MONTH_LENGTHS = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
LEAP_MONTH_LENGTHS = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
EXPECTED_COLUMN_COUNT = 13
EXPECTED_VERTICAL_RULE_COUNT = EXPECTED_COLUMN_COUNT + 1
EXPECTED_DAY_COUNT = 31

# Printed yearbooks use more than one vertical rhythm.  Some editions leave a
# larger gap after every fifth day, others only after days 10 and 20, and a few
# use an almost uniform 31-row body.  Treat these as competing form models
# instead of forcing every page into the five-day layout.
ANCHOR_RHYTHM_MODELS = (
    ("five-day-groups", frozenset({5, 10, 15, 20, 25}), 1.82),
    ("ten-day-groups", frozenset({10, 20}), 1.82),
    ("uniform", frozenset(), 1.0),
)

NUMERIC_TRANSLATION = str.maketrans(
    {
        "O": "0",
        "o": "0",
        "Q": "0",
        "C": "0",
        "c": "0",
        "D": "0",
        "U": "0",
        "u": "0",
        "V": "0",
        "v": "0",
        "I": "1",
        "i": "1",
        "l": "1",
        "|": "1",
        "!": "1",
        "J": "1",
        "Z": "2",
        "z": "2",
        "S": "5",
        "s": "5",
        "G": "6",
        "T": "7",
        "t": "7",
        "B": "8",
        "b": "8",
        "g": "9",
        "q": "9",
    }
)


@dataclass
class AnchorCandidate:
    top: int
    bottom: int
    center: float
    ink_centroid: float
    support_columns: int
    day_column_support: bool
    score: float


@dataclass
class PageGeometry:
    page: int
    table_index: int
    width: int
    height: int
    table_left: int
    table_top: int
    table_right: int
    header_bottom: int
    statistics_top: int
    vertical_rules: list[int]
    horizontal_rules: list[int]
    candidate_anchors: list[AnchorCandidate]
    row_anchors: list[AnchorCandidate]
    anchor_method: str = "day-column-projection"
    column_center_x: list[float] = field(default_factory=list)
    row_curve_centers: list[list[float]] = field(default_factory=list)
    table_top_curve: list[float] = field(default_factory=list)
    header_curve: list[float] = field(default_factory=list)
    statistics_curve: list[float] = field(default_factory=list)
    vertical_rule_models: list[list[float]] = field(default_factory=list)
    month_decimal_anchors: list[list[dict[str, Any]]] = field(
        default_factory=list
    )
    virtual_month_anchors: list[dict[str, Any]] = field(default_factory=list)
    month_lengths: list[int] = field(default_factory=list)
    source_region_diagnostics: dict[str, Any] = field(default_factory=dict)


@dataclass
class TableRegion:
    """A table located on the photographed source page."""

    page: int
    table_index: int
    source_bbox: tuple[int, int, int, int]
    source_rules: list[int]
    source_components: list[dict[str, float]]
    image: np.ndarray
    observed_rule_flags: list[bool] = field(default_factory=list)
    rectified_rules: list[int] = field(default_factory=list)
    rectified_boundaries: dict[str, int] = field(default_factory=dict)
    diagnostics: dict[str, Any] = field(default_factory=dict)


@dataclass
class TableRegionDetection:
    """Accepted tables plus strong rejected candidates for resolution control."""

    regions: list[TableRegion]
    candidate_group_count: int
    strong_rejected_candidate_count: int
    rejected_groups: list[dict[str, Any]] = field(default_factory=list)


def clustered_positions(values: list[float], tolerance: float) -> list[float]:
    groups: list[list[float]] = []
    for value in sorted(values):
        if not groups or value - float(np.mean(groups[-1])) > tolerance:
            groups.append([value])
        else:
            groups[-1].append(value)
    return [float(np.mean(group)) for group in groups]


def vertical_rule_cluster_tolerance(image_width: int) -> float:
    """Merge double edges and nearby fragments of one photographed rule."""

    return max(5.0, image_width * 0.006)


def merge_vertical_component_fragments(
    components: list[dict[str, float]], image_height: int, image_width: int
) -> list[dict[str, float]]:
    """Join end-to-end pieces of one curved rule before lattice selection."""

    if len(components) < 2:
        return components
    # A photographed rule can drift sideways while remaining one continuous
    # curve.  The former 0.9%-page tolerance left the two ends of test22's
    # right border as separate "rules".  1.6% is still far below half of a
    # genuine month-column pitch, while accommodating ordinary page curl.
    x_tolerance = max(7.0, image_width * 0.016)
    minimum_gap = -image_height * 0.015
    # Only locally neighbouring pieces may join.  A wider gap can accidentally
    # connect the corresponding rule of two stacked tables on one page.
    maximum_gap = image_height * 0.018
    parent = list(range(len(components)))

    def find(index: int) -> int:
        while parent[index] != index:
            parent[index] = parent[parent[index]]
            index = parent[index]
        return index

    def union(left_index: int, right_index: int) -> None:
        left_root = find(left_index)
        right_root = find(right_index)
        if left_root != right_root:
            parent[right_root] = left_root

    for left_index, left in enumerate(components[:-1]):
        for right_index in range(left_index + 1, len(components)):
            right = components[right_index]
            if abs(left["x"] - right["x"]) > x_tolerance:
                continue
            vertical_gap = max(left["y"], right["y"]) - min(
                left["bottom"], right["bottom"]
            ) - 1.0
            if minimum_gap <= vertical_gap <= maximum_gap:
                union(left_index, right_index)

    groups: dict[int, list[dict[str, float]]] = {}
    for index, component in enumerate(components):
        groups.setdefault(find(index), []).append(component)
    merged: list[dict[str, float]] = []
    for fragments in groups.values():
        if len(fragments) == 1:
            item = dict(fragments[0])
            item["fragment_count"] = float(item.get("fragment_count", 1.0))
            item["contains_core"] = bool(
                item.get("contains_core", item.get("is_core", False))
            )
            merged.append(item)
            continue
        top = min(item["y"] for item in fragments)
        bottom = max(item["bottom"] for item in fragments)
        left_edge = min(item["left"] for item in fragments)
        right_edge = max(item["right"] for item in fragments)
        core_fragments = [
            item
            for item in fragments
            if bool(item.get("contains_core", item.get("is_core", False)))
        ]
        core_top = (
            min(float(item.get("core_top", item["y"])) for item in core_fragments)
            if core_fragments
            else None
        )
        core_bottom = (
            max(
                float(item.get("core_bottom", item["bottom"]))
                for item in core_fragments
            )
            if core_fragments
            else None
        )
        # Auxiliary pieces establish continuity, but their centroids can come
        # from nearby digit strokes.  The lattice x coordinate must remain
        # anchored to the original long/core pieces.
        x_fragments = core_fragments or fragments
        x_weights = np.asarray(
            [float(item["height"]) for item in x_fragments], dtype=float
        )
        merged.append(
            {
                "x": float(
                    np.average(
                        np.asarray([item["x"] for item in x_fragments], dtype=float),
                        weights=np.maximum(x_weights, 1.0),
                    )
                ),
                "y": float(top),
                "height": float(bottom - top + 1.0),
                "bottom": float(bottom),
                "center_y": float((top + bottom) / 2.0),
                "area": float(sum(item["area"] for item in fragments)),
                "left": float(left_edge),
                "right": float(right_edge),
                "width": float(right_edge - left_edge + 1.0),
                "fragment_count": float(len(fragments)),
                "contains_core": bool(
                    any(
                        item.get("contains_core", item.get("is_core", False))
                        for item in fragments
                    )
                ),
                "core_top": core_top,
                "core_bottom": core_bottom,
                "core_height": (
                    None
                    if core_top is None or core_bottom is None
                    else float(core_bottom - core_top + 1.0)
                ),
            }
        )
    return merged


def composite_vertical_rule_evidence(
    components: list[dict[str, float]],
    rule_x: float,
    match_tolerance: float,
    pitch: float,
) -> tuple[dict[str, float], float]:
    """Aggregate vertically separated pieces of one locally curved rule.

    Connected-component centroids are only a 2-D summary of a photographed
    curve.  Its upper and lower pieces may therefore have different x values.
    Evidence is combined inside a guarded tube narrower than half a column;
    neighbouring month rules can never enter the same tube.
    """

    if not components:
        raise RuntimeError("竖线证据为空。")
    tube_radius = max(match_tolerance, pitch * 0.34)
    pieces = [
        item
        for item in components
        if abs(float(item["x"]) - float(rule_x)) <= tube_radius
    ]
    if not pieces:
        nearest = min(
            components, key=lambda item: abs(float(item["x"]) - float(rule_x))
        )
        pieces = [nearest]
    # A photographed page edge can run nearly parallel to the real outer rule
    # inside the same x tube.  Such components overlap for most of their height
    # and are alternatives, not fragments of one curve.  Aggregate only pieces
    # that are vertically complementary; otherwise keep the component nearest
    # to the proposed track.  This retains split curved rules while preventing
    # a nearby book/page edge from stretching the true outer-border interval.
    if len(pieces) > 1:
        ordered_pieces = sorted(
            pieces,
            key=lambda item: abs(float(item["x"]) - float(rule_x)),
        )
        coherent = [ordered_pieces[0]]
        coherent_top = float(ordered_pieces[0]["y"])
        coherent_bottom = float(
            ordered_pieces[0].get(
                "bottom",
                coherent_top + float(ordered_pieces[0]["height"]) - 1.0,
            )
        )
        for item in ordered_pieces[1:]:
            item_top = float(item["y"])
            item_bottom = float(
                item.get("bottom", item_top + float(item["height"]) - 1.0)
            )
            overlap = max(
                0.0,
                min(coherent_bottom, item_bottom)
                - max(coherent_top, item_top)
                + 1.0,
            )
            shorter_height = max(
                1.0,
                min(
                    coherent_bottom - coherent_top + 1.0,
                    item_bottom - item_top + 1.0,
                ),
            )
            if overlap / shorter_height >= 0.55:
                continue
            vertical_gap = max(coherent_top, item_top) - min(
                coherent_bottom, item_bottom
            ) - 1.0
            if vertical_gap > pitch * 0.65:
                continue
            coherent.append(item)
            coherent_top = min(coherent_top, item_top)
            coherent_bottom = max(coherent_bottom, item_bottom)
        pieces = coherent
    minimum_distance = min(
        abs(float(item["x"]) - float(rule_x)) for item in pieces
    )
    weights = np.asarray(
        [max(1.0, float(item.get("height", 1.0))) for item in pieces],
        dtype=float,
    )
    top = min(float(item["y"]) for item in pieces)
    bottom = max(
        float(
            item.get(
                "bottom", float(item["y"]) + float(item["height"]) - 1.0
            )
        )
        for item in pieces
    )
    left_edge = min(float(item.get("left", item["x"])) for item in pieces)
    right_edge = max(float(item.get("right", item["x"])) for item in pieces)
    core_pieces = [
        item
        for item in pieces
        if item.get("core_top") is not None and item.get("core_bottom") is not None
    ]
    core_top = (
        min(float(item["core_top"]) for item in core_pieces)
        if core_pieces
        else None
    )
    core_bottom = (
        max(float(item["core_bottom"]) for item in core_pieces)
        if core_pieces
        else None
    )
    return (
        {
            "x": float(
                np.average(
                    np.asarray([float(item["x"]) for item in pieces]),
                    weights=weights,
                )
            ),
            "y": float(top),
            "height": float(bottom - top + 1.0),
            "bottom": float(bottom),
            "center_y": float((top + bottom) / 2.0),
            "area": float(sum(float(item.get("area", 0.0)) for item in pieces)),
            "left": float(left_edge),
            "right": float(right_edge),
            "width": float(right_edge - left_edge + 1.0),
            "fragment_count": float(
                sum(float(item.get("fragment_count", 1.0)) for item in pieces)
            ),
            "evidence_piece_count": float(len(pieces)),
            "contains_core": bool(core_pieces),
            "core_top": core_top,
            "core_bottom": core_bottom,
            "core_height": (
                None
                if core_top is None or core_bottom is None
                else float(core_bottom - core_top + 1.0)
            ),
        },
        float(minimum_distance),
    )


def vertical_rule_candidate_metrics(
    components: list[dict[str, float]],
    rules: list[float] | np.ndarray,
    match_tolerance: float,
) -> dict[str, float]:
    """Measure whether two endpoints look like one table's outer borders.

    Equal spacing alone cannot distinguish a true table lattice from a shifted
    lattice completed by a page edge, book spine, or nearby object.  Real outer
    borders cover almost the same vertical interval and normally extend beyond
    the internal month separators into the header/statistics bands.
    """

    rule_values = np.asarray(rules, dtype=float)
    if len(rule_values) != EXPECTED_VERTICAL_RULE_COUNT or not components:
        return {}

    gaps = np.diff(rule_values)
    pitch = float(np.median(gaps))

    def nearest(value: float) -> tuple[dict[str, float], float]:
        return composite_vertical_rule_evidence(
            components, value, match_tolerance, pitch
        )

    left_component, left_distance = nearest(float(rule_values[0]))
    right_component, right_distance = nearest(float(rule_values[-1]))

    def interval(component: dict[str, float]) -> tuple[float, float, float]:
        top = float(component["y"])
        bottom = float(
            component.get(
                "bottom", top + float(component["height"]) - 1.0
            )
        )
        height = max(1.0, bottom - top + 1.0)
        return top, bottom, height

    left_top, left_bottom, left_height = interval(left_component)
    right_top, right_bottom, right_height = interval(right_component)
    overlap = max(
        0.0, min(left_bottom, right_bottom) - max(left_top, right_top) + 1.0
    )
    union = max(left_bottom, right_bottom) - min(left_top, right_top) + 1.0
    interval_iou = overlap / max(1.0, union)
    height_balance = min(left_height, right_height) / max(
        left_height, right_height
    )

    internal_heights: list[float] = []
    for value in rule_values[1:-1]:
        component, distance = nearest(float(value))
        if distance <= match_tolerance:
            internal_heights.append(interval(component)[2])
    if len(internal_heights) >= 4:
        internal_height = float(np.median(internal_heights))
    else:
        internal_height = float(
            np.median([interval(component)[2] for component in components])
        )
    minimum_relative_height = min(left_height, right_height) / max(
        1.0, internal_height
    )

    month_gaps = gaps[1:]
    month_gap_median = float(np.median(month_gaps))
    month_gap_cv = float(
        np.std(month_gaps) / max(1.0, float(np.mean(month_gaps)))
    )
    month_axis = np.arange(len(month_gaps), dtype=float)
    if len(month_gaps) >= 3:
        month_trend = np.polyval(
            np.polyfit(month_axis, month_gaps, 1), month_axis
        )
        month_gap_trend_residual = float(
            np.sqrt(np.mean(np.square(month_gaps - month_trend)))
            / max(1.0, month_gap_median)
        )
    else:
        month_gap_trend_residual = month_gap_cv
    day_gap_ratio = float(gaps[0] / max(1.0, month_gap_median))

    return {
        "left_top": left_top,
        "left_bottom": left_bottom,
        "left_height": left_height,
        "right_top": right_top,
        "right_bottom": right_bottom,
        "right_height": right_height,
        "left_match_distance": left_distance,
        "right_match_distance": right_distance,
        "vertical_interval_iou": interval_iou,
        "height_balance": height_balance,
        "internal_height_median": internal_height,
        "minimum_relative_height": minimum_relative_height,
        "day_gap_ratio": day_gap_ratio,
        "month_gap_cv": month_gap_cv,
        "month_gap_trend_residual": month_gap_trend_residual,
        "left_evidence_piece_count": float(
            left_component.get("evidence_piece_count", 1.0)
        ),
        "right_evidence_piece_count": float(
            right_component.get("evidence_piece_count", 1.0)
        ),
    }


def passes_outer_border_topology(metrics: dict[str, float]) -> bool:
    """Fail closed when two candidate endpoints do not form one outer frame."""

    return bool(
        metrics
        and metrics["vertical_interval_iou"] >= 0.78
        and metrics["height_balance"] >= 0.80
        and metrics["minimum_relative_height"] >= 1.05
        and 0.70 <= metrics["day_gap_ratio"] <= 1.45
    )


def passes_curved_outer_border_topology(metrics: dict[str, float]) -> bool:
    """Guarded recovery when a curved outer border is only partly connected.

    The relaxed interval test is allowed only when all thirteen column gaps
    already form a very regular daily-flow lattice.  Thus a short photographed
    outer fragment may be accepted, but an arbitrary page/object edge cannot
    manufacture a table from spacing alone.
    """

    return bool(
        metrics
        and metrics["vertical_interval_iou"] >= 0.62
        and metrics["height_balance"] >= 0.62
        and metrics["minimum_relative_height"] >= 0.72
        and 0.70 <= metrics["day_gap_ratio"] <= 1.45
        and metrics["month_gap_cv"] <= 0.055
        and metrics["month_gap_trend_residual"] <= 0.045
    )


def passes_one_sided_outer_completion(
    matched: np.ndarray,
    nearest: np.ndarray,
    pitch: float,
    topology: dict[str, float],
) -> bool:
    """Allow a strongly regular lattice whose far outer edge faded away.

    Close photographs sometimes retain 10-13 consecutive rules from one outer
    border while the opposite edge and a few adjacent month rules disappear in
    glare or page curl.  This guarded path requires one *observed* outer border,
    a contiguous run of at least ten tracks, no more than one internal hole and
    at most four extrapolated tracks at the missing edge.  A shifted page edge
    therefore cannot pass merely by matching the interior pitch.
    """

    if len(matched) != EXPECTED_VERTICAL_RULE_COUNT or not topology:
        return False
    left_observed = bool(matched[0])
    right_observed = bool(matched[-1])
    if left_observed == right_observed:
        return False
    observed_indices = np.flatnonzero(matched)
    if observed_indices.size < 10:
        return False
    first = int(observed_indices[0])
    last = int(observed_indices[-1])
    if left_observed:
        trailing_missing = EXPECTED_VERTICAL_RULE_COUNT - 1 - last
        edge_missing = trailing_missing
        internal_missing = int(np.count_nonzero(~matched[: last + 1]))
        observed_height = float(topology["left_height"])
    else:
        edge_missing = first
        internal_missing = int(np.count_nonzero(~matched[first:]))
        observed_height = float(topology["right_height"])
    residual = float(
        np.median(nearest[matched]) / max(1.0, float(pitch))
    )
    relative_height = observed_height / max(
        1.0, float(topology["internal_height_median"])
    )
    return bool(
        1 <= edge_missing <= 4
        and internal_missing <= 1
        and residual <= 0.13
        and relative_height >= 1.02
        and 0.70 <= float(topology["day_gap_ratio"]) <= 1.45
        and float(topology["month_gap_cv"]) <= 0.065
        and float(topology["month_gap_trend_residual"]) <= 0.055
    )


def render_pdf_page(pdf_path: Path, page_index: int, width: int) -> np.ndarray:
    document = pdfium.PdfDocument(str(pdf_path))
    page = document[page_index]
    scale = width / page.get_width()
    image = page.render(scale=scale).to_pil().convert("RGB")
    page.close()
    document.close()
    array = np.asarray(image)
    return cv2.cvtColor(array, cv2.COLOR_RGB2BGR)


def adaptive_ink(gray: np.ndarray) -> np.ndarray:
    """Retain faint printing under uneven illumination in photographed pages."""

    return cv2.adaptiveThreshold(
        gray,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV,
        41,
        13,
    )


def select_or_complete_regular_vertical_rules(
    components: list[dict[str, float]], image_width: int
) -> tuple[list[int], list[bool], dict[str, Any]]:
    """Select a 14-rule lattice while retaining weak, fragmented rules.

    A photographed curved rule is often split into a shorter connected
    component.  The old detector discarded the whole table when fewer than 14
    *long* components survived.  Here the regular 13-column topology may fill
    at most four internal tracks, but both outer borders must still be directly
    observed.  This prevents desk/page edges from creating a table by
    extrapolation alone.
    """

    tolerance = vertical_rule_cluster_tolerance(image_width)
    positions = clustered_positions(
        [float(item["x"]) for item in components], tolerance=tolerance
    )
    if len(positions) < 10:
        raise RuntimeError(
            f"仅检测到{len(positions)}条可用竖线片段，不能安全补全14条规则。"
        )

    # Preserve the proven exact path whenever fourteen complete positions are
    # available.  The lattice fallback below is only used for fragmentation.
    try:
        exact = select_regular_vertical_rules(components, image_width)
        exact_pitch = float(np.median(np.diff(np.asarray(exact, dtype=float))))
        exact_topology = vertical_rule_candidate_metrics(
            components,
            exact,
            max(tolerance * 1.6, exact_pitch * 0.14),
        )
        observed = [
            min(abs(float(value) - position) for position in positions)
            <= max(tolerance * 1.5, exact_pitch * 0.14)
            for value in exact
        ]
        return exact, observed, {
            "method": (
                "complete-components"
                if passes_outer_border_topology(exact_topology)
                else "complete-components-curved-outer"
            ),
            "observed_position_count": len(positions),
            "inferred_rule_indices": [
                index for index, flag in enumerate(observed) if not flag
            ],
        }
    except RuntimeError:
        pass

    minimum_pitch = image_width * 0.025
    maximum_pitch = image_width * 0.105
    model_keys: set[tuple[int, int]] = set()
    models: list[tuple[float, float]] = []
    for left_index in range(len(positions) - 1):
        for right_index in range(left_index + 1, len(positions)):
            separation = positions[right_index] - positions[left_index]
            for logical_steps in range(1, EXPECTED_VERTICAL_RULE_COUNT):
                pitch = separation / logical_steps
                if not minimum_pitch <= pitch <= maximum_pitch:
                    continue
                for logical_left in range(
                    EXPECTED_VERTICAL_RULE_COUNT - logical_steps
                ):
                    start = positions[left_index] - logical_left * pitch
                    end = start + (EXPECTED_VERTICAL_RULE_COUNT - 1) * pitch
                    if start < -pitch * 0.15 or end > image_width + pitch * 0.15:
                        continue
                    key = (round(start / 1.5), round(pitch / 0.35))
                    if key not in model_keys:
                        model_keys.add(key)
                        models.append((start, pitch))

    candidates: list[
        tuple[tuple[float, ...], np.ndarray, np.ndarray, bool]
    ] = []
    observed_positions = np.asarray(positions, dtype=float)

    for start, pitch in models:
        grid = start + np.arange(EXPECTED_VERTICAL_RULE_COUNT, dtype=float) * pitch
        nearest = np.min(
            np.abs(grid[:, None] - observed_positions[None, :]), axis=1
        )
        match_tolerance = max(tolerance * 1.6, pitch * 0.14)
        matched = nearest <= match_tolerance
        match_count = int(np.count_nonzero(matched))
        if match_count < 10:
            continue
        topology = vertical_rule_candidate_metrics(
            components, grid, match_tolerance
        )
        ordinary_topology = bool(
            bool(matched[0])
            and bool(matched[-1])
            and (
                passes_outer_border_topology(topology)
                or (
                    match_count == EXPECTED_VERTICAL_RULE_COUNT
                    and passes_curved_outer_border_topology(topology)
                )
            )
        )
        one_sided_completion = bool(
            len(positions) <= EXPECTED_VERTICAL_RULE_COUNT - 1
            and passes_one_sided_outer_completion(
                matched, nearest, pitch, topology
            )
        )
        if not (ordinary_topology or one_sided_completion):
            continue
        residual = float(np.median(nearest[matched]) / max(1.0, pitch))
        span_ratio = float((grid[-1] - grid[0]) / image_width)
        if not 0.55 <= span_ratio <= 0.93:
            continue
        # Prefer more direct observations, then a low residual.  Span is only
        # a weak prior because close photographs legitimately change it.
        score = (
            float(one_sided_completion),
            float(-match_count),
            -topology["vertical_interval_iou"],
            -topology["height_balance"],
            -topology["minimum_relative_height"],
            float(EXPECTED_VERTICAL_RULE_COUNT - match_count),
            residual,
            abs(span_ratio - 0.74) * 0.12,
        )
        candidates.append((score, grid, matched, one_sided_completion))
    if not candidates:
        return _select_nonuniform_vertical_lattice(
            components, positions, image_width, tolerance
        )
    two_sided_candidates = [item for item in candidates if not item[3]]
    if two_sided_candidates:
        candidates = two_sided_candidates
    else:
        # A curved two-sided solution is stronger evidence than a linear
        # one-sided extrapolation.  Try it before accepting the guarded edge
        # completion, then fall back only when it truly cannot be formed.
        try:
            return _select_nonuniform_vertical_lattice(
                components, positions, image_width, tolerance
            )
        except RuntimeError:
            pass
    _, grid, matched, one_sided_completion = min(
        candidates, key=lambda item: item[0]
    )
    return (
        [round(float(value)) for value in grid],
        [bool(value) for value in matched.tolist()],
        {
            "method": (
                "one-sided-topology-completed-lattice"
                if one_sided_completion
                else "topology-completed-lattice"
            ),
            "observed_position_count": len(positions),
            "inferred_rule_indices": [
                index for index, flag in enumerate(matched.tolist()) if not flag
            ],
        },
    )


def _select_nonuniform_vertical_lattice(
    components: list[dict[str, float]],
    positions: list[float],
    image_width: int,
    tolerance: float,
) -> tuple[list[int], list[bool], dict[str, Any]]:
    """Recover a smoothly varying 14-track lattice from partial observations.

    Close photographed pages have perspective and curl, so column gaps can vary
    monotonically across the page.  A constant-pitch model then fails even when
    ten or more true rules remain.  Candidate observed outer borders are paired,
    interior observations are assigned monotonically to logical indices, and a
    quadratic x(index) model fills at most four missing tracks.  Page edges are
    naturally rejected by their larger normalized residual/topology penalty.
    """

    observed_positions = np.asarray(positions, dtype=float)
    candidate_models: list[
        tuple[tuple[float, ...], np.ndarray, np.ndarray, float]
    ] = []
    minimum_span = image_width * 0.55
    maximum_span = image_width * 0.93

    for left_index in range(len(positions) - 1):
        for right_index in range(left_index + 1, len(positions)):
            left = float(positions[left_index])
            right = float(positions[right_index])
            span = right - left
            segment = observed_positions[left_index : right_index + 1]
            if not minimum_span <= span <= maximum_span or len(segment) < 10:
                continue

            linear_grid = np.linspace(
                left, right, EXPECTED_VERTICAL_RULE_COUNT, dtype=float
            )
            mapping: dict[int, float] = {0: left, 13: right}
            for value in segment[1:-1]:
                logical_index = int(
                    np.argmin(np.abs(linear_grid - float(value)))
                )
                if not 1 <= logical_index <= 12:
                    continue
                previous = mapping.get(logical_index)
                if previous is None or abs(
                    float(value) - linear_grid[logical_index]
                ) < abs(previous - linear_grid[logical_index]):
                    mapping[logical_index] = float(value)
            if len(mapping) < 10:
                continue

            grid = linear_grid
            for _ in range(3):
                logical = np.asarray(sorted(mapping), dtype=float)
                values = np.asarray(
                    [mapping[int(index)] for index in logical], dtype=float
                )
                weights = np.ones(len(logical), dtype=float)
                weights[(logical == 0) | (logical == 13)] = 4.0
                coefficients = np.polyfit(logical, values, 2, w=weights)
                grid = np.polyval(
                    coefficients,
                    np.arange(EXPECTED_VERTICAL_RULE_COUNT, dtype=float),
                )
                grid += np.linspace(
                    left - float(grid[0]),
                    right - float(grid[-1]),
                    EXPECTED_VERTICAL_RULE_COUNT,
                    dtype=float,
                )
                gaps = np.diff(grid)
                median_gap = float(np.median(gaps))
                if (
                    median_gap <= 0
                    or float(np.min(gaps)) < median_gap * 0.55
                    or float(np.max(gaps)) > median_gap * 1.55
                ):
                    mapping = {}
                    break
                remapped: dict[int, float] = {0: left, 13: right}
                for value in segment[1:-1]:
                    logical_index = int(np.argmin(np.abs(grid - float(value))))
                    if not 1 <= logical_index <= 12:
                        continue
                    previous = remapped.get(logical_index)
                    if previous is None or abs(
                        float(value) - grid[logical_index]
                    ) < abs(previous - grid[logical_index]):
                        remapped[logical_index] = float(value)
                mapping = remapped
                if len(mapping) < 10:
                    break
            if len(mapping) < 10:
                continue

            gaps = np.diff(grid)
            pitch = float(np.median(gaps))
            match_tolerance = max(tolerance * 1.8, pitch * 0.20)
            nearest = np.min(
                np.abs(grid[:, None] - observed_positions[None, :]), axis=1
            )
            matched = nearest <= match_tolerance
            match_count = int(np.count_nonzero(matched))
            if match_count < 10 or not bool(matched[0]) or not bool(matched[-1]):
                continue
            topology = vertical_rule_candidate_metrics(
                components, grid, match_tolerance
            )
            if not (
                passes_outer_border_topology(topology)
                or (
                    match_count >= 12
                    and passes_curved_outer_border_topology(topology)
                )
            ):
                continue
            residual = float(np.median(nearest[matched]) / max(1.0, pitch))
            outer_height_ratio = (
                (topology["left_height"] + topology["right_height"])
                / 2.0
                / max(1.0, topology["internal_height_median"])
            )
            span_ratio = span / image_width
            score = (
                float(-match_count),
                residual,
                (1.0 - topology["vertical_interval_iou"]) * 0.30
                + (1.0 - topology["height_balance"]) * 0.15
                + abs(outer_height_ratio - 1.20) * 0.05,
                abs(span_ratio - 0.79) * 0.08,
            )
            candidate_models.append((score, grid, matched, residual))

    if not candidate_models:
        raise RuntimeError(
            f"{len(positions)}条竖线片段不能组成有双侧实证的14轨规则网格。"
        )
    _, grid, matched, residual = min(candidate_models, key=lambda item: item[0])
    return (
        [round(float(value)) for value in grid],
        [bool(value) for value in matched.tolist()],
        {
            "method": "nonuniform-quadratic-lattice",
            "observed_position_count": len(positions),
            "inferred_rule_indices": [
                index for index, flag in enumerate(matched.tolist()) if not flag
            ],
            "normalized_fit_residual": round(float(residual), 6),
            "candidate_model_count": len(candidate_models),
        },
    )


def detect_table_regions_with_diagnostics(
    image: np.ndarray, page_number: int
) -> TableRegionDetection:
    """Find tables and retain strong rejected groups for resolution control."""

    height, width = image.shape[:2]
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    ink = adaptive_ink(gray)
    vertical_long = cv2.morphologyEx(
        ink,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(
            cv2.MORPH_RECT, (1, max(35, round(height * 0.021)))
        ),
    )
    vertical_local = cv2.morphologyEx(
        ink,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(
            cv2.MORPH_RECT, (1, max(21, round(height * 0.012)))
        ),
    )
    vertical = cv2.bitwise_or(vertical_long, vertical_local)
    count, _, stats, _ = cv2.connectedComponentsWithStats(vertical, 8)
    components: list[dict[str, float]] = []
    maximum_component_width = max(40.0, width * 0.035)
    for index in range(1, count):
        x, y, component_width, component_height, area = stats[index]
        is_core = bool(
            height * 0.075 <= component_height <= height * 0.72
            and component_width <= maximum_component_width
            and component_width / max(1.0, component_height) <= 0.105
            and area >= component_height * 0.22
        )
        if (
            max(18.0, height * 0.009) <= component_height <= height * 0.72
            and component_width <= maximum_component_width
            and component_width / max(1.0, component_height) <= 0.16
            and area >= component_height * 0.55
        ):
            components.append(
                {
                    "x": float(x + component_width / 2),
                    "y": float(y),
                    "height": float(component_height),
                    "bottom": float(y + component_height - 1),
                    "center_y": float(y + (component_height - 1) / 2),
                    "area": float(area),
                    "left": float(x),
                    "right": float(x + component_width - 1),
                    "width": float(component_width),
                    "fragment_count": 1.0,
                    "contains_core": is_core,
                    "core_top": float(y) if is_core else None,
                    "core_bottom": (
                        float(y + component_height - 1) if is_core else None
                    ),
                    "core_height": float(component_height) if is_core else None,
                }
            )
    raw_component_count = len(components)
    components = merge_vertical_component_fragments(components, height, width)
    # Short local pieces are bridge evidence only.  After chaining, retain the
    # same long-rule standard used by the original detector so digit strokes do
    # not become lattice tracks.
    components = [
        item
        for item in components
        if (
            height * 0.075 <= float(item["height"]) <= height * 0.72
            and float(item.get("width", 1.0)) <= max(40.0, width * 0.045)
            and float(item.get("width", 1.0))
            / max(1.0, float(item["height"]))
            <= 0.105
            and float(item.get("area", 0.0))
            >= float(item["height"]) * 0.22
            and bool(item.get("contains_core", False))
        )
    ]

    groups: list[list[dict[str, float]]] = []
    for component in sorted(components, key=lambda item: item["center_y"]):
        if (
            not groups
            or abs(
                component["center_y"]
                - float(
                    np.median([item["center_y"] for item in groups[-1]])
                )
            )
            > height * 0.12
        ):
            groups.append([component])
        else:
            groups[-1].append(component)

    regions: list[TableRegion] = []
    rejected_groups: list[dict[str, Any]] = []
    candidate_group_count = 0
    strong_rejected_candidate_count = 0
    for group in groups:
        group_positions = clustered_positions(
            [item["x"] for item in group],
            tolerance=vertical_rule_cluster_tolerance(width),
        )
        group_span = (
            float(max(group_positions) - min(group_positions))
            if len(group_positions) >= 2
            else 0.0
        )
        median_group_height = float(
            np.median([item["height"] for item in group])
        )
        strong_candidate = bool(
            len(group_positions) >= 10
            and group_span >= width * 0.50
            and median_group_height >= height * 0.15
        )
        if len(group_positions) >= 8:
            candidate_group_count += 1
        try:
            rules, observed_flags, diagnostics = (
                select_or_complete_regular_vertical_rules(group, width)
            )
        except RuntimeError as error:
            if len(group_positions) >= 8:
                log(
                    f"[表区诊断] 第{page_number}页候选竖线组被拒绝：{error}"
                )
            rejected_groups.append(
                {
                    "component_count": len(group),
                    "position_count": len(group_positions),
                    "x_span": round(group_span, 3),
                    "median_component_height": round(median_group_height, 3),
                    "strong_candidate": strong_candidate,
                    "reason": str(error),
                }
            )
            if strong_candidate:
                strong_rejected_candidate_count += 1
            continue
        pitch = float(np.median(np.diff(np.asarray(rules, dtype=float))))
        selected: list[dict[str, float]] = []
        median_top = float(np.median([item["y"] for item in group]))
        median_bottom = float(np.median([item["bottom"] for item in group]))
        for rule, observed in zip(rules, observed_flags):
            evidence, evidence_distance = composite_vertical_rule_evidence(
                group,
                float(rule),
                max(vertical_rule_cluster_tolerance(width) * 1.6, pitch * 0.14),
                pitch,
            )
            if observed and evidence_distance <= max(7.0, pitch * 0.18):
                selected.append(evidence)
            else:
                selected.append(
                    {
                        "x": float(rule),
                        "y": median_top,
                        "height": median_bottom - median_top + 1,
                        "bottom": median_bottom,
                        "center_y": (median_top + median_bottom) / 2,
                        "area": 0.0,
                    }
                )
        outer_evidence = [
            selected[index]
            for index in (0, EXPECTED_VERTICAL_RULE_COUNT - 1)
            if observed_flags[index]
        ]
        permits_one_sided_outer = (
            diagnostics.get("method")
            == "one-sided-topology-completed-lattice"
        )
        if len(outer_evidence) < (1 if permits_one_sided_outer else 2):
            continue
        # Outer borders begin at the table top; internal separators usually
        # begin at the header bottom and must never define the source top.
        top = max(
            0,
            round(
                min(
                    float(item.get("core_top", item["y"]))
                    if item.get("core_top") is not None
                    else float(item["y"])
                    for item in outer_evidence
                )
            ),
        )
        outer_bottom = max(
            float(item.get("core_bottom", item["bottom"]))
            if item.get("core_bottom") is not None
            else float(item["bottom"])
            for item in outer_evidence
        )
        bottom = min(height - 1, round(outer_bottom))
        outer_topology = vertical_rule_candidate_metrics(
            group,
            rules,
            max(vertical_rule_cluster_tolerance(width) * 1.6, pitch * 0.14),
        )
        diagnostics.update(
            {
                "component_count": len(group),
                "raw_page_component_count": raw_component_count,
                "merged_page_component_count": len(components),
                "merged_fragment_count": int(
                    sum(item.get("fragment_count", 1.0) for item in group)
                ),
                "source_rule_x": [int(value) for value in rules],
                "observed_rule_flags": observed_flags,
                "source_bbox": [rules[0], top, rules[-1], bottom],
                "outer_border_topology": {
                    key: round(float(value), 6)
                    for key, value in outer_topology.items()
                },
            }
        )
        regions.append(
            TableRegion(
                page=page_number,
                table_index=0,
                source_bbox=(rules[0], top, rules[-1], bottom),
                source_rules=rules,
                source_components=selected,
                image=np.empty((0, 0, 3), dtype=np.uint8),
                observed_rule_flags=observed_flags,
                diagnostics=diagnostics,
            )
        )
    regions.sort(key=lambda item: item.source_bbox[1])
    for table_index, region in enumerate(regions, start=1):
        region.table_index = table_index
    return TableRegionDetection(
        regions=regions,
        candidate_group_count=candidate_group_count,
        strong_rejected_candidate_count=strong_rejected_candidate_count,
        rejected_groups=rejected_groups,
    )


def detect_table_regions(image: np.ndarray, page_number: int) -> list[TableRegion]:
    """Compatibility wrapper returning only accepted daily-flow tables."""

    return detect_table_regions_with_diagnostics(image, page_number).regions


def assign_region_boundary_scan_limits(
    regions: list[TableRegion], image_height: int
) -> None:
    """Give every table a non-overlapping vertical ownership interval.

    A photographed page can contain two daily-flow tables only a title-height
    apart.  Boundary discovery must therefore never use the generous padding
    of one table to vote on the neighbouring table.  The midpoint between two
    detected outer boxes is a neutral ownership boundary.  The top search is
    deliberately tighter than the bottom search: a false top normally belongs
    to the preceding table, while the photographed bottom rule can bow farther
    below the vertical-line endpoints.
    """

    if image_height < 1:
        raise ValueError("页面高度必须为正数。")
    ordered = sorted(regions, key=lambda item: item.source_bbox[1])
    for index, region in enumerate(ordered):
        _, source_top, _, source_bottom = region.source_bbox
        rules = np.asarray(region.source_rules, dtype=float)
        pitch = (
            float(np.median(np.diff(rules)))
            if len(rules) >= 2
            else max(12.0, (source_bottom - source_top) / 8.0)
        )
        top_slack = max(10, round(pitch * 0.22) + 4)
        bottom_slack = max(25, round(pitch * 0.45) + 4)
        ownership_top = 0
        if index > 0:
            previous_bottom = ordered[index - 1].source_bbox[3]
            if previous_bottom < source_top:
                ownership_top = round((previous_bottom + source_top) / 2.0)
        ownership_bottom = image_height - 1
        if index + 1 < len(ordered):
            following_top = ordered[index + 1].source_bbox[1]
            if source_bottom < following_top:
                ownership_bottom = round((source_bottom + following_top) / 2.0)
        scan_top = max(0, ownership_top, source_top - top_slack)
        scan_bottom = min(
            image_height - 1, ownership_bottom, source_bottom + bottom_slack
        )
        if scan_bottom <= scan_top:
            raise RuntimeError(
                f"第{region.page}页表{region.table_index}的边界扫描区间为空。"
            )
        region.diagnostics["boundary_scan_limits"] = [
            int(scan_top),
            int(scan_bottom),
        ]
        region.diagnostics["boundary_ownership_limits"] = [
            int(ownership_top),
            int(ownership_bottom),
        ]
        region.diagnostics["boundary_scan_slack"] = {
            "top": int(top_slack),
            "bottom": int(bottom_slack),
        }


def _clamped_curve_values(
    sample_x: np.ndarray, sample_y: np.ndarray, query_x: np.ndarray
) -> np.ndarray:
    """Evaluate a shape-preserving curve without unsafe edge extrapolation."""

    if len(sample_x) >= 3 and np.all(np.diff(sample_x) > 0):
        values = np.asarray(
            PchipInterpolator(sample_x, sample_y, extrapolate=True)(query_x),
            dtype=float,
        )
    else:
        values = np.interp(query_x, sample_x, sample_y).astype(float)
    values[query_x <= sample_x[0]] = sample_y[0]
    values[query_x >= sample_x[-1]] = sample_y[-1]
    return values


def _regularize_horizontal_samples(
    sample_x: np.ndarray,
    raw_y: np.ndarray,
    search_radius: int,
    pitch: float,
) -> np.ndarray:
    """Remove isolated glyph snaps while retaining a smoothly bowed rule."""

    values = raw_y.astype(float).copy()
    if len(values) < 4:
        return values
    keep = np.ones(len(values), dtype=bool)
    coefficients = np.polyfit(sample_x, values, 2)
    for _ in range(4):
        prediction = np.polyval(coefficients, sample_x)
        residual = values - prediction
        centered = residual[keep] - float(np.median(residual[keep]))
        mad = float(np.median(np.abs(centered))) if centered.size else 0.0
        threshold = max(3.5, search_radius * 0.26, mad * 3.2)
        revised = np.abs(residual) <= threshold
        if int(np.count_nonzero(revised)) < 5 or np.array_equal(revised, keep):
            break
        keep = revised
        coefficients = np.polyfit(sample_x[keep], values[keep], 2)
    prediction = np.polyval(coefficients, sample_x)
    replacement_threshold = max(4.0, search_radius * 0.30)
    values[np.abs(values - prediction) > replacement_threshold] = prediction[
        np.abs(values - prediction) > replacement_threshold
    ]

    # A photographed page can fold, so this is a local slope limiter rather
    # than a global straight/quadratic replacement.
    maximum_step = max(4.0, pitch * 0.085)
    for _ in range(3):
        forward = values.copy()
        backward = values.copy()
        for index in range(1, len(values)):
            forward[index] = float(
                np.clip(
                    forward[index],
                    forward[index - 1] - maximum_step,
                    forward[index - 1] + maximum_step,
                )
            )
        for index in range(len(values) - 2, -1, -1):
            backward[index] = float(
                np.clip(
                    backward[index],
                    backward[index + 1] - maximum_step,
                    backward[index + 1] + maximum_step,
                )
            )
        values = (forward + backward) / 2.0
    return values


def _track_horizontal_rule_samples(
    gray: np.ndarray,
    horizontal_evidence: np.ndarray,
    rules: np.ndarray,
    nominal_y: float,
    search_radius: int,
) -> tuple[np.ndarray, np.ndarray]:
    """Track one curved horizontal rule independently through all 13 cells."""

    height, width = gray.shape
    centers_x = (rules[:-1] + rules[1:]) / 2.0
    raw: list[float] = []
    for column in range(EXPECTED_COLUMN_COUNT):
        cell_width = max(8.0, rules[column + 1] - rules[column])
        x0 = max(0, round(rules[column] + cell_width * 0.12))
        x1 = min(width, round(rules[column + 1] - cell_width * 0.12) + 1)
        y0 = max(1, round(nominal_y) - search_radius)
        y1 = min(height - 1, round(nominal_y) + search_radius + 1)
        if x1 <= x0 or y1 <= y0:
            raw.append(float(nominal_y))
            continue
        scores: list[float] = []
        for y in range(y0, y1):
            line_window = horizontal_evidence[y - 1 : y + 2, x0:x1]
            gray_window = gray[y - 1 : y + 2, x0:x1]
            support = float(np.count_nonzero(line_window)) / max(1, line_window.size)
            darkness = float(np.mean(255 - gray_window.astype(np.float32))) / 255.0
            distance = abs(y - nominal_y) / max(1.0, search_radius)
            scores.append(support * 3.2 + darkness * 0.32 - distance * 0.08)
        raw.append(float(y0 + int(np.argmax(np.asarray(scores, dtype=float)))))
    raw_values = np.asarray(raw, dtype=float)
    pitch = float(np.median(np.diff(rules)))
    return raw_values, _regularize_horizontal_samples(
        centers_x, raw_values, search_radius, pitch
    )


def _detect_source_horizontal_boundaries_legacy(
    gray: np.ndarray, ink: np.ndarray, region: TableRegion
) -> tuple[np.ndarray, dict[str, np.ndarray], np.ndarray, dict[str, Any]]:
    """Legacy global-y seed finder retained as a guarded fallback."""

    image_height, image_width = gray.shape
    rules = np.asarray(region.source_rules, dtype=float)
    pitch = float(np.median(np.diff(rules)))
    _, source_top, _, source_bottom = region.source_bbox
    source_span = max(1, source_bottom - source_top)
    upper_padding = max(
        18, round(source_span * 0.16), round(pitch * 3.6)
    )
    lower_padding = max(
        18, round(source_span * 0.16), round(pitch * 3.0)
    )
    y0 = max(0, source_top - upper_padding)
    y1 = min(image_height, source_bottom + lower_padding + 1)
    scan_limits = region.diagnostics.get("boundary_scan_limits")
    if isinstance(scan_limits, (list, tuple)) and len(scan_limits) == 2:
        y0 = max(y0, int(scan_limits[0]))
        y1 = min(y1, int(scan_limits[1]) + 1)

    horizontal = cv2.morphologyEx(
        ink,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(
            cv2.MORPH_RECT, (max(13, round(pitch * 0.28)), 1)
        ),
    )
    horizontal = cv2.dilate(
        horizontal,
        cv2.getStructuringElement(cv2.MORPH_RECT, (1, 5)),
    )
    coverage = np.zeros(image_height, dtype=np.int16)
    for column in range(EXPECTED_COLUMN_COUNT):
        cell_width = max(8.0, rules[column + 1] - rules[column])
        x0 = max(0, round(rules[column] + cell_width * 0.05))
        x1 = min(image_width, round(rules[column + 1] - cell_width * 0.05) + 1)
        if x1 <= x0:
            continue
        row_counts = np.count_nonzero(horizontal[y0:y1, x0:x1], axis=1)
        minimum_pixels = max(5, round((x1 - x0) * 0.18))
        coverage[y0:y1] += (row_counts >= minimum_pixels).astype(np.int16)

    active_rows = np.flatnonzero(
        (coverage >= 5)
        & (np.arange(image_height) >= y0)
        & (np.arange(image_height) < y1)
    )
    groups: list[list[int]] = []
    maximum_gap = max(2, round(source_span * 0.003))
    for row in active_rows.tolist():
        if not groups or row - groups[-1][-1] > maximum_gap:
            groups.append([row])
        else:
            groups[-1].append(row)
    bands: list[float] = []
    band_scores: list[int] = []
    for group in groups:
        rows = np.asarray(group, dtype=int)
        weights = coverage[rows].astype(float)
        bands.append(float(np.average(rows, weights=np.maximum(weights, 1.0))))
        band_scores.append(int(np.max(coverage[rows])))
    if len(bands) < 4:
        raise RuntimeError(
            f"横向局部线段只形成{len(bands)}个结构带，不能确定表头和统计区。"
        )

    band_values = np.asarray(bands, dtype=float)
    band_gaps = np.diff(band_values)
    minimum_daily_gap = max(source_span * 0.28, pitch * 3.8)
    eligible_gaps = np.flatnonzero(band_gaps >= minimum_daily_gap)
    if eligible_gaps.size == 0:
        raise RuntimeError("没有检测到足够长的31日数据区横向空带。")
    split_index = int(
        eligible_gaps[
            np.argmax(band_gaps[eligible_gaps])
        ]
    )
    header_nominal = float(band_values[split_index])
    statistics_nominal = float(band_values[split_index + 1])
    top_ranked: list[tuple[float, int]] = []
    for index in range(split_index):
        ratio = (header_nominal - band_values[index]) / max(1.0, pitch)
        if 0.18 <= ratio <= 1.15:
            score = (
                abs(ratio - 0.50)
                - 0.35 * band_scores[index] / EXPECTED_COLUMN_COUNT
            )
            top_ranked.append((float(score), index))
    if not top_ranked:
        raise RuntimeError("表头之前没有符合表头高度拓扑的真实顶框。")
    top_score, top_index = min(top_ranked, key=lambda item: item[0])
    top_nominal = float(band_values[top_index])

    bottom_ranked: list[tuple[float, int]] = []
    for index in range(split_index + 2, len(band_values)):
        ratio = (band_values[index] - statistics_nominal) / max(1.0, pitch)
        if 1.20 <= ratio <= 3.15:
            score = (
                abs(ratio - 2.45) * 0.32
                - 1.15 * band_scores[index] / EXPECTED_COLUMN_COUNT
                + 0.08
                * abs(band_values[index] - source_bottom)
                / max(1.0, source_span)
            )
            bottom_ranked.append((float(score), index))
    if not bottom_ranked:
        raise RuntimeError("统计区之后没有符合统计区高度拓扑的真实底框。")
    bottom_score, bottom_index = min(bottom_ranked, key=lambda item: item[0])
    bottom_nominal = float(band_values[bottom_index])
    if not (
        top_nominal + max(8.0, source_span * 0.018) < header_nominal
        and header_nominal + source_span * 0.25 < statistics_nominal
        and statistics_nominal + max(6.0, source_span * 0.025) < bottom_nominal
    ):
        raise RuntimeError(
            "四条结构曲线次序不成立，拒绝以固定比例替代真实表格边界。"
        )

    detected_span = max(1.0, bottom_nominal - top_nominal)
    search_radius = max(8, round(max(pitch * 0.18, detected_span * 0.018)))
    curves: dict[str, np.ndarray] = {}
    raw_curves: dict[str, np.ndarray] = {}
    for name, nominal in (
        ("table_top", top_nominal),
        ("header_bottom", header_nominal),
        ("statistics_top", statistics_nominal),
        ("table_bottom", bottom_nominal),
    ):
        raw, regularized = _track_horizontal_rule_samples(
            gray, horizontal, rules, nominal, search_radius
        )
        raw_curves[name] = raw
        curves[name] = regularized

    centers_x = (rules[:-1] + rules[1:]) / 2.0
    diagnostics: dict[str, Any] = {
        "horizontal_band_centers": [round(value, 3) for value in bands],
        "horizontal_band_support_columns": band_scores,
        "daily_band_gap": round(float(band_gaps[split_index]), 3),
        "minimum_daily_band_gap": round(float(minimum_daily_gap), 3),
        "boundary_band_indices": {
            "table_top": top_index,
            "header_bottom": split_index,
            "statistics_top": split_index + 1,
            "table_bottom": bottom_index,
        },
        "boundary_band_scores": {
            "table_top": round(top_score, 6),
            "table_bottom": round(bottom_score, 6),
        },
        "boundary_scan_y": [y0, y1 - 1],
        "horizontal_support_threshold": 5,
        "boundary_search_radius": search_radius,
        "source_column_center_x": [round(float(value), 3) for value in centers_x],
        "source_boundary_raw_y": {
            name: [round(float(value), 3) for value in values]
            for name, values in raw_curves.items()
        },
        "source_boundary_regularized_y": {
            name: [round(float(value), 3) for value in values]
            for name, values in curves.items()
        },
    }
    return centers_x, curves, horizontal, diagnostics


def _curve_local_maxima(
    values: np.ndarray, radius: int, threshold: float
) -> list[int]:
    maxima: list[int] = []
    for index in range(radius, len(values) - radius):
        window = values[index - radius : index + radius + 1]
        if values[index] < threshold or values[index] != np.max(window):
            continue
        if (
            maxima
            and index - maxima[-1] <= radius
            and values[index] <= values[maxima[-1]]
        ):
            continue
        maxima.append(index)
    return maxima


def _horizontal_curve_response(
    gray: np.ndarray,
    ink: np.ndarray,
    rules: np.ndarray,
    y0: int,
    y1: int,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    """Score short locally sloped rule fragments in every month cell."""

    pitch = float(np.median(np.diff(rules)))
    horizontal = cv2.morphologyEx(
        ink,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(
            cv2.MORPH_RECT, (max(11, round(pitch * 0.18)), 1)
        ),
    )
    height = max(0, y1 - y0)
    scores = np.zeros((EXPECTED_COLUMN_COUNT, height), dtype=np.float32)
    support = np.zeros_like(scores)
    line_span = np.zeros_like(scores)
    for column in range(EXPECTED_COLUMN_COUNT):
        cell_width = max(8.0, rules[column + 1] - rules[column])
        x0 = max(0, round(rules[column] + cell_width * 0.10))
        x1 = min(gray.shape[1], round(rules[column + 1] - cell_width * 0.10) + 1)
        if x1 <= x0 or height < 5:
            continue
        strip = horizontal[y0:y1, x0:x1]
        darkness = (
            255 - gray[y0:y1, x0:x1].astype(np.float32)
        ) / 255.0
        xs = np.arange(x1 - x0, dtype=np.int32)
        centered_x = xs - (len(xs) - 1) / 2.0
        for local_y in range(2, height - 2):
            best_support = 0.0
            best_darkness = 0.0
            best_span = 0.0
            best_objective = -1.0
            # A boundary may move four pixels across one photographed cell.
            for delta in range(-4, 5):
                ys = np.rint(
                    local_y
                    + delta * centered_x / max(1.0, len(xs) - 1.0)
                ).astype(np.int32)
                valid = (ys >= 0) & (ys < height)
                values = strip[ys[valid], xs[valid]] > 0
                dark_values = darkness[ys[valid], xs[valid]]
                current_support = (
                    float(np.mean(values)) if values.size else 0.0
                )
                current_darkness = (
                    float(np.mean(dark_values)) if dark_values.size else 0.0
                )
                indices = np.flatnonzero(values)
                current_span = (
                    float(indices[-1] - indices[0] + 1) / values.size
                    if indices.size >= 2
                    else 0.0
                )
                objective = current_support + 0.20 * current_span
                if objective > best_objective:
                    best_objective = objective
                    best_support = current_support
                    best_darkness = current_darkness
                    best_span = current_span
            support[column, local_y] = best_support
            line_span[column, local_y] = best_span
            scores[column, local_y] = (
                3.8 * best_support
                + 0.75 * best_span
                + 0.30 * best_darkness
            )
        scores[column] = cv2.dilate(
            scores[column][None, :], np.ones((1, 3), np.uint8)
        )[0]
    return scores, support, line_span, horizontal


def _horizontal_curve_candidates(
    scores: np.ndarray,
    support: np.ndarray,
    line_span: np.ndarray,
    y0: int,
) -> list[list[dict[str, float]]]:
    candidates: list[list[dict[str, float]]] = []
    for column in range(scores.shape[0]):
        threshold = max(0.78, float(np.quantile(scores[column], 0.82)))
        peaks = _curve_local_maxima(scores[column], 2, threshold)
        candidates.append(
            [
                {
                    "y": float(y0 + index),
                    "score": float(scores[column, index]),
                    "support": float(support[column, index]),
                    "span": float(line_span[column, index]),
                }
                for index in peaks
                if (
                    support[column, index] >= 0.08
                    or line_span[column, index] >= 0.45
                )
            ]
        )
    return candidates


def _horizontal_trajectory_from_seed(
    candidates: list[list[dict[str, float]]],
    seed_column: int,
    seed_index: int,
    pitch: float,
) -> dict[str, Any]:
    """Trace one smooth boundary in both directions through optional gaps."""

    maximum_step = max(5.0, pitch * 0.105)
    y_values: list[float | None] = [None] * len(candidates)
    chosen: list[dict[str, float] | None] = [None] * len(candidates)
    seed = candidates[seed_column][seed_index]
    y_values[seed_column] = seed["y"]
    chosen[seed_column] = seed

    def extend(
        indices: list[int], previous_y: float, previous_slope: float
    ) -> None:
        misses = 0
        for column in indices:
            predicted = previous_y + previous_slope
            available = [
                item
                for item in candidates[column]
                if abs(item["y"] - predicted)
                <= maximum_step * (1.0 + 0.45 * misses)
            ]
            if not available:
                misses += 1
                if misses > 2:
                    break
                previous_y = predicted
                continue
            item = max(
                available,
                key=lambda value: (
                    value["score"]
                    - 0.16 * abs(value["y"] - predicted)
                    - 0.08
                    * abs((value["y"] - previous_y) - previous_slope)
                ),
            )
            slope = item["y"] - previous_y
            y_values[column] = item["y"]
            chosen[column] = item
            previous_slope = 0.60 * previous_slope + 0.40 * slope
            previous_y = item["y"]
            misses = 0

    extend(
        list(range(seed_column + 1, len(candidates))),
        float(seed["y"]),
        0.0,
    )
    extend(
        list(range(seed_column - 1, -1, -1)),
        float(seed["y"]),
        0.0,
    )
    observed = [index for index, item in enumerate(chosen) if item is not None]
    if not observed:
        return {}
    observed_x = np.asarray(observed, dtype=float)
    observed_y = np.asarray(
        [float(y_values[index]) for index in observed], dtype=float
    )
    full_y = np.interp(
        np.arange(len(candidates), dtype=float), observed_x, observed_y
    )
    first_differences = np.diff(full_y)
    second_differences = np.diff(first_differences)
    support_columns = sum(
        1
        for item in chosen
        if item is not None and item["support"] >= 0.12
    )
    span_columns = sum(
        1
        for item in chosen
        if item is not None and item["span"] >= 0.55
    )
    mean_score = float(
        np.mean([item["score"] for item in chosen if item is not None])
    )
    objective = (
        len(observed)
        + 0.55 * support_columns
        + 0.40 * span_columns
        + 0.70 * mean_score
        - 0.08 * float(np.sum(np.abs(first_differences)))
        - 0.10 * float(np.sum(np.abs(second_differences)))
        - 0.42 * (len(candidates) - len(observed))
    )
    return {
        "y": full_y,
        "observed_columns": len(observed),
        "support_columns": support_columns,
        "span_columns": span_columns,
        "mean_score": mean_score,
        "objective": objective,
        "median_y": float(np.median(full_y)),
        "range_y": float(np.max(full_y) - np.min(full_y)),
    }


def _enumerate_horizontal_trajectories(
    candidates: list[list[dict[str, float]]], pitch: float
) -> list[dict[str, Any]]:
    paths: list[dict[str, Any]] = []
    for column, items in enumerate(candidates):
        for index in range(len(items)):
            path = _horizontal_trajectory_from_seed(
                candidates, column, index, pitch
            )
            if not path or int(path["observed_columns"]) < 7:
                continue
            if (
                int(path["support_columns"]) < 5
                and int(path["span_columns"]) < 7
            ):
                continue
            duplicate_index = next(
                (
                    existing_index
                    for existing_index, existing in enumerate(paths)
                    if float(
                        np.median(np.abs(existing["y"] - path["y"]))
                    )
                    < 4.0
                ),
                None,
            )
            if duplicate_index is None:
                paths.append(path)
            elif float(path["objective"]) > float(
                paths[duplicate_index]["objective"]
            ):
                paths[duplicate_index] = path
    return sorted(paths, key=lambda item: float(item["median_y"]))


def _outer_vertical_contact_support(
    ink: np.ndarray,
    rules: np.ndarray,
    boundary_y: np.ndarray,
    pitch: float,
    *,
    direction: str,
) -> tuple[float, float]:
    """Measure whether a horizontal boundary actually meets both outer rules.

    A horizontal stroke alone occupies only a few rows.  A true table corner
    also has locally vertical ink extending into the table.  Row-wise support
    inside a narrow tube remains valid when the photographed border is curved.
    """

    vertical = cv2.morphologyEx(
        ink,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(
            cv2.MORPH_RECT, (1, max(7, round(pitch * 0.14)))
        ),
    )
    height, width = ink.shape
    supports: list[float] = []
    for x_value, y_value in (
        (float(rules[0]), float(boundary_y[0])),
        (float(rules[-1]), float(boundary_y[-1])),
    ):
        # The nominal rule is a whole-track centroid; a curled outer border can
        # be farther away at the table corner.  The tube remains well below
        # half a column, so a neighbouring month separator cannot satisfy it.
        radius_x = max(5, round(pitch * 0.23))
        if direction == "below":
            ya = round(y_value - pitch * 0.08)
            yb = round(y_value + pitch * 0.62)
        elif direction == "above":
            ya = round(y_value - pitch * 0.62)
            yb = round(y_value + pitch * 0.08)
        else:
            raise ValueError(f"未知交点方向：{direction}")
        xa = max(0, round(x_value) - radius_x)
        xb = min(width, round(x_value) + radius_x + 1)
        ya = max(0, ya)
        yb = min(height, yb + 1)
        if xa >= xb or ya >= yb:
            supports.append(0.0)
            continue
        row_support = np.any(vertical[ya:yb, xa:xb] != 0, axis=1)
        supports.append(float(np.count_nonzero(row_support) / len(row_support)))
    return float(supports[0]), float(supports[1])


def _top_boundary_inner_rule_consensus(
    ink: np.ndarray,
    region: TableRegion,
    rules: np.ndarray,
    curves: dict[str, np.ndarray],
    pitch: float,
) -> tuple[bool, dict[str, Any]]:
    """Validate a table top after raw outer components were polluted/clipped.

    Internal month separators begin at the header-bottom rule.  Their median
    start is therefore an independent reference that a photographed page edge
    cannot shift.  It is used only together with two real outer intersections.
    """

    internal_tops: list[float] = []
    for index, component in enumerate(region.source_components[1:-1], start=1):
        observed = (
            not region.observed_rule_flags
            or index >= len(region.observed_rule_flags)
            or bool(region.observed_rule_flags[index])
        )
        if observed and "y" in component:
            core_top = component.get("core_top")
            internal_tops.append(
                float(component["y"] if core_top is None else core_top)
            )
    header_median = float(np.median(curves["header_bottom"]))
    top_median = float(np.median(curves["table_top"]))
    contact_left, contact_right = _outer_vertical_contact_support(
        ink,
        rules,
        curves["table_top"],
        pitch,
        direction="below",
    )
    internal_top = (
        float(np.median(internal_tops)) if internal_tops else float("nan")
    )
    header_error = (
        abs(header_median - internal_top) if internal_tops else float("inf")
    )
    top_error = (
        abs(top_median - internal_top) if internal_tops else float("inf")
    )
    top_to_internal = (
        internal_top - top_median if internal_tops else float("nan")
    )
    reference_name = (
        "header-bottom" if header_error <= top_error else "table-top"
    )
    reference_error = min(header_error, top_error)
    reference_topology_valid = bool(
        (reference_name == "header-bottom" and pitch * 0.18 <= top_to_internal <= pitch * 1.20)
        or (reference_name == "table-top" and top_error <= max(8.0, pitch * 0.32))
    )
    accepted = bool(
        len(internal_tops) >= 7
        and reference_error <= max(8.0, pitch * 0.32)
        and reference_topology_valid
        and min(contact_left, contact_right) >= 0.20
    )
    diagnostics = {
        "accepted": accepted,
        "internal_rule_count": len(internal_tops),
        "internal_start_median": (
            None if not internal_tops else round(internal_top, 3)
        ),
        "header_median": round(header_median, 3),
        "header_to_internal_error": (
            None if not internal_tops else round(header_error, 3)
        ),
        "top_to_internal_error": (
            None if not internal_tops else round(top_error, 3)
        ),
        "internal_start_reference": reference_name,
        "internal_start_reference_error": (
            None if not internal_tops else round(reference_error, 3)
        ),
        "top_to_internal_distance": (
            None if not internal_tops else round(top_to_internal, 3)
        ),
        "outer_contact_support": [
            round(contact_left, 4),
            round(contact_right, 4),
        ],
    }
    return accepted, diagnostics


def _choose_structural_boundary_trajectories(
    paths: list[dict[str, Any]],
    source_top: int,
    source_bottom: int,
    pitch: float,
    outer_bottom_targets: tuple[float | None, float | None] | None = None,
) -> tuple[dict[str, dict[str, Any]], dict[str, float]]:
    """Jointly choose four curves from the largest 31-day structural gap."""

    if len(paths) < 4:
        raise RuntimeError(f"跨列横线轨迹只有{len(paths)}条。")
    minimum_daily_gap = max(
        (source_bottom - source_top) * 0.28, pitch * 3.8
    )
    adjacent_gaps = [
        float(np.median(paths[index + 1]["y"] - paths[index]["y"]))
        for index in range(len(paths) - 1)
    ]
    eligible = [
        index
        for index, gap in enumerate(adjacent_gaps)
        if (
            gap >= minimum_daily_gap
            and int(paths[index]["observed_columns"]) >= 9
            and int(paths[index + 1]["observed_columns"]) >= 9
        )
    ]
    if not eligible:
        raise RuntimeError("没有形成由完整曲线围成的31日数据长空带。")
    header_index = max(eligible, key=lambda index: adjacent_gaps[index])
    statistics_index = header_index + 1
    header = paths[header_index]
    statistics = paths[statistics_index]

    top_ranked: list[tuple[float, int]] = []
    for index in range(header_index):
        ratio = float(np.median(header["y"] - paths[index]["y"])) / pitch
        compact_header = bool(
            max(0.08, 8.0 / max(1.0, pitch)) <= ratio < 0.18
            and abs(float(np.median(paths[index]["y"])) - source_top)
            <= pitch * 0.28
            and int(paths[index]["observed_columns"]) >= 7
            and int(paths[index]["support_columns"]) >= 7
            and int(paths[index]["span_columns"]) >= 5
        )
        if 0.18 <= ratio <= 1.15 or compact_header:
            score = (
                abs(ratio - 0.50)
                - 0.012 * float(paths[index]["objective"])
                + (0.18 if compact_header else 0.0)
            )
            top_ranked.append((score, index))
    if not top_ranked:
        raise RuntimeError("表头底曲线之前没有符合高度拓扑的顶框曲线。")
    top_index = min(top_ranked)[1]

    bottom_ranked: list[
        tuple[float, int, float, float, float, float]
    ] = []
    source_span = max(1, source_bottom - source_top)
    outer_tolerance = max(10.0, pitch * 0.35)
    for index in range(statistics_index + 1, len(paths)):
        ratio = float(
            np.median(paths[index]["y"] - statistics["y"])
        ) / pitch
        if 1.20 <= ratio <= 3.15:
            left_error = 0.0
            right_error = 0.0
            if outer_bottom_targets is not None:
                if outer_bottom_targets[0] is not None:
                    left_error = abs(
                        float(paths[index]["y"][0]) - outer_bottom_targets[0]
                    )
                if outer_bottom_targets[1] is not None:
                    right_error = abs(
                        float(paths[index]["y"][-1]) - outer_bottom_targets[1]
                    )
            contact_left, contact_right = paths[index].get(
                "bottom_outer_contact_support", (0.0, 0.0)
            )
            # Raw endpoints can both belong to the photographed page edge, so
            # disagreement carries no penalty.  They contribute only a small
            # bonus when *both* sides independently agree with this curve.
            endpoint_pairs = [
                (error, target)
                for error, target in zip(
                    (left_error, right_error),
                    outer_bottom_targets or (None, None),
                )
                if target is not None
            ]
            endpoint_penalty = (
                -1.00
                if len(endpoint_pairs) == 2
                and all(error <= outer_tolerance for error, _ in endpoint_pairs)
                else 0.0
            )
            contact_penalty = 0.22 * (
                max(0.0, 0.32 - float(contact_left))
                + max(0.0, 0.32 - float(contact_right))
            ) / 0.32
            score = (
                abs(ratio - 2.45) * 0.32
                - 0.15 * float(paths[index]["objective"])
                + 0.08
                * abs(float(np.median(paths[index]["y"])) - source_bottom)
                / source_span
                + endpoint_penalty
                + contact_penalty
            )
            bottom_ranked.append(
                (
                    score,
                    index,
                    left_error,
                    right_error,
                    float(contact_left),
                    float(contact_right),
                )
            )
    if not bottom_ranked:
        raise RuntimeError("统计区之后没有符合高度拓扑的表底曲线。")
    (
        _,
        bottom_index,
        bottom_left_error,
        bottom_right_error,
        bottom_contact_left,
        bottom_contact_right,
    ) = min(bottom_ranked)
    selected = {
        "table_top": paths[top_index],
        "header_bottom": header,
        "statistics_top": statistics,
        "table_bottom": paths[bottom_index],
    }
    selected_top_ratio = float(
        np.median(header["y"] - selected["table_top"]["y"])
    ) / max(1.0, pitch)
    if selected_top_ratio < 0.18:
        # Very compact headers often expose the top rule in only part of the
        # width; trajectory interpolation can then touch the header at an
        # unsupported column.  Preserve both detected curves while enforcing
        # their known non-crossing form topology at those missing samples.
        compact_top = dict(selected["table_top"])
        compact_top["y"] = np.minimum(
            np.asarray(compact_top["y"], dtype=float),
            np.asarray(header["y"], dtype=float) - max(9.0, pitch * 0.105),
        )
        selected["table_top"] = compact_top
    if not (
        np.all(selected["table_top"]["y"] + 8 < header["y"])
        and np.all(header["y"] + pitch * 3.2 < statistics["y"])
        and np.all(statistics["y"] + 6 < selected["table_bottom"]["y"])
    ):
        raise RuntimeError("联合选择的四条边界曲线交叉或次序不成立。")
    gaps = {
        "header_height": float(
            np.median(header["y"] - selected["table_top"]["y"])
        ),
        "daily_height": float(np.median(statistics["y"] - header["y"])),
        "statistics_height": float(
            np.median(selected["table_bottom"]["y"] - statistics["y"])
        ),
        "minimum_daily_gap": float(minimum_daily_gap),
        "bottom_left_outer_error": float(bottom_left_error),
        "bottom_right_outer_error": float(bottom_right_error),
        "bottom_outer_tolerance": float(outer_tolerance),
        "bottom_outer_contact_left": float(bottom_contact_left),
        "bottom_outer_contact_right": float(bottom_contact_right),
        "bottom_verified_outer_side_count": float(
            sum(
                target is not None
                for target in (outer_bottom_targets or (None, None))
            )
        ),
    }
    return selected, gaps


def _detect_source_horizontal_boundaries_curved(
    gray: np.ndarray, ink: np.ndarray, region: TableRegion
) -> tuple[np.ndarray, dict[str, np.ndarray], np.ndarray, dict[str, Any]]:
    """Jointly discover four boundaries from all cross-column trajectories."""

    image_height, _ = gray.shape
    rules = np.asarray(region.source_rules, dtype=float)
    if len(rules) != EXPECTED_VERTICAL_RULE_COUNT:
        raise RuntimeError("边界曲线需要14条竖轨。")
    centers_x = (rules[:-1] + rules[1:]) / 2.0
    pitch = float(np.median(np.diff(rules)))
    _, source_top, _, source_bottom = region.source_bbox
    source_span = max(1.0, float(source_bottom - source_top))
    if "boundary_scan_limits" not in region.diagnostics:
        assign_region_boundary_scan_limits([region], image_height)
    scan_limits = region.diagnostics["boundary_scan_limits"]
    y0 = max(0, int(scan_limits[0]))
    y1 = min(image_height - 1, int(scan_limits[1]))
    if y1 - y0 < max(80.0, source_span * 0.55):
        raise RuntimeError("本表边界扫描区间过窄。")

    outer_bottom_targets: tuple[float | None, float | None] | None = None
    if (
        len(region.source_components) == EXPECTED_VERTICAL_RULE_COUNT
        and len(region.observed_rule_flags) == EXPECTED_VERTICAL_RULE_COUNT
        and region.observed_rule_flags[0]
        and region.observed_rule_flags[-1]
    ):
        internal_heights = [
            float(
                item["height"]
                if item.get("core_height") is None
                else item["core_height"]
            )
            for item in region.source_components[1:-1]
            if "height" in item
        ]
        median_internal_height = (
            float(np.median(internal_heights))
            if len(internal_heights) >= 4
            else None
        )

        def reliable_bottom_target(component: dict[str, float]) -> float | None:
            if "bottom" not in component:
                return None
            component_height = component.get("core_height")
            component_bottom = component.get("core_bottom")
            if component_height is None:
                component_height = component.get("height")
            if component_bottom is None:
                component_bottom = component.get("bottom")
            if median_internal_height is None or component_height is None:
                return float(component_bottom)
            relative_height = float(component_height) / max(
                1.0, median_internal_height
            )
            if 0.92 <= relative_height <= 1.75:
                return float(component_bottom)
            return None

        outer_bottom_targets = (
            reliable_bottom_target(region.source_components[0]),
            reliable_bottom_target(region.source_components[-1]),
        )

    ownership_limits = region.diagnostics.get(
        "boundary_ownership_limits", [0, image_height - 1]
    )
    ownership_bottom = min(image_height - 1, int(ownership_limits[1]))
    expanded_bottom = min(
        ownership_bottom,
        max(
            y1,
            round(source_bottom + pitch * 2.20),
            round(source_top + source_span * 1.28),
        ),
    )
    scan_attempts = [(y0, y1, "source-bbox")]
    if expanded_bottom >= y1 + max(18, round(pitch * 0.45)):
        scan_attempts.append((y0, expanded_bottom, "adaptive-downward"))

    selected: dict[str, dict[str, Any]] | None = None
    gap_diagnostics: dict[str, float] = {}
    horizontal = np.zeros_like(ink)
    candidates: list[list[dict[str, float]]] = []
    paths: list[dict[str, Any]] = []
    scan_errors: list[str] = []
    selected_scan_method = "source-bbox"
    for attempt_y0, attempt_y1, attempt_method in scan_attempts:
        scores, support, line_span, attempt_horizontal = (
            _horizontal_curve_response(
                gray, ink, rules, attempt_y0, attempt_y1 + 1
            )
        )
        attempt_candidates = _horizontal_curve_candidates(
            scores, support, line_span, attempt_y0
        )
        attempt_paths = _enumerate_horizontal_trajectories(
            attempt_candidates, pitch
        )
        for path in attempt_paths:
            path["bottom_outer_contact_support"] = (
                _outer_vertical_contact_support(
                    ink,
                    rules,
                    np.asarray(path["y"], dtype=float),
                    pitch,
                    direction="above",
                )
            )
        try:
            attempt_selected, attempt_gaps = (
                _choose_structural_boundary_trajectories(
                    attempt_paths,
                    source_top,
                    source_bottom,
                    pitch,
                    outer_bottom_targets,
                )
            )
        except RuntimeError as error:
            scan_errors.append(f"{attempt_method}={error}")
            continue
        selected = attempt_selected
        gap_diagnostics = attempt_gaps
        horizontal = attempt_horizontal
        candidates = attempt_candidates
        paths = attempt_paths
        y0, y1 = attempt_y0, attempt_y1
        selected_scan_method = attempt_method
        break
    if selected is None:
        raise RuntimeError("；".join(scan_errors) or "没有找到四条结构边界曲线。")
    has_outer_track_evidence = bool(
        len(region.source_components) == EXPECTED_VERTICAL_RULE_COUNT
        and all(
            "height" in region.source_components[index]
            for index in (0, EXPECTED_VERTICAL_RULE_COUNT - 1)
        )
    )
    if has_outer_track_evidence:
        bottom_contacts = (
            gap_diagnostics.get("bottom_outer_contact_left", 0.0),
            gap_diagnostics.get("bottom_outer_contact_right", 0.0),
        )
        bottom_path = selected["table_bottom"]
        internal_consensus = bool(
            int(bottom_path.get("observed_columns", 0)) >= 11
            and int(bottom_path.get("support_columns", 0)) >= 10
            and int(bottom_path.get("span_columns", 0)) >= 10
        )
        one_sided_closure = bool(
            max(bottom_contacts) >= 0.30 and internal_consensus
        )
        if min(bottom_contacts) < 0.20 and not one_sided_closure:
            raise RuntimeError(
                "表底曲线既未闭合双侧外竖轨，也没有单侧闭合加内部曲线共识。"
            )
    raw_curves = {
        name: np.asarray(path["y"], dtype=float)
        for name, path in selected.items()
    }
    curves = {
        name: _regularize_horizontal_samples(
            centers_x,
            values,
            max(8, round(pitch * 0.25)),
            pitch,
        )
        for name, values in raw_curves.items()
    }

    source_top_tolerance = max(5.0, pitch * 0.22)
    top_median = float(np.median(curves["table_top"]))
    raw_top_aligned = not (
        float(np.min(curves["table_top"]))
        < source_top - source_top_tolerance
        or top_median > source_top + pitch * 0.65
    )
    inner_consensus, top_consensus_diagnostics = (
        _top_boundary_inner_rule_consensus(
            ink,
            region,
            rules,
            curves,
            pitch,
        )
    )
    if not raw_top_aligned and not inner_consensus:
        raise RuntimeError("顶框偏离source_bbox外竖框起点，拒绝伪边界。")
    if not (
        np.all(curves["header_bottom"] > curves["table_top"] + 6.0)
        and np.all(
            curves["statistics_top"]
            > curves["header_bottom"] + pitch * 3.2
        )
        and np.all(curves["table_bottom"] > curves["statistics_top"] + 6.0)
    ):
        raise RuntimeError("四条边界曲线逐列次序或间距不成立。")

    trajectory_diagnostics = [
        {
            "median_y": round(float(path["median_y"]), 3),
            "observed_columns": int(path["observed_columns"]),
            "support_columns": int(path["support_columns"]),
            "span_columns": int(path["span_columns"]),
            "objective": round(float(path["objective"]), 4),
            "range_y": round(float(path["range_y"]), 3),
        }
        for path in paths
    ]
    selected_diagnostics = {
        name: {
            "observed_columns": int(path["observed_columns"]),
            "support_columns": int(path["support_columns"]),
            "span_columns": int(path["span_columns"]),
            "objective": round(float(path["objective"]), 4),
            "range_y": round(float(path["range_y"]), 3),
        }
        for name, path in selected.items()
    }
    diagnostics: dict[str, Any] = {
        "boundary_seed_method": "column-local-trajectory-gap",
        "boundary_scan_method": selected_scan_method,
        "boundary_scan_limits": [y0, y1],
        "boundary_scan_y": [y0, y1],
        "source_column_center_x": [
            round(float(value), 3) for value in centers_x
        ],
        "minimum_daily_band_gap": round(
            float(gap_diagnostics["minimum_daily_gap"]), 3
        ),
        "local_candidate_count_by_column": [
            len(items) for items in candidates
        ],
        "horizontal_trajectory_count": len(paths),
        "horizontal_trajectories": trajectory_diagnostics,
        "boundary_curve_evidence": selected_diagnostics,
        "top_boundary_reference": {
            "raw_source_bbox_aligned": raw_top_aligned,
            "selection_method": (
                "raw-outer-start"
                if raw_top_aligned
                else "inner-rule-start-and-outer-intersections"
            ),
            **top_consensus_diagnostics,
        },
        "boundary_gap_diagnostics": {
            key: round(float(value), 3)
            for key, value in gap_diagnostics.items()
        },
        "source_boundary_raw_y": {
            name: [round(float(value), 3) for value in values]
            for name, values in raw_curves.items()
        },
        "source_boundary_regularized_y": {
            name: [round(float(value), 3) for value in values]
            for name, values in curves.items()
        },
    }
    return centers_x, curves, horizontal, diagnostics


def _detect_source_horizontal_boundaries(
    gray: np.ndarray, ink: np.ndarray, region: TableRegion
) -> tuple[np.ndarray, dict[str, np.ndarray], np.ndarray, dict[str, Any]]:
    """Find four true source curves with a strictly guarded legacy fallback."""

    try:
        return _detect_source_horizontal_boundaries_curved(gray, ink, region)
    except RuntimeError as curved_error:
        try:
            centers_x, curves, horizontal, diagnostics = (
                _detect_source_horizontal_boundaries_legacy(gray, ink, region)
            )
        except RuntimeError as legacy_error:
            raise RuntimeError(
                f"逐列曲线边界失败：{curved_error}；"
                f"全局边界回退也失败：{legacy_error}"
            ) from curved_error
        rules = np.asarray(region.source_rules, dtype=float)
        pitch = float(np.median(np.diff(rules)))
        top_curve = curves["table_top"]
        header_curve = curves["header_bottom"]
        statistics_curve = curves["statistics_top"]
        bottom_curve = curves["table_bottom"]
        _, source_top, _, _ = region.source_bbox
        if (
            float(np.min(top_curve)) < source_top - max(5.0, pitch * 0.22)
            or float(np.median(top_curve)) > source_top + pitch * 0.65
            or not np.all(header_curve > top_curve + pitch * 0.10)
            or not np.all(statistics_curve > header_curve + pitch * 3.0)
            or not np.all(bottom_curve > statistics_curve + pitch * 0.24)
        ):
            raise RuntimeError(
                f"逐列曲线边界失败：{curved_error}；"
                "全局回退顶框偏离source_bbox或四曲线拓扑不成立。"
            ) from curved_error
        diagnostics["boundary_seed_method"] = "guarded-global-y-fallback"
        diagnostics["curved_seed_error"] = str(curved_error)
        diagnostics["boundary_scan_limits"] = region.diagnostics.get(
            "boundary_scan_limits", diagnostics.get("boundary_scan_y", [])
        )
        return centers_x, curves, horizontal, diagnostics


def _robust_vertical_polynomial(
    y_values: np.ndarray,
    x_values: np.ndarray,
    weights: np.ndarray,
    residual_floor: float,
) -> tuple[np.ndarray, np.ndarray, float]:
    keep = np.ones(len(y_values), dtype=bool)
    coefficients = np.polyfit(y_values, x_values, 2, w=weights)
    for _ in range(5):
        residual = x_values - np.polyval(coefficients, y_values)
        centered = residual[keep] - float(np.median(residual[keep]))
        mad = float(np.median(np.abs(centered))) if centered.size else 0.0
        threshold = max(residual_floor, mad * 3.2)
        revised = np.abs(residual) <= threshold
        if int(np.count_nonzero(revised)) < 6 or np.array_equal(revised, keep):
            break
        keep = revised
        coefficients = np.polyfit(
            y_values[keep], x_values[keep], 2, w=weights[keep]
        )
    retained_residual = x_values[keep] - np.polyval(coefficients, y_values[keep])
    rmse = float(np.sqrt(np.mean(retained_residual**2))) if retained_residual.size else 0.0
    return coefficients.astype(float), keep, rmse


def _fit_source_vertical_rule_models(
    gray: np.ndarray,
    ink: np.ndarray,
    region: TableRegion,
    source_top: float,
    source_bottom: float,
) -> tuple[list[np.ndarray], list[dict[str, Any]]]:
    """Trace fourteen ordered x(y) curves from short local vertical pieces."""

    image_height, image_width = gray.shape
    rules = np.asarray(region.source_rules, dtype=float)
    pitch = float(np.median(np.diff(rules)))
    table_height = max(1.0, source_bottom - source_top)
    vertical = cv2.morphologyEx(
        ink,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(
            cv2.MORPH_RECT, (1, max(9, round(table_height * 0.014)))
        ),
    )
    sample_y = np.linspace(source_top, source_bottom, 72)
    search_radius = max(8, round(pitch * 0.27))
    half_band = max(3, round(table_height * 0.0055))
    models: list[np.ndarray | None] = []
    diagnostics: list[dict[str, Any]] = []
    for rule_index, nominal_x in enumerate(rules):
        observed_y: list[float] = []
        observed_x: list[float] = []
        observed_weight: list[float] = []
        for y_value in sample_y:
            y = round(float(y_value))
            ya = max(0, y - half_band)
            yb = min(image_height, y + half_band + 1)
            xa = max(0, round(nominal_x) - search_radius)
            xb = min(image_width, round(nominal_x) + search_radius + 1)
            if ya >= yb or xa >= xb:
                continue
            line_support = np.count_nonzero(vertical[ya:yb, xa:xb], axis=0).astype(float)
            line_support /= max(1, yb - ya)
            darkness = np.mean(
                255 - gray[ya:yb, xa:xb].astype(np.float32), axis=0
            ) / 255.0
            candidate_x = np.arange(xa, xb, dtype=float)
            distance = np.abs(candidate_x - nominal_x) / max(1.0, search_radius)
            score = line_support * 2.6 + darkness * 0.30 - distance * 0.16
            selected = int(np.argmax(score))
            if line_support[selected] < 0.22:
                continue
            observed_y.append(float(y_value))
            observed_x.append(float(xa + selected))
            observed_weight.append(float(0.5 + line_support[selected]))
        model: np.ndarray | None = None
        retained_count = 0
        rmse: float | None = None
        if len(observed_y) >= 8:
            model, keep, fitted_rmse = _robust_vertical_polynomial(
                np.asarray(observed_y, dtype=float),
                np.asarray(observed_x, dtype=float),
                np.asarray(observed_weight, dtype=float),
                max(2.0, pitch * 0.055),
            )
            retained_count = int(np.count_nonzero(keep))
            rmse = fitted_rmse
            probe = np.polyval(model, sample_y)
            if (
                retained_count < 6
                or float(np.max(np.abs(probe - nominal_x))) > search_radius * 1.35
            ):
                model = None
        models.append(model)
        diagnostics.append(
            {
                "rule_index": rule_index,
                "nominal_x": round(float(nominal_x), 3),
                "candidate_observation_count": len(observed_y),
                "retained_observation_count": retained_count,
                "rmse": None if rmse is None else round(float(rmse), 3),
                "inferred": model is None,
            }
        )

    good_indices = [index for index, model in enumerate(models) if model is not None]
    if len(good_indices) < 8:
        raise RuntimeError(
            f"14条竖向曲轨中只有{len(good_indices)}条有可靠局部线段证据。"
        )
    correction_models = []
    for index in good_indices:
        base = np.asarray([0.0, 0.0, rules[index]], dtype=float)
        correction_models.append(np.asarray(models[index], dtype=float) - base)
    shared_correction = np.median(np.stack(correction_models), axis=0)
    for index, model in enumerate(models):
        if model is None:
            models[index] = shared_correction + np.asarray(
                [0.0, 0.0, rules[index]], dtype=float
            )

    final_models = [np.asarray(model, dtype=float) for model in models]
    # Reject/correct a track that would cross its neighbours.  Its replacement
    # inherits only the common page bend, never a nearby digit stroke.
    probe_y = np.linspace(source_top, source_bottom, 32)
    for index in range(EXPECTED_VERTICAL_RULE_COUNT):
        values = np.polyval(final_models[index], probe_y)
        invalid = False
        if index > 0:
            previous = np.polyval(final_models[index - 1], probe_y)
            invalid = invalid or bool(np.any(values - previous < pitch * 0.42))
        if index + 1 < EXPECTED_VERTICAL_RULE_COUNT:
            following = np.polyval(final_models[index + 1], probe_y)
            invalid = invalid or bool(np.any(following - values < pitch * 0.42))
        if invalid:
            final_models[index] = shared_correction + np.asarray(
                [0.0, 0.0, rules[index]], dtype=float
            )
            diagnostics[index]["inferred"] = True
            diagnostics[index]["rejection_reason"] = "non-crossing topology"
    for index, model in enumerate(final_models):
        diagnostics[index]["coefficients"] = [
            round(float(value), 10) for value in model
        ]
    return final_models, diagnostics


def _ordered_rule_positions(
    values: np.ndarray, pitch: float, image_width: int
) -> np.ndarray:
    ordered = np.clip(values.astype(float), 0.0, image_width - 1.0)
    minimum_gap = max(4.0, pitch * 0.42)
    for _ in range(2):
        for index in range(1, len(ordered)):
            ordered[index] = max(ordered[index], ordered[index - 1] + minimum_gap)
        for index in range(len(ordered) - 2, -1, -1):
            ordered[index] = min(ordered[index], ordered[index + 1] - minimum_gap)
        ordered = np.clip(ordered, 0.0, image_width - 1.0)
    return ordered


def rectify_table_region(image: np.ndarray, region: TableRegion) -> np.ndarray:
    """Flatten a photographed surface with a local, non-rigid curve mesh."""

    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    ink = adaptive_ink(gray)
    rules = np.asarray(region.source_rules, dtype=float)
    pitch = float(np.median(np.diff(rules)))
    centers_x, sampled_curves, _, boundary_diagnostics = (
        _detect_source_horizontal_boundaries(gray, ink, region)
    )
    median_boundaries = {
        name: float(np.median(values)) for name, values in sampled_curves.items()
    }
    vertical_models, vertical_diagnostics = _fit_source_vertical_rule_models(
        gray,
        ink,
        region,
        median_boundaries["table_top"],
        median_boundaries["table_bottom"],
    )

    source_boundary_rules: dict[str, np.ndarray] = {}
    for name, values in sampled_curves.items():
        intersection_y = _clamped_curve_values(centers_x, values, rules)
        # Two fixed-point refinements approximate each intersection between the
        # horizontal curve and its corresponding x(y) rule track.
        for _ in range(2):
            intersection_x = np.asarray(
                [
                    np.polyval(vertical_models[index], intersection_y[index])
                    for index in range(EXPECTED_VERTICAL_RULE_COUNT)
                ],
                dtype=float,
            )
            intersection_y = _clamped_curve_values(
                centers_x, values, intersection_x
            )
        source_boundary_rules[name] = intersection_y

    minimum_header_height = max(8.0, pitch * 0.10)
    minimum_statistics_height = max(6.0, pitch * 0.08)
    top_rules = source_boundary_rules["table_top"]
    header_rules = np.maximum(
        source_boundary_rules["header_bottom"], top_rules + minimum_header_height
    )
    statistics_rules = np.maximum(
        source_boundary_rules["statistics_top"],
        header_rules + max(80.0, (median_boundaries["statistics_top"] - median_boundaries["header_bottom"]) * 0.72),
    )
    bottom_rules = np.maximum(
        source_boundary_rules["table_bottom"],
        statistics_rules + minimum_statistics_height,
    )
    source_boundary_rules = {
        "table_top": top_rules,
        "header_bottom": header_rules,
        "statistics_top": statistics_rules,
        "table_bottom": bottom_rules,
    }

    left, _, right, _ = region.source_bbox
    output_width = max(200, right - left + 1)
    output_rules = np.asarray(
        [round(float(value - left)) for value in rules], dtype=int
    )
    output_rules[0] = 0
    output_rules[-1] = output_width - 1
    header_height = max(12, round(float(np.median(header_rules - top_rules))))
    daily_height = max(180, round(float(np.median(statistics_rules - header_rules))))
    statistics_height = max(
        12, round(float(np.median(bottom_rules - statistics_rules)))
    )
    output_boundaries = {
        "table_top": 0,
        "header_bottom": header_height,
        "statistics_top": header_height + daily_height,
        "table_bottom": header_height + daily_height + statistics_height,
    }
    output_height = output_boundaries["table_bottom"] + 1
    output_x = np.arange(output_width, dtype=float)
    map_x = np.empty((output_height, output_width), dtype=np.float32)
    map_y = np.empty((output_height, output_width), dtype=np.float32)
    bands = (
        ("table_top", "header_bottom"),
        ("header_bottom", "statistics_top"),
        ("statistics_top", "table_bottom"),
    )
    for upper_name, lower_name in bands:
        output_upper = output_boundaries[upper_name]
        output_lower = output_boundaries[lower_name]
        denominator = max(1, output_lower - output_upper)
        for output_y in range(output_upper, output_lower + 1):
            fraction = (output_y - output_upper) / denominator
            source_y_rules = (
                source_boundary_rules[upper_name] * (1.0 - fraction)
                + source_boundary_rules[lower_name] * fraction
            )
            source_x_rules = np.asarray(
                [
                    np.polyval(vertical_models[index], source_y_rules[index])
                    for index in range(EXPECTED_VERTICAL_RULE_COUNT)
                ],
                dtype=float,
            )
            source_x_rules = _ordered_rule_positions(
                source_x_rules, pitch, image.shape[1]
            )
            map_x[output_y] = np.interp(
                output_x, output_rules, source_x_rules
            ).astype(np.float32)
            map_y[output_y] = np.interp(
                output_x, output_rules, source_y_rules
            ).astype(np.float32)

    # A visually plausible control mesh can still fold between control lines.
    # Check the dense mapping orientation before any OCR/cell extraction uses it.
    sampling_step = max(1, min(output_height, output_width) // 420)
    qa_map_x = map_x[::sampling_step, ::sampling_step].astype(np.float64)
    qa_map_y = map_y[::sampling_step, ::sampling_step].astype(np.float64)
    map_x_dy, map_x_dx = np.gradient(qa_map_x)
    map_y_dy, map_y_dx = np.gradient(qa_map_y)
    jacobian = map_x_dx * map_y_dy - map_x_dy * map_y_dx
    finite_jacobian = jacobian[np.isfinite(jacobian)]
    if finite_jacobian.size == 0:
        raise RuntimeError("非刚性映射没有有限的雅可比值。")
    median_jacobian = float(np.median(finite_jacobian))
    minimum_jacobian = float(np.min(finite_jacobian))
    first_percentile_jacobian = float(np.percentile(finite_jacobian, 1))
    if (
        median_jacobian <= 0
        or minimum_jacobian <= 0
        or first_percentile_jacobian < median_jacobian * 0.05
    ):
        raise RuntimeError(
            "非刚性曲线网格发生局部翻折，拒绝把错误映射传给OCR。"
        )

    rectified = cv2.remap(
        image,
        map_x,
        map_y,
        cv2.INTER_CUBIC,
        borderMode=cv2.BORDER_REPLICATE,
    )
    rectified_gray = cv2.cvtColor(rectified, cv2.COLOR_BGR2GRAY)
    background = cv2.GaussianBlur(rectified_gray, (0, 0), 31)
    normalized = cv2.divide(
        rectified_gray, np.maximum(background, 1), scale=245
    )

    region.source_bbox = (
        left,
        round(median_boundaries["table_top"]),
        right,
        round(median_boundaries["table_bottom"]),
    )
    region.rectified_rules = output_rules.astype(int).tolist()
    region.rectified_boundaries = output_boundaries
    region.diagnostics.update(boundary_diagnostics)
    region.diagnostics.update(
        {
            "source_boundary_y_at_rules": {
                name: [round(float(value), 3) for value in values]
                for name, values in source_boundary_rules.items()
            },
            "source_vertical_rule_tracks": vertical_diagnostics,
            "rectified_rules": region.rectified_rules,
            "rectified_boundaries": region.rectified_boundaries,
            "rectification_model": "piecewise-curve-mesh",
            "mapping_jacobian": {
                "sampling_step": sampling_step,
                "minimum": round(minimum_jacobian, 6),
                "first_percentile": round(first_percentile_jacobian, 6),
                "median": round(median_jacobian, 6),
            },
        }
    )
    return cv2.cvtColor(normalized, cv2.COLOR_GRAY2BGR)


def select_regular_vertical_rules(
    components: list[dict[str, float]], image_width: int
) -> list[int]:
    tolerance = vertical_rule_cluster_tolerance(image_width)
    x_positions = clustered_positions(
        [item["x"] for item in components], tolerance=tolerance
    )
    if len(x_positions) < EXPECTED_VERTICAL_RULE_COUNT:
        raise RuntimeError(
            f"仅检测到{len(x_positions)}条长竖线，"
            f"日流量表需要{EXPECTED_VERTICAL_RULE_COUNT}条。"
        )

    candidates: list[tuple[float, list[float]]] = []
    for start in range(
        len(x_positions) - EXPECTED_VERTICAL_RULE_COUNT + 1
    ):
        rules = x_positions[
            start : start + EXPECTED_VERTICAL_RULE_COUNT
        ]
        gaps = np.diff(np.asarray(rules, dtype=float))
        mean_gap = float(np.mean(gaps))
        if mean_gap <= image_width * 0.025:
            continue
        span = rules[-1] - rules[0]
        if span < image_width * 0.55:
            continue
        edge_penalty = abs(span / image_width - 0.79) * 0.12
        topology = vertical_rule_candidate_metrics(
            components,
            rules,
            max(tolerance * 1.6, mean_gap * 0.14),
        )
        if not (
            passes_outer_border_topology(topology)
            or passes_curved_outer_border_topology(topology)
        ):
            continue
        topology_penalty = (
            (1.0 - topology["vertical_interval_iou"]) * 0.10
            + (1.0 - topology["height_balance"]) * 0.06
            + max(0.0, 1.15 - topology["minimum_relative_height"])
            * 0.04
        )
        candidates.append(
            (
                topology["month_gap_trend_residual"]
                + topology["month_gap_cv"] * 0.20
                + topology_penalty
                + edge_penalty,
                rules,
            )
        )
    if not candidates:
        raise RuntimeError("没有找到由14条近似等距竖线组成的日流量表。")
    candidates.sort(key=lambda item: item[0])
    chosen = candidates[0][1]
    gaps = np.diff(np.asarray(chosen, dtype=float))
    month_gaps = gaps[1:]
    if float(np.max(month_gaps) / max(np.min(month_gaps), 1.0)) > 1.35:
        raise RuntimeError("日流量表月列间距不稳定，不能安全切分12个月。")
    return [round(value) for value in chosen]


def detect_rule_geometry(image: np.ndarray) -> tuple[list[int], list[int]]:
    height, width = image.shape[:2]
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    ink = cv2.threshold(gray, 190, 255, cv2.THRESH_BINARY_INV)[1]

    vertical = cv2.morphologyEx(
        ink,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(
            cv2.MORPH_RECT, (1, max(70, round(height * 0.043)))
        ),
    )
    count, _, stats, _ = cv2.connectedComponentsWithStats(vertical, 8)
    components: list[dict[str, float]] = []
    for index in range(1, count):
        x, y, component_width, component_height, area = stats[index]
        if (
            component_height >= height * 0.23
            and component_width <= max(14, width * 0.012)
            and area >= component_height * 0.55
        ):
            components.append(
                {
                    "x": float(x + component_width / 2),
                    "y": float(y),
                    "height": float(component_height),
                }
            )
    vertical_rules = select_regular_vertical_rules(components, width)

    table_span = vertical_rules[-1] - vertical_rules[0]
    horizontal = cv2.morphologyEx(
        ink,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(
            cv2.MORPH_RECT, (max(140, round(table_span * 0.18)), 1)
        ),
    )
    count, _, stats, _ = cv2.connectedComponentsWithStats(horizontal, 8)
    horizontal_y: list[float] = []
    for index in range(1, count):
        x, y, component_width, component_height, _ = stats[index]
        if (
            component_width >= table_span * 0.78
            and component_height <= max(12, height * 0.008)
            and x <= vertical_rules[0] + table_span * 0.08
            and x + component_width
            >= vertical_rules[-1] - table_span * 0.08
        ):
            horizontal_y.append(float(y + component_height / 2))
    horizontal_rules = [
        round(value)
        for value in clustered_positions(
            horizontal_y, tolerance=max(3.0, height * 0.002)
        )
    ]
    if len(horizontal_rules) < 3:
        raise RuntimeError("没有检测到表头、正文和统计区所需的长横线。")
    return vertical_rules, horizontal_rules


def find_table_vertical_bounds(
    image: np.ndarray, vertical_rules: list[int]
) -> tuple[int, int]:
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    ink = cv2.threshold(gray, 190, 255, cv2.THRESH_BINARY_INV)[1]
    left = max(0, vertical_rules[0] - 3)
    right = min(image.shape[1], vertical_rules[-1] + 4)
    region = ink[:, left:right]
    vertical = cv2.morphologyEx(
        region,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(
            cv2.MORPH_RECT, (1, max(70, round(image.shape[0] * 0.043)))
        ),
    )
    ys, _ = np.nonzero(vertical)
    if ys.size == 0:
        raise RuntimeError("无法确定日流量表竖线的上下范围。")
    return int(np.min(ys)), int(np.max(ys))


def active_bands(active: np.ndarray, offset: int) -> list[tuple[int, int]]:
    bands: list[tuple[int, int]] = []
    start: int | None = None
    for index, is_active in enumerate(active.tolist()):
        if is_active and start is None:
            start = index
        elif not is_active and start is not None:
            if index - start >= 2:
                bands.append((offset + start, offset + index - 1))
            start = None
    if start is not None and len(active) - start >= 2:
        bands.append((offset + start, offset + len(active) - 1))
    return bands


def column_support(
    body_ink: np.ndarray,
    vertical_rules: list[int],
    body_left: int,
    body_top: int,
    top: int,
    bottom: int,
) -> tuple[int, bool]:
    local_top = max(0, top - body_top)
    local_bottom = min(body_ink.shape[0], bottom - body_top + 1)
    support = 0
    day_support = False
    for column in range(EXPECTED_COLUMN_COUNT):
        left = max(0, vertical_rules[column] - body_left + 3)
        right = min(
            body_ink.shape[1],
            vertical_rules[column + 1] - body_left - 3,
        )
        cell = body_ink[local_top:local_bottom, left:right]
        area = int(np.count_nonzero(cell))
        required = max(4, round(cell.size * 0.006))
        if area >= required:
            support += 1
            if column == 0:
                day_support = True
    return support, day_support


def build_anchor_candidates(
    image: np.ndarray,
    vertical_rules: list[int],
    body_top: int,
    body_bottom: int,
) -> list[AnchorCandidate]:
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    fixed_ink = cv2.threshold(gray, 190, 255, cv2.THRESH_BINARY_INV)[1]
    otsu_ink = cv2.threshold(
        gray, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU
    )[1]
    body_left = vertical_rules[0]
    body_right = vertical_rules[-1]
    choices: list[tuple[float, np.ndarray, np.ndarray, list[tuple[int, int]]]] = []
    threshold_fractions = (
        0.009,
        0.011,
        0.013,
        0.015,
        0.020,
        0.025,
        0.030,
        0.035,
        0.040,
        0.045,
        0.050,
        0.060,
        0.070,
        0.080,
    )
    for page_ink in (fixed_ink, otsu_ink):
        body = page_ink[body_top:body_bottom, body_left:body_right].copy()
        vertical_mask = cv2.morphologyEx(
            body,
            cv2.MORPH_OPEN,
            cv2.getStructuringElement(
                cv2.MORPH_RECT,
                (1, max(40, round((body_bottom - body_top) * 0.13))),
            ),
        )
        body[vertical_mask > 0] = 0
        projection = np.count_nonzero(body, axis=1)
        for fraction in threshold_fractions:
            threshold = max(10, round(body.shape[1] * fraction))
            bands = active_bands(projection >= threshold, body_top)
            bands = [
                band
                for band in bands
                if 2 <= band[1] - band[0] + 1 <= image.shape[0] * 0.030
            ]
            choices.append((fraction, body, projection, bands))
    _, body, projection, chosen_bands = min(
        choices,
        key=lambda item: (
            abs(len(item[3]) - EXPECTED_DAY_COUNT),
            0 if len(item[3]) >= EXPECTED_DAY_COUNT else 1,
            abs(item[0] - 0.025),
        ),
    )

    candidates: list[AnchorCandidate] = []
    for top, bottom in chosen_bands:
        local_top = top - body_top
        local_bottom = bottom - body_top + 1
        weights = projection[local_top:local_bottom].astype(float)
        y_values = np.arange(top, bottom + 1, dtype=float)
        ink_centroid = float(
            np.average(y_values, weights=weights)
            if float(np.sum(weights)) > 0
            else (top + bottom) / 2
        )
        support, day_support = column_support(
            body,
            vertical_rules,
            body_left,
            body_top,
            top,
            bottom,
        )
        score = support + (3.0 if day_support else 0.0)
        candidates.append(
            AnchorCandidate(
                top=top,
                bottom=bottom,
                center=(top + bottom) / 2,
                ink_centroid=ink_centroid,
                support_columns=support,
                day_column_support=day_support,
                score=score,
            )
        )
    return candidates


def build_day_column_candidates(
    image: np.ndarray,
    vertical_rules: list[int],
    body_top: int,
    body_bottom: int,
) -> list[AnchorCandidate]:
    """Use the printed day numbers as a strong row-order anchor after dewarping."""

    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    page_ink = cv2.threshold(
        gray, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU
    )[1]
    body_left = vertical_rules[0]
    body_right = vertical_rules[-1]
    body = page_ink[body_top:body_bottom, body_left:body_right].copy()
    vertical_mask = cv2.morphologyEx(
        body,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(
            cv2.MORPH_RECT, (1, max(40, round(body.shape[0] * 0.13)))
        ),
    )
    body[vertical_mask > 0] = 0
    day_left = max(0, vertical_rules[0] - body_left + 3)
    day_right = min(
        body.shape[1], vertical_rules[1] - body_left - 3
    )
    projection = np.count_nonzero(body[:, day_left:day_right], axis=1)
    choices: list[tuple[int, int, int, list[tuple[int, int]]]] = []
    for threshold in range(1, 9):
        bands = active_bands(projection >= threshold, body_top)
        bands = [
            band
            for band in bands
            if 2 <= band[1] - band[0] + 1 <= image.shape[0] * 0.030
        ]
        choices.append(
            (
                abs(len(bands) - EXPECTED_DAY_COUNT),
                0 if len(bands) >= EXPECTED_DAY_COUNT else 1,
                threshold,
                bands,
            )
        )
    _, _, _, chosen_bands = min(choices)
    candidates: list[AnchorCandidate] = []
    for top, bottom in chosen_bands:
        local_top = top - body_top
        local_bottom = bottom - body_top + 1
        weights = projection[local_top:local_bottom].astype(float)
        y_values = np.arange(top, bottom + 1, dtype=float)
        ink_centroid = float(
            np.average(y_values, weights=weights)
            if float(np.sum(weights)) > 0
            else (top + bottom) / 2
        )
        support, day_support = column_support(
            body,
            vertical_rules,
            body_left,
            body_top,
            top,
            bottom,
        )
        candidates.append(
            AnchorCandidate(
                top=top,
                bottom=bottom,
                center=(top + bottom) / 2,
                ink_centroid=ink_centroid,
                support_columns=support,
                day_column_support=day_support,
                score=support + (3.0 if day_support else 0.0),
            )
        )
    return candidates


def build_multicolumn_support_candidates(
    image: np.ndarray,
    vertical_rules: list[int],
    body_top: int,
    body_bottom: int,
) -> list[AnchorCandidate]:
    """Recover row anchors from ink that agrees across many month columns.

    Photograph noise can join or split digits in the narrow day column.  A true
    data row, however, normally contains ink at nearly the same y position in
    most of the 13 columns.  This fallback counts that cross-column agreement
    and therefore does not assume that a row boundary is perfectly horizontal.
    """

    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    page_ink = cv2.threshold(
        gray, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU
    )[1]
    body_left = vertical_rules[0]
    body_right = vertical_rules[-1]
    body = page_ink[body_top:body_bottom, body_left:body_right].copy()
    vertical_mask = cv2.morphologyEx(
        body,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(
            cv2.MORPH_RECT, (1, max(40, round(body.shape[0] * 0.13)))
        ),
    )
    body[vertical_mask > 0] = 0

    projections: list[np.ndarray] = []
    for column in range(EXPECTED_COLUMN_COUNT):
        left = max(0, vertical_rules[column] - body_left + 5)
        right = min(
            body.shape[1], vertical_rules[column + 1] - body_left - 5
        )
        projections.append(np.count_nonzero(body[:, left:right], axis=1))
    column_projections = np.asarray(projections, dtype=np.int32)

    option_candidates: list[
        tuple[tuple[float, int, int, int], list[AnchorCandidate]]
    ] = []
    closest_candidates: tuple[
        tuple[int, int, int, int], list[AnchorCandidate]
    ] | None = None
    for pixel_threshold in (1, 2, 3, 4, 5, 6, 8, 10, 12):
        support_profile = np.count_nonzero(
            column_projections >= pixel_threshold, axis=0
        )
        for required_columns in (8, 7, 9, 6, 10):
            bands = active_bands(
                support_profile >= required_columns, body_top
            )
            bands = [
                band
                for band in bands
                if 2 <= band[1] - band[0] + 1 <= image.shape[0] * 0.030
            ]
            candidates: list[AnchorCandidate] = []
            for top, bottom in bands:
                local_top = top - body_top
                local_bottom = bottom - body_top + 1
                weights = support_profile[local_top:local_bottom].astype(float)
                y_values = np.arange(top, bottom + 1, dtype=float)
                ink_centroid = float(
                    np.average(y_values, weights=weights)
                    if float(np.sum(weights)) > 0
                    else (top + bottom) / 2
                )
                support, day_support = column_support(
                    body,
                    vertical_rules,
                    body_left,
                    body_top,
                    top,
                    bottom,
                )
                candidates.append(
                    AnchorCandidate(
                        top=top,
                        bottom=bottom,
                        center=(top + bottom) / 2,
                        ink_centroid=ink_centroid,
                        support_columns=support,
                        day_column_support=day_support,
                        score=support + (3.0 if day_support else 0.0),
                    )
                )

            closest_key = (
                abs(len(candidates) - EXPECTED_DAY_COUNT),
                0 if len(candidates) >= EXPECTED_DAY_COUNT else 1,
                abs(required_columns - 8),
                pixel_threshold,
            )
            if closest_candidates is None or closest_key < closest_candidates[0]:
                closest_candidates = (closest_key, candidates)

            try:
                selected = choose_31_anchors(
                    candidates, body_top, body_bottom
                )
            except RuntimeError:
                continue

            centers = np.asarray(
                [candidate.center for candidate in selected], dtype=float
            )
            # Score the best supported form rhythm.  This keeps the threshold
            # search neutral between five-day, ten-day and uniform editions.
            residual = best_anchor_rhythm_residual(centers)
            option_candidates.append(
                (
                    (
                        residual,
                        abs(len(candidates) - EXPECTED_DAY_COUNT),
                        abs(required_columns - 8),
                        pixel_threshold,
                    ),
                    candidates,
                )
            )

    if option_candidates:
        return min(option_candidates, key=lambda item: item[0])[1]
    if closest_candidates is not None:
        return closest_candidates[1]
    return []


def choose_31_anchors(
    candidates: list[AnchorCandidate],
    body_top: int,
    body_bottom: int,
    minimum_support_columns: int = 5,
) -> list[AnchorCandidate]:
    if len(candidates) < EXPECTED_DAY_COUNT:
        raise RuntimeError(
            f"墨迹投影只形成{len(candidates)}个候选行，少于31行；"
            "拒绝按固定行距凭空补行。"
        )
    candidates = sorted(candidates, key=lambda item: item.center)
    centers = np.asarray([item.center for item in candidates], dtype=float)
    positive_gaps = np.diff(centers)
    positive_gaps = positive_gaps[positive_gaps > 0]
    if not positive_gaps.size:
        raise RuntimeError("候选锚点没有有效的纵向间隔。")
    # Split glyph bands create abnormally small gaps.  The middle/upper part of
    # the gap distribution is a much safer pitch seed than the lower tail.
    lower = float(np.percentile(positive_gaps, 38))
    upper = float(np.percentile(positive_gaps, 82))
    ordinary = positive_gaps[
        (positive_gaps >= lower) & (positive_gaps <= upper)
    ]
    robust_pitch = float(np.median(ordinary)) if ordinary.size else float(
        np.median(positive_gaps)
    )
    robust_pitch = max(robust_pitch, (body_bottom - body_top) / 58.0)

    negative = -1.0e12
    path_options: list[
        tuple[tuple[float, float, float, float], list[AnchorCandidate]]
    ] = []
    n = len(candidates)
    for _, group_breaks, break_factor in ANCHOR_RHYTHM_MODELS:
        rhythm = anchor_rhythm_coordinates(group_breaks, break_factor)
        rhythm_units = max(1.0, float(rhythm[-1] - rhythm[0]))
        span_pitch = float((centers[-1] - centers[0]) / rhythm_units)
        pitch_seeds = {
            max(1.0, robust_pitch * factor)
            for factor in (0.82, 0.92, 1.0, 1.08, 1.18)
        }
        pitch_seeds.update(
            max(1.0, span_pitch * factor) for factor in (0.88, 1.0, 1.12)
        )
        for base_pitch in sorted(pitch_seeds):
            dp = np.full((EXPECTED_DAY_COUNT, n), negative, dtype=float)
            parent = np.full((EXPECTED_DAY_COUNT, n), -1, dtype=int)
            for index, candidate in enumerate(candidates):
                start_penalty = (
                    abs(candidate.center - body_top) / base_pitch * 0.10
                )
                dp[0, index] = candidate.score - start_penalty
            for selected_index in range(1, EXPECTED_DAY_COUNT):
                expected_gap = base_pitch * (
                    break_factor
                    if selected_index in group_breaks
                    else 1.0
                )
                minimum_gap = expected_gap * 0.58
                maximum_gap = expected_gap * 1.70
                for current in range(selected_index, n):
                    for previous in range(selected_index - 1, current):
                        if dp[selected_index - 1, previous] <= negative / 2:
                            continue
                        gap = (
                            candidates[current].center
                            - candidates[previous].center
                        )
                        if not minimum_gap <= gap <= maximum_gap:
                            continue
                        normalized_error = abs(gap - expected_gap) / max(
                            1.0, expected_gap
                        )
                        gap_penalty = (
                            normalized_error * 3.2
                            + normalized_error * normalized_error * 2.0
                        )
                        value = (
                            dp[selected_index - 1, previous]
                            + candidates[current].score
                            - gap_penalty
                        )
                        if value > dp[selected_index, current]:
                            dp[selected_index, current] = value
                            parent[selected_index, current] = previous
            end_scores = dp[-1].copy()
            for index, candidate in enumerate(candidates):
                end_scores[index] -= (
                    abs(body_bottom - candidate.center) / base_pitch * 0.10
                )
            end = int(np.argmax(end_scores))
            if end_scores[end] <= negative / 2:
                continue
            indices = [end]
            for selected_index in range(EXPECTED_DAY_COUNT - 1, 0, -1):
                end = int(parent[selected_index, end])
                if end < 0:
                    indices = []
                    break
                indices.append(end)
            if not indices:
                continue
            chosen_option = [
                candidates[index] for index in reversed(indices)
            ]
            chosen_centers = np.asarray(
                [item.center for item in chosen_option], dtype=float
            )
            slope, intercept = np.polyfit(rhythm, chosen_centers, 1)
            fitted = slope * rhythm + intercept
            rhythm_residual = float(
                np.sqrt(np.mean(np.square(chosen_centers - fitted)))
                / max(abs(float(slope)), 1.0)
            )
            weak_count = float(
                sum(
                    item.support_columns < minimum_support_columns
                    or not item.day_column_support
                    for item in chosen_option
                )
            )
            endpoint_penalty = float(
                (
                    abs(chosen_centers[0] - body_top)
                    + abs(body_bottom - chosen_centers[-1])
                )
                / max(1.0, abs(float(slope)))
            )
            mean_support = float(
                np.mean([item.score for item in chosen_option])
            )
            path_options.append(
                (
                    (
                        weak_count,
                        rhythm_residual,
                        endpoint_penalty * 0.025,
                        -mean_support,
                    ),
                    chosen_option,
                )
            )

    if not path_options:
        raise RuntimeError("31行锚点动态规划没有形成有效路径。")
    _, chosen = min(path_options, key=lambda item: item[0])

    if len(chosen) != EXPECTED_DAY_COUNT:
        raise RuntimeError("最终行锚点数量不是31。")
    if any(
        chosen[index].center >= chosen[index + 1].center
        for index in range(len(chosen) - 1)
    ):
        raise RuntimeError("最终行锚点没有严格递增。")
    weak = [
        index + 1
        for index, item in enumerate(chosen)
        if item.support_columns < minimum_support_columns
        or (
            not item.day_column_support
            and item.support_columns
            < max(8, minimum_support_columns + 3)
        )
    ]
    if weak:
        raise RuntimeError(
            "以下日序锚点缺少足够的多列/日列证据："
            + "、".join(map(str, weak))
        )
    return chosen


def anchor_rhythm_coordinates(
    group_breaks: frozenset[int] | set[int], break_factor: float
) -> np.ndarray:
    """Return normalized row centers for one printed grouping convention."""

    values = [0.0]
    for previous_day in range(1, EXPECTED_DAY_COUNT):
        values.append(
            values[-1]
            + (break_factor if previous_day in group_breaks else 1.0)
        )
    return np.asarray(values, dtype=float)


def best_anchor_rhythm_residual(centers: np.ndarray) -> float:
    """Measure a 31-row path against the best supported form rhythm."""

    if len(centers) != EXPECTED_DAY_COUNT:
        return float("inf")
    residuals: list[float] = []
    for _, group_breaks, break_factor in ANCHOR_RHYTHM_MODELS:
        pattern = anchor_rhythm_coordinates(group_breaks, break_factor)
        slope, intercept = np.polyfit(pattern, centers, 1)
        residuals.append(
            float(
                np.sqrt(
                    np.mean(
                        np.square(centers - (slope * pattern + intercept))
                    )
                )
                / max(abs(float(slope)), 1.0)
            )
        )
    return min(residuals)


def detect_page_geometry(
    image: np.ndarray, page_number: int, table_index: int = 1
) -> PageGeometry:
    vertical_rules, horizontal_rules = detect_rule_geometry(image)
    vertical_top, _ = find_table_vertical_bounds(image, vertical_rules)
    table_top = min(
        horizontal_rules,
        key=lambda value: abs(value - vertical_top),
    )
    rules_below_top = [
        value
        for value in horizontal_rules
        if value > table_top + image.shape[0] * 0.012
    ]
    if len(rules_below_top) < 2:
        raise RuntimeError("无法确定表头底线和统计区起始线。")
    header_bottom = rules_below_top[0]
    statistics_candidates = [
        value
        for value in rules_below_top[1:]
        if value > header_bottom + image.shape[0] * 0.20
    ]
    if not statistics_candidates:
        raise RuntimeError("无法确定逐日数据与统计区之间的分界线。")
    statistics_top = statistics_candidates[0]
    candidates = build_anchor_candidates(
        image,
        vertical_rules,
        header_bottom + 2,
        statistics_top - 1,
    )
    anchors = choose_31_anchors(
        candidates, header_bottom + 2, statistics_top - 1
    )
    return PageGeometry(
        page=page_number,
        table_index=table_index,
        width=image.shape[1],
        height=image.shape[0],
        table_left=vertical_rules[0],
        table_top=table_top,
        table_right=vertical_rules[-1],
        header_bottom=header_bottom,
        statistics_top=statistics_top,
        vertical_rules=vertical_rules,
        horizontal_rules=horizontal_rules,
        candidate_anchors=candidates,
        row_anchors=anchors,
    )


def darkest_rule_row(gray: np.ndarray, top: int, bottom: int) -> int:
    top = max(0, top)
    bottom = min(gray.shape[0], bottom)
    if bottom <= top:
        raise RuntimeError("横线搜索范围为空。")
    darkness = np.sum(255 - gray[top:bottom].astype(np.int32), axis=1)
    return top + int(np.argmax(darkness))


def validate_anchor_rhythm(
    anchors: list[AnchorCandidate], body_top: int, body_bottom: int
) -> None:
    """Reject a 31-row path that has visibly snapped two days to one glyph."""

    if len(anchors) != EXPECTED_DAY_COUNT:
        raise RuntimeError("锚点节律校验收到的行数不是31。")
    centers = np.asarray([item.center for item in anchors], dtype=float)
    gaps = np.diff(centers)
    ordinary = gaps[gaps <= np.percentile(gaps, 70)]
    base_pitch = float(np.median(ordinary)) if ordinary.size else 1.0
    if base_pitch <= 0:
        raise RuntimeError("31行锚点没有有效行距。")
    if float(np.min(gaps)) < base_pitch * 0.40:
        raise RuntimeError(
            "相邻日序锚点过近，疑似两个虚拟/实锚点吸附到同一处墨迹。"
        )
    if float(np.max(gaps)) > base_pitch * 2.35:
        raise RuntimeError("相邻日序锚点间出现异常大断层。")
    if centers[0] - body_top < base_pitch * 0.32:
        raise RuntimeError("第1日锚点贴住表头分界线。")
    if body_bottom - centers[-1] < base_pitch * 0.32:
        raise RuntimeError("第31日锚点贴住统计区分界线。")


def detect_rectified_geometry(
    image: np.ndarray, region: TableRegion
) -> PageGeometry:
    """Locate columns and 31 ink anchors after photographic dewarping."""

    height, width = image.shape[:2]
    left_source = region.source_bbox[0]
    if len(region.rectified_rules) == EXPECTED_VERTICAL_RULE_COUNT:
        vertical_rules = [
            min(width - 1, max(0, int(value)))
            for value in region.rectified_rules
        ]
    else:
        vertical_rules = [
            min(width - 1, max(0, round(x - left_source)))
            for x in region.source_rules
        ]
    vertical_rules[0] = 0
    vertical_rules[-1] = width - 1
    if region.rectified_boundaries:
        table_top = int(region.rectified_boundaries.get("table_top", 0))
        header_bottom = int(region.rectified_boundaries["header_bottom"])
        statistics_top = int(region.rectified_boundaries["statistics_top"])
    else:
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        table_top = 0
        header_bottom = darkest_rule_row(
            gray, round(height * 0.035), round(height * 0.095)
        )
        statistics_top = round(height * 0.780)
    if not (
        0 <= table_top < header_bottom < statistics_top < height
        and statistics_top - header_bottom >= height * 0.35
    ):
        raise RuntimeError("非刚性矫正后的四条边界次序或逐日区高度异常。")
    candidates = build_day_column_candidates(
        image, vertical_rules, header_bottom + 2, statistics_top - 2
    )
    anchor_method = "day-column-projection"
    try:
        anchors = choose_31_anchors(
            candidates,
            header_bottom + 2,
            statistics_top - 2,
            minimum_support_columns=2,
        )
        validate_anchor_rhythm(
            anchors, header_bottom + 2, statistics_top - 2
        )
    except RuntimeError as day_error:
        candidates = build_multicolumn_support_candidates(
            image, vertical_rules, header_bottom + 2, statistics_top - 2
        )
        try:
            anchors = choose_31_anchors(
                candidates, header_bottom + 2, statistics_top - 2
            )
            validate_anchor_rhythm(
                anchors, header_bottom + 2, statistics_top - 2
            )
        except RuntimeError as support_error:
            raise RuntimeError(
                "日列锚点失败，跨列一致性锚点也失败："
                f"日列={day_error}；跨列={support_error}"
            ) from support_error
        anchor_method = "multicolumn-support-projection"
    horizontal_rules = [table_top, header_bottom, statistics_top, height - 1]
    return PageGeometry(
        page=region.page,
        table_index=region.table_index,
        width=width,
        height=height,
        table_left=0,
        table_top=table_top,
        table_right=width - 1,
        header_bottom=header_bottom,
        statistics_top=statistics_top,
        vertical_rules=vertical_rules,
        horizontal_rules=horizontal_rules,
        candidate_anchors=candidates,
        row_anchors=anchors,
        anchor_method=anchor_method,
        source_region_diagnostics=dict(region.diagnostics),
    )


def anchor_row_bounds(
    anchors: list[AnchorCandidate],
    index: int,
    body_top: int,
    body_bottom: int,
) -> tuple[int, int]:
    if index == 0:
        top = body_top
    else:
        top = round((anchors[index - 1].center + anchors[index].center) / 2)
    if index == len(anchors) - 1:
        bottom = body_bottom
    else:
        bottom = round((anchors[index].center + anchors[index + 1].center) / 2)
    return top, bottom


def _fit_statistics_guided_row_curves_legacy(
    image: np.ndarray, geometry: PageGeometry
) -> None:
    """Track a non-crossing row-anchor chain through all 13 columns.

    The printed day column supplies the 1-31 order.  Each month column then
    contributes a new local ink anchor.  Forward and backward passes keep a
    difficult column from propagating its error through the remaining months;
    a regularized fit removes isolated glyph errors while retaining folds.
    """

    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    page_ink = cv2.threshold(
        gray, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU
    )[1]
    centers_x = [
        (geometry.vertical_rules[index] + geometry.vertical_rules[index + 1])
        / 2
        for index in range(EXPECTED_COLUMN_COUNT)
    ]
    base_centers = np.asarray(
        [anchor.center for anchor in geometry.row_anchors], dtype=float
    )
    day_progress = (base_centers - geometry.header_bottom) / max(
        1.0, geometry.statistics_top - geometry.header_bottom
    )
    statistics_samples = fit_statistics_separator_samples(
        gray, page_ink, geometry, centers_x, row_pitch
    )

    statistics_array = np.asarray(statistics_samples, dtype=float)
    coarse_by_day = base_centers[:, None] + day_progress[:, None] * (
        statistics_array[None, :] - geometry.statistics_top
    )
    observations = np.full(
        (EXPECTED_DAY_COUNT, EXPECTED_COLUMN_COUNT), np.nan, dtype=float
    )
    weights = np.zeros_like(observations)
    projections: list[np.ndarray] = []
    column_widths: list[int] = []
    for column in range(EXPECTED_COLUMN_COUNT):
        left = vertical_rule_x(geometry, column, body_midpoint) + 5
        right = vertical_rule_x(geometry, column + 1, body_midpoint) - 5
        column_width = max(1, right - left)
        column_widths.append(column_width)
        expected = coarse_by_day[:, column]
        projection = np.count_nonzero(page_ink[:, left:right], axis=1)
        projections.append(projection)
        for day_index in range(EXPECTED_DAY_COUNT):
            if day_index == 0:
                local_top = geometry.header_bottom + 2
            else:
                local_top = round((expected[day_index - 1] + expected[day_index]) / 2)
            if day_index == EXPECTED_DAY_COUNT - 1:
                local_bottom = round(statistics_samples[column] - 2)
            else:
                local_bottom = round((expected[day_index] + expected[day_index + 1]) / 2)
            local_top = max(0, local_top)
            local_bottom = min(geometry.height, local_bottom)
            if local_bottom <= local_top + 1:
                continue
            values = projection[local_top:local_bottom].astype(float)
            peak = float(np.max(values)) if values.size else 0.0
            total = float(np.sum(values))
            if peak < 2 or total < max(6.0, column_width * 0.055):
                continue
            threshold = max(1.0, peak * 0.16)
            selected = values >= threshold
            if not np.any(selected):
                continue
            y_values = np.arange(local_top, local_bottom, dtype=float)
            selected_weights = values * selected
            centroid = float(
                np.average(y_values, weights=selected_weights)
            )
            observations[day_index, column] = centroid
            weights[day_index, column] = min(
                3.0, 0.5 + peak / max(1.0, column_width * 0.10)
            )

    minimum_chain_gap = max(
        2.0, float(np.median(np.diff(base_centers))) * 0.20
    )

    def observe_anchor_column(
        column: int, predicted: np.ndarray
    ) -> tuple[np.ndarray, np.ndarray]:
        """Attach ordered predictions to local ink inside disjoint row bands."""

        tracked = predicted.astype(float).copy()
        confidence = np.zeros(EXPECTED_DAY_COUNT, dtype=float)
        projection = projections[column]
        column_width = column_widths[column]
        for day_index in range(EXPECTED_DAY_COUNT):
            if day_index == 0:
                local_top = geometry.header_bottom + 2
            else:
                local_top = round(
                    (predicted[day_index - 1] + predicted[day_index]) / 2
                )
            if day_index == EXPECTED_DAY_COUNT - 1:
                local_bottom = round(statistics_samples[column] - 2)
            else:
                local_bottom = round(
                    (predicted[day_index] + predicted[day_index + 1]) / 2
                )
            local_top = max(0, local_top)
            local_bottom = min(geometry.height, local_bottom)
            if local_bottom <= local_top + 1:
                continue
            values = projection[local_top:local_bottom].astype(float)
            peak = float(np.max(values)) if values.size else 0.0
            total = float(np.sum(values))
            if peak < 2 or total < max(6.0, column_width * 0.055):
                continue
            selected = values >= max(1.0, peak * 0.16)
            if not np.any(selected):
                continue
            y_values = np.arange(local_top, local_bottom, dtype=float)
            selected_weights = values * selected
            centroid = float(np.average(y_values, weights=selected_weights))
            evidence = min(
                3.0, 0.5 + peak / max(1.0, column_width * 0.10)
            )
            # Retain a small motion prior so a tall glyph cannot pull the chain
            # all the way to the edge of the neighbouring row's band.
            attachment = min(0.86, 0.55 + evidence * 0.09)
            tracked[day_index] = (
                attachment * centroid + (1.0 - attachment) * predicted[day_index]
            )
            confidence[day_index] = evidence

        for day_index in range(1, EXPECTED_DAY_COUNT):
            tracked[day_index] = max(
                tracked[day_index],
                tracked[day_index - 1] + minimum_chain_gap,
            )
        for day_index in range(EXPECTED_DAY_COUNT - 2, -1, -1):
            tracked[day_index] = min(
                tracked[day_index],
                tracked[day_index + 1] - minimum_chain_gap,
            )
        return tracked, confidence

    # Forward propagation follows the physical reading direction from the
    # proven day-number anchors.  The statistics boundary supplies only the
    # expected inter-column deformation, not an invented row position.
    forward = np.zeros_like(observations)
    forward_confidence = np.zeros_like(weights)
    forward[:, 0] = base_centers
    forward_confidence[:, 0] = 3.0
    for column in range(1, EXPECTED_COLUMN_COUNT):
        predicted = forward[:, column - 1] + (
            coarse_by_day[:, column] - coarse_by_day[:, column - 1]
        )
        forward[:, column], forward_confidence[:, column] = (
            observe_anchor_column(column, predicted)
        )

    # A reverse pass prevents a single weak/dirty month from becoming a
    # permanent accumulated offset.  Its right-edge seed prefers direct ink
    # but safely falls back to the coarse deformation estimate for blank cells.
    backward = np.zeros_like(observations)
    backward_confidence = np.zeros_like(weights)
    right_valid = np.isfinite(observations[:, -1])
    backward[:, -1] = np.where(
        right_valid, observations[:, -1], coarse_by_day[:, -1]
    )
    backward_confidence[:, -1] = np.where(
        right_valid, weights[:, -1], 0.15
    )
    for column in range(EXPECTED_COLUMN_COUNT - 2, -1, -1):
        predicted = backward[:, column + 1] + (
            coarse_by_day[:, column] - coarse_by_day[:, column + 1]
        )
        backward[:, column], backward_confidence[:, column] = (
            observe_anchor_column(column, predicted)
        )

    # Fuse direct, forward and reverse evidence.  Direct ink remains strongest;
    # the two chains mainly rescue large bends and bridge empty calendar cells.
    for day_index in range(EXPECTED_DAY_COUNT):
        for column in range(EXPECTED_COLUMN_COUNT):
            values: list[float] = []
            evidence_weights: list[float] = []
            if np.isfinite(observations[day_index, column]):
                values.append(float(observations[day_index, column]))
                evidence_weights.append(float(weights[day_index, column]))
            if forward_confidence[day_index, column] > 0:
                values.append(float(forward[day_index, column]))
                evidence_weights.append(
                    float(forward_confidence[day_index, column] * 0.65)
                )
            if backward_confidence[day_index, column] > 0:
                values.append(float(backward[day_index, column]))
                evidence_weights.append(
                    float(backward_confidence[day_index, column] * 0.65)
                )
            if values:
                observations[day_index, column] = float(
                    np.average(values, weights=evidence_weights)
                )
                weights[day_index, column] = min(
                    5.0, float(np.sum(evidence_weights))
                )

    count = EXPECTED_COLUMN_COUNT
    second_difference = np.zeros((count - 2, count), dtype=float)
    for index in range(count - 2):
        second_difference[index, index : index + 3] = (1.0, -2.0, 1.0)
    smoothness = second_difference.T @ second_difference
    fitted = np.zeros_like(observations)
    for day_index in range(EXPECTED_DAY_COUNT):
        coarse = coarse_by_day[day_index]
        observed = observations[day_index]
        row_weights = weights[day_index].copy()
        valid = np.isfinite(observed)
        observed_filled = np.where(valid, observed, coarse)
        # Reject a candidate that escaped its date band because a neighbouring
        # row has unusually tall digits.
        if day_index == 0:
            gap_limit = (base_centers[1] - base_centers[0]) * 0.48
        elif day_index == EXPECTED_DAY_COUNT - 1:
            gap_limit = (base_centers[-1] - base_centers[-2]) * 0.48
        else:
            gap_limit = min(
                base_centers[day_index] - base_centers[day_index - 1],
                base_centers[day_index + 1] - base_centers[day_index],
            ) * 0.48
        outlier = np.abs(observed_filled - coarse) > max(3.0, gap_limit)
        row_weights[outlier] = 0.0
        prior_weight = 0.20
        matrix = (
            np.diag(row_weights + prior_weight)
            + 2.8 * smoothness
        )
        target = row_weights * observed_filled + prior_weight * coarse
        fitted[day_index] = np.linalg.solve(matrix, target)

    # Paper deformation changes continuously down the page, whereas glyph
    # centroids vary slightly with the digits themselves.  Smooth only the
    # residual deformation field, then pin it to the ordered day column.
    residual = (fitted - coarse_by_day).astype(np.float32)
    residual = cv2.GaussianBlur(
        residual, (3, 5), sigmaX=0.8, sigmaY=1.0
    ).astype(float)
    residual -= residual[:, :1]
    fitted = coarse_by_day + residual

    # Enforce strict top-to-bottom order in every column.  Normally this is a
    # no-op; it is a safety net for a deep fold or a nearly blank date row.
    minimum_gap = minimum_chain_gap
    for column in range(EXPECTED_COLUMN_COUNT):
        for day_index in range(1, EXPECTED_DAY_COUNT):
            fitted[day_index, column] = max(
                fitted[day_index, column],
                fitted[day_index - 1, column] + minimum_gap,
            )
        for day_index in range(EXPECTED_DAY_COUNT - 2, -1, -1):
            fitted[day_index, column] = min(
                fitted[day_index, column],
                fitted[day_index + 1, column] - minimum_gap,
            )

    geometry.column_center_x = [float(value) for value in centers_x]
    geometry.row_curve_centers = fitted.tolist()
    geometry.statistics_curve = statistics_samples


def _fit_independent_sequence_row_curves_experimental(
    image: np.ndarray, geometry: PageGeometry
) -> None:
    """Build row curves from independently ordered ink anchors in every column.

    The day column proves the 1-31 identities.  Each month column detects its
    own complete top-to-bottom ink sequence and matches that sequence
    monotonically to the day identities.  The statistics boundary is measured
    only for diagnostics and never drives a daily-row position.
    """

    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    page_ink = cv2.threshold(
        gray, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU
    )[1]
    centers_x = [
        (geometry.vertical_rules[index] + geometry.vertical_rules[index + 1])
        / 2
        for index in range(EXPECTED_COLUMN_COUNT)
    ]
    base_centers = np.asarray(
        [anchor.center for anchor in geometry.row_anchors], dtype=float
    )
    ordinary_gaps = np.diff(base_centers)
    ordinary_gaps = ordinary_gaps[
        ordinary_gaps <= np.percentile(ordinary_gaps, 70)
    ]
    row_pitch = float(np.median(ordinary_gaps)) if ordinary_gaps.size else 10.0
    row_pitch = max(4.0, row_pitch)

    # At this stage the scalar table boundary is safer than a raw darkest-row
    # search: the latter can lock onto a rule *inside* the statistics block.
    statistics_samples = [
        float(geometry.statistics_top)
    ] * EXPECTED_COLUMN_COUNT

    observations = np.full(
        (EXPECTED_DAY_COUNT, EXPECTED_COLUMN_COUNT), np.nan, dtype=float
    )
    weights = np.zeros_like(observations)
    observations[:, 0] = base_centers
    weights[:, 0] = 5.0

    def valid_day_options(column: int) -> tuple[int, ...]:
        if column == 0:
            return (31,)
        if column == 2:
            return (28, 29)
        if column in (4, 6, 9, 11):
            return (30,)
        return (31,)

    def select_ordered_subset(
        candidates: list[tuple[float, float]], target_count: int
    ) -> list[tuple[float, float]] | None:
        """Select a rhythm-preserving monotone subset without inventing rows."""

        count = len(candidates)
        if count < target_count:
            return None
        if count == target_count:
            return candidates
        centers = np.asarray([item[0] for item in candidates], dtype=float)
        candidate_gaps = np.diff(centers)
        pitch_like = candidate_gaps[
            (candidate_gaps >= row_pitch * 0.45)
            & (candidate_gaps <= row_pitch * 1.55)
        ]
        scale = (
            float(np.median(pitch_like)) / row_pitch
            if pitch_like.size
            else 1.0
        )
        scale = min(1.30, max(0.70, scale))
        expected = base_centers[:target_count]
        infinity = 1.0e12
        dp = np.full((target_count, count), infinity, dtype=float)
        parent = np.full((target_count, count), -1, dtype=int)
        for current in range(count):
            dp[0, current] = (
                abs(centers[current] - expected[0]) / row_pitch * 0.25
                + current * 0.20
            )
        for day_index in range(1, target_count):
            expected_gap = (
                expected[day_index] - expected[day_index - 1]
            ) * scale
            for current in range(day_index, count):
                for previous in range(day_index - 1, current):
                    actual_gap = centers[current] - centers[previous]
                    gap_cost = (
                        abs(actual_gap - expected_gap) / row_pitch
                    ) ** 2 * 3.0
                    skipped = current - previous - 1
                    value = dp[day_index - 1, previous] + gap_cost + skipped * 0.25
                    if value < dp[day_index, current]:
                        dp[day_index, current] = value
                        parent[day_index, current] = previous
        end_costs = dp[-1].copy()
        for current in range(count):
            end_costs[current] += (count - current - 1) * 0.20
        end = int(np.argmin(end_costs))
        if not np.isfinite(end_costs[end]) or end_costs[end] >= infinity / 2:
            return None
        indices = [end]
        for day_index in range(target_count - 1, 0, -1):
            end = int(parent[day_index, end])
            if end < 0:
                return None
            indices.append(end)
        return [candidates[index] for index in reversed(indices)]

    for column in range(1, EXPECTED_COLUMN_COUNT):
        left = geometry.vertical_rules[column] + 6
        right = geometry.vertical_rules[column + 1] - 6
        column_width = max(1, right - left)
        projection = np.count_nonzero(page_ink[:, left:right], axis=1)
        body_top = max(
            geometry.header_bottom + 2,
            round(base_centers[0] - row_pitch * 0.70),
        )
        body_bottom = min(
            geometry.height,
            round(base_centers[-1] + row_pitch * 1.35),
        )
        options: list[
            tuple[tuple[float, float, int, int], list[tuple[float, float]], int]
        ] = []
        for threshold in range(1, 13):
            bands = active_bands(
                projection[body_top:body_bottom] >= threshold,
                body_top,
            )
            candidates: list[tuple[float, float]] = []
            for top, bottom in bands:
                height = bottom - top + 1
                values = projection[top : bottom + 1].astype(float)
                peak = float(np.max(values)) if values.size else 0.0
                total = float(np.sum(values))
                if not (2 <= height <= row_pitch * 0.88):
                    continue
                # A printed horizontal rule spans most of a month column; a
                # number never does.  Reject the rule before row matching.
                if peak >= column_width * 0.72:
                    continue
                y_values = np.arange(top, bottom + 1, dtype=float)
                centroid = float(
                    np.average(y_values, weights=values)
                    if total > 0
                    else (top + bottom) / 2
                )
                evidence = min(
                    4.0,
                    0.5
                    + peak / max(1.0, column_width * 0.10)
                    + total / max(1.0, column_width * row_pitch * 0.18),
                )
                candidates.append((centroid, evidence))

            for target_count in valid_day_options(column):
                selected = select_ordered_subset(candidates, target_count)
                if selected is None:
                    continue
                selected_centers = np.asarray(
                    [item[0] for item in selected], dtype=float
                )
                expected = base_centers[:target_count]
                degree = 2 if target_count >= 5 else 1
                coefficients = np.polyfit(expected, selected_centers, degree)
                residual = float(
                    np.sqrt(
                        np.mean(
                            (
                                selected_centers
                                - np.polyval(coefficients, expected)
                            )
                            ** 2
                        )
                    )
                    / row_pitch
                )
                options.append(
                    (
                        (
                            abs(len(candidates) - target_count),
                            residual,
                            threshold,
                            -target_count,
                        ),
                        selected,
                        target_count,
                    )
                )

        if options:
            _, selected, target_count = min(options, key=lambda item: item[0])
            for day_index, (center, evidence) in enumerate(selected):
                observations[day_index, column] = center
                weights[day_index, column] = evidence
            continue

        # Conservative fallback: estimate one column-local offset from ink and
        # recenter every day independently.  It does not use neighbouring
        # columns or the statistics line, so any failure remains local.
        target_count = max(valid_day_options(column))
        shift_candidates = np.linspace(-row_pitch * 1.5, row_pitch * 1.5, 61)
        shift_scores: list[float] = []
        half_window = max(3, round(row_pitch * 0.28))
        for shift in shift_candidates:
            score = 0.0
            for expected in base_centers[:target_count] + shift:
                top = max(0, round(expected) - half_window)
                bottom = min(geometry.height, round(expected) + half_window + 1)
                score += float(np.max(projection[top:bottom]))
            shift_scores.append(score)
        shift = float(shift_candidates[int(np.argmax(shift_scores))])
        expected_centers = base_centers + shift
        for day_index in range(target_count):
            if day_index == 0:
                top = geometry.header_bottom + 2
            else:
                top = round(
                    (expected_centers[day_index - 1] + expected_centers[day_index])
                    / 2
                )
            if day_index == EXPECTED_DAY_COUNT - 1:
                bottom = body_bottom
            else:
                bottom = round(
                    (expected_centers[day_index] + expected_centers[day_index + 1])
                    / 2
                )
            values = projection[max(0, top) : min(geometry.height, bottom)].astype(float)
            if values.size == 0 or float(np.max(values)) < 2:
                continue
            selected_rows = values >= max(1.0, float(np.max(values)) * 0.16)
            selected_weights = values * selected_rows
            if float(np.sum(selected_weights)) <= 0:
                continue
            y_values = np.arange(max(0, top), min(geometry.height, bottom), dtype=float)
            observations[day_index, column] = float(
                np.average(y_values, weights=selected_weights)
            )
            weights[day_index, column] = 1.0

    displacement = observations - base_centers[:, None]
    filled_displacement = np.zeros_like(displacement)
    day_indices = np.arange(EXPECTED_DAY_COUNT, dtype=float)
    for column in range(EXPECTED_COLUMN_COUNT):
        valid = np.isfinite(displacement[:, column])
        if np.count_nonzero(valid) >= 2:
            filled_displacement[:, column] = np.interp(
                day_indices,
                day_indices[valid],
                displacement[valid, column],
            )
        elif np.count_nonzero(valid) == 1:
            filled_displacement[:, column] = float(displacement[valid, column][0])

    # Smooth only the independently observed displacement field.  Direct ink
    # keeps most of the weight; smoothing bridges calendar blanks and removes
    # digit-shape centroid jitter without propagating a statistics-line error.
    smooth_displacement = cv2.GaussianBlur(
        filled_displacement.astype(np.float32),
        (3, 5),
        sigmaX=0.65,
        sigmaY=0.9,
    ).astype(float)
    fitted_displacement = smooth_displacement.copy()
    valid = np.isfinite(displacement)
    fitted_displacement[valid] = (
        displacement[valid] * 0.72 + smooth_displacement[valid] * 0.28
    )
    fitted_displacement[:, 0] = 0.0
    fitted = base_centers[:, None] + fitted_displacement

    minimum_gap = max(2.0, row_pitch * 0.20)
    for column in range(EXPECTED_COLUMN_COUNT):
        for day_index in range(1, EXPECTED_DAY_COUNT):
            fitted[day_index, column] = max(
                fitted[day_index, column],
                fitted[day_index - 1, column] + minimum_gap,
            )
        for day_index in range(EXPECTED_DAY_COUNT - 2, -1, -1):
            fitted[day_index, column] = min(
                fitted[day_index, column],
                fitted[day_index + 1, column] - minimum_gap,
            )

    geometry.column_center_x = [float(value) for value in centers_x]
    geometry.row_curve_centers = fitted.tolist()
    geometry.statistics_curve = statistics_samples


def fit_dynamic_vertical_rule_models(
    image: np.ndarray, geometry: PageGeometry, row_pitch: float
) -> None:
    """Fit every photographed table rule as x(y), not as a fixed x value."""

    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    median_column_width = float(np.median(np.diff(geometry.vertical_rules)))
    search_radius = max(6, round(median_column_width * 0.105))
    half_band = max(4, round(row_pitch * 0.24))
    top = max(1, geometry.table_top)
    bottom = min(geometry.height - 2, geometry.statistics_top)
    sample_y = np.linspace(top, bottom, 90).astype(int)
    # A curved rule appears in a probabilistic Hough transform as several
    # short, near-vertical segments.  Their endpoints seed the later robust
    # x(y) fit; they are not mistaken for one global straight line.
    edges = cv2.Canny(gray[top:bottom], 55, 150, apertureSize=3)
    hough_lines = cv2.HoughLinesP(
        edges,
        1,
        np.pi / 360,
        threshold=max(12, round(row_pitch * 0.8)),
        minLineLength=max(12, round(row_pitch * 1.6)),
        maxLineGap=max(5, round(row_pitch * 0.75)),
    )
    hough_points: list[list[tuple[float, float, float]]] = [
        [] for _ in geometry.vertical_rules
    ]
    if hough_lines is not None:
        for line in np.asarray(hough_lines).reshape(-1, 4):
            x1, local_y1, x2, local_y2 = map(float, line)
            y1 = local_y1 + top
            y2 = local_y2 + top
            vertical_span = abs(y2 - y1)
            if vertical_span < row_pitch * 1.4:
                continue
            if abs(x2 - x1) > max(3.0, vertical_span * 0.24):
                continue
            midpoint_x = (x1 + x2) / 2
            rule_index = int(
                np.argmin(
                    np.abs(
                        np.asarray(geometry.vertical_rules, dtype=float)
                        - midpoint_x
                    )
                )
            )
            if (
                abs(midpoint_x - geometry.vertical_rules[rule_index])
                > search_radius * 1.25
            ):
                continue
            segment_weight = min(255.0 * vertical_span, 255.0 * row_pitch * 5)
            hough_points[rule_index].extend(
                [
                    (y1, x1, segment_weight),
                    ((y1 + y2) / 2, midpoint_x, segment_weight),
                    (y2, x2, segment_weight),
                ]
            )
    fitted_models: list[np.ndarray | None] = []
    for rule_index, nominal_x in enumerate(geometry.vertical_rules):
        observed_y: list[float] = []
        observed_x: list[float] = []
        observed_weight: list[float] = []
        corridor_left = max(0, nominal_x - search_radius)
        corridor_right = min(geometry.width, nominal_x + search_radius + 1)
        for center_y in sample_y:
            band_top = max(0, center_y - half_band)
            band_bottom = min(geometry.height, center_y + half_band + 1)
            band = 255 - gray[band_top:band_bottom, corridor_left:corridor_right]
            if band.size == 0:
                continue
            darkness = np.sum(band.astype(np.float32), axis=0)
            xs = np.arange(corridor_left, corridor_right, dtype=float)
            # The rule must remain close to its nominal component center.  A
            # mild penalty prevents a repeated digit stroke from winning one
            # local band while still permitting genuine slant/curvature.
            penalized = darkness - np.abs(xs - nominal_x) * (
                max(1.0, float(np.percentile(darkness, 80))) * 0.018
            )
            local_index = int(np.argmax(penalized))
            peak = float(darkness[local_index])
            if peak < (band_bottom - band_top) * 255 * 0.28:
                continue
            observed_y.append(float(center_y))
            observed_x.append(float(xs[local_index]))
            observed_weight.append(peak)

        for hough_y, hough_x, hough_weight in hough_points[rule_index]:
            observed_y.append(float(hough_y))
            observed_x.append(float(hough_x))
            observed_weight.append(float(hough_weight))

        if len(observed_y) < 12:
            fitted_models.append(None)
            continue
        ys = np.asarray(observed_y, dtype=float)
        xs = np.asarray(observed_x, dtype=float)
        weights = np.sqrt(np.asarray(observed_weight, dtype=float))
        inliers = np.ones(len(ys), dtype=bool)
        coefficients = np.asarray([0.0, 0.0, float(nominal_x)])
        for _ in range(5):
            degree = 2 if int(np.count_nonzero(inliers)) >= 18 else 1
            fitted_coefficients = np.polyfit(
                ys[inliers], xs[inliers], degree, w=weights[inliers]
            )
            if degree == 1:
                fitted_coefficients = np.asarray(
                    [0.0, fitted_coefficients[0], fitted_coefficients[1]]
                )
            predicted = np.polyval(fitted_coefficients, ys)
            residual = np.abs(xs - predicted)
            updated = residual <= max(2.0, search_radius * 0.22)
            coefficients = fitted_coefficients
            if int(np.count_nonzero(updated)) < 10 or np.array_equal(updated, inliers):
                break
            inliers = updated
        check_y = np.linspace(top, bottom, 25)
        check_x = np.polyval(coefficients, check_y)
        if (
            not np.all(np.isfinite(check_x))
            or float(np.max(np.abs(check_x - nominal_x))) > search_radius * 1.05
        ):
            fitted_models.append(None)
        else:
            fitted_models.append(np.asarray(coefficients, dtype=float))

    # A failed local track must inherit the photographed bend/slant from its
    # neighbouring rules, not collapse to a vertical x=constant line.  Linear
    # interpolation of correction polynomials preserves ordered spacing and is
    # especially important near the curved right edge of close photographs.
    good_indices = [
        index for index, model in enumerate(fitted_models) if model is not None
    ]
    if good_indices:
        corrections = {
            index: np.asarray(fitted_models[index], dtype=float)
            - np.asarray(
                [0.0, 0.0, float(geometry.vertical_rules[index])],
                dtype=float,
            )
            for index in good_indices
        }
        shared_correction = np.median(
            np.stack(list(corrections.values())), axis=0
        )
    else:
        corrections = {}
        shared_correction = np.zeros(3, dtype=float)
    for index, model in enumerate(fitted_models):
        if model is not None:
            continue
        left_candidates = [value for value in good_indices if value < index]
        right_candidates = [value for value in good_indices if value > index]
        left_index = max(left_candidates) if left_candidates else None
        right_index = min(right_candidates) if right_candidates else None
        if left_index is not None and right_index is not None:
            fraction = (index - left_index) / (right_index - left_index)
            correction = (
                corrections[left_index] * (1.0 - fraction)
                + corrections[right_index] * fraction
            )
        elif left_index is not None:
            correction = corrections[left_index]
        elif right_index is not None:
            correction = corrections[right_index]
        else:
            correction = shared_correction
        fitted_models[index] = correction + np.asarray(
            [0.0, 0.0, float(geometry.vertical_rules[index])], dtype=float
        )

    geometry.vertical_rule_models = [
        [float(value) for value in np.asarray(model, dtype=float)]
        for model in fitted_models
    ]


def vertical_rule_x(geometry: PageGeometry, rule_index: int, y: float) -> float:
    if (
        geometry.vertical_rule_models
        and rule_index < len(geometry.vertical_rule_models)
    ):
        return float(np.polyval(geometry.vertical_rule_models[rule_index], y))
    return float(geometry.vertical_rules[rule_index])


def fit_horizontal_boundary_samples(
    gray: np.ndarray,
    geometry: PageGeometry,
    centers_x: list[float],
    nominal_y: float,
    search_radius: int,
) -> list[float]:
    """Track one photographed horizontal rule as y(x) from local evidence.

    A bowed/folded rule is represented by short near-horizontal Hough segments
    plus a local darkness peak in every logical column.  The returned samples
    preserve column order and are later interpolated with PCHIP; no rectangular
    boundary is imposed on the daily body.
    """

    height, width = gray.shape[:2]
    strip_top = max(0, round(nominal_y) - search_radius)
    strip_bottom = min(height, round(nominal_y) + search_radius + 1)
    if strip_bottom <= strip_top + 2:
        return [float(nominal_y)] * EXPECTED_COLUMN_COUNT
    median_width = float(np.median(np.diff(geometry.vertical_rules)))
    edges = cv2.Canny(gray[strip_top:strip_bottom], 45, 145, apertureSize=3)
    lines = cv2.HoughLinesP(
        edges,
        1,
        np.pi / 360,
        threshold=max(10, round(median_width * 0.13)),
        minLineLength=max(12, round(median_width * 0.34)),
        maxLineGap=max(5, round(median_width * 0.12)),
    )
    hough_y: list[list[float]] = [[] for _ in range(EXPECTED_COLUMN_COUNT)]
    if lines is not None:
        for x1, local_y1, x2, local_y2 in np.asarray(lines).reshape(-1, 4):
            x1, x2 = float(x1), float(x2)
            y1, y2 = float(local_y1 + strip_top), float(local_y2 + strip_top)
            span = abs(x2 - x1)
            if span < median_width * 0.30:
                continue
            if abs(y2 - y1) > max(4.0, span * 0.18):
                continue
            low_x, high_x = sorted((x1, x2))
            for column, center_x in enumerate(centers_x):
                if low_x - 3 <= center_x <= high_x + 3:
                    ratio = (center_x - x1) / (x2 - x1) if x2 != x1 else 0.5
                    hough_y[column].append(y1 + ratio * (y2 - y1))

    samples: list[float] = []
    for column, center_x in enumerate(centers_x):
        reference = (
            float(np.median(hough_y[column]))
            if hough_y[column]
            else float(nominal_y)
        )
        local_top = max(strip_top, round(reference) - search_radius)
        local_bottom = min(strip_bottom, round(reference) + search_radius + 1)
        half_width = max(8, round(median_width * 0.42))
        left = max(0, round(center_x) - half_width)
        right = min(width, round(center_x) + half_width + 1)
        candidates = list(range(local_top, local_bottom))
        if not candidates or right <= left:
            samples.append(float(reference))
            continue
        scores: list[float] = []
        for y in candidates:
            row = 255 - gray[y, left:right].astype(np.int32)
            darkness = float(np.sum(row))
            continuity = float(np.count_nonzero(row >= 70))
            prior = abs(y - reference) * max(1.0, median_width) * 255 * 0.05
            scores.append(darkness + continuity * 255 * 1.4 - prior)
        samples.append(float(candidates[int(np.argmax(scores))]))

    values = np.asarray(samples, dtype=float)
    local_median = cv2.medianBlur(values.astype(np.float32)[None, :], 3).reshape(-1)
    bad = np.abs(values - local_median) > max(4.0, search_radius * 0.38)
    values[bad] = local_median[bad]
    slope_cap = max(3.0, search_radius * 0.35)
    for _ in range(3):
        for column in range(1, EXPECTED_COLUMN_COUNT):
            values[column] = np.clip(
                values[column],
                values[column - 1] - slope_cap,
                values[column - 1] + slope_cap,
            )
        for column in range(EXPECTED_COLUMN_COUNT - 2, -1, -1):
            values[column] = np.clip(
                values[column],
                values[column + 1] - slope_cap,
                values[column + 1] + slope_cap,
            )
    return values.astype(float).tolist()


def boundary_values_at_rules(
    geometry: PageGeometry, values: list[float], fallback: float
) -> list[float]:
    if values and geometry.column_center_x:
        return curve_values_at_rules(geometry, values)
    return [float(fallback)] * len(geometry.vertical_rules)


def fit_statistics_separator_samples(
    gray: np.ndarray,
    page_ink: np.ndarray,
    geometry: PageGeometry,
    centers_x: list[float],
    row_pitch: float,
    lower_bounds: list[float] | None = None,
) -> list[float]:
    """Split the statistics section before any month-row anchoring."""

    search_radius = max(9, round(geometry.height * 0.024), round(row_pitch * 1.15))
    strip_top = max(0, geometry.statistics_top - search_radius)
    strip_bottom = min(geometry.height, geometry.statistics_top + search_radius + 1)
    strip_ink = page_ink[strip_top:strip_bottom]
    median_width = float(np.median(np.diff(geometry.vertical_rules)))
    horizontal = cv2.morphologyEx(
        strip_ink,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(
            cv2.MORPH_RECT, (max(11, round(median_width * 0.34)), 1)
        ),
    )

    # Near-horizontal Hough segments provide a second, independent boundary
    # observation when a bowed separator occupies several local slopes.
    edges = cv2.Canny(gray[strip_top:strip_bottom], 50, 145, apertureSize=3)
    lines = cv2.HoughLinesP(
        edges,
        1,
        np.pi / 360,
        threshold=max(12, round(median_width * 0.16)),
        minLineLength=max(14, round(median_width * 0.42)),
        maxLineGap=max(5, round(median_width * 0.12)),
    )
    hough_y_by_column: list[list[float]] = [
        [] for _ in range(EXPECTED_COLUMN_COUNT)
    ]
    if lines is not None:
        for x1, y1, x2, y2 in np.asarray(lines).reshape(-1, 4):
            x1 = float(x1)
            x2 = float(x2)
            y1 = float(y1 + strip_top)
            y2 = float(y2 + strip_top)
            horizontal_span = abs(x2 - x1)
            if horizontal_span < median_width * 0.35:
                continue
            if abs(y2 - y1) > max(4.0, horizontal_span * 0.16):
                continue
            low_x, high_x = sorted((x1, x2))
            for column, center_x in enumerate(centers_x):
                if low_x - 3 <= center_x <= high_x + 3:
                    ratio = (center_x - x1) / (x2 - x1) if x2 != x1 else 0.5
                    hough_y_by_column[column].append(y1 + ratio * (y2 - y1))

    samples: list[float] = []
    for column in range(EXPECTED_COLUMN_COUNT):
        lower_bound = (
            float(lower_bounds[column])
            if lower_bounds is not None and column < len(lower_bounds)
            else float(strip_top)
        )
        # Several horizontal rules exist in the statistics block.  The daily
        # body's lower edge is the *first* physical rule below the last valid
        # daily glyph, not the darkest rule in the search strip.  Using the
        # median/all-row darkness here used to select the rule below the
        # maximum-value row and made a day-31 crop contain both values.
        eligible_hough = sorted(
            value
            for value in hough_y_by_column[column]
            if value >= lower_bound - 1.5
        )
        reference_y = (
            float(eligible_hough[0])
            if eligible_hough
            else float(geometry.statistics_top)
        )
        search_top = max(strip_top, round(reference_y) - search_radius)
        search_bottom = min(strip_bottom, round(reference_y) + search_radius + 1)
        scores: list[float] = []
        candidate_y = list(range(search_top, search_bottom))
        strong_rule_rows: list[int] = []
        for y in candidate_y:
            left = round(vertical_rule_x(geometry, column, y)) + 3
            right = round(vertical_rule_x(geometry, column + 1, y)) - 3
            left = max(0, left)
            right = min(geometry.width, right)
            if right <= left:
                scores.append(0.0)
                continue
            local_y = y - strip_top
            line_support = float(np.count_nonzero(horizontal[local_y, left:right]))
            ink_support = float(np.count_nonzero(page_ink[y, left:right]))
            width = max(1, right - left)
            darkness = float(
                np.sum(255 - gray[y, left:right].astype(np.int32))
            )
            hough_bonus = (
                max(0.0, 1.0 - abs(y - reference_y) / 3.0)
                * (right - left)
                * 255
                * 0.35
                if hough_y_by_column[column]
                else 0.0
            )
            scores.append(line_support * 255 * 2.2 + darkness + hough_bonus)
            if (
                y >= lower_bound - 1.0
                and (
                    line_support / width >= 0.28
                    or ink_support / width >= 0.68
                )
            ):
                strong_rule_rows.append(y)
        if not scores:
            samples.append(float(geometry.statistics_top))
        elif strong_rule_rows:
            # Select the first consecutive rule band, then the strongest pixel
            # inside that band.  This retains a locally slanted/bowed divider
            # while preventing any later statistics-row rule from winning.
            first_band = [strong_rule_rows[0]]
            for y in strong_rule_rows[1:]:
                if y <= first_band[-1] + 1:
                    first_band.append(y)
                else:
                    break
            samples.append(
                float(
                    max(
                        first_band,
                        key=lambda y: scores[y - search_top],
                    )
                )
            )
        else:
            # With no sufficiently long physical rule, keep the estimate near
            # the initial table boundary instead of drifting to darker digits.
            adjusted = [
                score
                - abs(y - geometry.statistics_top)
                * max(1.0, median_width)
                * 255.0
                * 0.22
                for y, score in zip(candidate_y, scores)
            ]
            samples.append(float(candidate_y[int(np.argmax(adjusted))]))
    return samples


def detect_month_decimal_track(
    page_ink: np.ndarray,
    geometry: PageGeometry,
    column: int,
    row_pitch: float,
    body_bottom: int,
) -> tuple[tuple[float, float] | None, list[dict[str, float]]]:
    """Find the repeated decimal-point track inside one flow column.

    This follows the proven water-level anchor idea in
    ``flood_hydro_factor_pdf_ex2.py``: dot-sized components are first found
    independently, then only components agreeing with one slightly slanted
    x(y) track survive.  Requiring a repeated track prevents isolated dirt or
    a broken digit stroke from becoming a date anchor.
    """

    body_top = (
        max(
            geometry.header_bottom + 2,
            round(geometry.header_curve[column]) + 2,
        )
        if geometry.header_curve and column < len(geometry.header_curve)
        else geometry.header_bottom + 2
    )
    boundary_samples = [float(body_top), float(body_bottom - 1)]
    left = max(
        0,
        int(
            np.floor(
                min(vertical_rule_x(geometry, column, y) for y in boundary_samples)
            )
        ),
    )
    right = min(
        geometry.width,
        int(
            np.ceil(
                max(
                    vertical_rule_x(geometry, column + 1, y)
                    for y in boundary_samples
                )
            )
        )
        + 1,
    )
    width = max(
        1.0,
        vertical_rule_x(geometry, column + 1, (body_top + body_bottom) / 2)
        - vertical_rule_x(geometry, column, (body_top + body_bottom) / 2),
    )
    source = page_ink[body_top:body_bottom, left:right].copy()
    for local_y, absolute_y in enumerate(range(body_top, body_bottom)):
        allowed_left = round(vertical_rule_x(geometry, column, absolute_y)) + 5
        allowed_right = (
            round(vertical_rule_x(geometry, column + 1, absolute_y)) - 5
        )
        source[local_y, : max(0, allowed_left - left)] = 0
        source[local_y, max(0, allowed_right - left) :] = 0
    count, _, stats, centroids = cv2.connectedComponentsWithStats(source, 8)
    candidates: list[dict[str, float]] = []
    maximum_side = max(4, round(row_pitch * 0.27))
    maximum_area = max(18, round(row_pitch * row_pitch * 0.055))
    for component_index in range(1, count):
        x, y, component_width, component_height, area = map(
            int, stats[component_index]
        )
        center_x, center_y = map(float, centroids[component_index])
        fill_ratio = area / max(1, component_width * component_height)
        if not (
            2 <= component_width <= maximum_side
            and 1 <= component_height <= maximum_side
            and 3 <= area <= maximum_area
            and 0.32 <= fill_ratio <= 1.0
            and component_width <= component_height * 2.4
            and component_height <= component_width * 2.4
        ):
            continue
        absolute_y = center_y + body_top
        dynamic_left = vertical_rule_x(geometry, column, absolute_y) + 5
        dynamic_right = vertical_rule_x(geometry, column + 1, absolute_y) - 5
        dynamic_width = max(1.0, dynamic_right - dynamic_left)
        absolute_x = center_x + left
        relative_x = (absolute_x - dynamic_left) / dynamic_width
        if not 0.18 <= relative_x <= 0.88:
            continue
        candidates.append(
            {
                "x": absolute_x,
                "y": absolute_y,
                "area": float(area),
                "fill": float(fill_ratio),
            }
        )
    if len(candidates) < 2:
        return None, []

    # Search a narrow family of vertical/slightly slanted tracks.  Counting
    # distinct y bands, rather than raw components, prevents several fragments
    # from the same glyph winning the vote.
    tolerance = max(1.5, row_pitch * 0.095)
    slope_limit = max(0.012, width * 0.075 / max(1.0, body_bottom - body_top))
    best_score = -1.0
    best_inliers: list[dict[str, float]] = []
    y_origin = float(body_top)
    slopes = np.linspace(-slope_limit, slope_limit, 51)
    for slope in slopes:
        intercept_values = [
            item["x"] - slope * (item["y"] - y_origin)
            for item in candidates
        ]
        for seed in intercept_values:
            inliers = [
                item
                for item, intercept in zip(candidates, intercept_values)
                if abs(intercept - seed) <= tolerance
            ]
            ordered = sorted(inliers, key=lambda item: item["y"])
            distinct: list[dict[str, float]] = []
            for item in ordered:
                if (
                    not distinct
                    or item["y"] - distinct[-1]["y"] >= row_pitch * 0.42
                ):
                    distinct.append(item)
                elif item["fill"] * item["area"] > (
                    distinct[-1]["fill"] * distinct[-1]["area"]
                ):
                    distinct[-1] = item
            score = sum(1.0 + item["fill"] * 0.22 for item in distinct)
            if score > best_score:
                best_score = score
                best_inliers = distinct
    if len(best_inliers) < 2:
        return None, []

    y_values = np.asarray([item["y"] for item in best_inliers], dtype=float)
    x_values = np.asarray([item["x"] for item in best_inliers], dtype=float)
    slope, absolute_intercept = np.polyfit(y_values, x_values, 1)
    residuals = np.abs(x_values - (slope * y_values + absolute_intercept))
    keep = residuals <= tolerance
    if int(np.count_nonzero(keep)) < 2:
        return None, []
    slope, absolute_intercept = np.polyfit(y_values[keep], x_values[keep], 1)
    points = [
        item for item, retained in zip(best_inliers, keep.tolist()) if retained
    ]
    return (float(slope), float(absolute_intercept)), points


def align_observed_sequence_to_days(
    observations: list[dict[str, float]],
    base_centers: np.ndarray,
    preliminary: np.ndarray,
    printed_days: int,
    row_pitch: float,
    offset_range: tuple[float, float],
) -> tuple[dict[int, dict[str, float]], float] | None:
    """Align a sparse observed sequence against the exact day-column rhythm.

    Some source tables intentionally insert an extra vertical gap after every
    five dates.  Therefore a 48-pixel gap is not necessarily a missing day.
    Dynamic programming compares every observed gap with the corresponding
    *actual day-column anchor gap*, preserving such layout rhythms and making
    skipped decimal points local rather than cumulative.
    """

    ordered = sorted(observations, key=lambda item: item["y"])
    observation_count = len(ordered)
    if observation_count < 2:
        return None
    # Only a small excess is credible.  A denser track is more likely a digit
    # edge, ruled line, or halftone rhythm than a set of decimal points.
    if observation_count > printed_days + 2:
        return None
    y_values = np.asarray([item["y"] for item in ordered], dtype=float)
    infinity = float("inf")
    offset_min, offset_max = offset_range

    def offset_penalty(value: float) -> float:
        if value < offset_min:
            return (offset_min - value) / row_pitch
        if value > offset_max:
            return (value - offset_max) / row_pitch
        return 0.0

    def finalize_path(
        observation_path: list[int], day_path: list[int]
    ) -> tuple[dict[int, dict[str, float]], float] | None:
        if len(day_path) < 2:
            return None
        transition_errors: list[float] = []
        for index in range(1, len(day_path)):
            observed_gap = (
                y_values[observation_path[index]]
                - y_values[observation_path[index - 1]]
            )
            expected_gap = (
                base_centers[day_path[index]]
                - base_centers[day_path[index - 1]]
            )
            date_span = max(1, day_path[index] - day_path[index - 1])
            transition_errors.append(
                abs(observed_gap - expected_gap) / date_span
            )
        if float(np.median(transition_errors)) > row_pitch * 0.34:
            return None
        mapping = {
            int(day_index): ordered[observation_index]
            for observation_index, day_index in zip(
                observation_path, day_path
            )
        }
        raw_offset = float(
            np.median(
                [
                    item["y"] - preliminary[day_index]
                    for day_index, item in mapping.items()
                ]
            )
        )
        return mapping, raw_offset

    if observation_count <= printed_days:
        # Preserve the proven sparse-sequence solver exactly: when the track is
        # not overfull, every observation is evidence and may skip only dates.
        costs = np.full(
            (observation_count, printed_days), infinity, dtype=float
        )
        previous = np.full(
            (observation_count, printed_days), -1, dtype=int
        )
        for day_index in range(printed_days):
            residual = y_values[0] - preliminary[day_index]
            costs[0, day_index] = (
                day_index * 0.24 + offset_penalty(residual) * 2.2
            )
        for observation_index in range(1, observation_count):
            observed_gap = (
                y_values[observation_index]
                - y_values[observation_index - 1]
            )
            for day_index in range(observation_index, printed_days):
                residual = y_values[observation_index] - preliminary[day_index]
                point_cost = offset_penalty(residual) * 1.25
                for prior_day in range(observation_index - 1, day_index):
                    prior_cost = costs[observation_index - 1, prior_day]
                    if not np.isfinite(prior_cost):
                        continue
                    expected_gap = (
                        base_centers[day_index] - base_centers[prior_day]
                    )
                    gap_error = abs(observed_gap - expected_gap) / row_pitch
                    skipped_days = day_index - prior_day - 1
                    candidate = (
                        prior_cost
                        + gap_error * 2.35
                        + skipped_days * 0.27
                        + point_cost
                    )
                    if candidate < costs[observation_index, day_index]:
                        costs[observation_index, day_index] = candidate
                        previous[observation_index, day_index] = prior_day
        last_costs = costs[-1].copy()
        for day_index in range(printed_days):
            last_costs[day_index] += (
                printed_days - day_index - 1
            ) * 0.20
        last_day = int(np.argmin(last_costs))
        if not np.isfinite(last_costs[last_day]):
            return None
        day_path = [last_day]
        for observation_index in range(observation_count - 1, 0, -1):
            last_day = int(previous[observation_index, last_day])
            if last_day < 0:
                return None
            day_path.append(last_day)
        day_path.reverse()
        return finalize_path(list(range(observation_count)), day_path)

    # Overfull tracks use a bounded third state: the number of rejected
    # observations.  Existing sparse behaviour remains untouched, while up to
    # two statistics/dirt points can be discarded anywhere in the sequence.
    maximum_rejections = min(3, observation_count - 2)
    minimum_rejections = observation_count - printed_days
    rejection_penalty = 0.95
    costs = np.full(
        (
            observation_count,
            printed_days,
            maximum_rejections + 1,
        ),
        infinity,
        dtype=float,
    )
    previous_observation = np.full(costs.shape, -1, dtype=int)
    previous_day = np.full(costs.shape, -1, dtype=int)
    previous_rejections = np.full(costs.shape, -1, dtype=int)
    for observation_index in range(
        min(observation_count, maximum_rejections + 1)
    ):
        rejected = observation_index
        for day_index in range(printed_days):
            residual = y_values[observation_index] - preliminary[day_index]
            costs[observation_index, day_index, rejected] = (
                rejected * rejection_penalty
                + day_index * 0.24
                + offset_penalty(residual) * 2.2
            )
    for observation_index in range(1, observation_count):
        for day_index in range(1, printed_days):
            residual = y_values[observation_index] - preliminary[day_index]
            point_cost = offset_penalty(residual) * 1.25
            for prior_observation in range(observation_index):
                rejected_between = observation_index - prior_observation - 1
                if rejected_between > maximum_rejections:
                    continue
                observed_gap = (
                    y_values[observation_index]
                    - y_values[prior_observation]
                )
                for prior_day_index in range(day_index):
                    expected_gap = (
                        base_centers[day_index]
                        - base_centers[prior_day_index]
                    )
                    gap_error = abs(observed_gap - expected_gap) / row_pitch
                    skipped_days = day_index - prior_day_index - 1
                    for prior_rejected in range(
                        maximum_rejections - rejected_between + 1
                    ):
                        prior_cost = costs[
                            prior_observation,
                            prior_day_index,
                            prior_rejected,
                        ]
                        if not np.isfinite(prior_cost):
                            continue
                        rejected = prior_rejected + rejected_between
                        candidate = (
                            prior_cost
                            + rejected_between * rejection_penalty
                            + gap_error * 2.35
                            + skipped_days * 0.27
                            + point_cost
                        )
                        if candidate < costs[
                            observation_index, day_index, rejected
                        ]:
                            costs[
                                observation_index, day_index, rejected
                            ] = candidate
                            previous_observation[
                                observation_index, day_index, rejected
                            ] = prior_observation
                            previous_day[
                                observation_index, day_index, rejected
                            ] = prior_day_index
                            previous_rejections[
                                observation_index, day_index, rejected
                            ] = prior_rejected

    # Prefer the smallest rejection count that produces a geometrically valid
    # path.  This maximizes retained real dots before comparing score details.
    for total_rejections in range(
        minimum_rejections, maximum_rejections + 1
    ):
        endpoints: list[tuple[float, int, int, int]] = []
        for observation_index in range(observation_count):
            trailing = observation_count - observation_index - 1
            rejected_before = total_rejections - trailing
            if not 0 <= rejected_before <= maximum_rejections:
                continue
            for day_index in range(printed_days):
                cost = costs[
                    observation_index, day_index, rejected_before
                ]
                if not np.isfinite(cost):
                    continue
                endpoints.append(
                    (
                        float(
                            cost
                            + trailing * rejection_penalty
                            + (printed_days - day_index - 1) * 0.20
                        ),
                        observation_index,
                        day_index,
                        rejected_before,
                    )
                )
        for _, last_observation, last_day, rejected in sorted(endpoints):
            observation_path = [last_observation]
            day_path = [last_day]
            while True:
                prior_observation = int(
                    previous_observation[last_observation, last_day, rejected]
                )
                prior_day_index = int(
                    previous_day[last_observation, last_day, rejected]
                )
                prior_rejected = int(
                    previous_rejections[last_observation, last_day, rejected]
                )
                if (
                    prior_observation < 0
                    or prior_day_index < 0
                    or prior_rejected < 0
                ):
                    break
                observation_path.append(prior_observation)
                day_path.append(prior_day_index)
                last_observation = prior_observation
                last_day = prior_day_index
                rejected = prior_rejected
            observation_path.reverse()
            day_path.reverse()
            finalized = finalize_path(observation_path, day_path)
            if finalized is not None:
                return finalized
    return None


def detect_integer_token_centers_left_of_decimal(
    page_ink: np.ndarray,
    geometry: PageGeometry,
    column: int,
    decimal_model: tuple[float, float],
    row_pitch: float,
    body_bottom: int,
) -> list[dict[str, float]]:
    """Find integer/glyph rows strictly left of the fitted decimal track."""

    body_top = geometry.header_bottom + 2
    boundary_samples = [float(body_top), float(body_bottom - 1)]
    left = max(
        0,
        int(
            np.floor(
                min(vertical_rule_x(geometry, column, y) for y in boundary_samples)
            )
        ),
    )
    right = min(
        geometry.width,
        int(
            np.ceil(
                max(
                    vertical_rule_x(geometry, column + 1, y)
                    for y in boundary_samples
                )
            )
        )
        + 1,
    )
    masked = np.zeros((body_bottom - body_top, right - left), dtype=np.uint8)
    slope, intercept = decimal_model
    for local_y, absolute_y in enumerate(range(body_top, body_bottom)):
        decimal_x = round(slope * absolute_y + intercept)
        dynamic_left = round(vertical_rule_x(geometry, column, absolute_y)) + 6
        dynamic_right = (
            round(vertical_rule_x(geometry, column + 1, absolute_y)) - 6
        )
        start = min(dynamic_right, max(left, dynamic_left))
        limit = min(dynamic_right, right, max(start, decimal_x - 2))
        if limit > start:
            masked[local_y, start - left : limit - left] = page_ink[
                absolute_y, start:limit
            ]
    projection = np.count_nonzero(masked, axis=1).astype(float)
    active = (projection >= 1).astype(np.uint8)[:, None]
    active = cv2.morphologyEx(
        active,
        cv2.MORPH_CLOSE,
        cv2.getStructuringElement(cv2.MORPH_RECT, (1, 3)),
    ).reshape(-1).astype(bool)
    bands = active_bands(active, body_top)
    tokens: list[dict[str, float]] = []
    for top, bottom in bands:
        height = bottom - top + 1
        values = projection[top - body_top : bottom - body_top + 1]
        total = float(np.sum(values))
        if not (2 <= height <= row_pitch * 0.88 and total >= 5):
            continue
        ys = np.arange(top, bottom + 1, dtype=float)
        center_y = float(np.average(ys, weights=values))
        tokens.append(
            {
                "y": center_y,
                "top": float(top),
                "bottom": float(bottom),
                "ink": total,
            }
        )
    return tokens


def inspect_decimal_slot_occupancy(
    page_ink: np.ndarray,
    geometry: PageGeometry,
    column: int,
    decimal_model: tuple[float, float],
    center_y: float,
    row_pitch: float,
) -> dict[str, float | bool | str]:
    """Classify 119 versus 11.9 from the fitted slot and ink dispersion."""

    top = max(0, round(center_y - row_pitch * 0.40))
    bottom = min(geometry.height, round(center_y + row_pitch * 0.43) + 1)
    if bottom <= top:
        return {
            "layout": "ambiguous",
            "slot_ink": 0.0,
            "left_ink": 0.0,
            "right_ink": 0.0,
            "dispersion": 0.0,
        }
    decimal_x = decimal_model[0] * center_y + decimal_model[1]
    left = max(0, round(vertical_rule_x(geometry, column, center_y)) + 6)
    right = min(
        geometry.width,
        round(vertical_rule_x(geometry, column + 1, center_y)) - 6,
    )
    slot_half_width = max(2, round((right - left) * 0.018))
    slot_left = max(left, round(decimal_x) - slot_half_width)
    slot_right = min(right, round(decimal_x) + slot_half_width + 1)
    left_end = max(left, slot_left - 1)
    right_start = min(right, slot_right + 1)
    left_ink = float(np.count_nonzero(page_ink[top:bottom, left:left_end]))
    slot_ink = float(
        np.count_nonzero(page_ink[top:bottom, slot_left:slot_right])
    )
    right_ink = float(
        np.count_nonzero(page_ink[top:bottom, right_start:right])
    )
    ys, xs = np.nonzero(page_ink[top:bottom, left:right])
    del ys
    dispersion = (
        float(np.std(xs) / max(1.0, right - left)) if xs.size else 0.0
    )
    right_ratio = right_ink / max(1.0, left_ink + slot_ink + right_ink)
    slot_ratio = slot_ink / max(1.0, left_ink + slot_ink + right_ink)
    has_fraction_side = right_ink >= max(4.0, left_ink * 0.055)
    has_decimal_slot = slot_ink >= 2 and slot_ratio >= 0.012
    integer_left_only = left_ink >= 5 and right_ratio <= 0.025 and slot_ink <= 2
    if has_decimal_slot and has_fraction_side:
        layout = "decimal-slot-and-fraction"
    elif has_fraction_side and dispersion >= 0.115:
        layout = "faint-decimal-dispersed-fraction"
    elif integer_left_only:
        layout = "integer-left-only"
    else:
        layout = "ambiguous"
    return {
        "layout": layout,
        "slot_ink": slot_ink,
        "left_ink": left_ink,
        "right_ink": right_ink,
        "right_ink_ratio": float(right_ratio),
        "slot_ink_ratio": float(slot_ratio),
        "dispersion": dispersion,
    }


def apply_month_decimal_anchors(
    image: np.ndarray,
    page_ink: np.ndarray,
    geometry: PageGeometry,
    base_centers: np.ndarray,
    fitted: np.ndarray,
    row_pitch: float,
    body_bottom: int,
    statistics_samples: list[float],
) -> np.ndarray:
    """Strengthen every month with table-consensus dots and left tokens."""

    del image  # Kept in the signature for future grayscale diagnostics.
    diagnostics: list[list[dict[str, Any]]] = [
        [] for _ in range(EXPECTED_COLUMN_COUNT)
    ]
    virtual_diagnostics: list[dict[str, Any]] = []

    def maximum_printed_day(column: int) -> int:
        if 0 < column <= len(geometry.month_lengths):
            inferred = int(geometry.month_lengths[column - 1])
            if 28 <= inferred <= 31:
                return inferred
        if column == 2:
            return 29
        if column in (4, 6, 9, 11):
            return 30
        return 31

    def set_calendar_virtual_anchors(column: int, printed_days: int) -> None:
        for day_index in range(printed_days, EXPECTED_DAY_COUNT):
            layout_gap = base_centers[day_index] - base_centers[day_index - 1]
            fitted[day_index, column] = (
                fitted[day_index - 1, column] + layout_gap
            )
            virtual_diagnostics.append(
                {
                    "month": column,
                    "day": day_index + 1,
                    "row_center_y": float(fitted[day_index, column]),
                    "source": "calendar-layout-extrapolation",
                }
            )

    body_midpoint = (geometry.header_bottom + 2 + body_bottom) / 2
    body_span = max(1.0, body_bottom - geometry.header_bottom - 2)
    raw_tracks: dict[int, dict[str, Any]] = {}
    for column in range(1, EXPECTED_COLUMN_COUNT):
        column_body_bottom = min(
            body_bottom, round(statistics_samples[column]) - 2
        )
        model, points = detect_month_decimal_track(
            page_ink, geometry, column, row_pitch, column_body_bottom
        )
        if model is None or len(points) < 2:
            continue
        left = vertical_rule_x(geometry, column, body_midpoint) + 5
        right = vertical_rule_x(geometry, column + 1, body_midpoint) - 5
        width = max(1.0, right - left)
        midpoint_x = model[0] * body_midpoint + model[1]
        raw_tracks[column] = {
            "model": model,
            "points": points,
            "ratio": (midpoint_x - left) / width,
            "normalized_slope": model[0] * body_span / width,
            "weight": float(len(points)),
            "body_bottom": int(column_body_bottom),
        }
    if not raw_tracks:
        for column in range(1, EXPECTED_COLUMN_COUNT):
            set_calendar_virtual_anchors(
                column, maximum_printed_day(column)
            )
        geometry.month_decimal_anchors = diagnostics
        geometry.virtual_month_anchors = virtual_diagnostics
        return fitted

    def weighted_median(values: list[float], weights: list[float]) -> float:
        order = np.argsort(np.asarray(values, dtype=float))
        sorted_values = np.asarray(values, dtype=float)[order]
        sorted_weights = np.asarray(weights, dtype=float)[order]
        cutoff = float(np.sum(sorted_weights)) / 2
        index = int(np.searchsorted(np.cumsum(sorted_weights), cutoff))
        return float(sorted_values[min(index, len(sorted_values) - 1)])

    # Decimal alignment x can legitimately move when one month contains much
    # larger values than another.  Do not force a single x ratio.  A track is
    # trusted when it covers a meaningful portion of that month and its y
    # sequence matches the exact printed day-column rhythm.
    accepted_tracks: dict[int, dict[str, Any]] = {}
    validated_tracks: dict[int, dict[str, Any]] = {}
    dot_alignments: dict[int, dict[int, dict[str, float]]] = {}
    raw_dot_offsets: list[float] = []
    for column, item in raw_tracks.items():
        printed_days = maximum_printed_day(column)
        aligned = align_observed_sequence_to_days(
            item["points"],
            base_centers,
            fitted[:, column],
            printed_days,
            row_pitch,
            (row_pitch * 0.05, row_pitch * 0.48),
        )
        if aligned is None:
            continue
        mapping, raw_offset = aligned
        if len(mapping) < 2:
            continue
        validated_tracks[column] = item
        dot_alignments[column] = mapping
        minimum_dot_coverage = max(10, round(printed_days * 0.34))
        if len(mapping) < minimum_dot_coverage:
            continue
        accepted_tracks[column] = item
        raw_dot_offsets.append(raw_offset)

    synthesis_tracks = (
        list(accepted_tracks.values())
        if len(accepted_tracks) >= 2
        else list(raw_tracks.values())
    )
    consensus_ratio = weighted_median(
        [float(item["ratio"]) for item in synthesis_tracks],
        [float(item["weight"]) for item in synthesis_tracks],
    )
    consensus_slope = weighted_median(
        [float(item["normalized_slope"]) for item in synthesis_tracks],
        [float(item["weight"]) for item in synthesis_tracks],
    )
    if raw_dot_offsets:
        dot_offset = float(np.median(raw_dot_offsets))
        dot_offset = min(row_pitch * 0.45, max(row_pitch * 0.08, dot_offset))
    else:
        dot_offset = row_pitch * 0.25

    for column in range(1, EXPECTED_COLUMN_COUNT):
        printed_days = maximum_printed_day(column)
        left = geometry.vertical_rules[column] + 5
        right = geometry.vertical_rules[column + 1] - 5
        width = max(1.0, right - left)
        if column in validated_tracks:
            decimal_model = validated_tracks[column]["model"]
        else:
            slope = consensus_slope * width / body_span
            midpoint_x = left + consensus_ratio * width
            decimal_model = (
                float(slope),
                float(midpoint_x - slope * body_midpoint),
            )

        dot_mapping = dot_alignments.get(column, {})
        tokens = detect_integer_token_centers_left_of_decimal(
            page_ink,
            geometry,
            column,
            decimal_model,
            row_pitch,
            min(body_bottom, round(statistics_samples[column]) - 2),
        )
        token_alignment = align_observed_sequence_to_days(
            tokens,
            base_centers,
            fitted[:, column],
            printed_days,
            row_pitch,
            (-row_pitch * 0.34, row_pitch * 0.34),
        )
        token_mapping = token_alignment[0] if token_alignment is not None else {}

        observed: dict[
            int, tuple[float, str, float, float, dict[str, Any]]
        ] = {}
        for day_index, point in dot_mapping.items():
            center_y = float(point["y"] - dot_offset)
            occupancy = inspect_decimal_slot_occupancy(
                page_ink,
                geometry,
                column,
                decimal_model,
                center_y,
                row_pitch,
            )
            observed[day_index] = (
                center_y,
                "decimal-dot",
                float(point["x"]),
                float(point["y"]),
                occupancy,
            )
        for day_index, token in token_mapping.items():
            if day_index in observed:
                # The two independent mechanisms must agree on the same row.
                # A disagreement rejects the weaker token, never the date.
                if abs(token["y"] - observed[day_index][0]) > row_pitch * 0.36:
                    continue
                continue
            decimal_x = decimal_model[0] * token["y"] + decimal_model[1]
            occupancy = inspect_decimal_slot_occupancy(
                page_ink,
                geometry,
                column,
                decimal_model,
                float(token["y"]),
                row_pitch,
            )
            if occupancy["layout"] == "integer-left-only":
                kind = "integer-left-of-decimal-track"
            elif occupancy["layout"] in (
                "decimal-slot-and-fraction",
                "faint-decimal-dispersed-fraction",
            ):
                kind = "decimal-layout-occupancy"
            elif (
                any(value < day_index for value in dot_mapping)
                and any(value > day_index for value in dot_mapping)
            ):
                kind = "integer-between-decimal-anchors"
            else:
                kind = "left-token-via-table-decimal-consensus"
            observed[day_index] = (
                float(token["y"]),
                kind,
                float(decimal_x),
                float(token["y"]),
                occupancy,
            )

        observed_days = sorted(observed)
        if len(observed_days) < 3:
            set_calendar_virtual_anchors(column, printed_days)
            continue
        observed_values = np.asarray(
            [observed[index][0] for index in observed_days], dtype=float
        )
        observed_displacement = (
            observed_values - base_centers[np.asarray(observed_days, dtype=int)]
        )
        interpolated_displacement = np.interp(
            np.arange(EXPECTED_DAY_COUNT, dtype=float),
            np.asarray(observed_days, dtype=float),
            observed_displacement,
        )
        smooth_displacement = cv2.GaussianBlur(
            interpolated_displacement[:, None].astype(np.float32),
            (1, 5),
            sigmaX=0.0,
            sigmaY=0.75,
        ).reshape(-1).astype(float)
        consensus_curve = base_centers + smooth_displacement
        first_observed = observed_days[0]
        last_observed = observed_days[-1]
        for day_index in range(printed_days):
            if day_index in observed:
                fitted[day_index, column] = (
                    observed[day_index][0] * 0.84
                    + consensus_curve[day_index] * 0.16
                )
            elif first_observed < day_index < last_observed:
                fitted[day_index, column] = (
                    consensus_curve[day_index] * 0.72
                    + fitted[day_index, column] * 0.28
                )

        # Calendar-invalid trailing rows are geometry-only control points.
        # They must never attach to ink.  Continue the exact day-column layout
        # gap from the last real date so the final valid cell still receives a
        # stable lower boundary (Feb 30/31 and day 31 in 30-day months).
        set_calendar_virtual_anchors(column, printed_days)

        # A valid daily center cannot cross the statistics boundary.  This is
        # a form-level invariant, not a station-specific exception.
        maximum_center = body_bottom - max(2.0, row_pitch * 0.12)
        fitted[:printed_days, column] = np.minimum(
            fitted[:printed_days, column], maximum_center
        )
        for day_index in observed_days:
            center_y, method, point_x, point_y, occupancy = observed[day_index]
            diagnostics[column].append(
                {
                    "month": column,
                    "day": day_index + 1,
                    "kind": method,
                    "point_x": float(point_x),
                    "point_y": float(point_y),
                    "row_center_y": float(center_y),
                    "decimal_x": float(
                        decimal_model[0] * point_y + decimal_model[1]
                    ),
                    "dot_to_center_offset": float(dot_offset),
                    "slot_layout": str(occupancy["layout"]),
                    "slot_ink": float(occupancy["slot_ink"]),
                    "left_ink": float(occupancy["left_ink"]),
                    "right_ink": float(occupancy["right_ink"]),
                    "ink_dispersion": float(occupancy["dispersion"]),
                }
            )

    geometry.month_decimal_anchors = diagnostics
    geometry.virtual_month_anchors = virtual_diagnostics
    if any(diagnostics[month] for month in range(1, EXPECTED_COLUMN_COUNT)):
        if "month-decimal" not in geometry.anchor_method:
            geometry.anchor_method += "+month-decimal-and-integer"
    return fitted


def fit_multicolumn_row_curves(
    image: np.ndarray, geometry: PageGeometry
) -> None:
    """Independently re-anchor every month column to its local ink rows.

    A column-wide correlation first estimates only that column's coarse shift.
    Three local refinement passes then attach each ordered day to a real ink
    band inside a non-overlapping row window.  No neighbouring-column result or
    statistics boundary can move the search window, so drift cannot accumulate.
    """

    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    page_ink = cv2.threshold(
        gray, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU
    )[1]
    centers_x = [
        (geometry.vertical_rules[index] + geometry.vertical_rules[index + 1])
        / 2
        for index in range(EXPECTED_COLUMN_COUNT)
    ]
    base_centers = np.asarray(
        [anchor.center for anchor in geometry.row_anchors], dtype=float
    )
    ordinary_gaps = np.diff(base_centers)
    ordinary_gaps = ordinary_gaps[
        ordinary_gaps <= np.percentile(ordinary_gaps, 70)
    ]
    row_pitch = float(np.median(ordinary_gaps)) if ordinary_gaps.size else 10.0
    row_pitch = max(4.0, row_pitch)
    fit_dynamic_vertical_rule_models(image, geometry, row_pitch)
    geometry.column_center_x = [
        (
            vertical_rule_x(geometry, column, float(np.median(base_centers)))
            + vertical_rule_x(geometry, column + 1, float(np.median(base_centers)))
        )
        / 2
        for column in range(EXPECTED_COLUMN_COUNT)
    ]
    boundary_radius = max(7, round(row_pitch * 0.75))
    geometry.table_top_curve = fit_horizontal_boundary_samples(
        gray,
        geometry,
        geometry.column_center_x,
        float(geometry.table_top),
        boundary_radius,
    )
    geometry.header_curve = fit_horizontal_boundary_samples(
        gray,
        geometry,
        geometry.column_center_x,
        float(geometry.header_bottom),
        boundary_radius,
    )
    # Daily ink must never borrow evidence from the first statistics row.
    # This is especially important for day 31, which has no following date to
    # close its search window.
    # This is only a generous global ceiling.  Each column is subsequently
    # bounded by its independently tracked statistics separator curve.
    body_bottom = min(
        geometry.height - 2,
        max(
            round(base_centers[-1] + row_pitch * 1.45),
            round(geometry.statistics_top + row_pitch * 1.20),
        ),
    )
    preliminary_lower_bounds = [
        float(base_centers[29 if column in (4, 6, 9, 11) else 30])
        + max(3.0, row_pitch * 0.42)
        for column in range(EXPECTED_COLUMN_COUNT)
    ]
    preliminary_lower_bounds[2] = (
        float(base_centers[28]) + max(3.0, row_pitch * 0.42)
    )
    statistics_samples = fit_statistics_separator_samples(
        gray,
        page_ink,
        geometry,
        geometry.column_center_x,
        row_pitch,
        preliminary_lower_bounds,
    )

    observations = np.full(
        (EXPECTED_DAY_COUNT, EXPECTED_COLUMN_COUNT), np.nan, dtype=float
    )
    weights = np.zeros_like(observations)
    observations[:, 0] = base_centers
    weights[:, 0] = 5.0

    def maximum_printed_day(column: int) -> int:
        if 0 < column <= len(geometry.month_lengths):
            inferred = int(geometry.month_lengths[column - 1])
            if 28 <= inferred <= 31:
                return inferred
        if column == 2:
            return 29
        if column in (4, 6, 9, 11):
            return 30
        return 31

    def refined_statistics_separator(
        current_centers: np.ndarray,
    ) -> list[float]:
        """Fit the daily/statistics divider from the current month surface.

        The first separator estimate is intentionally permissive because no
        month-local row surface exists yet.  Once one does, the last legal date
        supplies an independent lower bound for every column.  Running this
        refinement before decimal anchoring keeps the first statistics row out
        of the dot track; running it again after anchoring keeps crop geometry
        synchronized with the final surface.
        """

        separator_lower_bounds: list[float] = []
        for column in range(EXPECTED_COLUMN_COUNT):
            printed_days = 31 if column == 0 else maximum_printed_day(column)
            last_center = current_centers[printed_days - 1, column]
            ink_lower_bound = last_center + max(3.0, row_pitch * 0.42)
            separator_lower_bounds.append(float(ink_lower_bound))
        samples = fit_statistics_separator_samples(
            gray,
            page_ink,
            geometry,
            centers_x,
            row_pitch,
            separator_lower_bounds,
        )
        for column in range(EXPECTED_COLUMN_COUNT):
            samples[column] = min(
                geometry.height - 2.0,
                max(float(samples[column]), separator_lower_bounds[column]),
            )
        separator = np.asarray(samples, dtype=float)
        lower_bounds = np.asarray(separator_lower_bounds, dtype=float)
        local_median = cv2.medianBlur(
            separator.astype(np.float32)[None, :], 3
        ).reshape(-1).astype(float)
        suspicious = np.abs(separator - local_median) > row_pitch * 0.42
        separator[suspicious] = np.maximum(
            lower_bounds[suspicious], local_median[suspicious]
        )
        separator_cap = max(3.0, row_pitch * 0.42)
        for _ in range(3):
            forward = separator.copy()
            backward = separator.copy()
            for column in range(1, EXPECTED_COLUMN_COUNT):
                forward[column] = max(
                    lower_bounds[column],
                    float(
                        np.clip(
                            forward[column],
                            forward[column - 1] - separator_cap,
                            forward[column - 1] + separator_cap,
                        )
                    ),
                )
            for column in range(EXPECTED_COLUMN_COUNT - 2, -1, -1):
                backward[column] = max(
                    lower_bounds[column],
                    float(
                        np.clip(
                            backward[column],
                            backward[column + 1] - separator_cap,
                            backward[column + 1] + separator_cap,
                        )
                    ),
                )
            separator = np.maximum(lower_bounds, (forward + backward) / 2)
        return separator.astype(float).tolist()

    for column in range(1, EXPECTED_COLUMN_COUNT):
        reference_y = float(np.median(base_centers))
        left = vertical_rule_x(geometry, column, reference_y) + 6
        right = vertical_rule_x(geometry, column + 1, reference_y) - 6
        column_width = max(1, round(right - left))
        column_body_top = max(
            geometry.header_bottom + 2,
            round(geometry.header_curve[column]) + 2
            if geometry.header_curve
            else geometry.header_bottom + 2,
        )
        column_body_bottom = min(body_bottom, round(statistics_samples[column]) - 2)
        projection = np.zeros(geometry.height, dtype=np.int32)
        for y in range(column_body_top, max(column_body_top, column_body_bottom)):
            dynamic_left = max(
                0, round(vertical_rule_x(geometry, column, y)) + 6
            )
            dynamic_right = min(
                geometry.width,
                round(vertical_rule_x(geometry, column + 1, y)) - 6,
            )
            if dynamic_right > dynamic_left:
                projection[y] = int(
                    np.count_nonzero(page_ink[y, dynamic_left:dynamic_right])
                )
        printed_days = maximum_printed_day(column)

        # Determine a robust column-local shift from the repeated 31-row
        # rhythm.  The window is deliberately narrower than half a row, so a
        # neighbouring date cannot improve the correlation score.
        half_score_window = max(2, round(row_pitch * 0.28))
        # Repeated daily rows make a full-row shift look deceptively similar
        # in correlation.  An independent month seed must stay inside the
        # identity interval of its day-column anchor; otherwise day 1 can lock
        # onto day 2 and shift the entire month.  Larger real bends are then
        # recovered gradually by the per-day deformation refinement below.
        shift_candidates = np.linspace(-row_pitch * 0.65, row_pitch * 0.65, 105)
        shift_scores: list[float] = []
        for shift in shift_candidates:
            score = 0.0
            supported = 0
            for center in base_centers[:printed_days] + shift:
                top = max(column_body_top, round(center) - half_score_window)
                bottom = min(
                    column_body_bottom,
                    round(center) + half_score_window + 1,
                )
                if bottom <= top:
                    continue
                peak = float(np.max(projection[top:bottom]))
                score += min(peak, column_width * 0.35)
                supported += peak >= 2
            identity_penalty = abs(shift) / row_pitch * column_width * 0.08
            shift_scores.append(
                score + supported * column_width * 0.015 - identity_penalty
            )
        shift = float(shift_candidates[int(np.argmax(shift_scores))])
        predicted = base_centers + shift
        final_centers = np.full(EXPECTED_DAY_COUNT, np.nan, dtype=float)
        final_weights = np.zeros(EXPECTED_DAY_COUNT, dtype=float)

        for _ in range(3):
            measured = np.full(EXPECTED_DAY_COUNT, np.nan, dtype=float)
            evidence_weights = np.zeros(EXPECTED_DAY_COUNT, dtype=float)
            for day_index in range(printed_days):
                if day_index == 0:
                    local_top = column_body_top
                else:
                    local_top = round(
                        (predicted[day_index - 1] + predicted[day_index]) / 2
                    )
                if day_index == EXPECTED_DAY_COUNT - 1:
                    local_bottom = column_body_bottom
                else:
                    local_bottom = round(
                        (predicted[day_index] + predicted[day_index + 1]) / 2
                    )
                local_top = max(0, local_top)
                local_bottom = min(column_body_bottom, local_bottom)
                if local_bottom <= local_top + 1:
                    continue
                values = projection[local_top:local_bottom].astype(float)
                peak = float(np.max(values)) if values.size else 0.0
                if peak < 2:
                    continue
                active = values >= max(1.0, peak * 0.12)
                bands = active_bands(active, local_top)
                choices: list[tuple[float, float, float]] = []
                for top, bottom in bands:
                    height = bottom - top + 1
                    band_values = projection[top : bottom + 1].astype(float)
                    band_peak = float(np.max(band_values))
                    total = float(np.sum(band_values))
                    if not (2 <= height <= row_pitch * 0.90):
                        continue
                    if band_peak >= column_width * 0.72 or total <= 0:
                        continue
                    y_values = np.arange(top, bottom + 1, dtype=float)
                    centroid = float(
                        np.average(y_values, weights=band_values)
                    )
                    evidence = min(
                        4.0,
                        0.5
                        + band_peak / max(1.0, column_width * 0.10)
                        + total / max(1.0, column_width * row_pitch * 0.18),
                    )
                    selection_cost = (
                        abs(centroid - predicted[day_index]) / row_pitch
                        - evidence * 0.035
                    )
                    choices.append((selection_cost, centroid, evidence))
                if choices:
                    _, centroid, evidence = min(choices, key=lambda item: item[0])
                    measured[day_index] = centroid
                    evidence_weights[day_index] = evidence

            valid = np.isfinite(measured)
            if np.count_nonzero(valid) < max(5, printed_days // 2):
                break
            displacement = measured[valid] - base_centers[valid]
            interpolated = np.interp(
                np.arange(EXPECTED_DAY_COUNT, dtype=float),
                np.flatnonzero(valid).astype(float),
                displacement,
            )
            smooth = cv2.GaussianBlur(
                interpolated[:, None].astype(np.float32),
                (1, 5),
                sigmaX=0.0,
                sigmaY=0.9,
            ).reshape(-1).astype(float)
            predicted = base_centers + smooth
            final_centers = measured
            final_weights = evidence_weights

        valid = np.isfinite(final_centers)
        observations[valid, column] = final_centers[valid]
        weights[valid, column] = final_weights[valid]

    displacement = observations - base_centers[:, None]
    filled_displacement = np.zeros_like(displacement)
    day_indices = np.arange(EXPECTED_DAY_COUNT, dtype=float)
    for column in range(EXPECTED_COLUMN_COUNT):
        valid = np.isfinite(displacement[:, column])
        if np.count_nonzero(valid) >= 2:
            filled_displacement[:, column] = np.interp(
                day_indices,
                day_indices[valid],
                displacement[valid, column],
            )
        elif np.count_nonzero(valid) == 1:
            filled_displacement[:, column] = float(displacement[valid, column][0])

    # A paper fold changes smoothly down the page.  A one-row displacement
    # jump is therefore a wrong glyph/row attachment, not real geometry.
    vertical_reference = np.zeros_like(filled_displacement)
    for column in range(EXPECTED_COLUMN_COUNT):
        for day_index in range(EXPECTED_DAY_COUNT):
            top = max(0, day_index - 2)
            bottom = min(EXPECTED_DAY_COUNT, day_index + 3)
            vertical_reference[day_index, column] = float(
                np.median(filled_displacement[top:bottom, column])
            )
    vertical_reference = cv2.GaussianBlur(
        vertical_reference.astype(np.float32),
        (1, 5),
        sigmaX=0.0,
        sigmaY=0.85,
    ).astype(float)
    valid = np.isfinite(displacement)
    outlier = valid & (
        np.abs(displacement - vertical_reference) > max(2.5, row_pitch * 0.24)
    )
    cleaned_displacement = vertical_reference.copy()
    trusted = valid & ~outlier
    cleaned_displacement[trusted] = (
        displacement[trusted] * 0.72 + vertical_reference[trusted] * 0.28
    )
    smooth_displacement = cv2.GaussianBlur(
        cleaned_displacement.astype(np.float32),
        (3, 3),
        sigmaX=0.50,
        sigmaY=0.65,
    ).astype(float)
    fitted_displacement = (
        cleaned_displacement * 0.82 + smooth_displacement * 0.18
    )
    fitted_displacement[:, 0] = 0.0
    fitted = base_centers[:, None] + fitted_displacement

    # The permissive separator used for the first projection can include the
    # first statistics row on a strongly curved page.  Tighten it now, before
    # decimal components are enumerated, using the independent last-valid-day
    # centers already available in every month.
    statistics_samples = refined_statistics_separator(fitted)
    pre_anchor_fitted = fitted.copy()
    fitted = apply_month_decimal_anchors(
        image,
        page_ink,
        geometry,
        base_centers,
        fitted,
        row_pitch,
        body_bottom,
        statistics_samples,
    )

    # A photographed sheet bends smoothly in both directions.  Decimal/token
    # evidence is rejected locally when it creates a one-day or one-column
    # displacement jump; such a jump is a sequence attachment error, never a
    # physical fold.  The repair uses the 2-D neighbourhood and the independent
    # ink projection prior, so no station/page-specific thresholds are needed.
    displacement = fitted - base_centers[:, None]
    prior_displacement = pre_anchor_fitted - base_centers[:, None]
    for _ in range(3):
        neighbourhood = cv2.medianBlur(
            displacement.astype(np.float32), 3
        ).astype(float)
        residual = np.abs(displacement - neighbourhood)
        invalid = residual > max(3.0, row_pitch * 0.32)
        invalid[:, 0] = False
        displacement[invalid] = (
            neighbourhood[invalid] * 0.68
            + prior_displacement[invalid] * 0.32
        )
    vertical_cap = max(3.0, row_pitch * 0.36)
    horizontal_cap = max(3.0, row_pitch * 0.48)
    for column in range(1, EXPECTED_COLUMN_COUNT):
        for day_index in range(1, EXPECTED_DAY_COUNT):
            delta = displacement[day_index, column] - displacement[
                day_index - 1, column
            ]
            if abs(delta) > vertical_cap:
                candidate = neighbourhood[day_index, column]
                displacement[day_index, column] = (
                    candidate * 0.75
                    + prior_displacement[day_index, column] * 0.25
                )
    for day_index in range(EXPECTED_DAY_COUNT):
        for column in range(1, EXPECTED_COLUMN_COUNT):
            delta = displacement[day_index, column] - displacement[
                day_index, column - 1
            ]
            if abs(delta) > horizontal_cap:
                displacement[day_index, column] = (
                    neighbourhood[day_index, column] * 0.72
                    + prior_displacement[day_index, column] * 0.28
                )
    # Project onto a bounded-slope surface in both directions.  Averaging the
    # forward and backward projections avoids propagating whichever side of a
    # detected step happened to be visited first.
    horizontal_surface_cap = max(2.5, row_pitch * 0.34)
    for day_index in range(EXPECTED_DAY_COUNT):
        forward = displacement[day_index].copy()
        backward = displacement[day_index].copy()
        for column in range(1, EXPECTED_COLUMN_COUNT):
            forward[column] = np.clip(
                forward[column],
                forward[column - 1] - horizontal_surface_cap,
                forward[column - 1] + horizontal_surface_cap,
            )
        for column in range(EXPECTED_COLUMN_COUNT - 2, -1, -1):
            backward[column] = np.clip(
                backward[column],
                backward[column + 1] - horizontal_surface_cap,
                backward[column + 1] + horizontal_surface_cap,
            )
        displacement[day_index] = (forward + backward) / 2
    vertical_surface_cap = max(2.5, row_pitch * 0.34)
    for column in range(1, EXPECTED_COLUMN_COUNT):
        forward = displacement[:, column].copy()
        backward = displacement[:, column].copy()
        for day_index in range(1, EXPECTED_DAY_COUNT):
            forward[day_index] = np.clip(
                forward[day_index],
                forward[day_index - 1] - vertical_surface_cap,
                forward[day_index - 1] + vertical_surface_cap,
            )
        for day_index in range(EXPECTED_DAY_COUNT - 2, -1, -1):
            backward[day_index] = np.clip(
                backward[day_index],
                backward[day_index + 1] - vertical_surface_cap,
                backward[day_index + 1] + vertical_surface_cap,
            )
        displacement[:, column] = (forward + backward) / 2
    displacement[:, 0] = 0.0
    for day_index in range(EXPECTED_DAY_COUNT):
        for column in range(1, EXPECTED_COLUMN_COUNT):
            displacement[day_index, column] = np.clip(
                displacement[day_index, column],
                displacement[day_index, column - 1] - horizontal_surface_cap,
                displacement[day_index, column - 1] + horizontal_surface_cap,
            )
    fitted = base_centers[:, None] + displacement

    minimum_gap = max(2.0, row_pitch * 0.20)
    for column in range(EXPECTED_COLUMN_COUNT):
        for day_index in range(1, EXPECTED_DAY_COUNT):
            fitted[day_index, column] = max(
                fitted[day_index, column],
                fitted[day_index - 1, column] + minimum_gap,
            )
        for day_index in range(EXPECTED_DAY_COUNT - 2, -1, -1):
            fitted[day_index, column] = min(
                fitted[day_index, column],
                fitted[day_index + 1, column] - minimum_gap,
            )

    # Refit once more after month anchors and surface regularization so the
    # exported crop boundary follows the same final geometry.
    statistics_samples = refined_statistics_separator(fitted)

    center_reference_y = float(np.median(base_centers))
    geometry.column_center_x = [
        (
            vertical_rule_x(geometry, column, center_reference_y)
            + vertical_rule_x(geometry, column + 1, center_reference_y)
        )
        / 2
        for column in range(EXPECTED_COLUMN_COUNT)
    ]
    geometry.row_curve_centers = fitted.tolist()
    geometry.statistics_curve = statistics_samples


def curve_values_at_rules(
    geometry: PageGeometry, values: list[float]
) -> list[float]:
    if not geometry.column_center_x or not values:
        return [float(values[0] if values else geometry.statistics_top)] * len(
            geometry.vertical_rules
        )
    source_x = np.asarray(geometry.column_center_x, dtype=float)
    source_y = np.asarray(values, dtype=float)
    query_x = np.asarray(geometry.vertical_rules, dtype=float)
    if len(source_x) >= 3 and np.all(np.diff(source_x) > 0):
        # Shape-preserving cubic interpolation follows photographic bending
        # without the overshoot produced by an unconstrained cubic spline.
        interpolated = PchipInterpolator(
            source_x, source_y, extrapolate=True
        )(query_x).astype(float)
        interpolated[query_x <= source_x[0]] = source_y[0]
        interpolated[query_x >= source_x[-1]] = source_y[-1]
        return interpolated.tolist()
    return np.interp(query_x, source_x, source_y).tolist()


def curved_cell_polygon(
    geometry: PageGeometry, month: int, day_index: int
) -> list[list[float]]:
    row_centers = geometry.row_curve_centers
    if not row_centers:
        top, bottom = anchor_row_bounds(
            geometry.row_anchors,
            day_index,
            geometry.header_bottom + 2,
            geometry.statistics_top - 1,
        )
        return [
            [geometry.vertical_rules[month], top],
            [geometry.vertical_rules[month + 1], top],
            [geometry.vertical_rules[month + 1], bottom],
            [geometry.vertical_rules[month], bottom],
        ]
    if day_index == 0:
        top_values = (
            [float(value + 2.0) for value in geometry.header_curve]
            if len(geometry.header_curve) == EXPECTED_COLUMN_COUNT
            else [float(geometry.header_bottom + 2)] * EXPECTED_COLUMN_COUNT
        )
    else:
        top_values = [
            (row_centers[day_index - 1][column] + row_centers[day_index][column])
            / 2
            for column in range(EXPECTED_COLUMN_COUNT)
        ]
    if day_index == EXPECTED_DAY_COUNT - 1:
        bottom_values = [
            row_centers[-1][column]
            + max(
                4.0,
                (row_centers[-1][column] - row_centers[-2][column]) * 0.52,
            )
            for column in range(EXPECTED_COLUMN_COUNT)
        ]
    else:
        bottom_values = [
            (row_centers[day_index][column] + row_centers[day_index + 1][column])
            / 2
            for column in range(EXPECTED_COLUMN_COUNT)
        ]
    top_rules = curve_values_at_rules(geometry, top_values)
    bottom_rules = curve_values_at_rules(geometry, bottom_values)
    # Calendar-invalid trailing centers are deliberate geometry controls.  A
    # month-end crop must close at the midpoint to the following virtual row,
    # exactly like every other date.  Extending it to the statistics separator
    # creates a 2-5x-tall crop that contains several rows and the first summary
    # rule (the former source of systematic Feb/Apr/Jun/Sep/Nov failures).
    top_left_x = vertical_rule_x(geometry, month, top_rules[month])
    top_right_x = vertical_rule_x(geometry, month + 1, top_rules[month + 1])
    bottom_right_x = vertical_rule_x(
        geometry, month + 1, bottom_rules[month + 1]
    )
    bottom_left_x = vertical_rule_x(geometry, month, bottom_rules[month])
    return [
        [top_left_x, top_rules[month]],
        [top_right_x, top_rules[month + 1]],
        [bottom_right_x, bottom_rules[month + 1]],
        [bottom_left_x, bottom_rules[month]],
    ]


def prepare_curved_cell(
    image: np.ndarray, polygon: list[list[float]], horizontal_inset: int
) -> np.ndarray | None:
    top_left = np.asarray(polygon[0], dtype=np.float32)
    top_right = np.asarray(polygon[1], dtype=np.float32)
    bottom_right = np.asarray(polygon[2], dtype=np.float32)
    bottom_left = np.asarray(polygon[3], dtype=np.float32)
    top_left[0] += horizontal_inset
    bottom_left[0] += horizontal_inset
    top_right[0] -= horizontal_inset
    bottom_right[0] -= horizontal_inset
    top_width = float(np.linalg.norm(top_right - top_left))
    bottom_width = float(np.linalg.norm(bottom_right - bottom_left))
    if min(top_width, bottom_width) <= 3:
        return None
    output_width = max(4, round((top_width + bottom_width) / 2))
    output_height = max(
        4,
        round(
            (
                float(np.linalg.norm(bottom_left - top_left))
                + float(np.linalg.norm(bottom_right - top_right))
            )
            / 2
        ),
    )
    fractions_x = np.linspace(0.0, 1.0, output_width, dtype=np.float32)[
        None, :, None
    ]
    fractions_y = np.linspace(0.0, 1.0, output_height, dtype=np.float32)[:, None]
    top_edge = top_left[None, None, :] * (1.0 - fractions_x) + top_right[
        None, None, :
    ] * fractions_x
    bottom_edge = bottom_left[None, None, :] * (
        1.0 - fractions_x
    ) + bottom_right[None, None, :] * fractions_x
    fractions_y_3d = fractions_y[:, :, None]
    mapped = top_edge * (1.0 - fractions_y_3d) + bottom_edge * fractions_y_3d
    map_x = mapped[:, :, 0].astype(np.float32)
    map_y = mapped[:, :, 1].astype(np.float32)
    cell = cv2.remap(
        image,
        map_x,
        map_y,
        cv2.INTER_CUBIC,
        borderMode=cv2.BORDER_REPLICATE,
    )
    return prepare_cell(cell, (0, 0, output_width, output_height))


def normalized_numeric_text(text: str) -> str:
    normalized = (
        str(text)
        .strip()
        .translate(NUMERIC_TRANSLATION)
        .replace("，", ".")
        .replace(",", ".")
        .replace("：", ".")
        .replace(":", ".")
    )
    return re.sub(r"\s+", "", normalized)


def excel_literal_text(value: Any) -> str:
    """Keep OCR diagnostics as text even when they resemble Excel formulas/errors."""

    text = "" if value is None else str(value)
    if text.startswith(("=", "+", "-", "@", "#")):
        return "'" + text
    return text


def parse_flow(text: str) -> float | None:
    cleaned = normalized_numeric_text(text)
    matches = re.findall(r"\d+(?:\.\d+)?", cleaned)
    if not matches:
        return None
    candidate = max(matches, key=lambda item: (len(item), "." in item))
    try:
        return float(candidate)
    except ValueError:
        return None


def cell_has_ink(cell: np.ndarray | None) -> bool:
    if cell is None or cell.size == 0:
        return False
    gray = cv2.cvtColor(cell, cv2.COLOR_BGR2GRAY)
    ink = cv2.threshold(gray, 190, 255, cv2.THRESH_BINARY_INV)[1]
    return int(np.count_nonzero(ink)) >= max(8, round(ink.size * 0.008))


def numeric_glyph_shape_evidence(cell: np.ndarray | None) -> bool:
    """Reject border fragments/bleed-through before asking OCR about Feb 29."""

    if cell is None or cell.size == 0:
        return False
    gray = cv2.cvtColor(cell, cv2.COLOR_BGR2GRAY)
    ink = cv2.threshold(
        gray, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU
    )[1]
    height, width = ink.shape
    inset_y = max(1, round(height * 0.04))
    inset_x = max(1, round(width * 0.03))
    core = ink[inset_y : height - inset_y, inset_x : width - inset_x]
    ys, xs = np.nonzero(core)
    if ys.size < max(8, round(core.size * 0.006)):
        return False
    vertical_span = float(np.max(ys) - np.min(ys) + 1) / max(1, core.shape[0])
    centroid_y = float(np.mean(ys)) / max(1, core.shape[0] - 1)
    lower_extent = float(np.max(ys)) / max(1, core.shape[0] - 1)
    ink_ratio = float(ys.size / max(1, core.size))
    return bool(
        vertical_span >= 0.30
        and 0.25 <= centroid_y <= 0.74
        and lower_extent >= 0.46
        and 0.006 <= ink_ratio <= 0.58
    )


def recognize_numeric_cell_evidence(
    cell: np.ndarray,
    recognizers: list[TextRecognition | None],
) -> list[dict[str, Any]]:
    """Return OCR votes that contain a plausible numeric flow value."""

    votes: list[dict[str, Any]] = []
    for recognizer in recognizers:
        if recognizer is None:
            continue
        try:
            prediction = next(
                iter(recognizer.predict(input=[cell], batch_size=1))
            )
            text, confidence = result_value(prediction)
        except Exception:
            continue
        normalized = normalized_numeric_text(str(text))
        value = parse_flow(normalized)
        digit_count = len(re.findall(r"\d", normalized))
        if value is None or digit_count < 1 or float(confidence) < 0.35:
            continue
        votes.append(
            {
                "text": str(text),
                "confidence": float(confidence),
                "value": float(value),
            }
        )
    return votes


def infer_month_lengths(
    image: np.ndarray,
    geometry: PageGeometry,
    primary_recognizer: TextRecognition | None = None,
    secondary_recognizer: TextRecognition | None = None,
) -> list[int]:
    """Infer leap year from a tight, curve-aware Feb-29 numeric cell.

    Mere ink is not enough: a dust spot, grid fragment or bleed-through once
    turned a single 1975 table into 366 rows.  The cell must contain a centered
    digit-like component and at least one OCR model must produce a number.
    """

    column_width = geometry.vertical_rules[3] - geometry.vertical_rules[2]
    cell_inset = max(3, round(column_width * 0.060))
    polygon = curved_cell_polygon(geometry, 2, 28)
    february_cell = prepare_curved_cell(image, polygon, cell_inset)
    shape_evidence = numeric_glyph_shape_evidence(february_cell)
    votes = (
        recognize_numeric_cell_evidence(
            february_cell,
            [primary_recognizer, secondary_recognizer],
        )
        if shape_evidence and february_cell is not None
        else []
    )
    leap = bool(votes)
    geometry.source_region_diagnostics["calendar_inference"] = {
        "method": "curve-aware-february-29-numeric-evidence",
        "shape_evidence": bool(shape_evidence),
        "ocr_votes": votes,
        "calendar": "leap" if leap else "common",
    }
    return (
        LEAP_MONTH_LENGTHS.copy()
        if leap
        else COMMON_MONTH_LENGTHS.copy()
    )


def segment_title_crops(
    image: np.ndarray, geometry: PageGeometry
) -> tuple[np.ndarray, list[np.ndarray]]:
    height = geometry.height
    span = geometry.table_right - geometry.table_left
    top = max(0, geometry.table_top - round(height * 0.060))
    bottom = max(top + 10, geometry.table_top - round(height * 0.010))
    left = max(0, geometry.table_left + round(span * 0.08))
    right = min(geometry.width, geometry.table_right - round(span * 0.08))
    crop = image[top:bottom, left:right]
    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
    ink = cv2.threshold(gray, 190, 255, cv2.THRESH_BINARY_INV)[1]
    ink = cv2.morphologyEx(
        ink,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2)),
    )
    closed = cv2.morphologyEx(
        ink,
        cv2.MORPH_CLOSE,
        cv2.getStructuringElement(
            cv2.MORPH_RECT, (max(7, round(crop.shape[1] * 0.010)), 3)
        ),
    )
    projection = np.count_nonzero(closed, axis=0)
    runs: list[tuple[int, int]] = []
    start: int | None = None
    for index, active in enumerate((projection >= 2).tolist()):
        if active and start is None:
            start = index
        elif not active and start is not None:
            if index - start >= 5:
                runs.append((start, index - 1))
            start = None
    if start is not None and len(projection) - start >= 5:
        runs.append((start, len(projection) - 1))

    merged: list[list[int]] = []
    merge_gap = max(9, round(crop.shape[1] * 0.014))
    for run_left, run_right in runs:
        if not merged or run_left - merged[-1][1] > merge_gap:
            merged.append([run_left, run_right])
        else:
            merged[-1][1] = run_right
    parts: list[np.ndarray] = []
    for run_left, run_right in merged:
        part = crop[:, max(0, run_left - 4) : min(crop.shape[1], run_right + 5)]
        if part.shape[1] >= 8:
            parts.append(part)
    return crop, parts


def parse_region_title_text(
    title_text: str, page: int, table_index: int
) -> dict[str, str]:
    """Parse table number, river, station and optional in/outflow scope."""

    compact = re.sub(r"\s+", "", str(title_text))
    compact = compact.replace("(", "（").replace(")", "）")
    # In small photographed headings OCR commonly confuses 逐 with 遂.  Limit
    # the correction to the fixed table-title phrase so genuine station names
    # are never rewritten.
    compact = re.sub(r"遂日(?=平均流量|平均)", "逐日", compact)
    table_match = re.match(r"(\d{1,5})", compact)
    table_number = table_match.group(1) if table_match else ""
    prefix = re.sub(r"^\d+", "", compact)
    prefix = re.split(
        r"(?:逐|遂)(?:日)?(?:平均|均)?(?:流量|流)?"
        r"|日平均|平均流量|集水面积|流量以",
        prefix,
    )[0]
    scope_match = re.search(
        r"(?P<scope>总?(?:入库|出库)(?:流量)?)$", prefix
    )
    flow_scope = scope_match.group("scope") if scope_match else ""
    if scope_match:
        prefix = prefix[: scope_match.start()]

    river = ""
    station = ""
    station_matches = list(
        re.finditer(r"([\u4e00-\u9fff·（）]+站)", prefix)
    )
    station_label = station_matches[-1].group(1) if station_matches else ""
    if station_label:
        river_match = re.match(
            r"(.+?(?:江|河|溪|沟|湖|渠|水库))(.+?站)$",
            station_label,
        )
        if river_match:
            river_candidate, station_candidate = river_match.groups()
            if (
                len(river_candidate) >= 2
                and not station_candidate.startswith("（")
            ):
                river = river_candidate
                station = station_candidate
            else:
                station = station_label
        else:
            station = station_label
    if not station:
        station = f"第{page}页表{table_index}站点"
    return {
        "table_number": table_number,
        "river": river,
        "station": station,
        "flow_scope": flow_scope,
    }


def station_name_quality(station: str) -> float:
    """Score structural completeness without using a station-name whitelist."""

    compact = re.sub(r"\s+", "", str(station)).replace("(", "（").replace(")", "）")
    if not compact or compact.startswith("第") or not compact.endswith("站"):
        return -100.0
    if compact.startswith("）") or compact.count("（") != compact.count("）"):
        return -40.0
    if "）（" in compact:
        return -20.0
    han_count = len(re.findall(r"[\u4e00-\u9fff]", compact[:-1]))
    if han_count < 2:
        return -30.0
    # Length is only a weak completeness cue.  Giving it dominant weight makes
    # a missed river suffix ("那龙 合山站") look like one long station name and
    # beat a clearer candidate that correctly separates "那龙河 / 合山站".
    score = float(4.0 + min(han_count, 6) * 0.20 + min(len(compact), 14) * 0.05)
    if "水库" in compact:
        score += 1.5
    if "（" in compact and "）" in compact:
        score += 0.8
    return score


def title_candidate_score(
    text: str,
    confidence: float,
    page: int,
    table_index: int,
) -> tuple[float, dict[str, str]]:
    parsed = parse_region_title_text(text, page, table_index)
    compact = re.sub(r"\s+", "", str(text))
    score = station_name_quality(parsed["station"])
    if re.search(r"(?:逐|遂)(?:日)?(?:平均|均)?(?:流量|流)?", compact):
        score += 3.0
    if parsed["table_number"]:
        score += 0.8
    if parsed["river"]:
        score += 1.5
    if parsed["flow_scope"]:
        score += 0.5
    score += max(0.0, min(1.0, float(confidence))) * 1.2
    return score, parsed


def enhanced_title_crop_variants(crop: np.ndarray) -> list[np.ndarray]:
    """Create scale/contrast variants for faint or partially clipped headings."""

    if crop.size == 0:
        return [crop]
    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
    clahe = cv2.createCLAHE(clipLimit=2.2, tileGridSize=(8, 4)).apply(gray)
    enhanced = cv2.cvtColor(clahe, cv2.COLOR_GRAY2BGR)
    enlarged = cv2.resize(
        enhanced,
        None,
        fx=1.55,
        fy=1.55,
        interpolation=cv2.INTER_CUBIC,
    )
    return [crop, enhanced, enlarged]


def choose_title_recognition(
    recognized: list[tuple[str, float]],
    page: int,
    table_index: int,
) -> tuple[str, float, dict[str, str]]:
    if not recognized:
        parsed = parse_region_title_text("", page, table_index)
        return "", 0.0, parsed
    candidates: list[
        tuple[float, int, str, float, dict[str, str]]
    ] = []
    for text, confidence in recognized:
        score, parsed = title_candidate_score(
            str(text), float(confidence), page, table_index
        )
        candidates.append(
            (score, len(re.sub(r"\s+", "", str(text))), str(text), float(confidence), parsed)
        )
    _, _, text, confidence, parsed = max(
        candidates, key=lambda item: (item[0], item[1], item[3])
    )
    return text, confidence, parsed


def recognize_page_heading(
    image: np.ndarray,
    geometry: PageGeometry,
    recognizer: TextRecognition,
) -> dict[str, Any]:
    title_crop, parts = segment_title_crops(image, geometry)
    crops = enhanced_title_crop_variants(title_crop) + parts
    predictions = list(recognizer.predict(input=crops, batch_size=len(crops)))
    recognized = [result_value(prediction) for prediction in predictions]
    full_text, full_confidence, parsed = choose_title_recognition(
        recognized, geometry.page, geometry.table_index
    )
    part_results = [
        {"text": text, "confidence": confidence}
        for text, confidence in recognized
        if str(text).strip()
    ]
    station = parsed["station"]
    river = parsed["river"]
    table_number = parsed["table_number"]
    flow_scope = parsed["flow_scope"]
    for index, item in enumerate(part_results):
        text = re.sub(r"\s+", "", str(item["text"]))
        part_parsed = parse_region_title_text(
            text, geometry.page, geometry.table_index
        )
        if not table_number:
            table_number = part_parsed["table_number"]
            if not table_number and re.fullmatch(r"\d{1,5}", text):
                table_number = text
        if not flow_scope and part_parsed["flow_scope"]:
            flow_scope = part_parsed["flow_scope"]
        if station.startswith("第") and not part_parsed["station"].startswith("第"):
            station = part_parsed["station"]
            river = part_parsed["river"]
        if station.startswith("第") and text.endswith("站") and "流量" not in text:
            station = text
        if not river and index > 0 and not station.startswith("第"):
            previous = re.sub(
                r"\s+", "", str(part_results[index - 1]["text"])
            )
            if re.fullmatch(r"[\u4e00-\u9fff·（）()]+", previous):
                river = previous

    station = re.sub(r"\s+", "", station).replace("(", "（").replace(")", "）")
    if not station or station.startswith("第"):
        station = f"第{geometry.page}页站点"

    span = geometry.table_right - geometry.table_left
    metadata_crop = image[
        max(0, geometry.table_top - round(geometry.height * 0.035)) : geometry.table_top,
        geometry.table_left + round(span * 0.62) : geometry.table_right,
    ]
    metadata_prediction = next(
        iter(recognizer.predict(input=[metadata_crop], batch_size=1))
    )
    metadata_text, metadata_confidence = result_value(metadata_prediction)
    area_match = re.search(
        r"集水面积\D*(\d+(?:\.\d+)?)", metadata_text.replace(" ", "")
    )
    catchment_area = float(area_match.group(1)) if area_match else None
    return {
        "table_number": table_number,
        "river": river,
        "station": station,
        "flow_scope": flow_scope,
        "title_text": full_text,
        "title_confidence": full_confidence,
        "title_parts": part_results,
        "metadata_text": metadata_text,
        "metadata_confidence": metadata_confidence,
        "catchment_area_km2": catchment_area,
        "unit": "m³/s",
    }


def recognize_region_heading(
    source_image: np.ndarray,
    region: TableRegion,
    recognizer: TextRecognition,
) -> dict[str, Any]:
    """Read the title belonging to one of several tables on a photo page."""

    page_height, page_width = source_image.shape[:2]
    left, top, right, _ = region.source_bbox
    span = right - left
    title_crops = [
        source_image[
            max(0, top - round(page_height * 0.038)) : max(
                1, top - round(page_height * 0.008)
            ),
            max(0, left - round(span * 0.01)) : min(
                page_width, left + round(span * 0.76)
            ),
        ],
        source_image[
            max(0, top - round(page_height * 0.052)) : max(
                1, top - round(page_height * 0.002)
            ),
            max(0, left - round(span * 0.025)) : min(
                page_width, left + round(span * 0.90)
            ),
        ],
    ]
    title_variants: list[np.ndarray] = []
    for title_crop in title_crops:
        title_variants.extend(enhanced_title_crop_variants(title_crop))
    metadata_crops = [
        source_image[
            max(0, top - round(page_height * 0.025)) : max(
                1, top - round(page_height * 0.003)
            ),
            max(0, left + round(span * 0.62)) : min(page_width, right),
        ],
        source_image[
            max(0, top - round(page_height * 0.033)) : top,
            max(0, left + round(span * 0.72)) : min(page_width, right),
        ],
        source_image[
            max(0, top - round(page_height * 0.030)) : max(
                1, top - round(page_height * 0.005)
            ),
            max(0, left + round(span * 0.72)) : min(page_width, right),
        ],
    ]
    title_predictions = list(
        recognizer.predict(
            input=title_variants, batch_size=len(title_variants)
        )
    )
    title_results = [result_value(item) for item in title_predictions]
    title_text, title_confidence, parsed = choose_title_recognition(
        title_results, region.page, region.table_index
    )
    metadata_predictions = list(
        recognizer.predict(input=metadata_crops, batch_size=len(metadata_crops))
    )
    metadata_results = [result_value(item) for item in metadata_predictions]
    metadata_text = " | ".join(text for text, _ in metadata_results)
    metadata_confidence = max(
        (confidence for _, confidence in metadata_results), default=0.0
    )
    metadata_compact = re.sub(r"\s+", "", str(metadata_text))
    area_candidates = re.findall(
        r"(\d+(?:\.\d+)?)\s*(?=k?m|k㎡|平方)",
        metadata_compact,
        re.I,
    )
    if not area_candidates:
        area_candidates = re.findall(
            r"集水面积\D*(\d+(?:\.\d+)?)", metadata_compact
        )
    area_text = max(area_candidates, key=len) if area_candidates else ""
    catchment_area = float(area_text) if area_text else None
    return {
        "table_number": parsed["table_number"],
        "river": parsed["river"],
        "station": parsed["station"],
        "flow_scope": parsed["flow_scope"],
        "title_text": title_text,
        "title_confidence": title_confidence,
        "title_parts": [
            {"text": text, "confidence": confidence}
            for text, confidence in title_results
            if str(text).strip()
        ],
        "metadata_text": metadata_text,
        "metadata_confidence": metadata_confidence,
        "catchment_area_km2": catchment_area,
        "unit": "m³/s",
    }


def geometry_to_json(geometry: PageGeometry) -> dict[str, Any]:
    def anchor_payload(anchor: AnchorCandidate) -> dict[str, Any]:
        return {
            "top": anchor.top,
            "bottom": anchor.bottom,
            "bbox_center_y": round(anchor.center, 3),
            "ink_centroid_y": round(anchor.ink_centroid, 3),
            "support_columns": anchor.support_columns,
            "day_column_support": anchor.day_column_support,
            "score": round(anchor.score, 3),
        }

    return {
        "page": geometry.page,
        "table_index": geometry.table_index,
        "image_width": geometry.width,
        "image_height": geometry.height,
        "table_bbox": [
            geometry.table_left,
            geometry.table_top,
            geometry.table_right,
            geometry.statistics_top,
        ],
        "header_bottom": geometry.header_bottom,
        "statistics_top": geometry.statistics_top,
        "table_top_curve_y": [
            round(value, 3) for value in geometry.table_top_curve
        ],
        "header_curve_y": [round(value, 3) for value in geometry.header_curve],
        "vertical_rules": geometry.vertical_rules,
        "vertical_rule_models_x_by_y": [
            [round(float(value), 9) for value in coefficients]
            for coefficients in geometry.vertical_rule_models
        ],
        "horizontal_rules": geometry.horizontal_rules,
        "source_region_diagnostics": geometry.source_region_diagnostics,
        "anchor_method": geometry.anchor_method,
        "candidate_anchor_count": len(geometry.candidate_anchors),
        "candidate_anchors": [
            anchor_payload(item) for item in geometry.candidate_anchors
        ],
        "selected_anchor_count": len(geometry.row_anchors),
        "selected_anchors": [
            {"day": index + 1, **anchor_payload(item)}
            for index, item in enumerate(geometry.row_anchors)
        ],
        "curve_model": "month-decimal-track-with-left-integer-fallback",
        "column_center_x": [
            round(value, 3) for value in geometry.column_center_x
        ],
        "statistics_curve_y": [
            round(value, 3) for value in geometry.statistics_curve
        ],
        "row_curve_centers": [
            {
                "day": day + 1,
                "y_by_column": [round(value, 3) for value in values],
            }
            for day, values in enumerate(geometry.row_curve_centers)
        ],
        "month_decimal_anchors": [
            {
                "month": month,
                "anchors": [
                    {
                        **item,
                        "point_x": round(float(item["point_x"]), 3),
                        "point_y": round(float(item["point_y"]), 3),
                        "row_center_y": round(
                            float(item["row_center_y"]), 3
                        ),
                        "decimal_x": round(float(item["decimal_x"]), 3),
                        "dot_to_center_offset": round(
                            float(item["dot_to_center_offset"]), 3
                        ),
                    }
                    for item in anchors
                ],
            }
            for month, anchors in enumerate(geometry.month_decimal_anchors)
            if month > 0 and anchors
        ],
        "virtual_calendar_anchors": [
            {
                **item,
                "row_center_y": round(float(item["row_center_y"]), 3),
            }
            for item in geometry.virtual_month_anchors
        ],
    }


def apply_corrected_labels(
    samples: list[dict[str, Any]], corrected_path: Path
) -> int:
    workbook = load_workbook(corrected_path, data_only=False)
    if "标签校正" not in workbook.sheetnames:
        raise ValueError("标签校正工作簿缺少“标签校正”工作表。")
    sheet = workbook["标签校正"]
    headers = {
        str(cell.value).strip(): index + 1
        for index, cell in enumerate(sheet[1])
        if cell.value is not None
    }
    required = {"样本ID", "正确标签", "纳入训练"}
    missing = required - set(headers)
    if missing:
        raise ValueError("标签校正工作簿缺少列：" + "、".join(sorted(missing)))
    corrections: dict[str, str] = {}
    for row in range(2, sheet.max_row + 1):
        sample_id = str(sheet.cell(row, headers["样本ID"]).value or "").strip()
        correct = str(sheet.cell(row, headers["正确标签"]).value or "").strip()
        include = str(sheet.cell(row, headers["纳入训练"]).value or "").strip()
        if sample_id and correct and include == "是":
            corrections[sample_id] = correct
    workbook.close()
    applied = 0
    for sample in samples:
        correct = corrections.get(sample["sample_id"])
        if not correct:
            continue
        value = parse_flow(correct)
        if value is None:
            raise ValueError(
                f"样本{sample['sample_id']}的正确标签不是有效流量：{correct}"
            )
        sample["correct_label"] = correct
        sample["include"] = "是"
        sample["selected_text"] = correct
        sample["value"] = value
        sample["value_source"] = "corrected-label"
        sample["note"] = "已采用人工校正标签"
        applied += 1
    return applied


def recognize_daily_cells(
    image: np.ndarray,
    geometry: PageGeometry,
    heading: dict[str, Any],
    primary_recognizer: TextRecognition,
    secondary_recognizer: TextRecognition | None,
    training_images_dir: Path,
    batch_size: int,
    low_confidence: float,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[int]]:
    # Fit once with a provisional Feb-29 control point so the calendar probe
    # can use the same non-rigid row/column surface as ordinary cells.
    geometry.month_lengths = LEAP_MONTH_LENGTHS.copy()
    fit_multicolumn_row_curves(image, geometry)
    month_lengths = infer_month_lengths(
        image,
        geometry,
        primary_recognizer,
        secondary_recognizer,
    )
    if month_lengths != geometry.month_lengths:
        geometry.month_lengths = list(month_lengths)
        # Refit February so day 29 becomes a geometry-only virtual anchor in a
        # common year and cannot attach itself to summary-row ink.
        fit_multicolumn_row_curves(image, geometry)
    else:
        geometry.month_lengths = list(month_lengths)
    training_images_dir.mkdir(parents=True, exist_ok=True)
    crops: list[np.ndarray] = []
    samples: list[dict[str, Any]] = []

    for day_index, day in enumerate(range(1, EXPECTED_DAY_COUNT + 1)):
        anchor = geometry.row_anchors[day_index]
        for month in range(1, 13):
            if day > month_lengths[month - 1]:
                continue
            column_width = (
                geometry.vertical_rules[month + 1]
                - geometry.vertical_rules[month]
            )
            cell_inset = max(3, round(column_width * 0.060))
            polygon = curved_cell_polygon(geometry, month, day_index)
            prepared = prepare_curved_cell(image, polygon, cell_inset)
            xs = [point[0] for point in polygon]
            ys = [point[1] for point in polygon]
            bounds = [
                round(min(xs)) + cell_inset,
                round(min(ys)),
                round(max(xs)) - cell_inset,
                round(max(ys)),
            ]
            sample_id = (
                f"p{geometry.page:03d}_t{geometry.table_index:02d}_"
                f"m{month:02d}_d{day:02d}_flow"
            )
            image_name = f"{sample_id}.png"
            image_path = training_images_dir / image_name
            if prepared is None:
                prepared = np.full((80, 80, 3), 255, dtype=np.uint8)
            cv2.imwrite(str(image_path), prepared)
            crops.append(prepared)
            samples.append(
                {
                    "sample_id": sample_id,
                    "page": geometry.page,
                    "panel": geometry.table_index,
                    "row": day,
                    "month": month,
                    "day": day,
                    "column": "流量(m³/s)",
                    "bounds": bounds,
                    "polygon": [
                        [round(x, 3), round(y, 3)] for x, y in polygon
                    ],
                    "anchor_y": round(anchor.center, 3),
                    "anchor_ink_centroid_y": round(anchor.ink_centroid, 3),
                    "anchor_support_columns": anchor.support_columns,
                    "image": f"images/{image_name}",
                    "station": heading["station"],
                }
            )

    primary_predictions = list(
        primary_recognizer.predict(input=crops, batch_size=batch_size)
    )
    secondary_predictions: list[Any] = []
    if secondary_recognizer is not None:
        secondary_predictions = list(
            secondary_recognizer.predict(input=crops, batch_size=batch_size)
        )
    if len(primary_predictions) != len(samples):
        raise RuntimeError("主OCR模型返回的流量候选数量与单元格数量不一致。")
    if secondary_recognizer is not None and len(secondary_predictions) != len(samples):
        raise RuntimeError("辅助OCR模型返回的流量候选数量与单元格数量不一致。")

    rows: list[dict[str, Any]] = []
    for index, sample in enumerate(samples):
        primary_text, primary_confidence = result_value(primary_predictions[index])
        primary_value = parse_flow(primary_text)
        secondary_text = ""
        secondary_confidence = 0.0
        secondary_value: float | None = None
        if secondary_recognizer is not None:
            secondary_text, secondary_confidence = result_value(
                secondary_predictions[index]
            )
            secondary_value = parse_flow(secondary_text)

        conflict = (
            primary_value is not None
            and secondary_value is not None
            and abs(primary_value - secondary_value) > 1.0e-9
        )
        primary_has_leading_artifact = bool(
            re.match(r"^[|!Il]", str(primary_text).strip())
        )
        primary_digit_count = len(re.findall(r"\d", str(primary_text)))
        secondary_digit_count = len(re.findall(r"\d", str(secondary_text)))
        secondary_has_clearer_digits = (
            conflict
            and secondary_digit_count >= primary_digit_count + 1
            and secondary_confidence >= primary_confidence - 0.20
        )
        if primary_value is not None and (
            secondary_value is None
            or not conflict
            or (
                primary_confidence >= secondary_confidence
                and not primary_has_leading_artifact
                and not secondary_has_clearer_digits
            )
        ):
            selected_value = primary_value
            selected_text = primary_text
            value_source = "primary"
        elif secondary_value is not None:
            selected_value = secondary_value
            selected_text = secondary_text
            value_source = "secondary"
        else:
            selected_value = None
            selected_text = ""
            value_source = "unreadable"

        confidences = [primary_confidence]
        if secondary_recognizer is not None:
            confidences.append(secondary_confidence)
        average_confidence = float(np.mean(confidences))
        notes: list[str] = []
        include = "是"
        if selected_value is None:
            notes.append("两个模型均没有形成有效数值")
            include = "待确认"
        if conflict:
            notes.append("主模型与辅助模型数值冲突")
            include = "待确认"
        if average_confidence < low_confidence:
            notes.append("平均置信度低于阈值")
            include = "待确认"

        sample.update(
            {
                "ocr_text": primary_text,
                "primary_text": primary_text,
                "primary_value": primary_value,
                "primary_confidence": primary_confidence,
                "secondary_text": secondary_text,
                "secondary_value": secondary_value,
                "secondary_confidence": secondary_confidence,
                "selected_text": selected_text,
                "normalized_label": (
                    "" if selected_value is None else f"{selected_value:g}"
                ),
                "confidence": average_confidence,
                "correct_label": (
                    "" if selected_value is None else f"{selected_value:g}"
                ),
                "include": include,
                "note": "；".join(notes),
                "value": selected_value,
                "value_source": value_source,
                "model_conflict": conflict,
            }
        )
        rows.append(
            {
                "页": geometry.page,
                "表区": geometry.table_index,
                "月": sample["month"],
                "日": sample["day"],
                "流量(m³/s)": selected_value,
                "平均置信度": average_confidence,
                "主模型文本": primary_text,
                "主模型置信度": primary_confidence,
                "辅助模型文本": secondary_text,
                "辅助模型置信度": secondary_confidence,
                "值来源": value_source,
                "模型冲突": conflict,
                "待确认": include == "待确认",
                "备注": "；".join(notes),
                "样本ID": sample["sample_id"],
                "行锚点Y": sample["anchor_y"],
                "锚点墨迹重心Y": sample["anchor_ink_centroid_y"],
                "锚点支持列数": sample["anchor_support_columns"],
                "单元格边界": sample["bounds"],
                "单元格多边形": sample["polygon"],
                "图片": sample["image"],
            }
        )
    return rows, samples, month_lengths


def synchronize_rows_from_samples(
    rows: list[dict[str, Any]], samples: list[dict[str, Any]]
) -> None:
    by_id = {sample["sample_id"]: sample for sample in samples}
    for row in rows:
        sample = by_id[row["样本ID"]]
        row["流量(m³/s)"] = sample["value"]
        row["值来源"] = sample["value_source"]
        row["待确认"] = sample["include"] == "待确认"
        row["备注"] = sample["note"]


def draw_source_curve_mesh_preview(
    image: np.ndarray, regions: list[TableRegion], output_path: Path
) -> None:
    """Show the curved source-space mesh before it is flattened."""

    preview = image.copy()
    for region in regions:
        diagnostics = region.diagnostics
        boundary_values = diagnostics.get("source_boundary_y_at_rules", {})
        track_diagnostics = diagnostics.get("source_vertical_rule_tracks", [])
        if not boundary_values or len(track_diagnostics) != EXPECTED_VERTICAL_RULE_COUNT:
            left, top, right, bottom = region.source_bbox
            cv2.rectangle(preview, (left, top), (right, bottom), (255, 0, 255), 2)
            continue
        models = [
            np.asarray(item["coefficients"], dtype=float)
            for item in track_diagnostics
        ]
        colors = {
            "table_top": (255, 0, 255),
            "header_bottom": (0, 0, 255),
            "statistics_top": (0, 0, 255),
            "table_bottom": (255, 0, 255),
        }
        for name, color in colors.items():
            y_values = np.asarray(boundary_values.get(name, []), dtype=float)
            if len(y_values) != EXPECTED_VERTICAL_RULE_COUNT:
                continue
            points = np.asarray(
                [
                    [round(float(np.polyval(models[index], y))), round(float(y))]
                    for index, y in enumerate(y_values)
                ],
                dtype=np.int32,
            )
            cv2.polylines(preview, [points], False, color, 2, cv2.LINE_AA)
        top_values = np.asarray(boundary_values.get("table_top", []), dtype=float)
        statistics_values = np.asarray(
            boundary_values.get("statistics_top", []), dtype=float
        )
        if (
            len(top_values) == EXPECTED_VERTICAL_RULE_COUNT
            and len(statistics_values) == EXPECTED_VERTICAL_RULE_COUNT
        ):
            for index, model in enumerate(models):
                sample_y = np.linspace(top_values[index], statistics_values[index], 90)
                points = np.asarray(
                    [
                        [round(float(np.polyval(model, y))), round(float(y))]
                        for y in sample_y
                    ],
                    dtype=np.int32,
                )
                cv2.polylines(
                    preview, [points], False, (255, 200, 0), 1, cv2.LINE_AA
                )
            label_x = round(float(np.polyval(models[0], top_values[0])))
            label_y = max(20, round(float(top_values[0])) - 10)
            cv2.putText(
                preview,
                f"T{region.table_index}",
                (label_x, label_y),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.7,
                (0, 0, 255),
                2,
                cv2.LINE_AA,
            )
    cv2.imwrite(str(output_path), preview)


def draw_anchor_preview(
    image: np.ndarray, geometry: PageGeometry, output_path: Path
) -> None:
    preview = image.copy()
    top_rules = boundary_values_at_rules(
        geometry, geometry.table_top_curve, geometry.table_top
    )
    bottom_rules = boundary_values_at_rules(
        geometry, geometry.statistics_curve, geometry.statistics_top
    )
    top_points = np.asarray(
        [
            [round(vertical_rule_x(geometry, index, y)), round(y)]
            for index, y in enumerate(top_rules)
        ],
        dtype=np.int32,
    )
    bottom_points = np.asarray(
        [
            [round(vertical_rule_x(geometry, index, y)), round(y)]
            for index, y in enumerate(bottom_rules)
        ],
        dtype=np.int32,
    )
    cv2.polylines(preview, [top_points], False, (255, 0, 255), 2, cv2.LINE_AA)
    cv2.polylines(preview, [bottom_points], False, (255, 0, 255), 2, cv2.LINE_AA)
    for rule_index in (0, len(geometry.vertical_rules) - 1):
        side_y = np.linspace(top_rules[rule_index], bottom_rules[rule_index], 80)
        side_points = np.asarray(
            [
                [round(vertical_rule_x(geometry, rule_index, y)), round(y)]
                for y in side_y
            ],
            dtype=np.int32,
        )
        cv2.polylines(
            preview, [side_points], False, (255, 0, 255), 2, cv2.LINE_AA
        )
    for rule_index in range(len(geometry.vertical_rules)):
        sample_y = np.linspace(
            geometry.table_top, geometry.statistics_top, 80
        )
        points = np.asarray(
            [
                [round(vertical_rule_x(geometry, rule_index, y)), round(y)]
                for y in sample_y
            ],
            dtype=np.int32,
        )
        cv2.polylines(preview, [points], False, (255, 200, 0), 1, cv2.LINE_AA)
    header_rules = boundary_values_at_rules(
        geometry, geometry.header_curve, geometry.header_bottom
    )
    header_points = np.asarray(
        [
            [round(vertical_rule_x(geometry, index, y)), round(y)]
            for index, y in enumerate(header_rules)
        ],
        dtype=np.int32,
    )
    cv2.polylines(preview, [header_points], False, (0, 0, 255), 2, cv2.LINE_AA)
    if geometry.statistics_curve:
        statistics_points = np.asarray(
            [
                [round(x), round(y)]
                for x, y in zip(
                    geometry.column_center_x, geometry.statistics_curve
                )
            ],
            dtype=np.int32,
        )
        cv2.polylines(
            preview, [statistics_points], False, (0, 0, 255), 2, cv2.LINE_AA
        )
    else:
        cv2.line(
            preview,
            (geometry.table_left, geometry.statistics_top),
            (geometry.table_right, geometry.statistics_top),
            (0, 0, 255),
            2,
        )
    selected_centers = {
        round(item.center): index + 1
        for index, item in enumerate(geometry.row_anchors)
    }
    for candidate in geometry.candidate_anchors:
        y = round(candidate.center)
        cv2.circle(preview, (geometry.table_left - 12, y), 4, (0, 215, 255), -1)
    for day_index, anchor in enumerate(geometry.row_anchors):
        y = round(anchor.center)
        day = selected_centers[y]
        if geometry.row_curve_centers:
            points = np.asarray(
                [
                    [
                        round(
                            (
                                vertical_rule_x(geometry, column, curve_y)
                                + vertical_rule_x(
                                    geometry, column + 1, curve_y
                                )
                            )
                            / 2
                        ),
                        round(curve_y),
                    ]
                    for column, curve_y in enumerate(
                        geometry.row_curve_centers[day_index]
                    )
                ],
                dtype=np.int32,
            )
            cv2.polylines(
                preview, [points], False, (0, 180, 0), 1, cv2.LINE_AA
            )
            for point_x, point_y in points:
                cv2.circle(
                    preview,
                    (int(point_x), int(point_y)),
                    2,
                    (255, 80, 0),
                    -1,
                )
        else:
            cv2.line(
                preview,
                (geometry.table_left, y),
                (geometry.table_right, y),
                (0, 180, 0),
                1,
            )
        cv2.circle(preview, (geometry.table_left - 12, y), 5, (0, 160, 0), -1)
        cv2.putText(
            preview,
            str(day),
            (max(2, geometry.table_left - 55), y + 4),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.38,
            (0, 100, 0),
            1,
            cv2.LINE_AA,
        )
    for month_anchors in geometry.month_decimal_anchors:
        for item in month_anchors:
            point = (
                round(float(item["point_x"])),
                round(float(item["point_y"])),
            )
            if item["kind"] == "decimal-dot":
                # Magenta rings are directly observed decimal points.
                cv2.circle(preview, point, 4, (255, 0, 255), 1, cv2.LINE_AA)
            else:
                # Yellow crosses are integer tokens found to the left of the
                # fitted decimal track and bounded by neighbouring dates.
                cv2.drawMarker(
                    preview,
                    point,
                    (0, 215, 255),
                    cv2.MARKER_CROSS,
                    7,
                    1,
                    cv2.LINE_AA,
                )
    for item in geometry.virtual_month_anchors:
        month = int(item["month"])
        point = (
            round(
                (
                    vertical_rule_x(
                        geometry, month, float(item["row_center_y"])
                    )
                    + vertical_rule_x(
                        geometry, month + 1, float(item["row_center_y"])
                    )
                )
                / 2
            ),
            round(float(item["row_center_y"])),
        )
        # Hollow cyan diamonds are geometry-only calendar control points; they
        # are deliberately not attached to any printed ink.
        cv2.drawMarker(
            preview,
            point,
            (255, 255, 0),
            cv2.MARKER_DIAMOND,
            8,
            1,
            cv2.LINE_AA,
        )
    cv2.imwrite(str(output_path), preview)


def draw_cell_preview(
    image: np.ndarray,
    geometry: PageGeometry,
    rows: list[dict[str, Any]],
    output_path: Path,
) -> None:
    preview = image.copy()
    for row in rows:
        color = (0, 0, 255) if row["待确认"] else (0, 170, 0)
        polygon = row.get("单元格多边形")
        if polygon:
            points = np.asarray(
                [[round(x), round(y)] for x, y in polygon], dtype=np.int32
            )
            cv2.polylines(
                preview, [points], True, color, 1, cv2.LINE_AA
            )
        else:
            left, top, right, bottom = row["单元格边界"]
            cv2.rectangle(preview, (left, top), (right, bottom), color, 1)
    cv2.rectangle(
        preview,
        (geometry.table_left, geometry.header_bottom),
        (geometry.table_right, geometry.statistics_top),
        (255, 0, 255),
        2,
    )
    cv2.imwrite(str(output_path), preview)


def save_workbook_with_lock_fallback(workbook: Workbook, path: Path) -> Path:
    """Save normally, or use a stable update suffix when Excel holds a lock."""

    try:
        workbook.save(path)
        return path
    except PermissionError:
        fallback = path.with_name(f"{path.stem}_更新{path.suffix}")
        try:
            workbook.save(fallback)
        except PermissionError:
            for occurrence in range(2, 100):
                candidate = path.with_name(
                    f"{path.stem}_更新_{occurrence:02d}{path.suffix}"
                )
                try:
                    workbook.save(candidate)
                    fallback = candidate
                    break
                except PermissionError:
                    continue
            else:
                raise RuntimeError(
                    f"原文件及98个更新文件都被占用，无法保存：{path}"
                )
        log(f"[输出诊断] {path.name}正被占用，已另存为{fallback.name}")
        return fallback


def write_station_workbook(
    output_dir: Path,
    station: str,
    output_name: str,
    heading: dict[str, Any],
    rows: list[dict[str, Any]],
    geometries: list[PageGeometry],
    month_lengths: list[int],
) -> Path:
    decimal_places = 1
    recognized_texts = [
        normalized_numeric_text(str(row.get("主模型文本", ""))) for row in rows
    ]
    if any(re.search(r"\.\d{2,}", text) for text in recognized_texts):
        decimal_places = 2
    flow_number_format = "0." + "0" * decimal_places
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "合并表"
    headers = ["月", "日", "流量(m³/s)", "平均置信度"]
    sheet.append(headers)
    pending_fill = PatternFill("solid", fgColor="FCE8E6")
    low_fill = PatternFill("solid", fgColor="FFF2CC")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for row in rows:
        sheet.append(
            [
                row["月"],
                row["日"],
                row["流量(m³/s)"],
                row["平均置信度"],
            ]
        )
        excel_row = sheet.max_row
        sheet.cell(excel_row, 1).number_format = "0"
        sheet.cell(excel_row, 2).number_format = "0"
        sheet.cell(excel_row, 3).number_format = flow_number_format
        sheet.cell(excel_row, 4).number_format = "0.00000000"
        if row["待确认"]:
            for cell in sheet[excel_row]:
                cell.fill = pending_fill
        elif row["平均置信度"] < 0.85:
            for cell in sheet[excel_row]:
                cell.fill = low_fill
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for width, column in zip((8, 8, 16, 18), range(1, 5)):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:D{sheet.max_row}"
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 25

    matrix = workbook.create_sheet("月日矩阵")
    matrix.merge_cells("A1:M1")
    title_bits = [
        heading.get("table_number", ""),
        heading.get("river", ""),
        station,
        heading.get("flow_scope", ""),
        "逐日平均流量表",
    ]
    matrix["A1"] = "  ".join(bit for bit in title_bits if bit)
    matrix["A1"].font = Font(bold=True, size=16, color="1F2937")
    matrix["A1"].alignment = Alignment(horizontal="center")
    matrix.merge_cells("J2:M2")
    area = heading.get("catchment_area_km2")
    matrix["J2"] = (
        f"集水面积  {area:g} km²，流量  m³/s"
        if isinstance(area, (int, float))
        else "流量  m³/s"
    )
    matrix["J2"].alignment = Alignment(horizontal="right")
    matrix.append(["日 \\ 月", *MONTH_NAMES])
    matrix_header_row = 3
    by_key = {(row["月"], row["日"]): row for row in rows}
    for day in range(1, 32):
        values: list[Any] = [day]
        for month in range(1, 13):
            row = by_key.get((month, day))
            values.append(None if row is None else row["流量(m³/s)"])
        matrix.append(values)
    for cell in matrix[matrix_header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row_number in range(4, 35):
        for column in range(1, 14):
            cell = matrix.cell(row_number, column)
            cell.alignment = Alignment(horizontal="center")
            if column > 1:
                cell.number_format = flow_number_format
    for column in range(1, 14):
        matrix.column_dimensions[get_column_letter(column)].width = 11
    matrix.column_dimensions["A"].width = 10
    matrix.freeze_panes = "B4"
    matrix.sheet_view.showGridLines = False
    matrix.row_dimensions[1].height = 28

    anchor_sheet = workbook.create_sheet("锚点校核")
    anchor_headers = [
        "页",
        "表区",
        "日",
        "外接框中心Y",
        "墨迹重心Y",
        "支持列数",
        "日列支持",
        "与前行间距",
        "曲线最大偏移",
        "曲线跨列范围",
        "状态",
    ]
    anchor_sheet.append(anchor_headers)
    for geometry in geometries:
        previous: float | None = None
        for day, anchor in enumerate(geometry.row_anchors, start=1):
            gap = None if previous is None else anchor.center - previous
            curve = (
                geometry.row_curve_centers[day - 1]
                if day - 1 < len(geometry.row_curve_centers)
                else [anchor.center]
            )
            maximum_deviation = max(abs(value - anchor.center) for value in curve)
            cross_column_range = max(curve) - min(curve)
            status = (
                "通过"
                if anchor.support_columns >= 5 and anchor.day_column_support
                else "待确认"
            )
            anchor_sheet.append(
                [
                    geometry.page,
                    geometry.table_index,
                    day,
                    anchor.center,
                    anchor.ink_centroid,
                    anchor.support_columns,
                    "是" if anchor.day_column_support else "否",
                    gap,
                    maximum_deviation,
                    cross_column_range,
                    status,
                ]
            )
            previous = anchor.center
    for cell in anchor_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row_number in range(2, anchor_sheet.max_row + 1):
        for column in (4, 5, 8, 9, 10):
            anchor_sheet.cell(row_number, column).number_format = "0.000"
    for column, width in enumerate(
        (8, 8, 8, 16, 16, 12, 12, 16, 16, 16, 12), start=1
    ):
        anchor_sheet.column_dimensions[get_column_letter(column)].width = width
    anchor_sheet.freeze_panes = "A2"
    anchor_sheet.auto_filter.ref = f"A1:K{anchor_sheet.max_row}"
    anchor_sheet.sheet_view.showGridLines = False

    pending = workbook.create_sheet("待确认项")
    pending_headers = [
        "输出行",
        "页",
        "月",
        "日",
        "建议值",
        "主模型文本",
        "辅助模型文本",
        "平均置信度",
        "原因",
        "样本ID",
        "图片",
    ]
    pending.append(pending_headers)
    for output_row, row in enumerate(rows, start=1):
        if not row["待确认"]:
            continue
        pending.append(
            [
                output_row,
                row["页"],
                row["月"],
                row["日"],
                row["流量(m³/s)"],
                excel_literal_text(row["主模型文本"]),
                excel_literal_text(row["辅助模型文本"]),
                row["平均置信度"],
                row["备注"],
                row["样本ID"],
                row["图片"],
            ]
        )
        pending.cell(pending.max_row, 11).hyperlink = row["图片"]
        pending.cell(pending.max_row, 11).style = "Hyperlink"
    for cell in pending[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for column, width in enumerate(
        (10, 8, 8, 8, 14, 18, 18, 16, 52, 30, 44), start=1
    ):
        pending.column_dimensions[get_column_letter(column)].width = width
    pending.freeze_panes = "A2"
    pending.auto_filter.ref = f"A1:K{max(1, pending.max_row)}"
    pending.sheet_view.showGridLines = False

    workbook_path = output_dir / f"{safe_filename(output_name)}.xlsx"
    return save_workbook_with_lock_fallback(workbook, workbook_path)


def station_output_name(station: str, occurrence: int, total: int) -> str:
    """Return a stable filename label while preserving the Chinese station name."""

    if total <= 1:
        return station
    return f"{station}_{occurrence:02d}"


def write_training_review_workbook(
    samples: list[dict[str, Any]], training_dir: Path
) -> Path:
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "使用说明"
    instructions.append(["逐日流量OCR训练样本校正说明"])
    instructions.append(["1", "打开“标签校正”工作表。"])
    instructions.append(["2", "点击图片文件名查看实际识别的流量单元格。"])
    instructions.append(["3", "只修改“正确标签”和“纳入训练”两列。"])
    instructions.append(["4", "正确标签必须与图片完全一致，并保留小数点。"])
    instructions.append(["5", "确认无误请选择“是”；无法判断请选择“待确认”。"])
    instructions.append(["6", "程序使用多列墨迹锚点重建31个日序，空白格不拟合坐标。"])
    instructions.append(["7", "可用 --corrected-labels 指向校正后的工作簿重新输出。"])
    instructions.column_dimensions["A"].width = 18
    instructions.column_dimensions["B"].width = 100
    instructions.merge_cells("A1:B1")
    instructions["A1"].font = Font(bold=True, size=15, color="FFFFFF")
    instructions["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    instructions.row_dimensions[1].height = 28
    for row in range(2, instructions.max_row + 1):
        instructions.cell(row, 1).font = Font(bold=True, color="1F4E78")
        instructions.cell(row, 2).alignment = Alignment(wrap_text=True)
        instructions.row_dimensions[row].height = 26
    instructions.sheet_view.showGridLines = False

    sheet = workbook.create_sheet("标签校正")
    headers = [
        "样本ID",
        "页",
        "表区",
        "行序号",
        "月",
        "日",
        "字段",
        "主模型原文",
        "主模型置信度",
        "辅助模型原文",
        "辅助模型置信度",
        "规范化标签",
        "正确标签",
        "纳入训练",
        "图片文件",
        "备注",
    ]
    sheet.append(headers)
    for sample in samples:
        sheet.append(
            [
                sample["sample_id"],
                sample["page"],
                sample["panel"],
                sample["row"],
                sample["month"],
                sample["day"],
                sample["column"],
                excel_literal_text(sample["primary_text"]),
                sample["primary_confidence"],
                excel_literal_text(sample["secondary_text"]),
                sample["secondary_confidence"],
                excel_literal_text(sample["normalized_label"]),
                excel_literal_text(sample["correct_label"]),
                sample["include"],
                sample["image"],
                excel_literal_text(sample.get("note", "")),
            ]
        )
        image_cell = sheet.cell(sheet.max_row, 15)
        image_cell.hyperlink = sample["image"]
        image_cell.style = "Hyperlink"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    low_fill = PatternFill("solid", fgColor="FFF2CC")
    review_fill = PatternFill("solid", fgColor="FCE4D6")
    editable_fill = PatternFill("solid", fgColor="E2F0D9")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 9).number_format = "0.00000000"
        sheet.cell(row, 11).number_format = "0.00000000"
        sheet.cell(row, 13).number_format = "@"
        sheet.cell(row, 13).fill = editable_fill
        sheet.cell(row, 14).fill = editable_fill
        confidence = min(
            sheet.cell(row, 9).value or 0.0,
            sheet.cell(row, 11).value or 1.0,
        )
        if confidence < 0.85:
            for column in range(1, 13):
                sheet.cell(row, column).fill = low_fill
        if sheet.cell(row, 14).value == "待确认":
            sheet.cell(row, 14).fill = review_fill
    validation = DataValidation(
        type="list", formula1='"是,否,待确认"', allow_blank=False
    )
    sheet.add_data_validation(validation)
    validation.add(f"N2:N{sheet.max_row}")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:P{sheet.max_row}"
    sheet.row_dimensions[1].height = 26
    widths = [30, 8, 8, 10, 8, 8, 18, 18, 15, 18, 15, 18, 18, 14, 48, 52]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.sheet_view.showGridLines = False
    path = training_dir / "labels_review.xlsx"
    return save_workbook_with_lock_fallback(workbook, path)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "识别按日×月排版的逐日平均流量扫描表；"
            "通过多列墨迹锚点重建31个日序，并排除下方统计区。"
        ),
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("pdf", type=Path, help="输入PDF文件")
    parser.add_argument("-o", "--output", type=Path, help="输出目录")
    parser.add_argument("--device", default="gpu:0", help="Paddle推理设备")
    parser.add_argument("--render-width", type=int, default=1500)
    parser.add_argument("--batch-size", type=int, default=64)
    parser.add_argument("--low-confidence", type=float, default=0.85)
    parser.add_argument(
        "--primary-model-name",
        default="PP-OCRv5_server_rec",
        help="标题和流量主识别模型",
    )
    parser.add_argument(
        "--secondary-model-name",
        default="en_PP-OCRv5_mobile_rec",
        help="流量辅助数字模型",
    )
    parser.add_argument(
        "--disable-secondary-model",
        action="store_true",
        help="禁用辅助数字模型",
    )
    parser.add_argument(
        "--corrected-labels",
        type=Path,
        help="可选的labels_review校正工作簿",
    )
    return parser


def build_table_failure_payload(
    region: TableRegion,
    stage: str,
    error: BaseException,
    render_width: int,
) -> dict[str, Any]:
    """Create a durable diagnostic when one table fails independently."""

    return {
        "page": region.page,
        "table_index": region.table_index,
        "stage": stage,
        "error_type": type(error).__name__,
        "message": str(error),
        "render_width": int(render_width),
        "source_bbox": [int(value) for value in region.source_bbox],
        "source_rules": [int(value) for value in region.source_rules],
        "observed_rule_flags": [bool(value) for value in region.observed_rule_flags],
        "boundary_scan_limits": region.diagnostics.get(
            "boundary_scan_limits", []
        ),
        "region_diagnostics": dict(region.diagnostics),
        "traceback": traceback.format_exc().splitlines()[-24:],
    }


def build_page_detection_failure_payload(
    page_number: int,
    detection: TableRegionDetection,
    render_width: int,
) -> dict[str, Any]:
    """Record a missing table without aborting the remaining PDF pages."""

    return {
        "page": int(page_number),
        "table_index": 0,
        "stage": "table_detection",
        "error_type": "RuntimeError",
        "message": (
            f"第{page_number}页没有检测到由14条竖线组成的逐日流量表。"
        ),
        "render_width": int(render_width),
        "candidate_group_count": int(detection.candidate_group_count),
        "strong_rejected_candidate_count": int(
            detection.strong_rejected_candidate_count
        ),
        "rejected_groups": list(detection.rejected_groups),
    }


def main() -> None:
    args = build_parser().parse_args()
    pdf_path = args.pdf.expanduser().resolve()
    if not pdf_path.is_file():
        raise FileNotFoundError(f"输入PDF不存在：{pdf_path}")
    script_dir = Path(__file__).resolve().parent
    output_dir = (
        args.output.expanduser().resolve()
        if args.output
        else script_dir / "output_dailyflowrate" / f"{pdf_path.stem}_result"
    )
    structure_dir = output_dir / "structure"
    previews_dir = output_dir / "previews"
    training_dir = output_dir / "training_samples"
    training_images_dir = training_dir / "images"
    for directory in (
        output_dir,
        structure_dir,
        previews_dir,
        training_images_dir,
    ):
        directory.mkdir(parents=True, exist_ok=True)

    verify_device(args.device)
    log("[1/4] 初始化印刷体文字主模型")
    primary_recognizer = TextRecognition(
        model_name=args.primary_model_name,
        device=args.device,
        engine="paddle",
    )
    secondary_recognizer: TextRecognition | None = None
    if not args.disable_secondary_model:
        log("[1/4] 初始化印刷数字辅助模型")
        secondary_recognizer = TextRecognition(
            model_name=args.secondary_model_name,
            device=args.device,
            engine="paddle",
        )

    document = pdfium.PdfDocument(str(pdf_path))
    page_count = len(document)
    document.close()
    if page_count < 1:
        raise RuntimeError("PDF没有页面。")

    all_rows: list[dict[str, Any]] = []
    all_samples: list[dict[str, Any]] = []
    dataset_groups: dict[str, dict[str, Any]] = {}
    table_failures: list[dict[str, Any]] = []

    for page_index in range(page_count):
        page_number = page_index + 1
        log(f"[2/4] 渲染并定位第{page_number}/{page_count}页")
        source_image = render_pdf_page(pdf_path, page_index, args.render_width)
        region_detection = detect_table_regions_with_diagnostics(
            source_image, page_number
        )
        regions = region_detection.regions
        initial_region_count = len(regions)
        needs_resolution_probe = bool(
            not regions
            or region_detection.strong_rejected_candidate_count > 0
        )
        if needs_resolution_probe and args.render_width < 2200:
            log(
                f"[2/4] 第{page_number}页初检得到{len(regions)}张表、"
                f"{region_detection.strong_rejected_candidate_count}个"
                "强候选未通过，自动提高到2200像素宽复检"
            )
            high_resolution_image = render_pdf_page(
                pdf_path, page_index, 2200
            )
            high_resolution_detection = (
                detect_table_regions_with_diagnostics(
                    high_resolution_image, page_number
                )
            )
            if len(high_resolution_detection.regions) > len(regions):
                source_image = high_resolution_image
                region_detection = high_resolution_detection
                regions = region_detection.regions
                log(
                    f"[2/4] 第{page_number}页采用2200像素结果："
                    f"表数由{initial_region_count}"
                    f"提升到{len(regions)}张。"
                )
            else:
                log(
                    f"[2/4] 第{page_number}页高分辨率没有增加完整表区，"
                    "保留原分辨率结果，避免无条件放大改变锚点节奏。"
                )
        if regions:
            assign_region_boundary_scan_limits(regions, source_image.shape[0])
        cv2.imwrite(
            str(structure_dir / f"page_{page_number:03d}_source.png"), source_image
        )
        work_items: list[tuple[np.ndarray, PageGeometry, dict[str, Any], str]] = []
        if len(regions) == 1:
            try:
                geometry = detect_page_geometry(source_image, page_number, 1)
                heading = recognize_page_heading(
                    source_image, geometry, primary_recognizer
                )
                work_items.append((source_image, geometry, heading, "original"))
            except RuntimeError as direct_error:
                log(
                    f"[表区诊断] 第{page_number}页直线网格路径未通过："
                    f"{direct_error}；改用非刚性曲线网格。"
                )
        if not work_items:
            if not regions:
                failure = build_page_detection_failure_payload(
                    page_number,
                    region_detection,
                    source_image.shape[1],
                )
                table_failures.append(failure)
                failure_path = structure_dir / (
                    f"page_{page_number:03d}_detection_failure.json"
                )
                failure_path.write_text(
                    json.dumps(failure, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
                draw_source_curve_mesh_preview(
                    source_image,
                    [],
                    previews_dir
                    / f"page_{page_number:03d}_source_curve_mesh.png",
                )
                log(
                    f"[表区失败] 第{page_number}页没有完整表区，"
                    f"已记录诊断并继续下一页：{failure_path}"
                )
                continue
            for region in regions:
                stage = "rectification"
                try:
                    rectified = rectify_table_region(source_image, region)
                    region.image = rectified
                    stage = "geometry"
                    geometry = detect_rectified_geometry(rectified, region)
                    stage = "heading"
                    heading = recognize_region_heading(
                        source_image, region, primary_recognizer
                    )
                    work_items.append(
                        (rectified, geometry, heading, "photo-dewarped")
                    )
                except Exception as error:
                    failure = build_table_failure_payload(
                        region, stage, error, source_image.shape[1]
                    )
                    table_failures.append(failure)
                    failure_path = structure_dir / (
                        f"page_{page_number:03d}_table_"
                        f"{region.table_index:02d}_failure.json"
                    )
                    failure_path.write_text(
                        json.dumps(failure, ensure_ascii=False, indent=2),
                        encoding="utf-8",
                    )
                    log(
                        f"[表区失败] 第{page_number}页表{region.table_index}"
                        f"在{stage}阶段失败：{error}；"
                        f"诊断={failure_path}"
                    )
                    continue

        draw_source_curve_mesh_preview(
            source_image,
            regions,
            previews_dir / f"page_{page_number:03d}_source_curve_mesh.png",
        )
        log(f"[2/4] 第{page_number}页检测到{len(work_items)}张站表")
        for table_image, geometry, heading, geometry_mode in work_items:
            table_index = geometry.table_index
            log(
                f"[2/4] 第{page_number}页表{table_index}："
                f"站点={heading['station']}，候选锚点="
                f"{len(geometry.candidate_anchors)}，最终锚点="
                f"{len(geometry.row_anchors)}"
            )
            try:
                rows, samples, month_lengths = recognize_daily_cells(
                    table_image,
                    geometry,
                    heading,
                    primary_recognizer,
                    secondary_recognizer,
                    training_images_dir,
                    args.batch_size,
                    args.low_confidence,
                )
            except Exception as error:
                matching_region = next(
                    (
                        item
                        for item in regions
                        if item.table_index == table_index
                    ),
                    None,
                )
                if matching_region is not None:
                    failure = build_table_failure_payload(
                        matching_region,
                        "cell_recognition",
                        error,
                        source_image.shape[1],
                    )
                else:
                    failure = {
                        "page": page_number,
                        "table_index": table_index,
                        "stage": "cell_recognition",
                        "error_type": type(error).__name__,
                        "message": str(error),
                        "render_width": source_image.shape[1],
                        "traceback": traceback.format_exc().splitlines()[-24:],
                    }
                table_failures.append(failure)
                failure_path = structure_dir / (
                    f"page_{page_number:03d}_table_"
                    f"{table_index:02d}_failure.json"
                )
                failure_path.write_text(
                    json.dumps(failure, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
                log(
                    f"[表区失败] 第{page_number}页表{table_index}"
                    f"在单元格识别阶段失败：{error}；诊断={failure_path}"
                )
                continue
            geometry_json = geometry_to_json(geometry)
            geometry_json["heading"] = heading
            geometry_json["month_lengths"] = month_lengths
            geometry_json["geometry_mode"] = geometry_mode
            geometry_path = structure_dir / (
                f"page_{page_number:03d}_table_{table_index:02d}_geometry.json"
            )
            geometry_path.write_text(
                json.dumps(geometry_json, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            cv2.imwrite(
                str(
                    structure_dir
                    / f"page_{page_number:03d}_table_{table_index:02d}_rectified.png"
                ),
                table_image,
            )
            draw_anchor_preview(
                table_image,
                geometry,
                previews_dir
                / f"page_{page_number:03d}_table_{table_index:02d}_anchors.png",
            )
            draw_cell_preview(
                table_image,
                geometry,
                rows,
                previews_dir
                / f"page_{page_number:03d}_table_{table_index:02d}_cells.png",
            )
            table_crop = table_image[
                max(0, geometry.table_top) : min(
                    geometry.height, geometry.statistics_top + 2
                ),
                max(0, geometry.table_left) : min(
                    geometry.width, geometry.table_right + 1
                ),
            ]
            cv2.imwrite(
                str(
                    previews_dir
                    / f"page_{page_number:03d}_table_{table_index:02d}_table.png"
                ),
                table_crop,
            )
            entry = {
                "page": page_number,
                "table_index": table_index,
                "station": heading["station"],
                "row_count": len(rows),
                "month_lengths": month_lengths,
                "pending_count": sum(1 for row in rows if row["待确认"]),
                "geometry_mode": geometry_mode,
                "geometry": str(geometry_path),
            }
            # One detected table is one annual dataset.  Repeated Chinese
            # station names are deliberately not merged because adjacent PDF
            # pages can represent different years whose year OCR is deferred.
            dataset_key = f"page_{page_number:03d}_table_{table_index:02d}"
            if dataset_key in dataset_groups:
                raise RuntimeError(f"年度表键重复：{dataset_key}")
            group = {
                "station": heading["station"],
                "heading": heading,
                "rows": [],
                "samples": [],
                "geometries": [],
                "month_lengths": month_lengths,
                "tables": [],
            }
            dataset_groups[dataset_key] = group
            group["rows"].extend(rows)
            group["samples"].extend(samples)
            group["geometries"].append(geometry)
            group["tables"].append(entry)
            all_rows.extend(rows)
            all_samples.extend(samples)

    # The scan is traversed by day then month to preserve row-anchor locality.
    # Final hydrological records must be chronological by month then day.
    all_rows.sort(
        key=lambda row: (row["页"], row["表区"], row["月"], row["日"])
    )

    if args.corrected_labels:
        corrected_path = args.corrected_labels.expanduser().resolve()
        if not corrected_path.is_file():
            raise FileNotFoundError(f"标签校正表不存在：{corrected_path}")
        applied = apply_corrected_labels(all_samples, corrected_path)
        synchronize_rows_from_samples(all_rows, all_samples)
        log(f"[3/4] 已应用{applied}处人工校正标签")

    station_totals = Counter(
        str(group["station"]) for group in dataset_groups.values()
    )
    station_seen: Counter[str] = Counter()
    manifest_stations: list[dict[str, Any]] = []
    for dataset_key, group in dataset_groups.items():
        station = str(group["station"])
        station_seen[station] += 1
        occurrence = station_seen[station]
        output_name = station_output_name(
            station, occurrence, station_totals[station]
        )
        group["rows"].sort(
            key=lambda row: (row["页"], row["表区"], row["月"], row["日"])
        )
        workbook_path = write_station_workbook(
            output_dir,
            station,
            output_name,
            group["heading"],
            group["rows"],
            group["geometries"],
            group["month_lengths"],
        )
        station_json_path = output_dir / f"{safe_filename(output_name)}.json"
        station_json_path.write_text(
            json.dumps(
                {
                    "input_pdf": str(pdf_path),
                    "station": station,
                    "dataset_key": dataset_key,
                    "output_name": output_name,
                    "station_occurrence": occurrence,
                    "station_occurrence_count": station_totals[station],
                    "heading": group["heading"],
                    "row_count": len(group["rows"]),
                    "pending_count": sum(
                        1 for row in group["rows"] if row["待确认"]
                    ),
                    "tables": group["tables"],
                    "rows": group["rows"],
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        manifest_stations.append(
            {
                "station": station,
                "dataset_key": dataset_key,
                "output_name": output_name,
                "station_occurrence": occurrence,
                "station_occurrence_count": station_totals[station],
                "row_count": len(group["rows"]),
                "pending_count": sum(
                    1 for row in group["rows"] if row["待确认"]
                ),
                "excel": str(workbook_path),
                "json": str(station_json_path),
            }
        )
    review_path = write_training_review_workbook(all_samples, training_dir)
    manifest_path = output_dir / f"{pdf_path.stem}_stations.json"
    run_status = (
        "complete"
        if not table_failures
        else ("partial" if dataset_groups else "failed")
    )
    manifest_path.write_text(
        json.dumps(
            {
                "input_pdf": str(pdf_path),
                "status": run_status,
                "station_count": len(station_totals),
                "dataset_count": len(dataset_groups),
                "failed_table_count": len(table_failures),
                "failures": table_failures,
                "stations": manifest_stations,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    log(
        f"[4/4] {run_status}：{len(station_totals)}个站点，"
        f"{len(dataset_groups)}张年度表，"
        f"{len(all_rows)}条逐日流量，"
        f"{len(table_failures)}张表失败"
    )
    for item in manifest_stations:
        log(f"[完成] {item['output_name']} Excel：{item['excel']}")
        log(f"[完成] {item['output_name']} JSON：{item['json']}")
    log(f"[完成] 锚点/分格预览：{previews_dir}")
    log(f"[完成] 标签校正表：{review_path}")
    log(f"[完成] 站点清单：{manifest_path}")
    if not dataset_groups:
        raise RuntimeError(
            f"没有任何站表完成处理；已保存{len(table_failures)}份失败诊断。"
        )


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        log("已取消。")
        raise SystemExit(130)
