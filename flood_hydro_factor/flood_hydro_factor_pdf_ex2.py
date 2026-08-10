"""Process legacy scanned flood-hydrology factor tables with 5-7 fields.

This second extended variant keeps the handwritten dual-anchor reconstruction
of flood_hydro_factor_pdf_ex.py, dynamically detects both the repeated-panel
count and whether each panel has five, six, or seven physical columns, and
separates multiple station sections even when a new title appears mid-page.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import OrderedDict
from datetime import datetime
from pathlib import Path
from typing import Any

import cv2
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
# Cached local models should remain usable without a connectivity probe.
os.environ.setdefault("PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK", "True")

from paddleocr import TextRecognition
from scipy.signal import find_peaks

from flood_hydro_factor_pdf import (
    expand_water_level,
    extract_station_name,
    load_saved_structure,
    log,
    parse_decimal,
    parse_integer,
    run_structure_ocr,
    safe_filename,
    verify_device,
)
from reconstruct_grid_table import (
    load_page,
    normalize_numeric,
    prepare_cell,
    result_value,
)


LEGACY_COLUMN_NAMES = [
    "月",
    "日",
    "时分",
    "水位(m)",
    "流量(m³/s)",
    "含沙量(kg/m³)",
    "水面比降(万分率)",
]
SAMPLE_COLUMN_SLUGS = [
    "month",
    "day",
    "time",
    "water_level",
    "flow",
    "sediment",
    "water_slope",
]
CORE_LEGACY_COLUMN_INDICES = [0, 1, 2, 3, 4]
OPTIONAL_LEGACY_COLUMN_INDICES = [5, 6]
SUPPORTED_LEGACY_FIELD_COUNTS = (5, 6, 7)
SUPPORTED_LEGACY_PANEL_COUNTS = tuple(range(1, 7))
LEGACY_STATION_TITLE_GAP_MIN_PX = 90.0
LEGACY_HEADER_ALIASES = {
    0: ("月",),
    1: ("日",),
    2: ("时分", "时间", "時分", "附分"),
    3: ("水位", "水立"),
    4: ("流量", "流童", "流景"),
    5: ("含沙量", "含沙", "沙量", "含砂量", "含砂"),
    6: ("水面比降", "比降", "万分率", "万分中"),
}
DEFAULT_FINETUNED_NUMERIC_MODEL_DIR = (
    Path(__file__).resolve().parent
    / "models"
    / "hydrology_numeric_ppocrv5_mobile_v11"
)

# OCR crops include one pixel outside the fitted boundary so handwriting that
# touches a printed rule is not discarded before it can be analysed.  Edge
# cleaning below removes only locally observed, long vertical rule strokes;
# it no longer whitens a fixed-width band inside every cell.
LEGACY_OCR_CELL_OUTSET_PX = 1
LEGACY_EDGE_RULE_SEARCH_PX = 7
LEGACY_EDGE_RULE_MIN_HEIGHT_RATIO = 0.72
LEGACY_EDGE_RULE_FRAGMENT_MIN_HEIGHT_RATIO = 0.52
LEGACY_EDGE_RULE_MAX_HORIZONTAL_PER_VERTICAL = 0.35
LEGACY_EDGE_RULE_MAX_FIT_RESIDUAL_PX = 1.25
LEGACY_GEOMETRY_EDGE_CLEAR_PX = 3
LEGACY_DAY_PAIR_REVIEW_CONFIDENCE = 0.70
LEGACY_TIME_COLON_FALLBACK_RATIO = 0.46
LEGACY_LOCAL_ROW_MAX_OFFSET = 3
LEGACY_OWNED_CELL_MIN_AREA = 8
LEGACY_WATER_DOT_X_TOLERANCE = 2.0
LEGACY_WATER_DOT_TO_GLYPH_CENTER = -3
LEGACY_TIME_COLON_CROP_MARGIN = 1
LEGACY_TIME_COLON_SEARCH_LEFT_RATIO = 0.25
LEGACY_TIME_COLON_SEARCH_RIGHT_RATIO = 0.62
LEGACY_TIME_COLON_CLUSTER_TOLERANCE = 2.5
LEGACY_TIME_COLON_X_TOLERANCE = 2.0
LEGACY_TIME_SPLIT_MINUTE_CONSTRAINT_CONFIDENCE = 0.85
LEGACY_TIME_MINUTE_RECONCILIATION_PENALTY = 0.05
LEGACY_DUAL_ANCHOR_MATCH_MAX_OFFSET = 7
LEGACY_DUAL_ANCHOR_SKIP_COST = 1.2
LEGACY_WATER_SPECK_MAX_AREA = 6
LEGACY_WATER_DOT_SPARSE_RATIO = 0.55
LEGACY_WATER_TOKEN_MIN_AREA = 8
LEGACY_WATER_TOKEN_MIN_SEPARATION = 7
LEGACY_SLOPE_MIN_CONFIDENCE = 0.55
LEGACY_DAY_CELL_MIN_AREA = 24
LEGACY_MONTH_CELL_MIN_AREA = 20
LEGACY_DATE_COMPONENT_MIN_AREA = 5
LEGACY_TRUNCATED_DIGIT_PENALTY = 0.22
LEGACY_DIGIT_LENGTH_BONUS = 0.10
LEGACY_NUMERIC_PURITY_BONUS = 0.08
LEGACY_EDGE_PREFIX_GAP_RATIO = 0.18
LEGACY_EDGE_PREFIX_GAP_OUTLIER = 1.8
LEGACY_EDGE_PREFIX_GAP_PENALTY = 0.45

LEGACY_NUMERIC_TRANSLATION = str.maketrans(
    {
        "O": "0",
        "o": "0",
        "Q": "0",
        "C": "0",
        "c": "0",
        "D": "0",
        "U": "0",
        "u": "0",
        "V": "0",
        "v": "0",
        "a": "0",
        "α": "0",
        ")": "0",
        "I": "1",
        "i": "1",
        "l": "1",
        "|": "1",
        "!": "1",
        "J": "1",
        "Z": "2",
        "z": "2",
        "P": "2",
        "A": "4",
        "H": "4",
        "h": "4",
        "S": "5",
        "s": "5",
        "G": "6",
        "T": "7",
        "t": "7",
        "B": "8",
        "b": "8",
        "g": "9",
        "q": "9",
    }
)


def clustered_positions(values: list[float], tolerance: float = 5.0) -> list[float]:
    clusters: list[list[float]] = []
    for value in sorted(values):
        if not clusters or value - float(np.mean(clusters[-1])) > tolerance:
            clusters.append([value])
        else:
            clusters[-1].append(value)
    return [float(np.mean(cluster)) for cluster in clusters]


def infer_legacy_panel_layout(
    line_x: list[float],
) -> tuple[int, int, list[list[float]], dict[str, Any]]:
    """Jointly infer repeated-panel count and fields per panel from rules."""
    interval_count = len(line_x) - 1
    candidates: list[dict[str, Any]] = []
    for field_count in SUPPORTED_LEGACY_FIELD_COUNTS:
        if interval_count % field_count != 0:
            continue
        panel_count = interval_count // field_count
        if panel_count not in SUPPORTED_LEGACY_PANEL_COUNTS:
            continue
        panels = [
            line_x[
                panel_index * field_count:
                panel_index * field_count + field_count + 1
            ]
            for panel_index in range(panel_count)
        ]
        normalized_widths: list[np.ndarray] = []
        panel_widths: list[float] = []
        for boundaries in panels:
            widths = np.diff(np.asarray(boundaries, dtype=float))
            panel_width = float(np.sum(widths))
            if panel_width <= 0 or np.any(widths <= 3):
                break
            panel_widths.append(panel_width)
            normalized_widths.append(widths / panel_width)
        if len(normalized_widths) != panel_count:
            continue
        width_matrix = np.vstack(normalized_widths)
        width_template = np.median(width_matrix, axis=0)
        pattern_error = float(np.mean(np.abs(width_matrix - width_template)))
        panel_width_error = float(
            np.std(panel_widths) / max(float(np.mean(panel_widths)), 1.0)
        )
        candidates.append(
            {
                "panel_count": panel_count,
                "field_count": field_count,
                "panels": panels,
                "pattern_error": pattern_error,
                "panel_width_error": panel_width_error,
                "score": pattern_error + panel_width_error * 0.5,
            }
        )

    if not candidates:
        expected = sorted(
            {
                panel_count * field_count + 1
                for panel_count in SUPPORTED_LEGACY_PANEL_COUNTS
                for field_count in SUPPORTED_LEGACY_FIELD_COUNTS
            }
        )
        raise RuntimeError(
            "无法将长竖线划分为1至6个重复表区、每区5至7字段；"
            f"当前检测到{len(line_x)}条，支持的竖线总数为{expected}。"
        )

    candidates.sort(key=lambda item: (item["score"], -item["panel_count"]))
    chosen = candidates[0]
    return (
        int(chosen["panel_count"]),
        int(chosen["field_count"]),
        chosen["panels"],
        {
            "resolution": "repeated-column-width-pattern",
            "candidate_count": len(candidates),
            "chosen_score": round(float(chosen["score"]), 8),
            "pattern_error": round(float(chosen["pattern_error"]), 8),
            "panel_width_error": round(float(chosen["panel_width_error"]), 8),
            "candidates": [
                {
                    "panel_count": int(item["panel_count"]),
                    "field_count": int(item["field_count"]),
                    "score": round(float(item["score"]), 8),
                }
                for item in candidates
            ],
        },
    )


def detect_legacy_geometry(
    image: np.ndarray,
) -> tuple[tuple[int, int, int, int], list[list[float]], int, int]:
    """Recover a dynamic number of repeated five-, six-, or seven-field panels."""
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    ink = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY_INV)[1]

    vertical = cv2.morphologyEx(
        ink,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (1, 80)),
    )
    count, _, stats, _ = cv2.connectedComponentsWithStats(vertical, 8)
    raw_vertical_components: list[tuple[float, int, int]] = []
    for index in range(1, count):
        x, y, width, height, _ = stats[index]
        if height >= 80 and width <= 12:
            raw_vertical_components.append((x + width / 2, y, height))

    # A station title crossing the table interrupts every inner rule.  OpenCV
    # then returns the upper and lower halves as separate components, neither
    # of which may satisfy the former half-page height test.  Cluster by x and
    # merge their y coverage before deciding whether they form one table rule.
    vertical_clusters: list[list[tuple[float, int, int]]] = []
    for component in sorted(raw_vertical_components, key=lambda item: item[0]):
        if (
            not vertical_clusters
            or component[0]
            - float(np.mean([item[0] for item in vertical_clusters[-1]]))
            > 5.0
        ):
            vertical_clusters.append([component])
        else:
            vertical_clusters[-1].append(component)

    vertical_components: list[tuple[float, int, int]] = []
    for cluster in vertical_clusters:
        intervals = sorted((y, y + height) for _, y, height in cluster)
        merged_intervals: list[list[int]] = []
        for start, end in intervals:
            if not merged_intervals or start > merged_intervals[-1][1] + 2:
                merged_intervals.append([start, end])
            else:
                merged_intervals[-1][1] = max(merged_intervals[-1][1], end)
        coverage = sum(end - start for start, end in merged_intervals)
        top = min(start for start, _ in merged_intervals)
        bottom = max(end for _, end in merged_intervals)
        span = bottom - top
        if coverage >= image.shape[0] * 0.5 and span >= image.shape[0] * 0.5:
            vertical_components.append(
                (
                    float(np.mean([item[0] for item in cluster])),
                    top,
                    span,
                )
            )
    line_x = clustered_positions([item[0] for item in vertical_components])
    panel_count, field_count, panel_boundaries, _ = infer_legacy_panel_layout(
        line_x
    )

    horizontal = cv2.morphologyEx(
        ink,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (80, 1)),
    )
    count, _, stats, _ = cv2.connectedComponentsWithStats(horizontal, 8)
    long_horizontal: list[float] = []
    for index in range(1, count):
        x, y, width, height, _ = stats[index]
        if width >= image.shape[1] * 0.6 and height <= 12:
            long_horizontal.append(y + height / 2)
    line_y = clustered_positions(long_horizontal)
    if len(line_y) < 2:
        raise RuntimeError("没有检测到历史表格的上下外框。")

    table_left = round(line_x[0])
    table_right = round(line_x[-1])
    table_top = round(min(line_y))
    table_bottom = round(max(line_y))
    outer_tolerance = 10
    inner_starts = [
        max(cluster, key=lambda component: component[2])[1]
        for cluster in vertical_clusters
        if (
            table_left + outer_tolerance
            < float(np.mean([component[0] for component in cluster]))
            < table_right - outer_tolerance
            and max(component[1] + component[2] for component in cluster)
            - min(component[1] for component in cluster)
            >= (table_bottom - table_top) * 0.5
        )
    ]
    if not inner_starts:
        raise RuntimeError("无法从内部竖线确定正文起点。")
    vertical_start = float(np.median(inner_starts))
    inner_horizontal = [
        y
        for y in line_y
        if table_top + 20 <= y <= table_bottom - 20
    ]
    if inner_horizontal:
        header_rule_y = min(inner_horizontal)
        # On a continuation page the field rules already start at the table's
        # top border.  The full-width header rule is then the reliable boundary
        # between printed headers and handwritten data.  A first-page station
        # title makes the field rules start much lower, so retain that start.
        data_top = round(
            vertical_start
            if vertical_start >= header_rule_y + 20
            else header_rule_y + 8
        )
    else:
        data_top = round(vertical_start)
    data_top = refine_legacy_data_top_from_first_anchors(
        image,
        panel_boundaries,
        data_top,
        table_bottom,
    )
    return (
        (table_left, table_top, table_right, table_bottom),
        panel_boundaries,
        data_top,
        field_count,
    )


def refine_legacy_data_top_from_first_anchors(
    image: np.ndarray,
    panel_boundaries: list[list[float]],
    provisional_top: int,
    table_bottom: int,
) -> int:
    """Move the body start below a station band using paired first anchors."""
    paired_first_rows: list[float] = []
    for boundaries in panel_boundaries:
        try:
            water_rows, _, _ = find_rows_from_water_decimal_points(
                image,
                boundaries,
                provisional_top,
                table_bottom,
            )
        except RuntimeError:
            water_rows = []
        try:
            _, _, time_rows, _, _ = detect_legacy_time_anchors(
                image,
                boundaries,
                provisional_top,
                table_bottom,
            )
        except RuntimeError:
            time_rows = []
        if (
            water_rows
            and time_rows
            and abs(int(water_rows[0]) - int(time_rows[0])) <= 20
        ):
            paired_first_rows.append((water_rows[0] + time_rows[0]) / 2)

    if not paired_first_rows:
        return provisional_top
    first_row = float(np.percentile(paired_first_rows, 25))
    anchor_based_top = round(first_row - 14)
    return min(table_bottom - 20, max(provisional_top, anchor_based_top))


def normalize_legacy_header_text(text: str) -> str:
    return re.sub(
        r"[\s\(\)（）\[\]【】._:：/\\·,，。]+",
        "",
        str(text),
    )


def legacy_header_semantic_score(text: str, canonical_index: int) -> float:
    normalized = normalize_legacy_header_text(text)
    if not normalized:
        return 0.0
    score = 0.0
    for alias in LEGACY_HEADER_ALIASES[canonical_index]:
        normalized_alias = normalize_legacy_header_text(alias)
        if normalized_alias in normalized:
            score = max(score, 1.0 + 0.35 * len(normalized_alias))
        elif normalized_alias[::-1] in normalized:
            score = max(score, 0.75 + 0.25 * len(normalized_alias))
    if canonical_index == 5 and any(char in normalized for char in "沙砂"):
        score = max(score, 1.0)
    if canonical_index == 6:
        if "比降" in normalized:
            score = max(score, 2.0)
        elif "万分" in normalized:
            score = max(score, 1.25)
    return score


def recognize_dynamic_legacy_schema(
    image: np.ndarray,
    result: dict[str, Any],
    recognizer: TextRecognition,
    table_bbox: tuple[int, int, int, int],
    panel_boundaries: list[list[float]],
    data_top: int,
    field_count: int,
    batch_size: int,
    sixth_field: str = "auto",
) -> tuple[list[int], dict[str, Any]]:
    """Map physical columns to the canonical seven-field schema.

    The five leading columns are structural anchors and remain in the legacy
    order.  Header OCR determines which optional field occupies the sixth
    column; a seven-column table can also recover swapped optional columns.
    """
    table_top = table_bbox[1]
    source_ocr = result["overall_ocr_res"]
    evidence: dict[int, list[dict[str, Any]]] = {
        index: [] for index in range(field_count)
    }

    for panel_index, boundaries in enumerate(panel_boundaries):
        for physical_index in range(field_count):
            left = round(boundaries[physical_index]) + 1
            right = round(boundaries[physical_index + 1]) - 1
            fragments: list[tuple[float, float, str, float]] = []
            for text, box, score in zip(
                source_ocr["rec_texts"],
                source_ocr["rec_boxes"],
                source_ocr["rec_scores"],
            ):
                box_left, box_top, box_right, box_bottom = map(float, box)
                center_x = (box_left + box_right) / 2
                center_y = (box_top + box_bottom) / 2
                if (
                    left <= center_x <= right
                    and table_top - 4 <= center_y <= data_top - 2
                ):
                    fragments.append(
                        (box_top, box_left, str(text), float(score))
                    )
            if fragments:
                fragments.sort(key=lambda item: (item[0], item[1]))
                combined = "".join(item[2] for item in fragments)
                weighted_score = sum(item[3] for item in fragments) / len(
                    fragments
                )
                evidence[physical_index].append(
                    {
                        "panel": panel_index + 1,
                        "source": "whole-page-ocr",
                        "text": combined,
                        "score": round(weighted_score, 8),
                    }
                )

    # Whole-page OCR normally contains the printed header.  For an ambiguous
    # optional column, also recognize the physical header crop in both
    # orientations because old forms may print multi-character headers
    # vertically.
    optional_physical_indices = list(range(5, field_count))
    crop_images: list[np.ndarray] = []
    crop_keys: list[tuple[int, int, str]] = []
    for panel_index, boundaries in enumerate(panel_boundaries):
        for physical_index in optional_physical_indices:
            left = round(boundaries[physical_index]) + 2
            right = round(boundaries[physical_index + 1]) - 2
            crop = image[
                max(table_top + 2, 0):min(data_top - 2, image.shape[0]),
                max(left, 0):min(right, image.shape[1]),
            ].copy()
            if crop.size == 0:
                continue
            crop = clear_legacy_geometry_cell_edges(crop)
            variants = {
                "header-original": crop,
                "header-rotate-cw": cv2.rotate(crop, cv2.ROTATE_90_CLOCKWISE),
                "header-rotate-ccw": cv2.rotate(
                    crop, cv2.ROTATE_90_COUNTERCLOCKWISE
                ),
            }
            for variant, variant_image in variants.items():
                enlarged = cv2.resize(
                    variant_image,
                    None,
                    fx=2,
                    fy=2,
                    interpolation=cv2.INTER_CUBIC,
                )
                crop_images.append(enlarged)
                crop_keys.append((panel_index, physical_index, variant))
    if crop_images:
        for key, prediction in zip(
            crop_keys,
            recognizer.predict(
                input=crop_images,
                batch_size=min(batch_size, len(crop_images)),
            ),
        ):
            panel_index, physical_index, variant = key
            text, score = result_value(prediction)
            if text:
                evidence[physical_index].append(
                    {
                        "panel": panel_index + 1,
                        "source": variant,
                        "text": text,
                        "score": round(float(score), 8),
                    }
                )

    semantic_scores: dict[int, dict[int, float]] = {}
    for physical_index, items in evidence.items():
        semantic_scores[physical_index] = {}
        for canonical_index in range(7):
            semantic_scores[physical_index][canonical_index] = sum(
                legacy_header_semantic_score(item["text"], canonical_index)
                * max(0.20, float(item["score"]))
                for item in items
            )

    canonical_indices = CORE_LEGACY_COLUMN_INDICES.copy()
    resolution = "five-core-fields"
    if field_count == 6:
        if sixth_field == "sediment":
            optional_index = 5
            resolution = "command-line-sixth-field=sediment"
        elif sixth_field == "slope":
            optional_index = 6
            resolution = "command-line-sixth-field=slope"
        else:
            sediment_score = semantic_scores[5][5]
            slope_score = semantic_scores[5][6]
            if sediment_score == slope_score == 0:
                texts = [item["text"] for item in evidence[5]]
                raise RuntimeError(
                    "六字段表的第六列表头无法区分含沙量或水面比降；"
                    f"OCR候选={texts}。请用 --sixth-field sediment 或 "
                    "--sixth-field slope 明确指定。"
                )
            optional_index = 5 if sediment_score > slope_score else 6
            resolution = "header-ocr-sixth-field"
        canonical_indices.append(optional_index)
    elif field_count == 7:
        direct_score = semantic_scores[5][5] + semantic_scores[6][6]
        swapped_score = semantic_scores[5][6] + semantic_scores[6][5]
        if swapped_score > direct_score + 0.50:
            canonical_indices.extend([6, 5])
            resolution = "header-ocr-swapped-optional-fields"
        else:
            canonical_indices.extend([5, 6])
            resolution = "header-ocr-or-legacy-optional-order"

    metadata = {
        "physical_field_count": field_count,
        "canonical_indices": canonical_indices,
        "canonical_fields": [
            LEGACY_COLUMN_NAMES[index] for index in canonical_indices
        ],
        "resolution": resolution,
        "header_evidence": evidence,
        "semantic_scores": {
            str(physical_index + 1): {
                LEGACY_COLUMN_NAMES[canonical_index]: round(score, 6)
                for canonical_index, score in scores.items()
                if score > 0
            }
            for physical_index, scores in semantic_scores.items()
        },
    }
    return canonical_indices, metadata


def find_rows_from_ink(
    image: np.ndarray,
    panel_left: int,
    panel_right: int,
    data_top: int,
    data_bottom: int,
) -> list[int]:
    """Infer dense scan rows from horizontal ink peaks after removing rules."""
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    ink = (gray < 180).astype(np.uint8)
    vertical = cv2.morphologyEx(
        ink * 255,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (1, 60)),
    ) > 0
    ink[vertical] = 0
    projection = ink[
        data_top:data_bottom,
        panel_left + 5:panel_right - 5,
    ].sum(axis=1).astype(float)
    smooth = np.convolve(projection, np.ones(3) / 3, mode="same")
    peaks, _ = find_peaks(
        smooth,
        distance=8,
        prominence=3,
        height=3,
    )
    # Projection maxima on this heavy scan sit near the lower stroke edge.
    # Move them upward to the visual character center before cell cropping.
    rows = [int(value + data_top - 3) for value in peaks]
    while len(rows) > 1 and rows[-1] - rows[-2] > 40:
        rows.pop()
    if not rows:
        raise RuntimeError("图像投影没有检测到正文行。")

    completed: list[int] = []
    for current, following in zip(rows[:-1], rows[1:]):
        completed.append(current)
        gap = following - current
        if 19 <= gap <= 29:
            completed.append(round((current + following) / 2))
    completed.append(rows[-1])
    return completed


def find_rows_from_water_decimal_points(
    image: np.ndarray,
    boundaries: list[float],
    data_top: int,
    table_bottom: int,
) -> tuple[list[int], tuple[float, float], list[bool]]:
    """Return only water-level decimal points that are actually observed.

    Missing rows are deliberately *not* interpolated here.  A single anchor
    stream cannot tell a genuinely missing dot from a dirt speck or a locally
    larger line pitch.  Row completion is deferred until the independently
    detected time-colon sequence and multi-column ink evidence are available.
    """
    cell_left = round(boundaries[3]) + 2
    cell_right = round(boundaries[4]) - 2
    cell_width = max(1, cell_right - cell_left)
    column_source = clear_legacy_geometry_cell_edges(
        image[data_top:table_bottom, cell_left:cell_right].copy()
    )
    gray = cv2.cvtColor(column_source, cv2.COLOR_BGR2GRAY)
    ink = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY_INV)[1]
    count, _, stats, centroids = cv2.connectedComponentsWithStats(ink, 8)

    candidates: list[tuple[float, float]] = []
    for component_index in range(1, count):
        _, _, width, height, area = stats[component_index]
        center_x, center_y = map(float, centroids[component_index])
        if (
            cell_width * 0.30 <= center_x <= cell_width * 0.65
            and 2 <= width <= 6
            and 1 <= height <= 6
            and 4 <= area <= 24
        ):
            candidates.append((center_y + data_top, center_x))

    if len(candidates) < 3:
        raise RuntimeError("水位列没有检测到足够的小数点，无法确认行序。")

    y_values = np.array([item[0] for item in candidates], dtype=float)
    x_values = np.array([item[1] for item in candidates], dtype=float)
    inliers = np.ones(len(candidates), dtype=bool)
    for _ in range(5):
        slope, intercept = np.polyfit(y_values[inliers], x_values[inliers], 1)
        residuals = np.abs(x_values - (slope * y_values + intercept))
        updated = residuals <= LEGACY_WATER_DOT_X_TOLERANCE
        if int(np.count_nonzero(updated)) < 3 or np.array_equal(updated, inliers):
            break
        inliers = updated
    slope, intercept = np.polyfit(y_values[inliers], x_values[inliers], 1)

    observed_y = sorted(float(value) for value in y_values[inliers])
    clustered_y: list[list[float]] = []
    for value in observed_y:
        if not clustered_y or value - float(np.mean(clustered_y[-1])) > 4:
            clustered_y.append([value])
        else:
            clustered_y[-1].append(value)
    observed_anchors = [float(np.mean(cluster)) for cluster in clustered_y]
    if len(observed_anchors) < 3:
        raise RuntimeError("水位小数点纵坐标不足，无法建立行序。")

    anchors = [round(value) for value in observed_anchors]
    for index in range(1, len(anchors)):
        if anchors[index] - anchors[index - 1] < 7:
            raise RuntimeError(
                "水位小数点锚点过近："
                f"第{index}、{index + 1}行分别为{anchors[index - 1]}、{anchors[index]}。"
            )
    return anchors, (float(slope), float(intercept)), [True] * len(anchors)


def detect_legacy_water_fraction_token_anchors(
    image: np.ndarray,
    boundaries: list[float],
    water_decimal_model: tuple[float, float],
    data_top: int,
    table_bottom: int,
) -> list[int]:
    """Detect one glyph-center anchor for every written water fraction.

    Historical tables often write ``107.19`` once and then retain only the
    right-aligned fractions ``12`` and ``04``.  The sparse decimal points still
    define the horizontal split corridor, while digit ink to the right of that
    corridor supplies the per-row water sequence.
    """
    cell_left = round(boundaries[3]) + 2
    cell_right = round(boundaries[4]) - 2
    cell_width = max(1, cell_right - cell_left)
    source = clear_legacy_geometry_cell_edges(
        image[data_top:table_bottom, cell_left:cell_right].copy()
    )
    gray = cv2.cvtColor(source, cv2.COLOR_BGR2GRAY)
    ink = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY_INV)[1]
    vertical = cv2.morphologyEx(
        ink,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (1, 24)),
    )
    ink[vertical > 0] = 0
    count, _, stats, centroids = cv2.connectedComponentsWithStats(ink, 8)
    slope, intercept = water_decimal_model

    components: list[dict[str, Any]] = []
    for component_index in range(1, count):
        x, y, width, height, area = map(int, stats[component_index])
        center_x, center_y = map(float, centroids[component_index])
        absolute_y = center_y + data_top
        decimal_x = slope * absolute_y + intercept
        if (
            area >= 3
            and 2 <= height <= 22
            and width < cell_width * 0.62
            and center_x >= decimal_x + 1.5
            and center_x <= cell_width - 1
        ):
            components.append(
                {
                    "left": x,
                    "right": x + width,
                    "top": y + data_top,
                    "bottom": y + height + data_top,
                    "center_y": absolute_y,
                    "area": area,
                }
            )

    clusters: list[list[dict[str, Any]]] = []
    for component in sorted(components, key=lambda item: item["center_y"]):
        if (
            not clusters
            or float(component["center_y"])
            - float(
                np.average(
                    [item["center_y"] for item in clusters[-1]],
                    weights=[item["area"] for item in clusters[-1]],
                )
            )
            > 5.5
        ):
            clusters.append([component])
        else:
            clusters[-1].append(component)

    tokens: list[dict[str, Any]] = []
    for cluster in clusters:
        left = min(int(item["left"]) for item in cluster)
        right = max(int(item["right"]) for item in cluster)
        top = min(int(item["top"]) for item in cluster)
        bottom = max(int(item["bottom"]) for item in cluster)
        area = sum(int(item["area"]) for item in cluster)
        if (
            area >= LEGACY_WATER_TOKEN_MIN_AREA
            and 3 <= bottom - top <= 24
            and right - left < cell_width * 0.58
        ):
            tokens.append(
                {
                    "center_y": float(
                        np.average(
                            [item["center_y"] for item in cluster],
                            weights=[item["area"] for item in cluster],
                        )
                    ),
                    "area": area,
                }
            )

    anchors: list[int] = []
    selected: list[dict[str, Any]] = []
    for token in sorted(tokens, key=lambda item: item["center_y"]):
        center_y = round(float(token["center_y"]))
        if (
            anchors
            and center_y - anchors[-1] < LEGACY_WATER_TOKEN_MIN_SEPARATION
        ):
            if int(token["area"]) > int(selected[-1]["area"]):
                anchors[-1] = center_y
                selected[-1] = token
            continue
        anchors.append(center_y)
        selected.append(token)
    if len(anchors) < 3:
        raise RuntimeError("水位小数区没有检测到足够的数字字符组。")
    return anchors


def legacy_row_bands(
    centers: list[int],
    data_top: int,
    table_bottom: int,
) -> list[tuple[int, int]]:
    """Split rows at adjacent-anchor midpoints without assuming horizontality."""
    normal_gap = float(np.median(np.diff(centers))) if len(centers) > 1 else 12.0
    station_title_gap = max(
        LEGACY_STATION_TITLE_GAP_MIN_PX,
        normal_gap * 2.5,
    )
    bands: list[tuple[int, int]] = []
    for row_index, center_y in enumerate(centers):
        previous_gap = (
            center_y - centers[row_index - 1]
            if row_index > 0
            else normal_gap
        )
        next_gap = (
            centers[row_index + 1] - center_y
            if row_index + 1 < len(centers)
            else normal_gap
        )
        if previous_gap >= station_title_gap:
            previous_gap = normal_gap
        if next_gap >= station_title_gap:
            next_gap = normal_gap
        top = (
            round(center_y - previous_gap / 2)
        )
        bottom = (
            round(center_y + next_gap / 2)
        )
        bands.append((max(data_top, top), min(table_bottom, bottom)))
    return bands


def detect_legacy_time_colon_anchors(
    image: np.ndarray,
    boundaries: list[float],
    data_top: int,
    table_bottom: int,
) -> tuple[int, list[int], list[bool], list[tuple[int, int] | None]]:
    """Build an independent row sequence from the entire time column.

    This detector deliberately does not receive water-level rows.  It first
    finds paired-dot and scan-joined colon shapes in the full time column,
    fits their shared x trajectory, and only then orders them by y.  Pair
    candidates claim their two components exclusively so the lower dot of one
    row cannot be paired with the upper dot of the following row.
    """
    cell_left = round(boundaries[2]) + 2
    cell_right = round(boundaries[3]) - 2
    cell_width = max(1, cell_right - cell_left)
    column_source = clear_legacy_geometry_cell_edges(
        image[data_top:table_bottom, cell_left:cell_right].copy()
    )
    gray = cv2.cvtColor(column_source, cv2.COLOR_BGR2GRAY)
    ink = cv2.threshold(gray, 175, 255, cv2.THRESH_BINARY_INV)[1]
    count, _, stats, centroids = cv2.connectedComponentsWithStats(ink, 8)

    components: list[dict[str, Any]] = []
    for component_index in range(1, count):
        x, _, width, height, area = map(int, stats[component_index])
        center_x, center_y = map(float, centroids[component_index])
        if not (
            area >= 3
            and width <= 5
            and height <= 12
            and cell_width * LEGACY_TIME_COLON_SEARCH_LEFT_RATIO
            <= center_x
            <= cell_width * LEGACY_TIME_COLON_SEARCH_RIGHT_RATIO
        ):
            continue
        components.append(
            {
                "id": component_index,
                "left": x,
                "right": x + width,
                "width": width,
                "height": height,
                "area": area,
                "center_x": center_x,
                "center_y": center_y + data_top,
            }
        )

    dot_components = [
        component
        for component in components
        if component["height"] <= 5 and component["area"] <= 20
    ]
    pair_options: list[dict[str, Any]] = []
    for first_index, first in enumerate(dot_components):
        for second in dot_components[first_index + 1:]:
            upper, lower = sorted(
                (first, second), key=lambda component: component["center_y"]
            )
            vertical_gap = lower["center_y"] - upper["center_y"]
            if (
                abs(first["center_x"] - second["center_x"]) <= 2.5
                and 3 <= vertical_gap <= 10
            ):
                pair_options.append(
                    {
                        "kind": "pair",
                        "component_ids": (first["id"], second["id"]),
                        "left": min(first["left"], second["left"]),
                        "right": max(first["right"], second["right"]),
                        "center_x": (
                            first["center_x"] + second["center_x"]
                        ) / 2,
                        "center_y": (
                            first["center_y"] + second["center_y"]
                        ) / 2,
                        "vertical_gap": vertical_gap,
                        "area": first["area"] + second["area"],
                    }
                )

    short_pair_gaps = [
        float(option["vertical_gap"])
        for option in pair_options
        if option["vertical_gap"] <= 6.5
    ]
    typical_pair_gap = (
        float(np.median(short_pair_gaps)) if short_pair_gaps else 5.0
    )
    expected_x = cell_width * LEGACY_TIME_COLON_FALLBACK_RATIO
    used_component_ids: set[int] = set()
    paired_shapes: list[dict[str, Any]] = []
    for option in sorted(
        pair_options,
        key=lambda item: (
            abs(float(item["vertical_gap"]) - typical_pair_gap),
            abs(float(item["center_x"]) - expected_x),
            -int(item["area"]),
        ),
    ):
        if any(
            component_id in used_component_ids
            for component_id in option["component_ids"]
        ):
            continue
        paired_shapes.append(option)
        used_component_ids.update(option["component_ids"])

    joined_shapes = [
        {
            "kind": "joined",
            "component_ids": (component["id"],),
            "left": component["left"],
            "right": component["right"],
            "center_x": component["center_x"],
            "center_y": component["center_y"],
            "vertical_gap": 0.0,
            "area": component["area"],
        }
        for component in components
        if (
            6 <= component["height"] <= 11
            and 4 <= component["area"] <= 35
            and component["width"] <= 5
        )
    ]
    shapes = paired_shapes + joined_shapes
    if len(shapes) < 3:
        raise RuntimeError("时分列没有独立检测到足够的冒号，无法建立辅助行序。")

    x_clusters: list[list[dict[str, Any]]] = []
    for shape in sorted(shapes, key=lambda item: float(item["center_x"])):
        if (
            not x_clusters
            or float(shape["center_x"])
            - float(
                np.median(
                    [item["center_x"] for item in x_clusters[-1]]
                )
            )
            > LEGACY_TIME_COLON_CLUSTER_TOLERANCE
        ):
            x_clusters.append([shape])
        else:
            x_clusters[-1].append(shape)
    best_cluster = max(
        x_clusters,
        key=lambda cluster: (
            len(cluster),
            -abs(
                float(np.median([item["center_x"] for item in cluster]))
                - expected_x
            ),
        ),
    )
    y_values = np.array(
        [item["center_y"] for item in best_cluster], dtype=float
    )
    x_values = np.array(
        [item["center_x"] for item in best_cluster], dtype=float
    )
    slope, intercept = np.polyfit(y_values, x_values, 1)
    trajectory_shapes = [
        item
        for item in best_cluster
        if abs(
            float(item["center_x"])
            - (slope * float(item["center_y"]) + intercept)
        )
        <= LEGACY_TIME_COLON_X_TOLERANCE
    ]

    y_clusters: list[list[dict[str, Any]]] = []
    for shape in sorted(
        trajectory_shapes, key=lambda item: float(item["center_y"])
    ):
        if (
            not y_clusters
            or float(shape["center_y"])
            - float(
                np.mean([item["center_y"] for item in y_clusters[-1]])
            )
            > 4
        ):
            y_clusters.append([shape])
        else:
            y_clusters[-1].append(shape)
    selected_shapes = [
        min(
            cluster,
            key=lambda item: (
                0 if item["kind"] == "pair" else 1,
                abs(
                    float(item["center_x"])
                    - (slope * float(item["center_y"]) + intercept)
                ),
                -int(item["area"]),
            ),
        )
        for cluster in y_clusters
    ]
    minimum_support = max(3, round(len(selected_shapes) * 0.10))
    if len(selected_shapes) < minimum_support:
        raise RuntimeError("时分冒号纵向轨迹支持不足，无法建立辅助行序。")

    anchors = [round(float(item["center_y"])) for item in selected_shapes]
    colon_bands = [
        (
            max(0, int(item["left"]) - 1),
            min(cell_width, int(item["right"]) + 1),
        )
        for item in selected_shapes
    ]
    colon_x = round(
        float(np.median([item["center_x"] for item in selected_shapes]))
    )
    return colon_x, anchors, [True] * len(anchors), colon_bands


def detect_legacy_time_token_anchors(
    image: np.ndarray,
    boundaries: list[float],
    data_top: int,
    table_bottom: int,
) -> tuple[int, list[int], list[bool], list[tuple[int, int] | None]]:
    """Build a row sequence from complete numeric tokens in the time column.

    Some annual tables print only the hour (for example ``2`` or ``20``) and
    omit the ``:00`` suffix on every row.  Connected digit components on the
    same baseline are therefore grouped into one observed time token.  This
    detector is independent of water-level dots and does not invent a colon.
    """
    cell_left = round(boundaries[2]) + 2
    cell_right = round(boundaries[3]) - 2
    cell_width = max(1, cell_right - cell_left)
    column_source = clear_legacy_geometry_cell_edges(
        image[data_top:table_bottom, cell_left:cell_right].copy()
    )
    gray = cv2.cvtColor(column_source, cv2.COLOR_BGR2GRAY)
    ink = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY_INV)[1]
    vertical = cv2.morphologyEx(
        ink,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (1, 24)),
    )
    ink[vertical > 0] = 0
    count, _, stats, centroids = cv2.connectedComponentsWithStats(ink, 8)

    components: list[dict[str, Any]] = []
    for component_index in range(1, count):
        x, y, width, height, area = map(int, stats[component_index])
        center_x, center_y = map(float, centroids[component_index])
        if (
            area >= 3
            and 2 <= height <= 22
            and width < cell_width * 0.88
            and 1 <= center_x <= cell_width - 1
        ):
            components.append(
                {
                    "left": x,
                    "right": x + width,
                    "top": y + data_top,
                    "bottom": y + height + data_top,
                    "center_x": center_x,
                    "center_y": center_y + data_top,
                    "area": area,
                }
            )

    clusters: list[list[dict[str, Any]]] = []
    for component in sorted(components, key=lambda item: item["center_y"]):
        compatible = [
            cluster
            for cluster in clusters
            if abs(
                float(component["center_y"])
                - float(
                    np.average(
                        [item["center_y"] for item in cluster],
                        weights=[item["area"] for item in cluster],
                    )
                )
            )
            <= 5.5
        ]
        if compatible:
            compatible[-1].append(component)
        else:
            clusters.append([component])

    tokens: list[dict[str, Any]] = []
    for cluster in clusters:
        left = min(int(item["left"]) for item in cluster)
        right = max(int(item["right"]) for item in cluster)
        top = min(int(item["top"]) for item in cluster)
        bottom = max(int(item["bottom"]) for item in cluster)
        area = sum(int(item["area"]) for item in cluster)
        if (
            area >= LEGACY_OWNED_CELL_MIN_AREA
            and 3 <= bottom - top <= 24
            and right - left < cell_width * 0.90
        ):
            tokens.append(
                {
                    "center_x": float(
                        np.average(
                            [item["center_x"] for item in cluster],
                            weights=[item["area"] for item in cluster],
                        )
                    ),
                    "center_y": float(
                        np.average(
                            [item["center_y"] for item in cluster],
                            weights=[item["area"] for item in cluster],
                        )
                    ),
                    "area": area,
                }
            )

    anchors: list[int] = []
    selected_tokens: list[dict[str, Any]] = []
    for token in sorted(tokens, key=lambda item: item["center_y"]):
        y = round(float(token["center_y"]))
        if anchors and y - anchors[-1] < 7:
            if int(token["area"]) > int(selected_tokens[-1]["area"]):
                anchors[-1] = y
                selected_tokens[-1] = token
            continue
        anchors.append(y)
        selected_tokens.append(token)
    if len(anchors) < 3:
        raise RuntimeError("时分列没有检测到足够的数字字符组，无法建立辅助行序。")
    token_x = round(
        float(np.median([item["center_x"] for item in selected_tokens]))
    )
    return token_x, anchors, [True] * len(anchors), [None] * len(anchors)


def detect_legacy_time_anchors(
    image: np.ndarray,
    boundaries: list[float],
    data_top: int,
    table_bottom: int,
) -> tuple[
    str,
    int,
    list[int],
    list[bool],
    list[tuple[int, int] | None],
]:
    """Select colon-split or hour-only time anchoring for one panel."""
    token_result: tuple[
        int, list[int], list[bool], list[tuple[int, int] | None]
    ] | None = None
    colon_result: tuple[
        int, list[int], list[bool], list[tuple[int, int] | None]
    ] | None = None
    try:
        token_result = detect_legacy_time_token_anchors(
            image, boundaries, data_top, table_bottom
        )
    except RuntimeError:
        pass
    try:
        colon_result = detect_legacy_time_colon_anchors(
            image, boundaries, data_top, table_bottom
        )
    except RuntimeError:
        pass

    if colon_result is not None:
        token_count = len(token_result[1]) if token_result is not None else 0
        colon_ratio = len(colon_result[1]) / max(1, token_count)
        if token_result is None or colon_ratio >= 0.35:
            return ("colon_split", *colon_result)
    if token_result is not None:
        return ("hour_only", *token_result)
    raise RuntimeError(
        "时分列既没有稳定冒号轨迹，也没有足够的小时数字字符组。"
    )


def inspect_legacy_water_anchor_components(
    image: np.ndarray,
    boundaries: list[float],
    rows: list[int],
    water_decimal_model: tuple[float, float],
    data_top: int,
    table_bottom: int,
) -> list[dict[str, Any] | None]:
    """Measure the observed component behind each water-row candidate."""
    cell_left = round(boundaries[3]) + 2
    cell_right = round(boundaries[4]) - 2
    cell_width = max(1, cell_right - cell_left)
    column_source = clear_legacy_geometry_cell_edges(
        image[data_top:table_bottom, cell_left:cell_right].copy()
    )
    gray = cv2.cvtColor(column_source, cv2.COLOR_BGR2GRAY)
    ink = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY_INV)[1]
    count, _, stats, centroids = cv2.connectedComponentsWithStats(ink, 8)
    slope, intercept = water_decimal_model
    candidates: list[dict[str, Any]] = []
    for component_index in range(1, count):
        _, _, width, height, area = map(int, stats[component_index])
        center_x, center_y = map(float, centroids[component_index])
        absolute_y = center_y + data_top
        x_residual = abs(center_x - (slope * absolute_y + intercept))
        if (
            cell_width * 0.30 <= center_x <= cell_width * 0.65
            and 2 <= width <= 6
            and 1 <= height <= 6
            and 4 <= area <= 24
            and x_residual <= LEGACY_WATER_DOT_X_TOLERANCE
        ):
            candidates.append(
                {
                    "center_y": absolute_y,
                    "center_x": center_x,
                    "width": width,
                    "height": height,
                    "area": area,
                    "x_residual": x_residual,
                }
            )

    details: list[dict[str, Any] | None] = []
    for row in rows:
        nearby = [
            candidate
            for candidate in candidates
            if abs(float(candidate["center_y"]) - row) <= 3.5
        ]
        details.append(
            min(
                nearby,
                key=lambda candidate: (
                    abs(float(candidate["center_y"]) - row),
                    -int(candidate["area"]),
                ),
            )
            if nearby
            else None
        )
    return details


def align_legacy_anchor_sequences(
    water_rows: list[int],
    time_rows: list[int],
) -> list[tuple[str, int | None, int | None]]:
    """Monotonically align independent water-dot and time-colon sequences."""
    water_count = len(water_rows)
    time_count = len(time_rows)
    infinity = float("inf")
    costs = [
        [infinity] * (time_count + 1) for _ in range(water_count + 1)
    ]
    previous: list[list[tuple[int, int, str] | None]] = [
        [None] * (time_count + 1) for _ in range(water_count + 1)
    ]
    costs[0][0] = 0.0

    for water_index in range(water_count + 1):
        for time_index in range(time_count + 1):
            current_cost = costs[water_index][time_index]
            if not np.isfinite(current_cost):
                continue
            if water_index < water_count:
                candidate_cost = current_cost + LEGACY_DUAL_ANCHOR_SKIP_COST
                if candidate_cost < costs[water_index + 1][time_index]:
                    costs[water_index + 1][time_index] = candidate_cost
                    previous[water_index + 1][time_index] = (
                        water_index,
                        time_index,
                        "water_only",
                    )
            if time_index < time_count:
                candidate_cost = current_cost + LEGACY_DUAL_ANCHOR_SKIP_COST
                if candidate_cost < costs[water_index][time_index + 1]:
                    costs[water_index][time_index + 1] = candidate_cost
                    previous[water_index][time_index + 1] = (
                        water_index,
                        time_index,
                        "time_only",
                    )
            if water_index < water_count and time_index < time_count:
                offset = abs(
                    water_rows[water_index] - time_rows[time_index]
                )
                if offset <= LEGACY_DUAL_ANCHOR_MATCH_MAX_OFFSET:
                    candidate_cost = (
                        current_cost
                        + 0.1
                        + offset / LEGACY_DUAL_ANCHOR_MATCH_MAX_OFFSET
                    )
                    if candidate_cost < costs[water_index + 1][time_index + 1]:
                        costs[water_index + 1][time_index + 1] = candidate_cost
                        previous[water_index + 1][time_index + 1] = (
                            water_index,
                            time_index,
                            "matched",
                        )

    operations: list[tuple[str, int | None, int | None]] = []
    water_index, time_index = water_count, time_count
    while water_index or time_index:
        step = previous[water_index][time_index]
        if step is None:
            raise RuntimeError("主辅锚点行序无法完成单调映射。")
        previous_water, previous_time, action = step
        operations.append(
            (
                action,
                previous_water if action != "time_only" else None,
                previous_time if action != "water_only" else None,
            )
        )
        water_index, time_index = previous_water, previous_time
    operations.reverse()
    return operations


def detect_legacy_multicolumn_row_candidates(
    image: np.ndarray,
    boundaries: list[float],
    data_top: int,
    table_bottom: int,
) -> list[dict[str, Any]]:
    """Detect row occupancy independently in time, water and flow columns.

    This is intentionally weaker than character recognition: it merely asks
    whether several semantic data columns contain handwriting at a compatible
    y coordinate.  Table rules are removed and components are clustered once
    per column before the three streams are fused, preventing one fragmented
    digit from voting several times for a row.
    """
    column_observations: list[dict[str, Any]] = []
    physical_count = len(boundaries) - 1
    for physical_index in range(2, min(5, physical_count)):
        cell_left = round(boundaries[physical_index]) + 2
        cell_right = round(boundaries[physical_index + 1]) - 2
        source = clear_legacy_geometry_cell_edges(
            image[data_top:table_bottom, cell_left:cell_right].copy()
        )
        if source.size == 0:
            continue
        gray = cv2.cvtColor(source, cv2.COLOR_BGR2GRAY)
        ink = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY_INV)[1]
        vertical = cv2.morphologyEx(
            ink,
            cv2.MORPH_OPEN,
            cv2.getStructuringElement(cv2.MORPH_RECT, (1, 24)),
        )
        ink[vertical > 0] = 0
        count, _, stats, centroids = cv2.connectedComponentsWithStats(ink, 8)
        components: list[tuple[float, int]] = []
        for component_index in range(1, count):
            _, _, width, height, area = map(int, stats[component_index])
            _, center_y = map(float, centroids[component_index])
            if (
                area >= 3
                and 2 <= height <= 22
                and width < source.shape[1] * 0.88
            ):
                components.append((center_y + data_top, area))

        clusters: list[list[tuple[float, int]]] = []
        for component in sorted(components):
            if (
                not clusters
                or component[0]
                - float(
                    np.average(
                        [item[0] for item in clusters[-1]],
                        weights=[item[1] for item in clusters[-1]],
                    )
                )
                > 6.0
            ):
                clusters.append([component])
            else:
                clusters[-1].append(component)
        for cluster in clusters:
            area = sum(item[1] for item in cluster)
            if area < LEGACY_OWNED_CELL_MIN_AREA:
                continue
            column_observations.append(
                {
                    "y": float(
                        np.average(
                            [item[0] for item in cluster],
                            weights=[item[1] for item in cluster],
                        )
                    ),
                    "column": physical_index,
                    "area": area,
                }
            )

    fused: list[list[dict[str, Any]]] = []
    for observation in sorted(column_observations, key=lambda item: item["y"]):
        compatible = [
            cluster
            for cluster in fused
            if abs(
                float(observation["y"])
                - float(
                    np.average(
                        [item["y"] for item in cluster],
                        weights=[item["area"] for item in cluster],
                    )
                )
            )
            <= 6.0
            and int(observation["column"])
            not in {int(item["column"]) for item in cluster}
        ]
        if compatible:
            compatible[0].append(observation)
        else:
            fused.append([observation])

    candidates: list[dict[str, Any]] = []
    for cluster in fused:
        columns = sorted({int(item["column"]) for item in cluster})
        total_area = sum(int(item["area"]) for item in cluster)
        candidates.append(
            {
                "y": float(
                    np.average(
                        [item["y"] for item in cluster],
                        weights=[item["area"] for item in cluster],
                    )
                ),
                "columns": columns,
                "support": len(columns),
                "ink_area": total_area,
            }
        )
    return candidates


def legacy_multicolumn_support(
    candidates: list[dict[str, Any]],
    center_y: float,
    tolerance: float = 6.0,
) -> dict[str, Any] | None:
    """Return the strongest nearby multi-column occupancy candidate."""
    nearby = [
        candidate
        for candidate in candidates
        if abs(float(candidate["y"]) - center_y) <= tolerance
    ]
    return (
        max(
            nearby,
            key=lambda item: (
                int(item["support"]),
                int(item["ink_area"]),
                -abs(float(item["y"]) - center_y),
            ),
        )
        if nearby
        else None
    )


def legacy_weighted_median(samples: list[tuple[float, float]]) -> float:
    """Weighted median used for reliability-aware local line pitch."""
    ordered = sorted(samples)
    total = sum(weight for _, weight in ordered)
    threshold = total / 2
    cumulative = 0.0
    for value, weight in ordered:
        cumulative += weight
        if cumulative >= threshold:
            return float(value)
    return float(ordered[-1][0])


def legacy_pitch_samples(
    water_rows: list[int],
    time_rows: list[int],
    ink_candidates: list[dict[str, Any]],
    water_reliability: float,
    time_reliability: float,
) -> list[tuple[float, float, float]]:
    """Collect position, gap and adaptive source weight for local pitch."""
    samples: list[tuple[float, float, float]] = []
    sources = (
        (water_rows, max(0.20, water_reliability)),
        (time_rows, max(0.20, time_reliability)),
        (
            [round(float(item["y"])) for item in ink_candidates if item["support"] >= 2],
            0.75,
        ),
    )
    for positions, weight in sources:
        for first, second in zip(positions[:-1], positions[1:]):
            gap = second - first
            if 8 <= gap <= 26:
                samples.append(((first + second) / 2, float(gap), weight))
    return samples


def legacy_local_pitch(
    samples: list[tuple[float, float, float]],
    center_y: float,
) -> float:
    """Estimate row pitch near a gap instead of imposing a page-wide pitch."""
    local = [
        (gap, weight)
        for sample_y, gap, weight in samples
        if abs(sample_y - center_y) <= 180
    ]
    fallback = [(gap, weight) for _, gap, weight in samples]
    return legacy_weighted_median(local or fallback) if (local or fallback) else 14.0


def reconcile_legacy_dual_anchor_rows(
    image: np.ndarray,
    boundaries: list[float],
    water_rows: list[int],
    water_decimal_model: tuple[float, float],
    decimal_observed: list[bool],
    time_rows: list[int],
    colon_observed: list[bool],
    colon_bands: list[tuple[int, int] | None],
    data_top: int,
    table_bottom: int,
) -> tuple[
    list[int],
    list[int],
    list[bool],
    list[bool],
    list[tuple[int, int] | None],
    list[str],
    dict[str, Any],
]:
    """Build logical rows from observed anchors and joint ink hypotheses.

    Neither anchor stream is allowed to manufacture rows in isolation.  The
    two observed sequences are aligned first; unmatched candidates must pass a
    symmetric multi-column handwriting check.  Only then are large gaps tested
    with reliability-weighted, locally pitched hypotheses, and every inferred
    row must be backed by handwriting in at least two semantic columns.
    """
    water_details = inspect_legacy_water_anchor_components(
        image,
        boundaries,
        water_rows,
        water_decimal_model,
        data_top,
        table_bottom,
    )
    operations = align_legacy_anchor_sequences(water_rows, time_rows)
    matched_offsets = [
        water_rows[water_index] - time_rows[time_index]
        for action, water_index, time_index in operations
        if action == "matched"
        and water_index is not None
        and time_index is not None
    ]
    typical_offset = (
        float(np.median(matched_offsets)) if matched_offsets else 3.0
    )
    ink_candidates = detect_legacy_multicolumn_row_candidates(
        image, boundaries, data_top, table_bottom
    )
    strong_ink_candidates = [
        candidate for candidate in ink_candidates if candidate["support"] >= 2
    ]
    occupancy_count = max(1, len(strong_ink_candidates))
    water_reliability = min(1.0, len(water_rows) / occupancy_count)
    time_reliability = min(1.0, len(time_rows) / occupancy_count)
    pitch_samples = legacy_pitch_samples(
        water_rows,
        time_rows,
        strong_ink_candidates,
        water_reliability,
        time_reliability,
    )
    global_pitch = legacy_local_pitch(
        pitch_samples, (data_top + table_bottom) / 2
    )

    final_water_rows: list[int] = []
    final_time_rows: list[int] = []
    final_decimal_observed: list[bool] = []
    final_colon_observed: list[bool] = []
    final_colon_bands: list[tuple[int, int] | None] = []
    statuses: list[str] = []
    summary = {
        "independent_water_anchor_count": len(water_rows),
        "independent_time_anchor_count": len(time_rows),
        "matched_anchor_count": 0,
        "water_only_anchor_count": 0,
        "time_only_recovered_count": 0,
        "rejected_water_speck_count": 0,
        "rejected_time_candidate_count": 0,
        "rejected_unsupported_anchor_count": 0,
        "merged_close_anchor_count": 0,
        "multicolumn_candidate_count": len(strong_ink_candidates),
        "joint_inferred_row_count": 0,
        "rejected_unsupported_hypothesis_count": 0,
        "water_anchor_reliability": round(water_reliability, 4),
        "time_anchor_reliability": round(time_reliability, 4),
        "global_row_pitch": round(global_pitch, 3),
    }

    for action, water_index, time_index in operations:
        if action == "matched":
            assert water_index is not None and time_index is not None
            final_water_rows.append(water_rows[water_index])
            final_time_rows.append(time_rows[time_index])
            final_decimal_observed.append(decimal_observed[water_index])
            final_colon_observed.append(colon_observed[time_index])
            final_colon_bands.append(colon_bands[time_index])
            statuses.append("matched_observed_dual_anchor")
            summary["matched_anchor_count"] += 1
            continue

        if action == "water_only":
            assert water_index is not None
            detail = water_details[water_index]
            is_tiny_speck = bool(
                detail is not None
                and int(detail["area"]) <= LEGACY_WATER_SPECK_MAX_AREA
            )
            water_y = water_rows[water_index]
            fallback_time_y = round(water_y - typical_offset)
            ink_support = legacy_multicolumn_support(
                strong_ink_candidates, fallback_time_y
            )
            # A lone dot-sized component is not a row.  Conversely, a genuine
            # water dot may survive without a colon only when the same band has
            # handwriting in multiple data columns.
            if is_tiny_speck and (
                ink_support is None or int(ink_support["support"]) < 2
            ):
                summary["rejected_water_speck_count"] += 1
                summary["rejected_unsupported_anchor_count"] += 1
                continue
            if ink_support is None:
                summary["rejected_unsupported_anchor_count"] += 1
                continue
            final_water_rows.append(water_y)
            final_time_rows.append(
                min(table_bottom, max(data_top, fallback_time_y))
            )
            final_decimal_observed.append(decimal_observed[water_index])
            final_colon_observed.append(False)
            final_colon_bands.append(None)
            statuses.append("observed_water_only_multicolumn_confirmed")
            summary["water_only_anchor_count"] += 1
            continue

        assert action == "time_only" and time_index is not None
        time_y = time_rows[time_index]
        inferred_water_y = round(time_y + typical_offset)
        ink_support = legacy_multicolumn_support(
            strong_ink_candidates, time_y
        )
        if ink_support is None or int(ink_support["support"]) < 2:
            summary["rejected_time_candidate_count"] += 1
            summary["rejected_unsupported_anchor_count"] += 1
            continue
        final_water_rows.append(
            min(table_bottom, max(data_top, inferred_water_y))
        )
        final_time_rows.append(time_y)
        final_decimal_observed.append(False)
        final_colon_observed.append(colon_observed[time_index])
        final_colon_bands.append(colon_bands[time_index])
        statuses.append("observed_time_only_multicolumn_confirmed")
        summary["time_only_recovered_count"] += 1

    reconciled_records = sorted(
        zip(
            final_water_rows,
            final_time_rows,
            final_decimal_observed,
            final_colon_observed,
            final_colon_bands,
            statuses,
        ),
        key=lambda item: (item[0], item[1]),
    )
    merged_records: list[tuple[Any, ...]] = []
    for record in reconciled_records:
        local_merge_pitch = legacy_local_pitch(
            pitch_samples,
            float(record[1]),
        )
        minimum_separation = max(7, round(local_merge_pitch * 0.60))
        if (
            not merged_records
            or record[0] - merged_records[-1][0] >= minimum_separation
        ):
            merged_records.append(record)
            continue

        previous = merged_records.pop()
        close_pair = [previous, record]
        # Two independent observed anchors dominate a nearby single-anchor
        # candidate.  This prevents handwriting belonging to adjacent genuine
        # rows from lending false multi-column support to a speck between them.
        primary = max(
            close_pair,
            key=lambda item: (
                int(bool(item[2])) + int(bool(item[3])),
                int(bool(item[3])),
                int(bool(item[2])),
                -abs((item[0] - item[1]) - typical_offset),
            ),
        )
        observed_water = [item for item in close_pair if bool(item[2])]
        water_choice = primary if bool(primary[2]) else (
            min(
                observed_water,
                key=lambda item: abs((item[0] - item[1]) - typical_offset),
            )
            if observed_water
            else primary
        )
        observed_time = [item for item in close_pair if bool(item[3])]
        time_choice = primary if bool(primary[3]) else (
            min(
                observed_time,
                key=lambda item: abs(
                    (water_choice[0] - item[1]) - typical_offset
                ),
            )
            if observed_time
            else primary
        )
        merged_records.append(
            (
                water_choice[0],
                time_choice[1],
                any(bool(item[2]) for item in close_pair),
                any(bool(item[3]) for item in close_pair),
                time_choice[4] if bool(time_choice[3]) else None,
                "merged_close_dual_anchor",
            )
        )
        summary["merged_close_anchor_count"] += 1
    reconciled_records = merged_records

    # Both observed streams can miss the same row.  Fill such a gap only after
    # comparing competing interval-count hypotheses.  A hypothesis is valid
    # only when every inserted position maps to a distinct, two-column ink
    # candidate; unsupported equidistant coordinates are never emitted.
    completed_records: list[tuple[Any, ...]] = []
    for record_index, record in enumerate(reconciled_records):
        completed_records.append(record)
        if record_index + 1 >= len(reconciled_records):
            continue
        following = reconciled_records[record_index + 1]
        first_time_y = float(record[1])
        following_time_y = float(following[1])
        gap = following_time_y - first_time_y
        midpoint = (first_time_y + following_time_y) / 2
        local_pitch = legacy_local_pitch(pitch_samples, midpoint)
        if (
            gap < local_pitch * 1.55
            or gap >= max(LEGACY_STATION_TITLE_GAP_MIN_PX, local_pitch * 4.5)
        ):
            continue

        maximum_intervals = max(2, min(5, round(gap / 8)))
        hypotheses: list[tuple[float, list[dict[str, Any]], int]] = []
        for interval_count in range(2, maximum_intervals + 1):
            expected_step = gap / interval_count
            spacing_penalty = abs(expected_step - local_pitch) / max(local_pitch, 1)
            used_ids: set[int] = set()
            matched_candidates: list[dict[str, Any]] = []
            evidence_penalty = 0.0
            valid = True
            for interval_index in range(1, interval_count):
                expected_y = first_time_y + expected_step * interval_index
                options = [
                    candidate
                    for candidate in strong_ink_candidates
                    if id(candidate) not in used_ids
                    and first_time_y + 6 < float(candidate["y"]) < following_time_y - 6
                    and abs(float(candidate["y"]) - expected_y)
                    <= max(5.0, local_pitch * 0.36)
                ]
                if not options:
                    valid = False
                    break
                choice = min(
                    options,
                    key=lambda item: (
                        abs(float(item["y"]) - expected_y),
                        -int(item["support"]),
                        -int(item["ink_area"]),
                    ),
                )
                used_ids.add(id(choice))
                matched_candidates.append(choice)
                evidence_penalty += (
                    abs(float(choice["y"]) - expected_y)
                    / max(local_pitch, 1)
                    - 0.10 * (int(choice["support"]) - 2)
                )
            if valid:
                # More reliable observed streams increase the cost of inventing
                # extra intervals; rich three-column ink offsets that cost.
                insertion_penalty = (
                    (interval_count - 1)
                    * 0.08
                    * (water_reliability + time_reliability)
                )
                hypotheses.append(
                    (
                        spacing_penalty + evidence_penalty + insertion_penalty,
                        matched_candidates,
                        interval_count,
                    )
                )

        if not hypotheses:
            summary["rejected_unsupported_hypothesis_count"] += 1
            continue
        _, selected_candidates, _ = min(hypotheses, key=lambda item: item[0])
        for candidate in sorted(selected_candidates, key=lambda item: item["y"]):
            inferred_time_y = round(float(candidate["y"]))
            inferred_water_y = round(inferred_time_y + typical_offset)
            completed_records.append(
                (
                    min(table_bottom, max(data_top, inferred_water_y)),
                    min(table_bottom, max(data_top, inferred_time_y)),
                    False,
                    False,
                    None,
                    "joint_inferred_multicolumn_ink",
                )
            )
            summary["joint_inferred_row_count"] += 1
    reconciled_records = sorted(
        completed_records, key=lambda item: (item[0], item[1])
    )

    if reconciled_records:
        (
            final_water_rows,
            final_time_rows,
            final_decimal_observed,
            final_colon_observed,
            final_colon_bands,
            statuses,
        ) = (list(values) for values in zip(*reconciled_records))

    if len(final_water_rows) < 3:
        raise RuntimeError("双锚点核对后有效行数不足。")
    for row_index in range(1, len(final_water_rows)):
        if final_water_rows[row_index] - final_water_rows[row_index - 1] < 7:
            raise RuntimeError(
                "双锚点核对后的行序过近："
                f"{final_water_rows[row_index - 1]}、"
                f"{final_water_rows[row_index]}。"
            )
    return (
        final_water_rows,
        final_time_rows,
        final_decimal_observed,
        final_colon_observed,
        final_colon_bands,
        statuses,
        summary,
    )


def assign_legacy_cells_from_dual_anchors(
    image: np.ndarray,
    boundaries: list[float],
    canonical_indices: list[int],
    water_rows: list[int],
    time_rows: list[int],
    data_top: int,
    table_bottom: int,
) -> tuple[list[list[int | None]], dict[tuple[int, int], dict[str, Any]]]:
    """Use water anchors on the right and exclusive time anchors on the left.

    Empty cells never receive a fitted coordinate.  Day/month components are
    claimed top-to-bottom from the first auxiliary time row they touch, so the
    following row cannot reuse a nearby component from the previous date.
    """
    content_centers = [
        center_y + LEGACY_WATER_DOT_TO_GLYPH_CENTER for center_y in water_rows
    ]
    field_count = len(canonical_indices)
    local_rows: list[list[int | None]] = []
    for index in range(len(water_rows)):
        centers: list[int | None] = [None] * field_count
        centers[2] = time_rows[index]
        centers[3] = water_rows[index]
        local_rows.append(centers)
    owned_cells: dict[tuple[int, int], dict[str, Any]] = {}
    water_bands = legacy_row_bands(content_centers, data_top, table_bottom)
    time_bands = legacy_row_bands(time_rows, data_top, table_bottom)

    # Main anchor: water and every column to its right.  The original row-band
    # pixels go to OCR, but only after a real-ink existence check succeeds.
    for physical_index in range(3, field_count):
        canonical_index = canonical_indices[physical_index]
        cell_left, cell_right = legacy_ocr_cell_bounds(
            image.shape[1], boundaries, physical_index
        )
        for row_index, (band_top, band_bottom) in enumerate(water_bands):
            expected_y = content_centers[row_index]
            raw_cell_source = image[
                band_top:band_bottom, cell_left:cell_right
            ].copy()
            cell_source = clear_legacy_cell_edges(raw_cell_source)
            if cell_source.size == 0:
                continue
            if canonical_index == 6 and is_legacy_isolated_slash(cell_source):
                # A slash in this sparse column is a printed/handwritten
                # empty-field marker, not the numeric value 1.
                continue
            gray = cv2.cvtColor(cell_source, cv2.COLOR_BGR2GRAY)
            threshold = 180
            ink = cv2.threshold(gray, threshold, 255, cv2.THRESH_BINARY_INV)[1]
            count, labels, stats, _ = cv2.connectedComponentsWithStats(ink, 8)
            kept_mask = np.zeros_like(ink)
            for component_index in range(1, count):
                _, _, width, _, area = stats[component_index]
                if area >= 2 and width < cell_source.shape[1] * 0.90:
                    kept_mask[labels == component_index] = 255
            total_area = cv2.countNonZero(kept_mask)
            if total_area < LEGACY_OWNED_CELL_MIN_AREA:
                continue

            projection = (kept_mask > 0).sum(axis=1).astype(float)
            smooth = np.convolve(
                projection, np.ones(3) / 3, mode="same"
            )
            if canonical_index == 3:
                # The decimal point itself is the water-column row anchor.
                local_center = water_rows[row_index]
            else:
                local_center = band_top + int(np.argmax(smooth)) - 3
                local_center = min(
                    expected_y + LEGACY_LOCAL_ROW_MAX_OFFSET,
                    max(expected_y - LEGACY_LOCAL_ROW_MAX_OFFSET, local_center),
                )
            local_rows[row_index][physical_index] = int(local_center)
            raw_variant = cell_source.copy()
            cell_source[:1, :] = 255
            cell_source[-1:, :] = 255
            owned_cells[(row_index, physical_index)] = {
                "image": cell_source,
                "raw_image": raw_variant,
                "left": cell_left,
                "right": cell_right,
                "top": band_top,
                "bottom": band_bottom,
                "ink_area": total_area,
            }

    # Auxiliary anchor: the time column itself.  Its coordinate is the colon
    # corridor y; the whole time-cell ink is still preserved for OCR.
    time_left, time_right = legacy_ocr_cell_bounds(
        image.shape[1], boundaries, 2
    )
    for row_index, (band_top, band_bottom) in enumerate(time_bands):
        raw_cell_source = image[
            band_top:band_bottom, time_left:time_right
        ].copy()
        cell_source = clear_legacy_cell_edges(raw_cell_source)
        gray = cv2.cvtColor(cell_source, cv2.COLOR_BGR2GRAY)
        ink = cv2.threshold(gray, 175, 255, cv2.THRESH_BINARY_INV)[1]
        if cv2.countNonZero(ink) < LEGACY_OWNED_CELL_MIN_AREA:
            local_rows[row_index][2] = None
            continue
        raw_variant = cell_source.copy()
        cell_source[:1, :] = 255
        cell_source[-1:, :] = 255
        owned_cells[(row_index, 2)] = {
            "image": cell_source,
            "raw_image": raw_variant,
            "left": time_left,
            "right": time_right,
            "top": band_top,
            "bottom": band_bottom,
            "ink_area": cv2.countNonZero(ink),
        }

    # Day and month are sparse.  Assign global components from top to bottom;
    # once a component belongs to the first time-anchor row, it is unavailable
    # to all following rows.
    for column_index in (1, 0):
        cell_left, cell_right = legacy_ocr_cell_bounds(
            image.shape[1], boundaries, column_index
        )
        column_source = clear_legacy_cell_edges(
            image[data_top:table_bottom, cell_left:cell_right].copy()
        )
        gray = cv2.cvtColor(column_source, cv2.COLOR_BGR2GRAY)
        ink = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY_INV)[1]
        count, labels, stats, centroids = cv2.connectedComponentsWithStats(ink, 8)
        row_masks = [np.zeros_like(ink) for _ in time_rows]
        for component_index in range(1, count):
            _, _, width, _, area = stats[component_index]
            if (
                area < LEGACY_DATE_COMPONENT_MIN_AREA
                or width >= column_source.shape[1] * 0.90
            ):
                continue
            component_mask = (
                (labels == component_index).astype(np.uint8) * 255
            )
            overlaps = [
                cv2.countNonZero(
                    component_mask[
                        max(0, band_top - data_top):
                        min(ink.shape[0], band_bottom - data_top)
                    ]
                )
                for band_top, band_bottom in time_bands
            ]
            meaningful_area = max(4, round(area * 0.18))
            meaningful_rows = [
                index
                for index, overlap in enumerate(overlaps)
                if overlap >= meaningful_area
            ]
            component_y = float(centroids[component_index][1] + data_top)
            owner = min(
                range(len(time_rows)),
                key=lambda index: (abs(component_y - time_rows[index]), index),
            )
            if len(meaningful_rows) >= 2:
                local_gaps = []
                if owner > 0:
                    local_gaps.append(time_rows[owner] - time_rows[owner - 1])
                if owner + 1 < len(time_rows):
                    local_gaps.append(time_rows[owner + 1] - time_rows[owner])
                local_gap = min(local_gaps) if local_gaps else 24
                single_row_tolerance = max(
                    8.0,
                    min(12.0, float(local_gap) * 0.30),
                )
                if abs(component_y - time_rows[owner]) <= single_row_tolerance:
                    # A tall handwritten digit often crosses the midpoint of
                    # two time-derived row bands.  Its centroid still aligns
                    # with one auxiliary anchor, so keep the component whole
                    # and give it to that row only.  Splitting it would turn
                    # the lower strokes into a false date marker on a blank
                    # following row.
                    row_masks[owner] = cv2.bitwise_or(
                        row_masks[owner], component_mask
                    )
                    continue

                # Only a component whose centroid really lies between rows is
                # treated as vertically joined data.  Each resulting piece is
                # still owned by exactly one row.
                for split_owner in meaningful_rows:
                    band_top, band_bottom = time_bands[split_owner]
                    relative_top = max(0, band_top - data_top)
                    relative_bottom = min(ink.shape[0], band_bottom - data_top)
                    row_masks[split_owner][relative_top:relative_bottom] = cv2.bitwise_or(
                        row_masks[split_owner][relative_top:relative_bottom],
                        component_mask[relative_top:relative_bottom],
                    )
                continue

            if abs(component_y - time_rows[owner]) <= 8:
                row_masks[owner] = cv2.bitwise_or(
                    row_masks[owner], component_mask
                )

        for row_index, (band_top, band_bottom) in enumerate(time_bands):
            relative_top = max(0, band_top - data_top)
            relative_bottom = min(ink.shape[0], band_bottom - data_top)
            band_mask = row_masks[row_index][relative_top:relative_bottom]
            total_area = cv2.countNonZero(band_mask)
            minimum_area = (
                LEGACY_DAY_CELL_MIN_AREA
                if column_index == 1
                else LEGACY_MONTH_CELL_MIN_AREA
            )
            if total_area < minimum_area:
                continue
            source_crop = column_source[relative_top:relative_bottom]
            expanded = cv2.dilate(
                band_mask, np.ones((3, 3), np.uint8), iterations=1
            )
            owned_image = np.full_like(source_crop, 255)
            owned_image[expanded > 0] = source_crop[expanded > 0]
            projection = (band_mask > 0).sum(axis=1).astype(float)
            center_y = band_top + int(round(float(np.average(
                np.arange(projection.size), weights=projection
            ))))
            local_rows[row_index][column_index] = center_y
            owned_image[:1, :] = 255
            owned_image[-1:, :] = 255
            owned_cells[(row_index, column_index)] = {
                "image": owned_image,
                "raw_image": clear_legacy_cell_edges(
                    image[band_top:band_bottom, cell_left:cell_right].copy()
                ),
                "left": cell_left,
                "right": cell_right,
                "top": band_top,
                "bottom": band_bottom,
                "ink_area": total_area,
            }

    return local_rows, owned_cells


def legacy_ocr_cell_bounds(
    image_width: int,
    boundaries: list[float],
    column_index: int,
) -> tuple[int, int]:
    """Return a slightly wide OCR crop around fitted table boundaries."""
    left = max(
        0,
        round(boundaries[column_index]) - LEGACY_OCR_CELL_OUTSET_PX,
    )
    right = min(
        image_width,
        round(boundaries[column_index + 1])
        + LEGACY_OCR_CELL_OUTSET_PX
        + 1,
    )
    return left, max(left + 1, right)


def clear_legacy_geometry_cell_edges(cell: np.ndarray) -> np.ndarray:
    """Keep the established fixed edge suppression for anchor geometry only."""
    if cell.size == 0:
        return cell
    cleaned = cell.copy()
    margin = min(
        LEGACY_GEOMETRY_EDGE_CLEAR_PX,
        max(1, cleaned.shape[1] // 8),
    )
    cleaned[:, :margin] = 255
    cleaned[:, -margin:] = 255
    return cleaned


def clear_legacy_cell_edges(
    cell: np.ndarray,
    clear_left: bool = True,
    clear_right: bool = True,
) -> np.ndarray:
    """Remove only observed long rule strokes near the crop edges.

    A former fixed three-pixel white band could amputate a leading digit when
    handwriting touched the table rule.  The new mask requires a component to
    be vertically continuous through most of the row crop.  A second fitted
    track handles a narrow rule fragment that touches the outer crop boundary
    but is slightly slanted and therefore cannot survive a strict ``(1, N)``
    morphology kernel.  Only pixels close to the fitted track are erased; the
    function never whitens a fixed-width band inside the cell.
    """
    if cell.size == 0:
        return cell
    cleaned = cell.copy()
    height, width = cleaned.shape[:2]
    if height < 3 or width < 3 or not (clear_left or clear_right):
        return cleaned

    gray = cv2.cvtColor(cleaned, cv2.COLOR_BGR2GRAY)
    ink = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY_INV)[1]
    kernel_height = max(
        3,
        min(
            height,
            int(np.ceil(height * LEGACY_EDGE_RULE_MIN_HEIGHT_RATIO)),
        ),
    )
    vertical = cv2.morphologyEx(
        ink,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (1, kernel_height)),
    )
    search_width = min(
        LEGACY_EDGE_RULE_SEARCH_PX,
        max(1, width // 5),
    )
    allowed = np.zeros_like(vertical)
    if clear_left:
        allowed[:, :search_width] = 255
    if clear_right:
        allowed[:, width - search_width:] = 255
    candidates = cv2.bitwise_and(vertical, allowed)
    count, labels, stats, _ = cv2.connectedComponentsWithStats(candidates, 8)
    rule_mask = np.zeros_like(candidates)
    edge_tolerance = max(1, round(height * 0.12))
    for component_index in range(1, count):
        x, y, component_width, component_height, _ = map(
            int, stats[component_index]
        )
        touches_vertical_span = bool(
            y <= edge_tolerance
            and y + component_height >= height - edge_tolerance
        )
        near_allowed_edge = bool(
            (clear_left and x < search_width)
            or (
                clear_right
                and x + component_width > width - search_width
            )
        )
        if touches_vertical_span and near_allowed_edge:
            rule_mask[labels == component_index] = 255

    # Scanned table rules are not always perfectly vertical.  A one-pixel-wide
    # vertical opening misses a rule as soon as its x position drifts between
    # rows.  Inspect original connected ink at the actual crop boundary and fit
    # x = slope*y + intercept for narrow, tall components.  Requiring the
    # component to touch the *outer* boundary is important: a genuine digit 1
    # merely near the edge is preserved, while a clipped/slanted table rule is
    # still removed.
    component_count, component_labels, component_stats, _ = (
        cv2.connectedComponentsWithStats(ink, 8)
    )
    fitted_rule_mask = np.zeros_like(ink)
    minimum_fragment_height = max(
        4,
        int(
            np.ceil(
                height * LEGACY_EDGE_RULE_FRAGMENT_MIN_HEIGHT_RATIO
            )
        ),
    )
    boundary_tolerance = 1
    for component_index in range(1, component_count):
        x, y, component_width, component_height, component_area = map(
            int, component_stats[component_index]
        )
        if component_height < minimum_fragment_height:
            continue
        touches_left_boundary = bool(
            clear_left and x <= boundary_tolerance
        )
        touches_right_boundary = bool(
            clear_right
            and x + component_width >= width - boundary_tolerance
        )
        if not (touches_left_boundary or touches_right_boundary):
            continue
        if touches_left_boundary and x + component_width > search_width + 2:
            continue
        if (
            touches_right_boundary
            and x < width - search_width - 2
        ):
            continue

        # A table rule is narrow relative to its vertical extent.  This also
        # prevents a boundary-touching handwritten glyph joined to other ink
        # from being erased as one large component.
        maximum_component_width = max(
            2,
            int(np.ceil(component_height * 0.40)),
        )
        if component_width > maximum_component_width:
            continue
        if component_area > component_height * max(3, component_width):
            continue

        component_y, component_x = np.where(
            component_labels == component_index
        )
        unique_y = np.unique(component_y)
        if unique_y.size < minimum_fragment_height:
            continue
        row_centers = np.asarray(
            [
                float(np.median(component_x[component_y == row]))
                for row in unique_y
            ],
            dtype=np.float32,
        )
        if unique_y.size >= 2:
            slope, intercept = np.polyfit(
                unique_y.astype(np.float32), row_centers, 1
            )
        else:
            slope, intercept = 0.0, float(row_centers[0])
        if (
            not np.isfinite(slope)
            or not np.isfinite(intercept)
            or abs(float(slope))
            > LEGACY_EDGE_RULE_MAX_HORIZONTAL_PER_VERTICAL
        ):
            continue
        fitted_x = slope * unique_y + intercept
        residual = float(np.percentile(np.abs(row_centers - fitted_x), 90))
        if residual > LEGACY_EDGE_RULE_MAX_FIT_RESIDUAL_PX:
            continue

        fitted_at_top = float(slope * unique_y[0] + intercept)
        fitted_at_bottom = float(slope * unique_y[-1] + intercept)
        if touches_left_boundary and min(fitted_at_top, fitted_at_bottom) > 1.5:
            continue
        if (
            touches_right_boundary
            and max(fitted_at_top, fitted_at_bottom) < width - 2.5
        ):
            continue

        component_mask = np.uint8(
            component_labels == component_index
        ) * 255
        # Remove antialiased shoulders without widening the operation into a
        # rectangular white band.  The vertical extent remains that of the
        # observed fitted component.
        component_mask = cv2.dilate(
            component_mask,
            cv2.getStructuringElement(cv2.MORPH_RECT, (3, 1)),
        )
        fitted_rule_mask = cv2.bitwise_or(
            fitted_rule_mask, component_mask
        )

    rule_mask = cv2.bitwise_or(rule_mask, fitted_rule_mask)
    cleaned[rule_mask > 0] = 255
    return cleaned


def is_legacy_isolated_slash(cell: np.ndarray) -> bool:
    """Detect a lone diagonal slash used as an empty-field marker."""
    if cell.size == 0:
        return False
    gray = cv2.cvtColor(cell, cv2.COLOR_BGR2GRAY)
    ink = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY_INV)[1]
    count, labels, stats, _ = cv2.connectedComponentsWithStats(ink, 8)
    components = [
        index
        for index in range(1, count)
        if int(stats[index, cv2.CC_STAT_AREA]) >= 2
    ]
    if len(components) != 1:
        return False

    component_index = components[0]
    _, _, width, height, area = map(int, stats[component_index])
    if not (
        2 <= width <= 7
        and height >= 6
        and area <= max(24, round(height * 2.5))
    ):
        return False

    y_positions, x_positions = np.where(labels == component_index)
    if (
        np.unique(y_positions).size < 5
        or np.unique(x_positions).size < 2
    ):
        return False
    correlation = float(np.corrcoef(x_positions, y_positions)[0, 1])
    horizontal_per_vertical = float(np.polyfit(y_positions, x_positions, 1)[0])
    return (
        np.isfinite(correlation)
        and abs(correlation) >= 0.80
        and abs(horizontal_per_vertical) >= 0.18
    )


def prepare_legacy_time_part(
    image: np.ndarray,
    bounds: tuple[int, int, int, int],
) -> np.ndarray | None:
    """Prepare an hour or minute crop without altering connected glyphs."""
    left, top, right, bottom = bounds
    cell = image[top:bottom, left:right].copy()
    if cell.size == 0:
        return None
    cell = cv2.copyMakeBorder(
        cell, 2, 2, 2, 2, cv2.BORDER_CONSTANT, value=(255, 255, 255)
    )
    return prepare_cell(cell, (0, 0, cell.shape[1], cell.shape[0]))


def prepare_legacy_water_part(
    image: np.ndarray,
    bounds: tuple[int, int, int, int],
    clear_left: bool = False,
    clear_right: bool = False,
) -> np.ndarray | None:
    """Prepare one side of a water value after excluding the decimal band."""
    left, top, right, bottom = bounds
    cell = image[top:bottom, left:right].copy()
    if cell.size == 0:
        return None
    cell = clear_legacy_cell_edges(
        cell,
        clear_left=clear_left,
        clear_right=clear_right,
    )
    gray = cv2.cvtColor(cell, cv2.COLOR_BGR2GRAY)
    binary = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY)[1]
    if cv2.countNonZero(255 - binary) < 4:
        return None
    prepared = cv2.cvtColor(binary, cv2.COLOR_GRAY2BGR)
    prepared = cv2.copyMakeBorder(
        prepared, 2, 2, 2, 2, cv2.BORDER_CONSTANT, value=(255, 255, 255)
    )
    return prepare_cell(
        prepared, (0, 0, prepared.shape[1], prepared.shape[0])
    )

def prepare_legacy_cell(
    image: np.ndarray,
    bounds: tuple[int, int, int, int],
    column_index: int,
) -> np.ndarray | None:
    left, top, right, bottom = bounds
    cell = clear_legacy_cell_edges(image[top:bottom, left:right])
    if cell.size == 0:
        return None
    gray = cv2.cvtColor(cell, cv2.COLOR_BGR2GRAY)
    if column_index <= 1:
        threshold = 160
    elif column_index == 2:
        threshold = 175
    else:
        threshold = 180
    binary = cv2.threshold(gray, threshold, 255, cv2.THRESH_BINARY)[1]
    if column_index == 2:
        binary = cv2.erode(binary, np.ones((2, 2), np.uint8), iterations=1)
    if cv2.countNonZero(255 - binary) < 4:
        return None
    enlarged = cv2.resize(
        binary, None, fx=5, fy=5, interpolation=cv2.INTER_CUBIC
    )
    enlarged = cv2.copyMakeBorder(
        enlarged,
        15,
        15,
        20,
        20,
        cv2.BORDER_CONSTANT,
        value=255,
    )
    return cv2.cvtColor(enlarged, cv2.COLOR_GRAY2BGR)


def prepare_legacy_date_pair(
    image: np.ndarray, bounds: tuple[int, int, int, int]
) -> np.ndarray | None:
    left, top, right, bottom = bounds
    cell = clear_legacy_cell_edges(
        image[max(0, top - 4):bottom + 2, left:right]
    )
    if cell.size == 0:
        return None
    enlarged = cv2.resize(
        cell, None, fx=5, fy=5, interpolation=cv2.INTER_CUBIC
    )
    return cv2.copyMakeBorder(
        enlarged,
        15,
        15,
        20,
        20,
        cv2.BORDER_CONSTANT,
        value=(255, 255, 255),
    )


def prepare_legacy_raw_cell(
    image: np.ndarray,
    bounds: tuple[int, int, int, int],
    clear_edges: bool = True,
) -> np.ndarray | None:
    left, top, right, bottom = bounds
    cell = image[top:bottom, left:right].copy()
    if clear_edges:
        cell = clear_legacy_cell_edges(cell)
    if cell.size == 0:
        return None
    enlarged = cv2.resize(
        cell, None, fx=5, fy=5, interpolation=cv2.INTER_CUBIC
    )
    return cv2.copyMakeBorder(
        enlarged,
        15,
        15,
        20,
        20,
        cv2.BORDER_CONSTANT,
        value=(255, 255, 255),
    )


def estimate_legacy_digit_run_lower_bound(image: np.ndarray) -> int:
    """Return a conservative lower bound for visible horizontal glyph runs."""
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    ink = cv2.threshold(gray, 160, 255, cv2.THRESH_BINARY_INV)[1]
    active = np.count_nonzero(ink, axis=0) >= max(
        2, round(ink.shape[0] * 0.06)
    )
    minimum_width = max(2, round(ink.shape[0] * 0.035))
    runs: list[tuple[int, int]] = []
    start: int | None = None
    for x, present in enumerate(active):
        if present and start is None:
            start = x
        elif not present and start is not None:
            if x - start >= minimum_width:
                runs.append((start, x))
            start = None
    if start is not None and len(active) - start >= minimum_width:
        runs.append((start, len(active)))
    # Adjacent handwritten digits can touch, so this is deliberately a lower
    # bound rather than an exact character count.
    return min(len(runs), 8)


def estimate_legacy_horizontal_ink_spacing(
    image: np.ndarray,
) -> dict[str, float | int]:
    """Describe horizontal ink groups for detecting an isolated edge prefix."""
    if image.size == 0:
        return {
            "run_count": 0,
            "leading_gap_px": 0.0,
            "leading_gap_height_ratio": 0.0,
            "leading_gap_outlier_ratio": 0.0,
        }
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    ink = cv2.threshold(gray, 175, 255, cv2.THRESH_BINARY_INV)[1]
    y_positions, x_positions = np.where(ink > 0)
    if x_positions.size == 0:
        return {
            "run_count": 0,
            "leading_gap_px": 0.0,
            "leading_gap_height_ratio": 0.0,
            "leading_gap_outlier_ratio": 0.0,
        }
    ink_left = int(x_positions.min())
    ink_right = int(x_positions.max()) + 1
    ink_top = int(y_positions.min())
    ink_bottom = int(y_positions.max()) + 1
    content = ink[ink_top:ink_bottom, ink_left:ink_right]
    active = np.count_nonzero(content, axis=0) >= max(
        2, round(content.shape[0] * 0.05)
    )
    runs: list[tuple[int, int]] = []
    start: int | None = None
    for x, present in enumerate(active):
        if present and start is None:
            start = x
        elif not present and start is not None:
            if x - start >= 2:
                runs.append((start, x))
            start = None
    if start is not None and len(active) - start >= 2:
        runs.append((start, len(active)))
    gaps = [
        max(0, following[0] - previous[1])
        for previous, following in zip(runs[:-1], runs[1:])
    ]
    leading_gap = float(gaps[0]) if gaps else 0.0
    ink_height = max(1, ink_bottom - ink_top)
    other_gaps = [float(value) for value in gaps[1:] if value > 0]
    typical_other_gap = (
        float(np.median(other_gaps)) if other_gaps else 0.0
    )
    outlier_ratio = (
        leading_gap / typical_other_gap
        if typical_other_gap > 0
        else 0.0
    )
    return {
        "run_count": len(runs),
        "leading_gap_px": leading_gap,
        "leading_gap_height_ratio": leading_gap / ink_height,
        "leading_gap_outlier_ratio": outlier_ratio,
    }


def numericize_legacy_text(text: str) -> str:
    return text.translate(LEGACY_NUMERIC_TRANSLATION)


def normalize_legacy_minute_digits(text: str) -> str:
    """Map zero-like OCR glyphs in an unsplit minute image to digits."""
    compact = re.sub(r"\s+", "", text)
    # A pair of touching handwritten zeros is occasionally emitted as one
    # wide w/W glyph. This alias is minute-specific: applying it to general
    # numeric fields would turn legitimate text into a value. Do not map a
    # lone "8" here because reviewed samples show it can mean 10/11/18 too.
    if compact in {"w", "W"}:
        return "00"
    cleaned = numericize_legacy_text(text)
    return "".join(re.findall(r"\d", cleaned))


def normalize_legacy_numeric(text: str, column_index: int) -> str:
    cleaned = numericize_legacy_text(text)
    if column_index != 2:
        return normalize_numeric(cleaned, column_index)

    normalized = normalize_numeric(cleaned, column_index)
    if re.fullmatch(r"\d{2}:\d{2}", normalized):
        normalized_hour, normalized_minute = map(int, normalized.split(":"))
        if (
            0 <= normalized_hour <= 24
            and 0 <= normalized_minute <= 59
            and (normalized_hour < 24 or normalized_minute == 0)
        ):
            return normalized
    digits = "".join(re.findall(r"\d", cleaned))
    hour: int | None = None
    minute: int | None = None
    if len(digits) <= 2 and digits:
        hour, minute = int(digits), 0
    elif len(digits) == 3:
        hour, minute = int(digits[0]), int(digits[1:])
    elif len(digits) == 4:
        hour, minute = int(digits[:2]), int(digits[2:])
        if hour > 24 and digits.endswith("00"):
            hour, minute = int(digits[0]), 0
    if (
        hour is not None
        and minute is not None
        and 0 <= hour <= 24
        and 0 <= minute <= 59
        and (hour < 24 or minute == 0)
    ):
        return f"{hour:02d}:{minute:02d}"
    return ""


def legacy_candidate_is_plausible(
    normalized: str,
    column_index: int,
) -> bool:
    if not normalized:
        return False
    digits = "".join(re.findall(r"\d", normalized))
    if column_index == 0:
        return parse_integer(normalized, 1, 12) is not None
    if column_index == 1:
        return parse_integer(normalized, 1, 31) is not None
    if column_index == 2:
        return bool(re.fullmatch(r"\d{2}:\d{2}", normalized))
    if column_index == 7:
        return 1 <= len(digits) <= 4
    if column_index == 8:
        return bool(digits) and len(digits) <= 2 and int(digits) <= 24
    if column_index == 9:
        return bool(digits) and len(digits) <= 2 and int(digits) <= 59
    if column_index == 12:
        return 1 <= len(digits) <= 4
    if column_index == 13:
        return 1 <= len(digits) <= 2
    return bool(re.fullmatch(r"\d+(?:\.\d+)?", normalized))


def choose_legacy_numeric_candidate(
    candidates: list[tuple[str, float, str, str]],
    column_index: int,
    ink_run_lower_bound: int = 0,
    spacing_by_variant: dict[str, dict[str, float | int]] | None = None,
) -> tuple[tuple[str, float], str, str, str]:
    """Choose a numeric candidate while penalizing visibly truncated text."""
    evaluated: list[dict[str, Any]] = []
    variant_spacing = spacing_by_variant or {}
    for text, score, variant, model in candidates:
        normalized = normalize_legacy_numeric(text, column_index)
        mapped = numericize_legacy_text(text)
        significant = [char for char in mapped if not char.isspace()]
        numeric_characters = sum(
            char.isdigit() or char in ".:" for char in significant
        )
        purity = (
            numeric_characters / len(significant) if significant else 0.0
        )
        digit_count = len(re.findall(r"\d", normalized))
        plausible = legacy_candidate_is_plausible(
            normalized, column_index
        )
        completeness_gap = max(0, ink_run_lower_bound - digit_count)
        rank = (
            float(score)
            + LEGACY_DIGIT_LENGTH_BONUS * min(digit_count, 6)
            + LEGACY_NUMERIC_PURITY_BONUS * purity
            - LEGACY_TRUNCATED_DIGIT_PENALTY * completeness_gap
        )
        evaluated.append(
            {
                "text": text,
                "score": float(score),
                "variant": variant,
                "model": model,
                "normalized": normalized,
                "digit_count": digit_count,
                "plausible": plausible,
                "purity": purity,
                "rank": rank,
                "edge_prefix_risk": False,
                "edge_prefix_reference": "",
                "edge_prefix_gap_ratio": 0.0,
                "edge_prefix_gap_outlier": 0.0,
                "spacing": variant_spacing.get(variant, {}),
            }
        )

    # A wide/raw flow crop can retain one faint left-edge stroke that a
    # digit-biased model reads as an extra leading ``1``.  If the longer raw
    # result is exactly ``1 + primary`` but the independently prepared primary
    # image does not expose enough ink runs for that extra position, treat the
    # prefix as table-edge contamination instead of rewarding its extra digit.
    if column_index == 4:
        primary_plausible = [
            item
            for item in evaluated
            if item["variant"] == "primary" and item["plausible"]
        ]
        for item in evaluated:
            if item["variant"] != "raw" or not item["plausible"]:
                continue
            raw_value = str(item["normalized"])
            references = [
                primary
                for primary in primary_plausible
                if raw_value == "1" + str(primary["normalized"])
            ]
            spacing = dict(item.get("spacing") or {})
            gap_ratio = float(
                spacing.get("leading_gap_height_ratio") or 0.0
            )
            gap_outlier = float(
                spacing.get("leading_gap_outlier_ratio") or 0.0
            )
            strong_gap = bool(
                gap_ratio >= LEGACY_EDGE_PREFIX_GAP_RATIO
                and (
                    gap_outlier == 0.0
                    or gap_outlier >= LEGACY_EDGE_PREFIX_GAP_OUTLIER
                )
            )
            insufficient_ink_support = bool(
                ink_run_lower_bound < int(item["digit_count"])
            )
            if references:
                spacing_penalty = LEGACY_EDGE_PREFIX_GAP_PENALTY * min(
                    2.0,
                    gap_ratio / max(LEGACY_EDGE_PREFIX_GAP_RATIO, 1e-6),
                )
                item["rank"] = float(item["rank"]) - spacing_penalty
                item["edge_prefix_gap_ratio"] = gap_ratio
                item["edge_prefix_gap_outlier"] = gap_outlier
            if references and (strong_gap or insufficient_ink_support):
                reference = max(
                    references,
                    key=lambda candidate: (
                        candidate["score"],
                        candidate["rank"],
                    ),
                )
                item["edge_prefix_risk"] = True
                item["edge_prefix_reference"] = str(
                    reference["normalized"]
                )

    plausible_candidates = [
        item
        for item in evaluated
        if item["plausible"] and not item["edge_prefix_risk"]
    ]
    if not plausible_candidates:
        plausible_candidates = [
            item for item in evaluated if item["plausible"]
        ]
    if not evaluated:
        return ("", 0.0), "", "", ""

    if not plausible_candidates:
        chosen = max(evaluated, key=lambda item: item["score"])
        comparison_pool = evaluated
    else:
        server_candidates = [
            item for item in plausible_candidates if item["model"] == "server"
        ]
        numeric_candidates = [
            item for item in plausible_candidates if item["model"] == "numeric"
        ]
        best_server = (
            max(server_candidates, key=lambda item: (item["rank"], item["score"]))
            if server_candidates
            else None
        )
        best_numeric = (
            max(numeric_candidates, key=lambda item: (item["rank"], item["score"]))
            if numeric_candidates
            else None
        )
        if best_server is None:
            chosen = best_numeric
        elif best_numeric is None:
            chosen = best_server
        elif best_numeric["normalized"] == best_server["normalized"]:
            chosen = max(
                (best_server, best_numeric),
                key=lambda item: (item["rank"], item["score"]),
            )
        elif (
            best_numeric["digit_count"] > best_server["digit_count"]
            and best_numeric["score"] >= 0.45
            and best_numeric["score"] >= best_server["score"] - 0.30
        ):
            # The numeric model may repair a truncated server result.  It is
            # not allowed to replace a same-length conflicting value because
            # confidence scores are not calibrated across the two models.
            chosen = best_numeric
        else:
            chosen = best_server
        comparison_pool = plausible_candidates

    confidence_winner = max(comparison_pool, key=lambda item: item["score"])
    decision_note = ""
    rejected_edge_prefixes = [
        item for item in evaluated if item["edge_prefix_risk"]
    ]
    if rejected_edge_prefixes:
        rejected = max(
            rejected_edge_prefixes,
            key=lambda item: (item["score"], item["rank"]),
        )
        decision_note = (
            "边缘前导1防误判：备用宽图候选="
            f"{rejected['normalized']}（{rejected['model']}/"
            f"{rejected['variant']}，{rejected['score']:.3f}）仅比主图候选="
            f"{rejected['edge_prefix_reference']}多一个前导1，"
            f"主图独立墨迹段下限={ink_run_lower_bound}不足以支持"
            f"{rejected['digit_count']}位；首组间距/字高="
            f"{float(rejected['edge_prefix_gap_ratio']):.2f}，"
            f"相对其余间距={float(rejected['edge_prefix_gap_outlier']):.2f}；"
            f"保留={chosen['normalized']}，"
            "请在校正表复核"
        )
    if (
        not decision_note
        and
        chosen is not confidence_winner
        and chosen["digit_count"] > confidence_winner["digit_count"]
    ):
        decision_note = (
            f"数字完整性优先：采用{chosen['model']}/{chosen['variant']}="
            f"{chosen['text']}（{chosen['score']:.3f}，"
            f"{chosen['digit_count']}位），未采用高置信短候选="
            f"{confidence_winner['text']}（{confidence_winner['score']:.3f}，"
            f"{confidence_winner['digit_count']}位）"
        )
    length_warning = ""
    if (
        ink_run_lower_bound >= 2
        and chosen["digit_count"] < ink_run_lower_bound
    ):
        length_warning = (
            f"疑似长度缺失：图像至少有{ink_run_lower_bound}段数字墨迹，"
            f"候选仅{chosen['digit_count']}位，请重点校对"
        )
    return (
        (str(chosen["text"]), float(chosen["score"])),
        f"{chosen['variant']}:{chosen['model']}",
        decision_note,
        length_warning,
    )


def prepare_legacy_station_crop_variants(
    image: np.ndarray,
    bounds: tuple[int, int, int, int],
) -> list[np.ndarray]:
    """Build conservative station-title crops without amputating edge glyphs."""
    left, top, right, bottom = bounds
    left = max(0, min(image.shape[1], left))
    right = max(left, min(image.shape[1], right))
    top = max(0, min(image.shape[0], top))
    bottom = max(top, min(image.shape[0], bottom))
    source = image[top:bottom, left:right].copy()
    if source.size == 0 or source.shape[0] < 2 or source.shape[1] < 2:
        return []

    height, width = source.shape[:2]
    gray = cv2.cvtColor(source, cv2.COLOR_BGR2GRAY)
    ink = cv2.threshold(gray, 185, 255, cv2.THRESH_BINARY_INV)[1]
    horizontal = cv2.morphologyEx(
        ink,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(
            cv2.MORPH_RECT, (max(15, width // 4), 1)
        ),
    )
    vertical = cv2.morphologyEx(
        ink,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(
            cv2.MORPH_RECT, (1, max(9, round(height * 0.65)))
        ),
    )
    text_mask = cv2.bitwise_and(
        ink,
        cv2.bitwise_not(cv2.bitwise_or(horizontal, vertical)),
    )
    count, _, stats, _ = cv2.connectedComponentsWithStats(text_mask, 8)
    components: list[tuple[int, int, int, int]] = []
    for component_index in range(1, count):
        x, y, component_width, component_height, area = map(
            int, stats[component_index]
        )
        if (
            area >= 4
            and component_width >= 2
            and component_height >= 3
            and component_width < width * 0.35
            and component_height < height * 0.92
        ):
            components.append(
                (x, y, x + component_width, y + component_height)
            )

    crop_bounds: list[tuple[int, int, int, int]] = [
        (0, 0, width, height)
    ]
    if components:
        ink_left = min(item[0] for item in components)
        ink_top = min(item[1] for item in components)
        ink_right = max(item[2] for item in components)
        ink_bottom = max(item[3] for item in components)
        median_height = float(
            np.median([item[3] - item[1] for item in components])
        )
        for horizontal_factor in (1.5, 3.0):
            horizontal_pad = max(
                12, round(median_height * horizontal_factor)
            )
            vertical_pad = max(5, round(median_height * 0.65))
            crop_bounds.append(
                (
                    max(0, ink_left - horizontal_pad),
                    max(0, ink_top - vertical_pad),
                    min(width, ink_right + horizontal_pad),
                    min(height, ink_bottom + vertical_pad),
                )
            )

    variants: list[np.ndarray] = []
    seen: set[tuple[int, int, int, int]] = set()
    for crop_left, crop_top, crop_right, crop_bottom in crop_bounds:
        key = (crop_left, crop_top, crop_right, crop_bottom)
        if key in seen or crop_right - crop_left < 2 or crop_bottom - crop_top < 2:
            continue
        seen.add(key)
        crop = source[crop_top:crop_bottom, crop_left:crop_right].copy()
        # OCR must never see a title glyph flush against its input boundary.
        # This border is added after the conservative crop and does not remove
        # any source pixels.
        border_x = max(10, round(crop.shape[0] * 0.35))
        border_y = max(4, round(crop.shape[0] * 0.12))
        crop = cv2.copyMakeBorder(
            crop,
            border_y,
            border_y,
            border_x,
            border_x,
            cv2.BORDER_CONSTANT,
            value=(255, 255, 255),
        )
        variants.append(
            cv2.resize(crop, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
        )
    return variants


def choose_legacy_station_candidate(
    candidates: list[tuple[str, str, float]],
) -> tuple[str, str, float]:
    """Prefer completeness only within a credible confidence neighbourhood."""
    best_confidence = max(float(item[2]) for item in candidates)
    credible = [
        item
        for item in candidates
        if float(item[2]) >= max(0.45, best_confidence - 0.25)
    ]
    return max(
        credible,
        key=lambda item: (len(item[0]), float(item[2])),
    )


def recognize_legacy_station(
    image: np.ndarray,
    recognizer: TextRecognition,
    table_bbox: tuple[int, int, int, int],
    data_top: int,
) -> dict[str, Any]:
    left, top, right, _ = table_bbox
    width = right - left
    crops: list[np.ndarray] = []
    for upper_pad in (52, 62, 72):
        crops.extend(
            prepare_legacy_station_crop_variants(
                image,
                (
                    round(left),
                    max(round(top + upper_pad), 0),
                    round(right),
                    min(round(data_top - 4), image.shape[0]),
                ),
            )
        )
    # Retain the established centered views as complementary candidates.  They
    # remain useful when a very wide band makes small title glyphs too tiny for
    # recognition; the broad/padded variants above protect long titles from
    # the old ratio crop's truncation risk.
    for left_ratio, right_ratio, upper_pad in (
        (0.30, 0.30, 63),
        (0.28, 0.28, 65),
        (0.23, 0.23, 72),
        (0.18, 0.18, 78),
    ):
        crop = image[
            max(round(top + upper_pad), 0):min(
                round(data_top - 4), image.shape[0]
            ),
            max(round(left + width * left_ratio), 0):min(
                round(right - width * right_ratio), image.shape[1]
            ),
        ]
        if crop.size == 0 or crop.shape[0] < 2 or crop.shape[1] < 2:
            continue
        crops.append(
            cv2.resize(crop, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
        )

    candidates: list[tuple[str, str, float]] = []
    if not crops:
        raise RuntimeError("站名候选区域为空，无法识别页首站名。")
    for prediction in recognizer.predict(input=crops, batch_size=len(crops)):
        text, score = result_value(prediction)
        station = extract_station_name(text)
        if station:
            candidates.append((station, text, score))
    if not candidates:
        raise RuntimeError("历史表格标题已定位，但没有识别到有效站名。")
    station, text, score = choose_legacy_station_candidate(candidates)
    return {
        "station": station,
        "recognized_text": text,
        "confidence": round(float(score), 8),
    }


def recognize_legacy_station_markers_from_gaps(
    image: np.ndarray,
    recognizer: TextRecognition,
    table_left: float,
    table_right: float,
    data_top: float,
    panel_rows: list[list[int]],
) -> list[dict[str, Any]]:
    """OCR station-title gaps using broad and safely padded crop variants."""
    candidate_bands: list[tuple[int, int]] = []
    first_rows = [rows[0] for rows in panel_rows if rows]
    if first_rows:
        first_row = min(first_rows)
        if first_row - data_top >= 100:
            candidate_bands.append((round(data_top + 8), first_row - 8))

    gap_candidates: list[list[tuple[int, int]]] = []
    for rows in panel_rows:
        if len(rows) < 3:
            gap_candidates.append([])
            continue
        gaps = np.diff(rows)
        ordinary_gap = float(np.median(gaps))
        threshold = max(90.0, ordinary_gap * 2.5)
        gap_candidates.append(
            [
                (rows[index], rows[index + 1])
                for index, gap in enumerate(gaps)
                if gap >= threshold
            ]
        )
    for top, bottom in gap_candidates[0] if gap_candidates else []:
        center = (top + bottom) / 2
        confirmations = sum(
            any(
                abs((other_top + other_bottom) / 2 - center) <= 35
                for other_top, other_bottom in gaps
            )
            for gaps in gap_candidates
        )
        if confirmations >= min(2, len(panel_rows)):
            candidate_bands.append((top + 8, bottom - 8))

    markers: list[dict[str, Any]] = []
    for band_top, band_bottom in candidate_bands:
        if band_bottom - band_top < 25:
            continue
        crops = prepare_legacy_station_crop_variants(
            image,
            (
                round(table_left),
                round(band_top),
                round(table_right),
                round(band_bottom),
            ),
        )
        candidates: list[tuple[str, str, float]] = []
        if crops:
            for prediction in recognizer.predict(
                input=crops, batch_size=len(crops)
            ):
                text, score = result_value(prediction)
                station = extract_station_name(text)
                if station:
                    candidates.append((station, text, float(score)))
        if not candidates:
            continue
        station, text, score = choose_legacy_station_candidate(candidates)
        markers.append(
            {
                "station": station,
                "y": round((band_top + band_bottom) / 2),
                "recognized_text": text,
                "confidence": round(float(score), 8),
            }
        )
    return markers


def recognize_structure_station_markers(
    source_ocr: dict[str, Any],
    table_bbox: tuple[int, int, int, int],
) -> list[dict[str, Any]]:
    """Recover complete station titles from the whole-page OCR stream.

    A tightly resized handwritten-title crop can recognize ``郁江`` as only
    ``江`` even when the structure OCR has already read the complete printed
    title.  Station text is rare and explicitly ends in ``站``, so matching the
    whole-page boxes is a safer complementary path than widening crops forever.
    """
    table_left, table_top, table_right, table_bottom = table_bbox
    markers: list[dict[str, Any]] = []
    for text, box, score in zip(
        source_ocr.get("rec_texts", []),
        source_ocr.get("rec_boxes", []),
        source_ocr.get("rec_scores", []),
    ):
        if len(box) < 4:
            continue
        left, top, right, bottom = map(float, box[:4])
        center_x = (left + right) / 2
        center_y = (top + bottom) / 2
        if not (
            table_left <= center_x <= table_right
            and table_top <= center_y <= table_bottom
        ):
            continue
        station = extract_station_name(str(text))
        if station is None:
            continue
        markers.append(
            {
                "station": station,
                "y": round(center_y),
                "recognized_text": str(text),
                "confidence": round(float(score), 8),
                "source": "structure-ocr-station-line",
            }
        )
    return markers


def recognize_legacy_station_markers(
    image: np.ndarray,
    recognizer: TextRecognition,
    table_bbox: tuple[int, int, int, int],
    data_top: int,
    panel_rows: list[list[int]],
    initial_marker: dict[str, Any] | None,
    source_ocr: dict[str, Any],
) -> list[dict[str, Any]]:
    """Combine the historical header title with station titles inside a page."""
    table_left, table_top, table_right, _ = table_bbox
    markers: list[dict[str, Any]] = []
    if initial_marker is not None:
        markers.append(
            {
                **initial_marker,
                "y": max(table_top, data_top - 1),
                "source": "page-header",
            }
        )

    body_markers = recognize_legacy_station_markers_from_gaps(
        image,
        recognizer,
        table_left,
        table_right,
        data_top,
        panel_rows,
    )
    for marker in body_markers:
        candidate = {**marker, "source": "in-page-gap"}
        duplicate = any(
            existing["station"] == candidate["station"]
            and abs(int(existing["y"]) - int(candidate["y"])) <= 40
            for existing in markers
        )
        if not duplicate:
            markers.append(candidate)

    # Reconcile crop OCR with the already available whole-page OCR.  A source
    # line in the same title band replaces a shorter crop result, which fixes
    # lost leading characters without any station-name dictionary.
    for candidate in recognize_structure_station_markers(
        source_ocr, table_bbox
    ):
        nearby_indices = [
            index
            for index, existing in enumerate(markers)
            if abs(int(existing["y"]) - int(candidate["y"])) <= 70
        ]
        if not nearby_indices:
            markers.append(candidate)
            continue
        best_index = min(
            nearby_indices,
            key=lambda index: abs(
                int(markers[index]["y"]) - int(candidate["y"])
            ),
        )
        existing = markers[best_index]
        existing_station = str(existing["station"])
        candidate_station = str(candidate["station"])
        existing_confidence = float(existing.get("confidence", 0.0))
        candidate_confidence = float(candidate["confidence"])
        if (
            candidate_confidence >= existing_confidence + 0.18
            or (
                len(candidate_station) > len(existing_station)
                and candidate_confidence >= existing_confidence - 0.20
            )
            or (
                len(candidate_station) == len(existing_station)
                and candidate_confidence > existing_confidence
            )
        ):
            markers[best_index] = candidate
    return sorted(markers, key=lambda item: int(item["y"]))


def reconstruct_legacy_page(
    pdf_path: Path,
    result: dict[str, Any],
    recognizer: TextRecognition,
    numeric_recognizer: TextRecognition | None,
    preview_dir: Path,
    batch_size: int,
    training_images_dir: Path | None = None,
    sixth_field: str = "auto",
) -> tuple[list[dict[str, Any]], dict[str, Any], list[dict[str, Any]]]:
    width, height = int(result["width"]), int(result["height"])
    page_index = int(result.get("page_index") or 0)
    image = load_page(pdf_path, page_index, width, height)
    (
        table_bbox,
        panel_boundaries,
        data_top,
        physical_field_count,
    ) = detect_legacy_geometry(image)
    table_left, table_top, table_right, table_bottom = table_bbox
    canonical_indices, schema_metadata = recognize_dynamic_legacy_schema(
        image,
        result,
        recognizer,
        table_bbox,
        panel_boundaries,
        data_top,
        physical_field_count,
        batch_size,
        sixth_field,
    )
    try:
        initial_station_marker = recognize_legacy_station(
            image, recognizer, table_bbox, data_top
        )
    except RuntimeError:
        # Continuation pages do not always repeat the title.  They inherit the
        # preceding page's last station during the grouping stage.
        initial_station_marker = None

    overlay = image.copy()
    all_cells: list[np.ndarray] = []
    cell_keys: list[tuple[int, int, int]] = []
    cell_variants: list[str] = []
    primary_cell_images: dict[tuple[int, int, int], np.ndarray] = {}
    cell_ink_run_lower_bounds: dict[tuple[int, int, int], int] = {}
    cell_variant_spacing: dict[
        tuple[tuple[int, int, int], str], dict[str, float | int]
    ] = {}
    time_part_images: dict[tuple[int, int, str], np.ndarray] = {}
    water_part_images: dict[tuple[int, int, str], np.ndarray] = {}
    source_recognized: dict[tuple[int, int, int], tuple[str, float]] = {}
    panel_rows: list[list[int]] = []
    panel_local_row_centers: list[list[list[int | None]]] = []
    panel_time_modes: list[str] = []
    panel_water_anchor_modes: list[str] = []
    panel_water_fraction_rows: list[list[int]] = []
    panel_time_colon_x: list[int] = []
    panel_time_colon_observed: list[list[bool]] = []
    panel_time_colon_bands: list[list[tuple[int, int] | None]] = []
    panel_water_decimal_models: list[tuple[float, float]] = []
    panel_water_decimal_observed: list[list[bool]] = []
    panel_anchor_statuses: list[list[str]] = []
    panel_anchor_summaries: list[dict[str, int]] = []
    source_ocr = result["overall_ocr_res"]

    for panel_index, boundaries in enumerate(panel_boundaries):
        panel_left = round(boundaries[0])
        panel_right = round(boundaries[-1])
        (
            independent_water_rows,
            water_decimal_model,
            independent_decimal_observed,
        ) = find_rows_from_water_decimal_points(
            image, boundaries, data_top, table_bottom
        )
        raw_water_fraction_rows = detect_legacy_water_fraction_token_anchors(
            image,
            boundaries,
            water_decimal_model,
            data_top,
            table_bottom,
        )
        occupancy_candidates = detect_legacy_multicolumn_row_candidates(
            image, boundaries, data_top, table_bottom
        )
        strong_water_occupancy = [
            candidate
            for candidate in occupancy_candidates
            if int(candidate["support"]) >= 2
            and 3 in {int(value) for value in candidate["columns"]}
        ]
        water_fraction_rows = [
            center_y
            for center_y in raw_water_fraction_rows
            if legacy_multicolumn_support(
                strong_water_occupancy, center_y, tolerance=7.0
            )
            is not None
        ]
        water_dot_coverage = len(independent_water_rows) / max(
            1, len(water_fraction_rows)
        )
        water_anchor_mode = "decimal_dot"
        if (
            len(water_fraction_rows) >= 3
            and water_dot_coverage < LEGACY_WATER_DOT_SPARSE_RATIO
        ):
            water_anchor_mode = "sparse_fraction_token"
            observed_dot_rows = list(independent_water_rows)
            # Downstream geometry historically stores the decimal-point y and
            # shifts it three pixels upward to the glyph center.  Preserve that
            # contract by converting each observed fraction center to a
            # pseudo-dot coordinate; the observation flag remains truthful.
            independent_water_rows = [
                center_y - LEGACY_WATER_DOT_TO_GLYPH_CENTER
                for center_y in water_fraction_rows
            ]
            independent_decimal_observed = [
                any(abs(dot_y - row_y) <= 5 for dot_y in observed_dot_rows)
                for row_y in independent_water_rows
            ]
        (
            time_mode,
            time_colon_x,
            independent_time_rows,
            independent_colon_observed,
            independent_time_colon_bands,
        ) = detect_legacy_time_anchors(
            image, boundaries, data_top, table_bottom
        )
        (
            rows,
            time_rows,
            decimal_observed,
            colon_observed,
            time_colon_bands,
            anchor_statuses,
            anchor_summary,
        ) = reconcile_legacy_dual_anchor_rows(
            image,
            boundaries,
            independent_water_rows,
            water_decimal_model,
            independent_decimal_observed,
            independent_time_rows,
            independent_colon_observed,
            independent_time_colon_bands,
            data_top,
            table_bottom,
        )
        panel_rows.append(rows)
        panel_time_modes.append(time_mode)
        panel_water_anchor_modes.append(water_anchor_mode)
        panel_water_fraction_rows.append(water_fraction_rows)
        panel_water_decimal_models.append(water_decimal_model)
        panel_water_decimal_observed.append(decimal_observed)
        panel_time_colon_x.append(time_colon_x)
        panel_time_colon_observed.append(colon_observed)
        panel_time_colon_bands.append(time_colon_bands)
        panel_anchor_statuses.append(anchor_statuses)
        anchor_summary["water_anchor_mode"] = water_anchor_mode
        anchor_summary["independent_decimal_anchor_count"] = (
            sum(bool(value) for value in independent_decimal_observed)
            if water_anchor_mode == "sparse_fraction_token"
            else len(independent_water_rows)
        )
        anchor_summary["independent_water_fraction_token_count"] = len(
            water_fraction_rows
        )
        anchor_summary["water_dot_coverage"] = round(
            water_dot_coverage, 4
        )
        panel_anchor_summaries.append(anchor_summary)
        local_row_centers, owned_cells = assign_legacy_cells_from_dual_anchors(
            image,
            boundaries,
            canonical_indices,
            rows,
            time_rows,
            data_top,
            table_bottom,
        )
        panel_local_row_centers.append(local_row_centers)
        for x in boundaries:
            cv2.line(
                overlay,
                (round(x), data_top),
                (round(x), table_bottom),
                (255, 0, 255),
                1,
            )
        cv2.line(
            overlay,
            (
                round(boundaries[3])
                + 2
                + round(
                    water_decimal_model[0] * data_top
                    + water_decimal_model[1]
                ),
                data_top,
            ),
            (
                round(boundaries[3])
                + 2
                + round(
                    water_decimal_model[0] * table_bottom
                    + water_decimal_model[1]
                ),
                table_bottom,
            ),
            (0, 255, 255),
            1,
        )

        if water_anchor_mode == "sparse_fraction_token":
            for token_y in water_fraction_rows:
                token_x = (
                    round(boundaries[3])
                    + 2
                    + round(
                        water_decimal_model[0] * token_y
                        + water_decimal_model[1]
                    )
                    + 8
                )
                cv2.line(
                    overlay,
                    (token_x - 3, token_y),
                    (token_x + 3, token_y),
                    (255, 255, 0),
                    1,
                )

        final_water_set = set(rows)
        for rejected_y in independent_water_rows:
            if rejected_y in final_water_set:
                continue
            rejected_x = (
                round(boundaries[3])
                + 2
                + round(
                    water_decimal_model[0] * rejected_y
                    + water_decimal_model[1]
                )
            )
            cv2.drawMarker(
                overlay,
                (rejected_x, rejected_y),
                (0, 0, 255),
                cv2.MARKER_TILTED_CROSS,
                7,
                1,
            )
        final_observed_time_set = {
            row_y
            for row_y, is_observed in zip(time_rows, colon_observed)
            if is_observed
        }
        for rejected_y in independent_time_rows:
            if rejected_y in final_observed_time_set:
                continue
            cv2.drawMarker(
                overlay,
                (round(boundaries[2]) + 2 + time_colon_x, rejected_y),
                (255, 0, 255),
                cv2.MARKER_TILTED_CROSS,
                7,
                1,
            )

        for row_index, center_y in enumerate(rows):
            review_time_colon_band = panel_time_colon_bands[panel_index][
                row_index
            ]
            row_centers = local_row_centers[row_index]
            decimal_x = (
                round(boundaries[3])
                + 2
                + round(water_decimal_model[0] * center_y + water_decimal_model[1])
            )
            cv2.circle(
                overlay,
                (decimal_x, center_y),
                2,
                (
                    (0, 255, 0)
                    if decimal_observed[row_index]
                    else (0, 0, 255)
                ),
                -1,
            )
            colon_y = time_rows[row_index]
            colon_band = time_colon_bands[row_index]
            if colon_band is not None:
                band_left, band_right = colon_band
                time_origin_x = round(boundaries[2]) + 2
                cv2.rectangle(
                    overlay,
                    (time_origin_x + band_left, colon_y - 4),
                    (time_origin_x + max(band_left, band_right - 1), colon_y + 4),
                    (255, 128, 0),
                    1,
                )
            cv2.circle(
                overlay,
                (round(boundaries[2]) + 2 + time_colon_x, colon_y),
                2,
                (
                    (255, 255, 0)
                    if colon_observed[row_index]
                    else (255, 0, 0)
                ),
                -1,
            )
            row_points: list[tuple[int, int] | None] = []
            for physical_index, local_y in enumerate(row_centers):
                point = (
                    (
                        round(
                            (boundaries[physical_index]
                            + boundaries[physical_index + 1])
                            / 2
                        ),
                        int(local_y),
                    )
                    if local_y is not None
                    else None
                )
                row_points.append(point)
                if point is not None:
                    cv2.circle(overlay, point, 1, (0, 165, 255), -1)
            for first, second in zip(row_points[:-1], row_points[1:]):
                if first is not None and second is not None:
                    cv2.line(overlay, first, second, (0, 165, 255), 1)

            for physical_index, column_index in enumerate(canonical_indices):
                local_center_y = row_centers[physical_index]
                owned_cell = owned_cells.get((row_index, physical_index))
                if owned_cell is None or local_center_y is None:
                    continue
                cell_image = owned_cell["image"]
                raw_cell_image = owned_cell.get("raw_image", cell_image)
                cell_height, cell_width = cell_image.shape[:2]
                cell_left = int(owned_cell["left"])
                cell_right = int(owned_cell["right"])
                candidates: list[tuple[str, float]] = []
                for text, box, score in zip(
                    source_ocr["rec_texts"],
                    source_ocr["rec_boxes"],
                    source_ocr["rec_scores"],
                ):
                    box_left, box_top, box_right, box_bottom = box
                    center_x = (box_left + box_right) / 2
                    box_center_y = (box_top + box_bottom) / 2
                    if (
                        cell_left <= center_x <= cell_right
                        and abs(box_center_y - local_center_y) <= 5
                        and box_bottom - box_top <= 25
                    ):
                        candidates.append((str(text), float(score)))
                if candidates:
                    source_recognized[(panel_index, row_index, column_index)] = max(
                        candidates, key=lambda item: item[1]
                    )

                if column_index == 6:
                    prepared = prepare_cell(
                        cell_image,
                        (0, 0, cell_width, cell_height),
                    )
                else:
                    prepared = prepare_legacy_cell(
                        cell_image,
                        (0, 0, cell_width, cell_height),
                        column_index,
                    )
                if prepared is not None:
                    sample_key = (panel_index, row_index, column_index)
                    all_cells.append(prepared)
                    cell_keys.append(sample_key)
                    cell_variants.append("primary")
                    primary_cell_images[sample_key] = prepared.copy()
                    cell_ink_run_lower_bounds[sample_key] = (
                        estimate_legacy_digit_run_lower_bound(prepared)
                    )
                    cell_variant_spacing[(sample_key, "primary")] = (
                        estimate_legacy_horizontal_ink_spacing(prepared)
                    )
                if column_index in (0, 1, 2, 4):
                    raw_prepared = prepare_legacy_raw_cell(
                        raw_cell_image,
                        (
                            0,
                            0,
                            raw_cell_image.shape[1],
                            raw_cell_image.shape[0],
                        ),
                        # The time column has an independently detected colon
                        # and split minute image, so retaining a table rule in
                        # its wide/raw fallback creates false trailing ``1``
                        # candidates such as 20:01.  Use the adaptive fitted
                        # edge cleaner here; it removes only observed straight
                        # or slightly slanted rule tracks and does not whiten a
                        # fixed band.  Other legacy raw fallbacks retain their
                        # established untouched geometry.
                        clear_edges=(column_index == 2),
                    )
                    if raw_prepared is not None:
                        all_cells.append(raw_prepared)
                        cell_keys.append((panel_index, row_index, column_index))
                        cell_variants.append("raw")
                        cell_variant_spacing[
                            ((panel_index, row_index, column_index), "raw")
                        ] = estimate_legacy_horizontal_ink_spacing(raw_prepared)
                if column_index == 2:
                    colon_band = time_colon_bands[row_index]
                    time_coordinate_offset = (
                        round(boundaries[2]) + 2 - cell_left
                    )
                    if time_mode == "hour_only":
                        hour_right = cell_width
                        minute_left = cell_width
                        minute_right = cell_width
                    elif colon_band is not None:
                        colon_left, colon_right = colon_band
                        colon_left += time_coordinate_offset
                        colon_right += time_coordinate_offset
                        hour_right = min(
                            cell_width,
                            max(
                                2,
                                colon_left - LEGACY_TIME_COLON_CROP_MARGIN,
                            ),
                        )
                        minute_left = min(
                            cell_width,
                            colon_right + LEGACY_TIME_COLON_CROP_MARGIN,
                        )
                        minute_right = max(
                            minute_left,
                            cell_width,
                        )
                    else:
                        hour_right = min(
                            cell_width,
                            max(
                                2,
                                time_colon_x
                                + time_coordinate_offset
                                - 2,
                            ),
                        )
                        minute_left = min(
                            cell_width,
                            time_colon_x + time_coordinate_offset + 2,
                        )
                        minute_right = max(
                            minute_left,
                            cell_width,
                        )
                    hour_prepared = prepare_legacy_time_part(
                        cell_image,
                        (0, 0, hour_right, cell_height),
                    )
                    minute_prepared = (
                        None
                        if time_mode == "hour_only"
                        else prepare_legacy_time_part(
                            cell_image,
                            (minute_left, 0, minute_right, cell_height),
                        )
                    )
                    if hour_prepared is not None:
                        hour_key = (panel_index, row_index, 8)
                        all_cells.append(hour_prepared)
                        cell_keys.append(hour_key)
                        cell_variants.append("primary")
                        cell_ink_run_lower_bounds[hour_key] = (
                            estimate_legacy_digit_run_lower_bound(hour_prepared)
                        )
                        time_part_images[
                            (panel_index, row_index, "hour")
                        ] = hour_prepared.copy()
                    if minute_prepared is not None:
                        minute_key = (panel_index, row_index, 9)
                        all_cells.append(minute_prepared)
                        cell_keys.append(minute_key)
                        cell_variants.append("primary")
                        cell_ink_run_lower_bounds[minute_key] = (
                            estimate_legacy_digit_run_lower_bound(minute_prepared)
                        )
                        time_part_images[
                            (panel_index, row_index, "minute")
                        ] = minute_prepared.copy()
                if column_index == 3:
                    water_decimal_x = (
                        round(boundaries[3])
                        + 2
                        - cell_left
                        + round(
                            water_decimal_model[0] * local_center_y
                            + water_decimal_model[1]
                        )
                    )
                    integer_right = min(
                        cell_width,
                        max(2, water_decimal_x - 3),
                    )
                    fraction_left = min(
                        cell_width,
                        water_decimal_x + 3,
                    )
                    integer_prepared = prepare_legacy_water_part(
                        cell_image,
                        (0, 0, integer_right, cell_height),
                        clear_left=True,
                    )
                    fraction_prepared = prepare_legacy_water_part(
                        cell_image,
                        (fraction_left, 0, cell_width, cell_height),
                        clear_right=True,
                    )
                    if integer_prepared is not None:
                        integer_key = (panel_index, row_index, 12)
                        all_cells.append(integer_prepared)
                        cell_keys.append(integer_key)
                        cell_variants.append("primary")
                        cell_ink_run_lower_bounds[integer_key] = (
                            estimate_legacy_digit_run_lower_bound(integer_prepared)
                        )
                        water_part_images[
                            (panel_index, row_index, "integer")
                        ] = integer_prepared.copy()
                    if fraction_prepared is not None:
                        fraction_key = (panel_index, row_index, 13)
                        all_cells.append(fraction_prepared)
                        cell_keys.append(fraction_key)
                        cell_variants.append("primary")
                        cell_ink_run_lower_bounds[fraction_key] = (
                            estimate_legacy_digit_run_lower_bound(fraction_prepared)
                        )
                        water_part_images[
                            (panel_index, row_index, "fraction")
                        ] = fraction_prepared.copy()
            date_owned_cells = [
                owned_cells[(row_index, physical_index)]
                for physical_index in (0, 1)
                if (row_index, physical_index) in owned_cells
            ]
            if date_owned_cells:
                pair_left = min(int(cell["left"]) for cell in date_owned_cells)
                pair_right = max(int(cell["right"]) for cell in date_owned_cells)
                pair_top = min(cell["top"] for cell in date_owned_cells)
                pair_bottom = max(cell["bottom"] for cell in date_owned_cells)
                pair_image = np.full(
                    (pair_bottom - pair_top, pair_right - pair_left, 3),
                    255,
                    dtype=np.uint8,
                )
                for date_cell in date_owned_cells:
                    source = date_cell["image"]
                    x_offset = date_cell["left"] - pair_left
                    y_offset = date_cell["top"] - pair_top
                    target = pair_image[
                        y_offset:y_offset + source.shape[0],
                        x_offset:x_offset + source.shape[1],
                    ]
                    np.minimum(target, source, out=target)
                date_pair = prepare_legacy_date_pair(
                    pair_image,
                    (0, 0, pair_image.shape[1], pair_image.shape[0]),
                )
                if date_pair is not None:
                    all_cells.append(date_pair)
                    cell_keys.append((panel_index, row_index, 7))
                    cell_variants.append("auxiliary")

    station_markers = recognize_legacy_station_markers(
        image,
        recognizer,
        table_bbox,
        data_top,
        panel_rows,
        initial_station_marker,
        source_ocr,
    )
    for marker in station_markers:
        cv2.line(
            overlay,
            (table_left, int(marker["y"])),
            (table_right, int(marker["y"])),
            (0, 0, 255),
            2,
        )

    preview_path = preview_dir / f"grid_overlay_page_{page_index + 1}.png"
    cv2.imwrite(str(preview_path), overlay)

    candidate_records: dict[
        tuple[int, int, int], list[tuple[str, float, str, str]]
    ] = {}
    prediction_sources = [
        (
            "server",
            recognizer.predict(input=all_cells, batch_size=batch_size),
        )
    ]
    if numeric_recognizer is not None:
        prediction_sources.append(
            (
                "numeric",
                numeric_recognizer.predict(
                    input=all_cells, batch_size=batch_size
                ),
            )
        )
    for model, predictions in prediction_sources:
        for key, variant, prediction in zip(
            cell_keys,
            cell_variants,
            predictions,
        ):
            text, score = result_value(prediction)
            candidate_records.setdefault(key, []).append(
                (text, float(score), variant, model)
            )

    recognized: dict[tuple[int, int, int], tuple[str, float]] = {}
    recognized_variants: dict[tuple[int, int, int], str] = {}
    recognized_decisions: dict[tuple[int, int, int], str] = {}
    recognized_length_warnings: dict[tuple[int, int, int], str] = {}
    primary_recognized: dict[tuple[int, int, int], tuple[str, float]] = {}
    for key, candidates in candidate_records.items():
        column_index = key[2]
        ink_lower_bound = cell_ink_run_lower_bounds.get(key, 0)
        spacing_by_variant = {
            variant: cell_variant_spacing[(key, variant)]
            for variant in ("primary", "raw")
            if (key, variant) in cell_variant_spacing
        }
        chosen, source, decision, length_warning = (
            choose_legacy_numeric_candidate(
                candidates,
                column_index,
                ink_lower_bound,
                spacing_by_variant,
            )
        )
        recognized[key] = chosen
        recognized_variants[key] = source
        if decision:
            recognized_decisions[key] = decision
        if length_warning:
            recognized_length_warnings[key] = length_warning

        primary_candidates = [
            candidate
            for candidate in candidates
            if candidate[2] == "primary"
        ]
        if primary_candidates:
            primary_chosen, _, _, _ = choose_legacy_numeric_candidate(
                primary_candidates,
                column_index,
                ink_lower_bound,
                {
                    "primary": spacing_by_variant.get("primary", {})
                },
            )
            # This result is produced from the exact PNG exported to the
            # review workbook.  Keep it separately so OCR text/confidence in
            # labels_review.xlsx always describe the image beside them.
            primary_recognized[key] = primary_chosen

    output_rows: list[dict[str, Any]] = []
    training_samples: list[dict[str, Any]] = []
    if training_images_dir is not None:
        training_images_dir.mkdir(parents=True, exist_ok=True)
    for panel_index, rows in enumerate(panel_rows):
        time_mode = panel_time_modes[panel_index]
        for row_index, center_y in enumerate(rows):
            review_time_colon_band = panel_time_colon_bands[panel_index][
                row_index
            ]
            water_center_y = panel_local_row_centers[panel_index][row_index][3]
            review_water_decimal_x = round(
                panel_water_decimal_models[panel_index][0] * water_center_y
                + panel_water_decimal_models[panel_index][1]
            )
            raw_texts: list[str] = []
            normalized_values: list[str] = []
            value_scores: list[float] = []
            training_raw_texts: list[str] = []
            training_normalized_values: list[str] = []
            training_scores: list[float] = []
            recognition_notes: list[str] = []
            for column_index in range(7):
                sample_key = (panel_index, row_index, column_index)
                raw_text, score = recognized.get(sample_key, ("", 0.0))
                selected_variant = recognized_variants.get(sample_key, "")
                primary_text, primary_score = primary_recognized.get(
                    sample_key, ("", 0.0)
                )
                source_text, source_score = source_recognized.get(
                    sample_key, ("", 0.0)
                )
                cell_normalized = normalize_legacy_numeric(
                    raw_text, column_index
                )
                source_normalized = normalize_legacy_numeric(
                    source_text, column_index
                )
                used_source_fallback = False
                cell_digit_count = len(re.findall(r"\d", cell_normalized))
                source_digit_count = len(
                    re.findall(r"\d", source_normalized)
                )
                use_complete_date_source = bool(
                    column_index in (0, 1)
                    and source_normalized
                    and source_digit_count >= cell_digit_count
                    and source_score >= float(score) + 0.15
                )
                # Whole-page OCR boxes may cover only one digit of a complete
                # cell.  Their confidence is not comparable with a full-cell
                # sequence, so non-date fields only use them for an invalid
                # cell.  Month/day additionally require non-shorter text and a
                # clear confidence margin before a source candidate can win.
                if (
                    (not cell_normalized and source_normalized)
                    or use_complete_date_source
                ):
                    raw_text, score = source_text, source_score
                    used_source_fallback = True
                normalized_text = normalize_legacy_numeric(
                    raw_text, column_index
                )
                rejected_low_confidence_slope = bool(
                    column_index == 6
                    and normalized_text
                    and float(score) < LEGACY_SLOPE_MIN_CONFIDENCE
                )
                if rejected_low_confidence_slope:
                    normalized_text = ""
                raw_texts.append(raw_text)
                normalized_values.append(normalized_text)
                value_scores.append(float(score))
                training_raw_texts.append(primary_text)
                training_normalized_values.append(
                    normalize_legacy_numeric(primary_text, column_index)
                )
                training_scores.append(float(primary_score))

                note_parts: list[str] = []
                if source_text:
                    note_parts.append(
                        f"整页OCR候选={source_text}，置信度={source_score:.3f}"
                    )
                if rejected_low_confidence_slope:
                    note_parts.append(
                        "水面比降候选低于字段最低置信度，按空值处理："
                        f"{raw_text}（{float(score):.3f}）"
                    )
                if used_source_fallback:
                    if use_complete_date_source:
                        note_parts.append(
                            "月日整页候选位数不短且置信度明显更高，"
                            "最终输出采用整页OCR"
                        )
                    else:
                        note_parts.append(
                            "分格结果无效，最终输出采用整页OCR兜底"
                        )
                elif source_normalized and source_score > score:
                    note_parts.append(
                        "整页OCR虽置信度更高，但未覆盖有效分格结果"
                    )
                decision_note = recognized_decisions.get(sample_key, "")
                if decision_note:
                    note_parts.append(decision_note)
                length_warning = recognized_length_warnings.get(
                    sample_key, ""
                )
                if length_warning:
                    note_parts.append(length_warning)
                if selected_variant.startswith("raw:") and (
                    primary_text != raw_text
                    or abs(float(primary_score) - float(score)) > 1e-9
                ):
                    note_parts.append(
                        f"最终输出采用{selected_variant}备用图OCR={raw_text}，"
                        f"置信度={float(score):.3f}；"
                        "OCR原文仍对应导出的主图"
                    )
                recognition_notes.append("；".join(note_parts))

            values = normalized_values.copy()
            effective_scores = value_scores.copy()

            hour_text, hour_score = recognized.get(
                (panel_index, row_index, 8), ("", 0.0)
            )
            minute_text, minute_score = recognized.get(
                (panel_index, row_index, 9), ("", 0.0)
            )
            digit_translation = LEGACY_NUMERIC_TRANSLATION
            hour_digits = "".join(
                re.findall(r"\d", hour_text.translate(digit_translation))
            )
            minute_digits = normalize_legacy_minute_digits(minute_text)
            water_integer_text, water_integer_score = recognized.get(
                (panel_index, row_index, 12), ("", 0.0)
            )
            water_fraction_text, water_fraction_score = recognized.get(
                (panel_index, row_index, 13), ("", 0.0)
            )
            water_integer_digits = "".join(
                re.findall(
                    r"\d", water_integer_text.translate(digit_translation)
                )
            )
            water_fraction_digits = "".join(
                re.findall(
                    r"\d", water_fraction_text.translate(digit_translation)
                )
            )
            split_time = ""
            split_time_score = 0.0
            if 1 <= len(hour_digits) <= 2 and 0 <= int(hour_digits) <= 24:
                hour_value = int(hour_digits)
                if (
                    1 <= len(minute_digits) <= 2
                    and 0 <= int(minute_digits) <= 59
                    and (hour_value < 24 or int(minute_digits) == 0)
                ):
                    split_time = f"{hour_value:02d}:{int(minute_digits):02d}"
                    split_time_score = (float(hour_score) + float(minute_score)) / 2
                elif (panel_index, row_index, 9) not in recognized:
                    split_time = f"{hour_value:02d}:00"
                    split_time_score = float(hour_score)

            time_was_corrected = bool(
                split_time
                and split_time != values[2]
            )
            # Colon mode prefers the observed hour/minute split.  Hour-only
            # mode recognizes the complete numeric token as the hour and
            # supplies an implicit ``:00`` without manufacturing minute ink.
            if split_time:
                values[2] = split_time
                if split_time_score > 0:
                    effective_scores[2] = split_time_score
                if time_was_corrected:
                    recognition_notes[2] = ";".join(
                        part
                        for part in (
                            recognition_notes[2],
                            f"最终输出采用时/分拆分结果={split_time}",
                        )
                        if part
                    )

            water_dot_observed = bool(
                panel_water_decimal_observed[panel_index][row_index]
            )
            water_integer_available = bool(
                (panel_index, row_index, 12) in recognized
                and 1 <= len(water_integer_digits) <= 4
            )
            water_fraction_available = bool(
                (panel_index, row_index, 13) in recognized
                and len(water_fraction_digits) == 2
            )
            split_water = ""
            split_water_score = 0.0
            if water_fraction_available:
                if water_dot_observed and water_integer_available:
                    split_water = (
                        f"{water_integer_digits}.{water_fraction_digits}"
                    )
                    split_water_score = (
                        float(water_integer_score)
                        + float(water_fraction_score)
                    ) / 2
                elif not water_dot_observed:
                    # No observed decimal point means that the repeated
                    # integer prefix was omitted in the source table.  Keep
                    # only the fractional digits so expand_water_level() can
                    # inherit the last explicit integer.
                    split_water = water_fraction_digits
                    split_water_score = float(water_fraction_score)

            water_was_corrected = bool(
                split_water and split_water != values[3]
            )
            current_water_valid = bool(
                re.fullmatch(r"\d+\.\d{1,2}", values[3])
                or re.fullmatch(r"\d{1,2}", values[3])
            )
            # Split water recognition is a repair path, not an unconditional
            # replacement.  Its narrow crops can confuse similar handwritten
            # digits, so a syntactically valid whole-cell value remains the
            # automatic primary result.  Accepted manual labels are applied
            # later and always take precedence over both OCR branches.
            split_repairs_observed_decimal = bool(
                split_water
                and water_dot_observed
                and "." not in values[3]
            )
            if split_water and (
                not current_water_valid or split_repairs_observed_decimal
            ):
                values[3] = split_water
                if split_water_score > 0:
                    effective_scores[3] = split_water_score
                if water_was_corrected:
                    recognition_notes[3] = ";".join(
                        part
                        for part in (
                            recognition_notes[3],
                            (
                                "观测到小数点，整格候选仅剩小数部分；"
                                if split_repairs_observed_decimal
                                else ""
                            ),
                            f"最终输出采用水位拆分结果={split_water}",
                        )
                        if part
                    )

            pair_text, pair_score = recognized.get(
                (panel_index, row_index, 7), ("", 0.0)
            )
            pair_cleaned = pair_text.translate(LEGACY_NUMERIC_TRANSLATION)
            pair_digits = "".join(re.findall(r"\d", pair_cleaned))
            pair_day = ""
            month_value = parse_integer(values[0], 1, 12)
            if month_value is not None:
                month_prefix = str(month_value)
                if pair_digits.startswith(month_prefix):
                    day_digits = pair_digits[len(month_prefix):]
                    if 1 <= len(day_digits) <= 2 and 1 <= int(day_digits) <= 31:
                        pair_day = day_digits
            elif parse_integer(values[1], 1, 31) is None:
                if 1 <= len(pair_digits) <= 2 and 1 <= int(pair_digits) <= 31:
                    pair_day = pair_digits

            current_day = parse_integer(values[1], 1, 31)
            # On dense scans the narrow day cell and the wider month+day crop
            # can retain complementary digits.  Example: the day crop reads
            # ``1`` while the pair crop reads ``62`` (month 6 + day tail 2).
            # Merge the two one-digit day fragments when they form a valid
            # two-digit day and the wider crop is reasonably confident.
            if (
                current_day is not None
                and len(values[1]) == 1
                and len(pair_day) == 1
                and pair_day != values[1]
                and pair_score >= LEGACY_DAY_PAIR_REVIEW_CONFIDENCE
            ):
                merged_day = values[1] + pair_day
                if 10 <= int(merged_day) <= 31:
                    pair_day = merged_day

            day_was_corrected = bool(
                pair_day
                and (
                    current_day is None
                    or value_scores[1] < LEGACY_DAY_PAIR_REVIEW_CONFIDENCE
                )
                and pair_day != values[1]
            )
            if day_was_corrected:
                values[1] = pair_day
                if pair_score > 0:
                    effective_scores[1] = float(pair_score)

            # Keep the viable calendar/time alternatives until all pages and
            # panels have been merged.  Month/day are sparse fields, while
            # whole-cell time OCR and colon-split OCR can disagree; choosing
            # them here would discard the neighbouring-row evidence needed by
            # the sequence resolver in normalize_legacy_rows().
            context_candidates: dict[str, list[dict[str, Any]]] = {
                "月": [],
                "日": [],
                "时分": [],
            }

            def add_context_candidate(
                field: str,
                value: str,
                candidate_score: float,
                source: str,
            ) -> None:
                if not value:
                    return
                existing = next(
                    (
                        item
                        for item in context_candidates[field]
                        if item["value"] == value
                    ),
                    None,
                )
                candidate = {
                    "value": value,
                    "score": round(float(candidate_score), 8),
                    "source": source,
                }
                if existing is None:
                    context_candidates[field].append(candidate)
                elif float(candidate_score) > float(existing["score"]):
                    existing.update(candidate)

            def add_time_context_candidate(
                value: str,
                candidate_score: float,
                source: str,
            ) -> None:
                (
                    reconciled_value,
                    reconciled_score,
                    reconciled_source,
                ) = reconcile_legacy_time_candidate_with_split_minute(
                    value,
                    candidate_score,
                    source,
                    minute_digits,
                    minute_score,
                )
                add_context_candidate(
                    "时分",
                    reconciled_value,
                    reconciled_score,
                    reconciled_source,
                )

            for field, column_index in (("月", 0), ("日", 1), ("时分", 2)):
                for text, candidate_score, variant, model in candidate_records.get(
                    (panel_index, row_index, column_index), []
                ):
                    candidate_value = normalize_legacy_numeric(
                        text, column_index
                    )
                    candidate_source = f"{variant}:{model}"
                    if field == "时分":
                        add_time_context_candidate(
                            candidate_value,
                            candidate_score,
                            candidate_source,
                        )
                    else:
                        add_context_candidate(
                            field,
                            candidate_value,
                            candidate_score,
                            candidate_source,
                        )
                source_text, source_score = source_recognized.get(
                    (panel_index, row_index, column_index), ("", 0.0)
                )
                source_value = normalize_legacy_numeric(
                    source_text, column_index
                )
                if field == "时分":
                    add_time_context_candidate(
                        source_value,
                        source_score,
                        "whole-page",
                    )
                    add_time_context_candidate(
                        values[column_index],
                        effective_scores[column_index],
                        "row-selected",
                    )
                else:
                    add_context_candidate(
                        field,
                        source_value,
                        source_score,
                        "whole-page",
                    )
                    add_context_candidate(
                        field,
                        values[column_index],
                        effective_scores[column_index],
                        "row-selected",
                    )

            hour_alternatives: dict[int, tuple[float, str]] = {}
            for text, candidate_score, variant, model in candidate_records.get(
                (panel_index, row_index, 8), []
            ):
                digits = "".join(
                    re.findall(
                        r"\d", text.translate(LEGACY_NUMERIC_TRANSLATION)
                    )
                )
                if 1 <= len(digits) <= 2 and 0 <= int(digits) <= 24:
                    value = int(digits)
                    previous = hour_alternatives.get(value)
                    if previous is None or candidate_score > previous[0]:
                        hour_alternatives[value] = (
                            float(candidate_score),
                            f"{variant}:{model}",
                        )
            minute_alternatives: dict[int, tuple[float, str]] = {}
            for text, candidate_score, variant, model in candidate_records.get(
                (panel_index, row_index, 9), []
            ):
                digits = normalize_legacy_minute_digits(text)
                if 1 <= len(digits) <= 2 and 0 <= int(digits) <= 59:
                    value = int(digits)
                    previous = minute_alternatives.get(value)
                    if previous is None or candidate_score > previous[0]:
                        minute_alternatives[value] = (
                            float(candidate_score),
                            f"{variant}:{model}",
                        )
            if hour_digits and 0 <= int(hour_digits) <= 24:
                hour_alternatives.setdefault(
                    int(hour_digits), (float(hour_score), "split-selected")
                )
            if minute_digits and 0 <= int(minute_digits) <= 59:
                minute_alternatives.setdefault(
                    int(minute_digits),
                    (float(minute_score), "split-selected"),
                )
            if not minute_alternatives:
                minute_alternatives[0] = (0.0, "omitted-minute")
            for hour_value, (candidate_hour_score, hour_source) in (
                hour_alternatives.items()
            ):
                for minute_value, (candidate_minute_score, minute_source) in (
                    minute_alternatives.items()
                ):
                    if hour_value == 24 and minute_value != 0:
                        continue
                    add_context_candidate(
                        "时分",
                        f"{hour_value:02d}:{minute_value:02d}",
                        (
                            candidate_hour_score + candidate_minute_score
                        ) / 2,
                        (
                            f"hour-only({hour_source}+implicit-minute-zero)"
                            if time_mode == "hour_only"
                            else f"colon-split({hour_source}+{minute_source})"
                        ),
                    )

            if training_images_dir is not None:
                for column_index in range(7):
                    if column_index in (2, 3):
                        continue
                    sample_key = (panel_index, row_index, column_index)
                    sample_image = primary_cell_images.get(sample_key)
                    if sample_image is None:
                        continue
                    sample_id = (
                        f"p{page_index + 1:03d}_b{panel_index + 1}_"
                        f"r{row_index + 1:03d}_c{column_index + 1:02d}"
                    )
                    filename = (
                        f"{sample_id}_{SAMPLE_COLUMN_SLUGS[column_index]}.png"
                    )
                    image_path = training_images_dir / filename
                    cv2.imwrite(str(image_path), sample_image)
                    note_parts: list[str] = []
                    if column_index == 1 and day_was_corrected:
                        original = normalized_values[1] or "空"
                        note_parts.append(
                            f"月日联合识别建议：{original}→{values[1]}；"
                            f"联合OCR={pair_text}，置信度={pair_score:.3f}"
                        )
                    if recognition_notes[column_index]:
                        note_parts.append(recognition_notes[column_index])
                    note = "；".join(note_parts)
                    correct_label = values[column_index]
                    requires_edge_prefix_review = (
                        "边缘前导1防误判" in note
                    )
                    training_samples.append(
                        {
                            "sample_id": sample_id,
                            "page": page_index + 1,
                            "panel": panel_index + 1,
                            "row": row_index + 1,
                            "column": LEGACY_COLUMN_NAMES[column_index],
                            "ocr_text": training_raw_texts[column_index],
                            "normalized_label": training_normalized_values[
                                column_index
                            ],
                            "confidence": round(training_scores[column_index], 8),
                            "correct_label": correct_label,
                            "include": (
                                "是"
                                if correct_label
                                and not requires_edge_prefix_review
                                else "待确认"
                            ),
                            "image": f"images/{filename}",
                            "note": note,
                        }
                    )

                hour_label = ""
                minute_label = ""
                if split_time:
                    split_hour, split_minute = split_time.split(":")
                    hour_label = str(int(split_hour))
                    minute_label = split_minute
                hour_normalized = (
                    str(int(hour_digits))
                    if 1 <= len(hour_digits) <= 2
                    and 0 <= int(hour_digits) <= 24
                    else ""
                )
                minute_normalized = (
                    minute_digits
                    if 1 <= len(minute_digits) <= 2
                    and 0 <= int(minute_digits) <= 59
                    else ""
                )
                time_part_specs = [
                    (
                        "hour",
                        "h",
                        "时",
                        hour_text,
                        hour_normalized,
                        hour_label,
                        float(hour_score),
                        "time_hour",
                        (
                            "无冒号整格小时子图；训练标签不补前导零，"
                            "分钟00由版式隐含且不作为图像样本。"
                            if time_mode == "hour_only"
                            else "冒号左侧小时子图；训练标签不补前导零。"
                        ),
                    ),
                    (
                        "minute",
                        "m",
                        "分",
                        minute_text,
                        minute_normalized,
                        minute_label,
                        float(minute_score),
                        "time_minute",
                        "冒号右侧分钟子图；训练标签保留两位且不包含冒号。",
                    ),
                ]
                for (
                    part_key,
                    id_suffix,
                    field_name,
                    part_ocr_text,
                    part_normalized,
                    part_correct_label,
                    part_score,
                    filename_suffix,
                    part_note,
                ) in time_part_specs:
                    part_image = time_part_images.get(
                        (panel_index, row_index, part_key)
                    )
                    if part_image is None:
                        continue
                    sample_id = (
                        f"p{page_index + 1:03d}_b{panel_index + 1}_"
                        f"r{row_index + 1:03d}_c03{id_suffix}"
                    )
                    filename = f"{sample_id}_{filename_suffix}.png"
                    cv2.imwrite(str(training_images_dir / filename), part_image)
                    training_samples.append(
                        {
                            "sample_id": sample_id,
                            "page": page_index + 1,
                            "panel": panel_index + 1,
                            "row": row_index + 1,
                            "column": field_name,
                            "ocr_text": part_ocr_text,
                            "normalized_label": part_normalized,
                            "confidence": round(part_score, 8),
                            "correct_label": part_correct_label,
                            "include": (
                                "是" if part_correct_label else "待确认"
                            ),
                            "image": f"images/{filename}",
                        "note": (
                            f"{part_note} 时分编码模式={time_mode}，"
                            f"辅助锚点x={panel_time_colon_x[panel_index]}，"
                            f"动态冒号边界="
                            f"{review_time_colon_band if review_time_colon_band is not None else '无'}。 "
                            f"{recognized_decisions.get((panel_index, row_index, 8 if part_key == 'hour' else 9), '')} "
                            f"{recognized_length_warnings.get((panel_index, row_index, 8 if part_key == 'hour' else 9), '')}"
                        ),
                        }
                    )

                water_integer_label = (
                    water_integer_digits if water_dot_observed else ""
                )
                water_fraction_label = water_fraction_digits
                full_water_match = re.fullmatch(
                    r"(\d+)\.(\d+)", values[3]
                )
                if full_water_match:
                    water_integer_label = full_water_match.group(1)
                    water_fraction_label = full_water_match.group(2)
                water_part_specs = [
                    (
                        "integer",
                        "i",
                        "水位整数",
                        water_integer_text,
                        water_integer_digits,
                        water_integer_label,
                        float(water_integer_score),
                        "water_integer",
                        "小数点左侧整数子图；训练标签不包含小数点。",
                    ),
                    (
                        "fraction",
                        "f",
                        "水位小数",
                        water_fraction_text,
                        water_fraction_digits,
                        water_fraction_label,
                        float(water_fraction_score),
                        "water_fraction",
                        "小数点右侧小数子图；训练标签不包含小数点。",
                    ),
                ]
                for (
                    part_key,
                    id_suffix,
                    field_name,
                    part_ocr_text,
                    part_normalized,
                    part_correct_label,
                    part_score,
                    filename_suffix,
                    part_note,
                ) in water_part_specs:
                    part_image = water_part_images.get(
                        (panel_index, row_index, part_key)
                    )
                    if part_image is None:
                        continue
                    sample_id = (
                        f"p{page_index + 1:03d}_b{panel_index + 1}_"
                        f"r{row_index + 1:03d}_c04{id_suffix}"
                    )
                    filename = f"{sample_id}_{filename_suffix}.png"
                    cv2.imwrite(str(training_images_dir / filename), part_image)
                    training_samples.append(
                        {
                            "sample_id": sample_id,
                            "page": page_index + 1,
                            "panel": panel_index + 1,
                            "row": row_index + 1,
                            "column": field_name,
                            "ocr_text": part_ocr_text,
                            "normalized_label": part_normalized,
                            "confidence": round(part_score, 8),
                            "correct_label": part_correct_label,
                            "include": (
                                "是" if part_correct_label else "待确认"
                            ),
                            "image": f"images/{filename}",
                        "note": (
                            f"{part_note} 小数点x="
                            f"{review_water_decimal_x}。 "
                            f"{recognized_decisions.get((panel_index, row_index, 12 if part_key == 'integer' else 13), '')} "
                            f"{recognized_length_warnings.get((panel_index, row_index, 12 if part_key == 'integer' else 13), '')}"
                        ),
                        }
                    )

            scores = [
                score
                for value, score in zip(values, effective_scores)
                if value
            ]
            day_marker_min_digits = 0
            if 1 in canonical_indices:
                day_marker_min_digits = int(
                    cell_ink_run_lower_bounds.get(
                        (
                            panel_index,
                            row_index,
                            canonical_indices.index(1),
                        ),
                        0,
                    )
                )
            output_rows.append(
                {
                    "page": page_index + 1,
                    "panel": panel_index + 1,
                    "row": row_index + 1,
                    "row_center": center_y,
                    "values": values,
                    "value_scores": effective_scores,
                    "context_candidates": context_candidates,
                    # These booleans are the machine-readable form of the
                    # month/day ink boxes shown in the grid preview.  They
                    # distinguish an explicit printed marker from a blank row
                    # covered by the preceding marker, independently of what
                    # OCR text was selected later.
                    "_month_marker_ink": bool(
                        0 in canonical_indices
                        and (
                            panel_index,
                            row_index,
                            canonical_indices.index(0),
                        )
                        in primary_cell_images
                    ),
                    "_day_marker_ink": bool(
                        1 in canonical_indices
                        and (
                            panel_index,
                            row_index,
                            canonical_indices.index(1),
                        )
                        in primary_cell_images
                    ),
                    "_day_marker_min_digits": min(
                        2, day_marker_min_digits
                    ),
                    # Preserve the independently recognized water parts even
                    # when this logical row was recovered from the time
                    # anchor and the decimal-point observation flag is false.
                    # A station-level resolver can then validate the complete
                    # value against following explicit water levels.
                    "_water_split_candidate": (
                        f"{water_integer_digits}.{water_fraction_digits}"
                        if water_integer_available
                        and water_fraction_available
                        else ""
                    ),
                    "_water_split_score": round(
                        (
                            float(water_integer_score)
                            + float(water_fraction_score)
                        )
                        / 2,
                        8,
                    )
                    if water_integer_available and water_fraction_available
                    else 0.0,
                    "_water_decimal_observed": water_dot_observed,
                    "average_confidence": (
                        sum(scores) / len(scores) if scores else 0.0
                    ),
                }
            )

    metadata = {
        "page": page_index + 1,
        "legacy_dynamic_panel_count": len(panel_boundaries),
        "legacy_dynamic_field_count": physical_field_count,
        "schema": schema_metadata,
        "table_bbox": list(table_bbox),
        "data_top": data_top,
        "station_marker": station_markers[0] if station_markers else None,
        "station_markers": station_markers,
        "panels": [
            {
                "panel": index + 1,
                "row_count": len(rows),
                "row_anchor": (
                    "observed_dual_anchor_alignment_with_multicolumn_ink_hypotheses"
                ),
                "anchor_alignment": panel_anchor_summaries[index],
                "anchor_status_counts": {
                    status: panel_anchor_statuses[index].count(status)
                    for status in sorted(set(panel_anchor_statuses[index]))
                },
                "observed_decimal_anchor_count": int(
                    sum(panel_water_decimal_observed[index])
                ),
                "water_anchor_mode": panel_water_anchor_modes[index],
                "water_fraction_token_anchor_count": len(
                    panel_water_fraction_rows[index]
                ),
                "water_dot_coverage": panel_anchor_summaries[index].get(
                    "water_dot_coverage", 1.0
                ),
                "water_anchor_semantics": (
                    "fraction_glyph_center_with_sparse_decimal_attribute"
                    if panel_water_anchor_modes[index]
                    == "sparse_fraction_token"
                    else "observed_decimal_point_center"
                ),
                "joint_inferred_or_auxiliary_row_count": int(
                    len(rows) - sum(panel_water_decimal_observed[index])
                ),
                "time_encoding_mode": panel_time_modes[index],
                "auxiliary_anchor": (
                    "time_numeric_token_center_y"
                    if panel_time_modes[index] == "hour_only"
                    else "time_colon_corridor_y"
                ),
                "observed_colon_anchor_count": int(
                    sum(
                        band is not None
                        for band in panel_time_colon_bands[index]
                    )
                ),
                "observed_time_anchor_count": int(
                    sum(panel_time_colon_observed[index])
                ),
                "dynamic_colon_band_count": int(
                    sum(
                        band is not None
                        for band in panel_time_colon_bands[index]
                    )
                ),
                "time_split_strategy": (
                    "whole_numeric_token_as_hour_with_implicit_zero_minute"
                    if panel_time_modes[index] == "hour_only"
                    else "hour_before_colon_left_edge_and_minute_after_colon_right_edge"
                ),
                "ink_ownership": (
                    "water_right_presence_gate_and_time_left_first_claim"
                ),
                "local_row_alignment": True,
                "local_row_max_offset": LEGACY_LOCAL_ROW_MAX_OFFSET,
                "time_colon_x": panel_time_colon_x[index],
                "water_decimal_x_top": round(
                    panel_water_decimal_models[index][0] * data_top
                    + panel_water_decimal_models[index][1]
                ),
                "water_decimal_x_bottom": round(
                    panel_water_decimal_models[index][0] * table_bottom
                    + panel_water_decimal_models[index][1]
                ),
            }
            for index, rows in enumerate(panel_rows)
        ],
        "preview": str(preview_path),
    }
    return output_rows, metadata, training_samples


def legacy_days_in_month(month: int | None) -> int:
    """Return the maximum legal day without assuming a particular year."""
    if month == 2:
        return 29
    if month in {4, 6, 9, 11}:
        return 30
    return 31


def parse_legacy_clock(value: str) -> tuple[str, int] | None:
    """Parse a 24-hour clock, retaining historical 24:00 as day-end."""
    match = re.fullmatch(r"\s*(\d{1,2})\s*[:：]\s*(\d{1,2})\s*", value)
    if match is None:
        return None
    hour, minute = map(int, match.groups())
    if not (0 <= minute <= 59):
        return None
    if not (0 <= hour <= 23 or (hour == 24 and minute == 0)):
        return None
    return f"{hour:02d}:{minute:02d}", hour * 60 + minute


def apply_legacy_date_marker_coverage(
    rows: list[dict[str, Any]],
) -> None:
    """Assign month/day spans from preview geometry before clock decoding.

    ``_month_marker_ink`` and ``_day_marker_ink`` come from the same owned-cell
    geometry used to draw the preview.  A valid marker governs its own row and
    following blank rows until the next observed marker.  An observed marker
    whose value is unreadable starts an unknown span instead of silently
    inheriting through a real printed boundary.
    """
    current_month: int | None = None
    current_day: int | None = None
    month_span_id = 0
    day_span_id = 0
    day_marker_seen = False

    # Resolve values on the physical day-marker rows before expanding their
    # spans.  This pass deliberately uses only day OCR alternatives and the
    # ordered marker sequence, never clock order.  It is global within each
    # decoded month, because one damaged marker can otherwise hide the useful
    # 22 ... 24 context from a local three-marker repair.
    physical_day_markers = [
        index
        for index, row in enumerate(rows)
        if bool(row.get("_day_marker_ink", False))
        or "日" in set(row.get("_manual_corrected_fields", []))
    ]
    marker_groups: list[tuple[int, list[int]]] = []
    for row_index in physical_day_markers:
        month = parse_integer(rows[row_index]["values"][0], 1, 12)
        if month is None:
            continue
        if not marker_groups or marker_groups[-1][0] != month:
            marker_groups.append((month, [row_index]))
        else:
            marker_groups[-1][1].append(row_index)

    for month, marker_indices in marker_groups:
        maximum_day = legacy_days_in_month(month)
        marker_options: list[dict[int, tuple[float, str, str]]] = []
        for row_index in marker_indices:
            minimum_digits = int(
                rows[row_index].get("_day_marker_min_digits", 0) or 0
            )
            raw_options = legacy_context_options(
                rows[row_index],
                "日",
                lambda value, limit=maximum_day: parse_integer(
                    value, 1, limit
                ),
                str(rows[row_index]["values"][1]),
                1,
            )
            marker_options.append(
                {
                    int(value): detail
                    for value, detail in raw_options.items()
                    if len(str(int(value))) >= minimum_digits
                }
            )
        if not any(marker_options):
            continue
        manual_values = [
            (
                position,
                parse_integer(
                    rows[row_index]["values"][1], 1, maximum_day
                ),
            )
            for position, row_index in enumerate(marker_indices)
            if "日" in set(
                rows[row_index].get("_manual_corrected_fields", [])
            )
        ]
        if any(value is None for _, value in manual_values):
            raise RuntimeError(f"{month}月存在非法的人工日期标记。")
        if any(
            int(following[1]) < int(previous[1])
            for previous, following in zip(
                manual_values[:-1], manual_values[1:]
            )
        ):
            raise RuntimeError(f"{month}月的人工日期标记发生倒退。")

        # Select the maximum-cardinality nondecreasing chain of observed OCR
        # candidates.  A legal jump such as 19 -> 21 is retained; no continuity
        # cost is allowed to rewrite it.  Confidence only breaks ties between
        # chains that preserve the same number of physical observations.
        nodes: list[tuple[int, int, float, str, bool, bool]] = []
        for position, (row_index, options) in enumerate(
            zip(marker_indices, marker_options)
        ):
            manual = "日" in set(
                rows[row_index].get("_manual_corrected_fields", [])
            )
            selected_day = parse_integer(
                rows[row_index]["values"][1], 1, maximum_day
            )
            for day, detail in options.items():
                candidate_source = str(detail[2])
                nodes.append(
                    (
                        position,
                        int(day),
                        float(detail[0]),
                        candidate_source,
                        manual,
                        selected_day == int(day),
                    )
                )
        nodes.sort(key=lambda node: (node[0], node[1]))
        chain_scores: list[float] = []
        chain_previous = [-1] * len(nodes)
        for node_index, node in enumerate(nodes):
            position, day, confidence, source, manual, selected = node
            node_weight = (
                1000.0
                if manual
                else (10.0 if selected else 1.0)
                + confidence
            )
            best_score = node_weight
            best_previous = -1
            for previous_index, previous_node in enumerate(nodes[:node_index]):
                previous_position, previous_day = previous_node[:2]
                if previous_position >= position or previous_day > day:
                    continue
                candidate_score = (
                    chain_scores[previous_index] + node_weight
                )
                if candidate_score > best_score:
                    best_score = candidate_score
                    best_previous = previous_index
            chain_scores.append(best_score)
            chain_previous[node_index] = best_previous

        selected_nodes: dict[int, tuple[int, float, str]] = {}
        if nodes:
            node_index = max(
                range(len(nodes)),
                key=lambda index: chain_scores[index],
            )
            while node_index >= 0:
                position, day, confidence, source, _, _ = nodes[node_index]
                selected_nodes[position] = (day, confidence, source)
                node_index = chain_previous[node_index]

        # Manual markers are hard anchors.  If the unconstrained OCR chain did
        # not include one, retain it and discard only incompatible OCR anchors.
        for position, value in manual_values:
            day = int(value)
            selected_nodes[position] = (day, 1.0, "manual-marker")
        ordered_anchors = sorted(selected_nodes.items())
        for left_index, (left_position, left_detail) in enumerate(
            ordered_anchors
        ):
            left_day = int(left_detail[0])
            for right_position, right_detail in ordered_anchors[left_index + 1 :]:
                if int(right_detail[0]) < left_day:
                    selected_nodes.pop(right_position, None)

        decoded_days: list[int | None] = [None] * len(marker_indices)
        resolution_sources: list[str] = [""] * len(marker_indices)
        for position, detail in selected_nodes.items():
            decoded_days[position] = int(detail[0])
            resolution_sources[position] = str(detail[2])
        ordered_positions = sorted(selected_nodes)
        for left_position, right_position in zip(
            ordered_positions[:-1], ordered_positions[1:]
        ):
            left_day = int(selected_nodes[left_position][0])
            right_day = int(selected_nodes[right_position][0])
            position_gap = right_position - left_position
            day_gap = right_day - left_day
            if day_gap != position_gap:
                continue
            # Only an exact gap has a unique solution.  Wider day gaps are
            # legitimate missing-record days and remain untouched.
            for position in range(left_position + 1, right_position):
                decoded_days[position] = left_day + position - left_position
                resolution_sources[position] = (
                    "physical-day-marker-exact-gap-inference"
                )

        for position, (row_index, replacement) in enumerate(
            zip(marker_indices, decoded_days)
        ):
            current_day = parse_integer(
                rows[row_index]["values"][1], 1, maximum_day
            )
            if replacement is None:
                rows[row_index]["values"][1] = ""
                rows[row_index]["_date_day_marker_resolution"] = {
                    "from": current_day,
                    "to": None,
                    "month": month,
                    "marker_position": position + 1,
                    "marker_count": len(marker_indices),
                    "source": "unresolved-physical-day-marker",
                }
                continue
            if current_day == replacement:
                continue
            rows[row_index]["values"][1] = str(replacement)
            rows[row_index]["_date_day_marker_resolution"] = {
                "from": current_day,
                "to": replacement,
                "month": month,
                "marker_position": position + 1,
                "marker_count": len(marker_indices),
                "source": resolution_sources[position],
            }

    def nearby_day_marker_value(start_index: int) -> int | None:
        for candidate in rows[start_index : start_index + 8]:
            if not bool(candidate.get("_day_marker_ink", False)):
                continue
            selected = parse_integer(candidate["values"][1], 1, 31)
            if selected is not None:
                return selected
            options = legacy_context_options(
                candidate,
                "日",
                lambda value: parse_integer(value, 1, 31),
                str(candidate["values"][1]),
                1,
            )
            if options:
                return int(
                    max(options.items(), key=lambda item: item[1][0])[0]
                )
        return None

    for row_index, row in enumerate(rows):
        values = row["values"]
        manual_fields = set(row.get("_manual_corrected_fields", []))
        physical_month_marker = bool(
            row.get("_month_marker_ink", False)
        )
        marker_month = physical_month_marker or (
            "月" in manual_fields
        )
        marker_day = bool(row.get("_day_marker_ink", False)) or (
            "日" in manual_fields
        )
        observed_month = parse_integer(values[0], 1, 12)
        month_marker_source = (
            "manual-marker"
            if "月" in manual_fields
            else "explicit-marker"
            if physical_month_marker
            else ""
        )
        if (
            not marker_month
            and current_month is not None
            and observed_month is not None
            and observed_month != current_month
        ):
            # The global month candidate stream may locate a boundary whose
            # tiny month glyph was missed by row-level ink ownership.  Use the
            # first decoded value change as a secondary boundary; unlike clock
            # rollback, this evidence comes from the month OCR candidates.
            marker_month = True
            month_marker_source = "decoded-month-candidate-boundary"
            row["_date_month_boundary_without_owned_cell"] = True

        month_changed = False
        if marker_month:
            month_options = legacy_context_options(
                row,
                "月",
                lambda value: parse_integer(value, 1, 12),
                str(values[0]),
                0,
            )
            nearby_day = nearby_day_marker_value(row_index)
            expected_next_month = (
                current_month % 12 + 1
                if current_month is not None
                else None
            )
            day_reset_supports_new_month = bool(
                current_day is not None
                and current_day >= 25
                and nearby_day is not None
                and nearby_day <= 7
            )
            if day_reset_supports_new_month and expected_next_month is not None:
                if expected_next_month in month_options:
                    observed_month = expected_next_month
                    row["_date_month_marker_resolution"] = (
                        "月标记候选结合日标记回卷选择下一月"
                    )
                elif observed_month == current_month:
                    # The boundary ink is explicit even when its digit OCR
                    # repeats the previous month.  A late-day to early-day
                    # marker reset supplies independent structural evidence.
                    observed_month = expected_next_month
                    row["_date_month_marker_resolution"] = (
                        "显式月标记伴随日标记由月末回卷至月初"
                    )
            month_span_id += 1
            month_changed = (
                current_month is not None
                and observed_month is not None
                and observed_month != current_month
            )
            current_month = observed_month
        elif current_month is None and observed_month is not None:
            # Only the station-leading segment may rely on an inferred month;
            # once a printed marker has been seen, blank rows stay in its span.
            month_span_id += 1
            current_month = observed_month

        if current_month is not None:
            values[0] = str(current_month)
        row["_date_month_span_id"] = month_span_id or None
        row["_date_span_month"] = current_month
        row["_date_month_span_source"] = (
            month_marker_source
            if marker_month and current_month is not None
            else "unknown-marker"
            if marker_month
            else "covered-by-previous-marker"
            if month_span_id
            else "unresolved"
        )

        if month_changed:
            current_day = None
            day_marker_seen = False
        maximum_day = (
            legacy_days_in_month(current_month)
            if current_month is not None
            else 31
        )
        observed_day = parse_integer(values[1], 1, maximum_day)
        if marker_day:
            day_marker_seen = True
            day_span_id += 1
            current_day = observed_day
        elif (
            not day_marker_seen
            and current_day is None
            and observed_day is not None
        ):
            # This covers a station-leading day recovered bidirectionally by
            # the existing leading-edge logic.  It remains auditable as an
            # inferred span rather than being mislabeled an explicit marker.
            day_span_id += 1
            current_day = observed_day

        if current_day is not None:
            values[1] = str(current_day)
        elif day_marker_seen:
            values[1] = ""
        row["_date_day_span_id"] = day_span_id or None
        row["_date_span_day"] = current_day
        row["_date_day_span_source"] = (
            "explicit-marker"
            if marker_day and current_day is not None
            else "unknown-marker"
            if marker_day
            else "covered-by-unknown-marker"
            if day_marker_seen and current_day is None
            else "covered-by-previous-marker"
            if day_span_id and current_day is not None
            else "leading-inference"
            if day_span_id
            else "unresolved"
        )

    for span_key, start_key, end_key in (
        (
            "_date_month_span_id",
            "_date_month_span_start_row",
            "_date_month_span_end_row",
        ),
        (
            "_date_day_span_id",
            "_date_day_span_start_row",
            "_date_day_span_end_row",
        ),
    ):
        span_rows: dict[int, list[int]] = {}
        for row_index, row in enumerate(rows, start=1):
            span_id = row.get(span_key)
            if span_id is not None:
                span_rows.setdefault(int(span_id), []).append(row_index)
        for row_indexes in span_rows.values():
            start_row = min(row_indexes)
            end_row = max(row_indexes)
            for row_index in row_indexes:
                rows[row_index - 1][start_key] = start_row
                rows[row_index - 1][end_key] = end_row


def reconcile_legacy_time_candidate_with_split_minute(
    value: str,
    score: float,
    source: str,
    split_minute_digits: str,
    split_minute_score: float,
) -> tuple[str, float, str]:
    """Make a whole-time fallback agree with reliable colon-split minutes.

    A wide/raw time crop can retain the right table rule and read ``20:00`` as
    ``20:01``.  Chronology alone may then prefer that contaminated candidate
    over a duplicated or rolled-back selected clock.  The colon-split minute
    crop is geometrically more specific: when it contains exactly two legal
    digits at high confidence, keep a fallback candidate's useful hour but
    replace a conflicting minute.  The small score penalty ensures an already
    consistent observation still wins a tie.
    """
    parsed = parse_legacy_clock(value)
    if parsed is None:
        return value, float(score), source
    if (
        len(split_minute_digits) != 2
        or not split_minute_digits.isdigit()
        or float(split_minute_score)
        < LEGACY_TIME_SPLIT_MINUTE_CONSTRAINT_CONFIDENCE
    ):
        return value, float(score), source
    split_minute = int(split_minute_digits)
    if not 0 <= split_minute <= 59:
        return value, float(score), source

    normalized, _ = parsed
    hour, candidate_minute = map(int, normalized.split(":"))
    if candidate_minute == split_minute:
        return normalized, float(score), source
    if hour == 24 and split_minute != 0:
        return value, float(score), source

    reconciled = f"{hour:02d}:{split_minute:02d}"
    return (
        reconciled,
        max(0.0, float(score) - LEGACY_TIME_MINUTE_RECONCILIATION_PENALTY),
        (
            f"{source}|minute-reconciled-from-{candidate_minute:02d}"
            f"-to-{split_minute:02d}"
        ),
    )


def legacy_context_options(
    row: dict[str, Any],
    field: str,
    parser: Any,
    selected_value: str,
    score_index: int,
) -> dict[Any, tuple[float, str, str]]:
    """Return unique parsed alternatives keyed by their semantic value."""
    options: dict[Any, tuple[float, str, str]] = {}

    def add(value: str, score: float, source: str) -> None:
        parsed = parser(value)
        if parsed is None:
            return
        semantic = parsed[1] if isinstance(parsed, tuple) else parsed
        normalized = parsed[0] if isinstance(parsed, tuple) else str(parsed)
        previous = options.get(semantic)
        candidate = (float(score), normalized, source)
        if previous is None or candidate[0] > previous[0]:
            options[semantic] = candidate

    scores = row.get("value_scores") or []
    selected_score = (
        float(scores[score_index]) if score_index < len(scores) else 0.0
    )
    add(selected_value, selected_score, "row-selected")
    for candidate in row.get("context_candidates", {}).get(field, []):
        add(
            str(candidate.get("value") or ""),
            float(candidate.get("score") or 0.0),
            str(candidate.get("source") or "candidate"),
        )
    return options


def legacy_day_rollover_evidence(
    rows: list[dict[str, Any]],
) -> list[float]:
    """Estimate where a chronological day sequence crosses a month boundary.

    The evidence is independent of the month OCR stream.  A clear late-month
    to early-month rollback is strong evidence; small backwards moves are
    treated as likely day OCR noise and do not reset the running day.
    """
    evidence = [0.0] * len(rows)
    last_day: int | None = None
    for index, row in enumerate(rows):
        month_options = legacy_context_options(
            row,
            "月",
            lambda value: parse_integer(value, 1, 12),
            row["values"][0],
            0,
        )
        options = legacy_context_options(
            row,
            "日",
            lambda value: parse_integer(value, 1, 31),
            row["values"][1],
            1,
        )
        if not options:
            continue
        day, _ = max(options.items(), key=lambda item: item[1][0])
        day = int(day)
        if last_day is None:
            last_day = day
            continue
        drop = last_day - day
        # In this annual-table family a real month boundary writes a month
        # marker on the same row.  A day-only rollback is much more likely to
        # be a truncated/misread day and must not move the calendar state.
        if month_options and last_day >= 25 and day <= 7:
            evidence[index] = 1.0
            last_day = day
        elif month_options and last_day >= 20 and day <= 10 and drop >= 14:
            evidence[index] = min(0.95, 0.65 + drop / 60)
            last_day = day
        elif day >= last_day:
            last_day = day
    return evidence


def decode_legacy_month_sequence(
    rows: list[dict[str, Any]],
    locked: Any,
) -> tuple[list[int], list[float]]:
    """Globally decode the inherited month sequence with dynamic programming.

    OCR month markers are sparse emissions.  Day rollbacks are an independent
    transition signal.  The full station is optimized at once, so a later
    reliable marker can correct an earlier low-confidence marker instead of
    being rejected as a backwards month.
    """
    if not rows:
        return [], []
    rollover = legacy_day_rollover_evidence(rows)
    month_options: list[dict[int, tuple[float, str, str]]] = []
    for row in rows:
        options = legacy_context_options(
            row,
            "月",
            lambda value: parse_integer(value, 1, 12),
            row["values"][0],
            0,
        )
        month_options.append(
            {int(value): detail for value, detail in options.items()}
        )

    def emission_cost(row_index: int, month: int) -> float:
        options = month_options[row_index]
        if locked(rows[row_index], "月"):
            selected = parse_integer(rows[row_index]["values"][0], 1, 12)
            return 0.0 if selected == month else 1000.0
        if not options:
            return 0.0
        best_score = max(detail[0] for detail in options.values())
        if month in options:
            return (1.0 - options[month][0]) * 1.6
        return 4.0 + best_score * 6.0

    def transition_cost(
        previous: int,
        current: int,
        strength: float,
        has_month_marker: bool,
    ) -> float:
        forward = (current - previous) % 12
        if forward == 0:
            return 3.8 * strength
        if forward == 1:
            if strength > 0:
                return 0.05 + (1.0 - strength) * 2.35
            return 1.8 if has_month_marker else 12.0
        if 2 <= forward <= 6:
            return 14.0 + forward * 1.5 - strength
        return 22.0 + (12 - forward) * 0.6

    infinity = float("inf")
    costs = [[infinity] * 13 for _ in rows]
    previous_states = [[0] * 13 for _ in rows]
    for month in range(1, 13):
        costs[0][month] = emission_cost(0, month)
    for row_index in range(1, len(rows)):
        for month in range(1, 13):
            emission = emission_cost(row_index, month)
            for previous_month in range(1, 13):
                candidate = (
                    costs[row_index - 1][previous_month]
                    + transition_cost(
                        previous_month,
                        month,
                        rollover[row_index],
                        bool(month_options[row_index]),
                    )
                    + emission
                )
                if candidate < costs[row_index][month]:
                    costs[row_index][month] = candidate
                    previous_states[row_index][month] = previous_month

    final_month = min(range(1, 13), key=lambda month: costs[-1][month])
    decoded = [final_month]
    for row_index in range(len(rows) - 1, 0, -1):
        final_month = previous_states[row_index][final_month]
        decoded.append(final_month)
    decoded.reverse()
    return decoded, rollover


def decode_legacy_joint_day_time_sequence(
    rows: list[dict[str, Any]],
    locked: Any,
) -> list[tuple[int, str, str]]:
    """Decode day and clock jointly under a strictly increasing timeline.

    The earlier staged resolver allowed equal clocks on the inherited date.
    That could flatten several distinct observations into ``23 08:00``.  This
    dynamic program treats the complete month/day/minute tuple as the state:
    known timestamps must advance strictly, a clock rollback favours the next
    day, and an unresolved clock is left blank instead of being fabricated.
    """
    if not rows:
        return []

    months = [parse_integer(row["values"][0], 1, 12) for row in rows]
    if any(month is None for month in months):
        raise RuntimeError("联合日期时分解码缺少合法月份。")

    month_bases: list[int] = []
    cycle = 0
    previous_month = int(months[0])
    for parsed_month in months:
        month = int(parsed_month)
        if month < previous_month:
            cycle += 1
        days_before = sum(
            legacy_days_in_month(value) for value in range(1, month)
        )
        month_bases.append(cycle * 366 + days_before)
        previous_month = month

    day_options_by_row: list[dict[int, tuple[float, str, str]]] = []
    time_options_by_row: list[
        dict[int, tuple[float, str, str]]
    ] = []
    for row, parsed_month in zip(rows, months):
        maximum_day = legacy_days_in_month(int(parsed_month))
        day_options = legacy_context_options(
            row,
            "日",
            lambda value, limit=maximum_day: parse_integer(value, 1, limit),
            row["values"][1],
            1,
        )
        minimum_day_digits = int(
            row.get("_day_marker_min_digits", 0) or 0
        )
        if minimum_day_digits:
            day_options = {
                value: detail
                for value, detail in day_options.items()
                if len(str(int(value))) >= minimum_day_digits
            }
        time_options = legacy_context_options(
            row,
            "时分",
            parse_legacy_clock,
            row["values"][2],
            2,
        )
        day_options_by_row.append(
            {int(value): detail for value, detail in day_options.items()}
        )
        time_options_by_row.append(
            {int(value): detail for value, detail in time_options.items()}
        )

    def day_emission(row_index: int, day: int) -> float:
        options = day_options_by_row[row_index]
        covered_day = rows[row_index].get("_date_span_day")
        if covered_day is not None:
            return 0.0 if int(covered_day) == day else float("inf")
        if locked(rows[row_index], "日"):
            selected = parse_integer(
                rows[row_index]["values"][1],
                1,
                legacy_days_in_month(int(months[row_index])),
            )
            return 0.0 if selected == day else float("inf")
        if not options:
            return 0.0
        if day in options:
            return (1.0 - float(options[day][0])) * 1.8
        best_score = max(float(detail[0]) for detail in options.values())
        return 5.5 + best_score * 3.0

    def time_states(
        row_index: int,
    ) -> list[tuple[int | None, float, str]]:
        options = time_options_by_row[row_index]
        if locked(rows[row_index], "时分"):
            selected = parse_legacy_clock(rows[row_index]["values"][2])
            if selected is None:
                return [(None, 0.0, "manual-invalid-clock")]
            return [(selected[1], 0.0, selected[0])]
        states = [
            (
                int(minutes),
                (1.0 - float(detail[0])) * 2.0,
                str(detail[1]),
            )
            for minutes, detail in options.items()
        ]
        # Blank remains a safety state, but the DP state below carries the last
        # known timestamp across it.  Therefore a blank can never be used to
        # make a later clock regression appear legal.
        states.append((None, 6.0 if options else 1.5, ""))
        return states

    def transition_cost(
        previous_serial: int,
        previous_minutes: int | None,
        current_serial: int,
        current_minutes: int | None,
    ) -> float:
        day_delta = current_serial - previous_serial
        if day_delta < 0:
            return float("inf")
        if day_delta == 0:
            if (
                previous_minutes is not None
                and current_minutes is not None
                and current_minutes <= previous_minutes
            ):
                return float("inf")
            return 0.05 if None not in (previous_minutes, current_minutes) else 1.0
        if day_delta == 1:
            if previous_minutes is None or current_minutes is None:
                return 0.8
            return 0.05 if current_minutes <= previous_minutes else 0.65
        return 1.5 + 0.8 * (day_delta - 1)

    # State = (chosen_day, chosen_minutes, last_known_day_serial,
    #          last_known_minutes).  The latter pair is deliberately retained
    # when chosen_minutes is None, so chronology remains strict across blanks.
    JointState = tuple[int, int | None, int | None, int | None]
    costs: dict[JointState, float] = {}
    backpointers: list[dict[JointState, JointState | None]] = []
    time_labels: list[dict[int | None, str]] = []
    for row_index, parsed_month in enumerate(months):
        maximum_day = legacy_days_in_month(int(parsed_month))
        current_costs: dict[JointState, float] = {}
        current_previous: dict[JointState, JointState | None] = {}
        label_by_minutes: dict[int | None, str] = {}
        candidate_times = time_states(row_index)
        for minutes, _, label in candidate_times:
            label_by_minutes[minutes] = label
        candidate_days = range(1, maximum_day + 1)
        prepared_days = [
            (
                day,
                month_bases[row_index] + day,
                day_emission(row_index, day),
            )
            for day in candidate_days
        ]
        prepared_days = [
            item for item in prepared_days if np.isfinite(item[2])
        ]
        if row_index == 0:
            for day, current_serial, day_cost in prepared_days:
                for minutes, clock_cost, _ in candidate_times:
                    state: JointState = (
                        day,
                        minutes,
                        current_serial if minutes is not None else None,
                        minutes,
                    )
                    current_costs[state] = day_cost + clock_cost
                    current_previous[state] = None
        else:
            for previous_state, previous_cost in costs.items():
                (
                    previous_day,
                    previous_minutes,
                    previous_last_serial,
                    previous_last_minutes,
                ) = previous_state
                previous_row_serial = (
                    month_bases[row_index - 1] + previous_day
                )
                for day, current_serial, day_cost in prepared_days:
                    if current_serial < previous_row_serial:
                        continue
                    previous_span = rows[row_index - 1].get(
                        "_date_day_span_id"
                    )
                    current_span = rows[row_index].get(
                        "_date_day_span_id"
                    )
                    if (
                        previous_span is not None
                        and previous_span == current_span
                        and day != previous_day
                    ):
                        # Preview geometry has already established that these
                        # rows belong to one printed date coverage block.  An
                        # ambiguous or rolled-back clock may affect candidate
                        # confidence, but it cannot split that block into two
                        # dates.
                        continue
                    for minutes, clock_cost, _ in candidate_times:
                        if minutes is None:
                            last_serial = previous_last_serial
                            last_minutes = previous_last_minutes
                            step_cost = (
                                1.0
                                + 0.2
                                * max(0, current_serial - previous_row_serial)
                            )
                        else:
                            if previous_last_serial is not None:
                                if current_serial < previous_last_serial:
                                    continue
                                if (
                                    current_serial == previous_last_serial
                                    and previous_last_minutes is not None
                                    and minutes <= previous_last_minutes
                                ):
                                    continue
                            last_serial = current_serial
                            last_minutes = minutes
                            step_cost = transition_cost(
                                previous_last_serial
                                if previous_last_serial is not None
                                else previous_row_serial,
                                previous_last_minutes,
                                current_serial,
                                minutes,
                            )
                        state = (
                            day,
                            minutes,
                            last_serial,
                            last_minutes,
                        )
                        candidate_cost = (
                            previous_cost
                            + step_cost
                            + day_cost
                            + clock_cost
                        )
                        if candidate_cost < current_costs.get(
                            state, float("inf")
                        ):
                            current_costs[state] = candidate_cost
                            current_previous[state] = previous_state
        if not current_costs:
            diagnostic = {
                "row": row_index + 1,
                "values": rows[row_index].get("values", [])[:3],
                "month_span": rows[row_index].get("_date_month_span_id"),
                "day_span": rows[row_index].get("_date_day_span_id"),
                "span_day": rows[row_index].get("_date_span_day"),
                "day_source": rows[row_index].get(
                    "_date_day_span_source"
                ),
                "previous_values": (
                    rows[row_index - 1].get("values", [])[:3]
                    if row_index > 0
                    else []
                ),
                "month_markers": [
                    (
                        index + 1,
                        candidate.get("values", [])[:2],
                        candidate.get("_date_month_marker_resolution"),
                    )
                    for index, candidate in enumerate(rows)
                    if candidate.get("_month_marker_ink")
                ],
            }
            raise RuntimeError(
                f"第{row_index + 1}行无法形成严格递增的日期时分序列："
                f"{diagnostic}"
            )
        # Blank states can preserve several different prior stamps.  Keep a
        # generous beam to bound pathological scans without changing ordinary
        # rows, whose known-clock states are already compact.
        if len(current_costs) > 800:
            kept_states = sorted(current_costs, key=current_costs.get)[:800]
            current_costs = {
                state: current_costs[state] for state in kept_states
            }
            current_previous = {
                state: current_previous[state] for state in kept_states
            }
        costs = current_costs
        backpointers.append(current_previous)
        time_labels.append(label_by_minutes)

    state = min(costs, key=costs.get)
    decoded_states = [state]
    for row_index in range(len(rows) - 1, 0, -1):
        previous_state = backpointers[row_index][state]
        if previous_state is None:
            raise RuntimeError("联合日期时分序列回溯失败。")
        decoded_states.append(previous_state)
        state = previous_state
    decoded_states.reverse()

    decoded: list[tuple[int, str, str]] = []
    last_known_stamp: tuple[int, int] | None = None
    for row_index, state in enumerate(decoded_states):
        day, minutes = state[0], state[1]
        if minutes is not None:
            stamp = (month_bases[row_index] + day, minutes)
            if last_known_stamp is not None and stamp <= last_known_stamp:
                raise RuntimeError(
                    "联合日期时分解码内部校核失败：已知时间戳没有严格递增。"
                )
            last_known_stamp = stamp
        label = time_labels[row_index].get(minutes, "")
        source = (
            "strict-joint-calendar-clock-dp"
            if minutes is not None
            else "strict-chronology-unresolved-clock"
        )
        decoded.append((day, label, source))
    return decoded


def resolve_legacy_calendar_time_context(
    raw_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Choose month/day/time OCR alternatives using neighbouring rows.

    The resolver is deliberately conservative: it keeps a valid selected
    value unless it breaks the surrounding calendar/clock order and another
    observed candidate repairs that break.  A value is inferred without an
    OCR alternative only for an exact one-step sandwich such as 2, ?, 4.
    """
    rows = [
        {
            **row,
            "values": list(row["values"]),
            "_context_corrections": list(
                row.get("_context_corrections", [])
            ),
        }
        for row in raw_rows
    ]

    def locked(row: dict[str, Any], field: str) -> bool:
        return field in set(row.get("_manual_corrected_fields", []))

    def record(
        row: dict[str, Any],
        field: str,
        old_value: str,
        new_value: str,
        reason: str,
        source: str,
    ) -> None:
        if old_value == new_value:
            return
        row["_context_corrections"].append(
            {
                "field": field,
                "from": old_value,
                "to": new_value,
                "reason": reason,
                "source": source,
            }
        )

    # --- Month: decode the entire station instead of greedily inheriting one
    # marker.  Write the decoded state on every logical row so all downstream
    # day/time checks use the same auditable calendar path.
    original_month_values = [row["values"][0] for row in rows]
    decoded_months, rollover_evidence = decode_legacy_month_sequence(
        rows, locked
    )
    for index, (row, decoded_month) in enumerate(zip(rows, decoded_months)):
        original = original_month_values[index]
        explicit = parse_integer(original, 1, 12)
        row["values"][0] = str(decoded_month)
        if explicit is not None and explicit != decoded_month:
            record(
                row,
                "月",
                original,
                str(decoded_month),
                (
                    "全站月份序列联合解码；"
                    f"本行日期回卷证据={rollover_evidence[index]:.2f}"
                ),
                "global-calendar-dynamic-programming",
            )
    inherited_months: list[int | None] = list(decoded_months)

    # --- Leading day: a station has no previous row from which an invalid
    # first day can be inherited.  Recover only when the next two rows provide
    # a strict daily sequence and all three records share the same clock time.
    # This makes the inference bidirectional while avoiding a blind ``next-1``
    # rule on tables that contain several observations per day.
    if len(rows) >= 3 and not locked(rows[0], "日"):
        leading_day = parse_integer(rows[0]["values"][1], 1, 31)
        next_day = parse_integer(rows[1]["values"][1], 1, 31)
        following_day = parse_integer(rows[2]["values"][1], 1, 31)
        leading_time = parse_legacy_clock(rows[0]["values"][2])
        next_time = parse_legacy_clock(rows[1]["values"][2])
        following_time = parse_legacy_clock(rows[2]["values"][2])
        same_month = bool(
            inherited_months[0] is not None
            and inherited_months[0] == inherited_months[1]
            and inherited_months[1] == inherited_months[2]
        )
        repeated_daily_clock = bool(
            leading_time is not None
            and next_time is not None
            and following_time is not None
            and leading_time[1] == next_time[1] == following_time[1]
        )
        if (
            leading_day is None
            and next_day is not None
            and following_day == next_day + 1
            and next_day > 1
            and same_month
            and repeated_daily_clock
        ):
            recovered_day = next_day - 1
            old_value = rows[0]["values"][1]
            rows[0]["values"][1] = str(recovered_day)
            record(
                rows[0],
                "日",
                old_value,
                str(recovered_day),
                (
                    "站点首行无前文；后两行日期连续且"
                    f"三行时分相同={leading_time[0]}"
                ),
                "calendar-leading-edge-backcast",
            )

    # Multi-observation days need a different leading-edge proof.  When the
    # first explicit future day is preceded by a clock rollover (for example
    # 08:00, 14:00, 20:00, then day 21 at 08:00), all leading rows belong to
    # day 20.  Candidate voting is used here because the selected clock OCR can
    # itself be wrong while an alternate split candidate preserves the cycle.
    if rows and parse_integer(rows[0]["values"][1], 1, 31) is None:
        first_future_index: int | None = None
        first_future_day: int | None = None
        for index in range(1, min(len(rows), 11)):
            candidate_day = parse_integer(rows[index]["values"][1], 1, 31)
            if candidate_day is not None:
                first_future_index = index
                first_future_day = candidate_day
                break
        if (
            first_future_index is not None
            and first_future_day is not None
            and first_future_day > 1
            and not locked(rows[0], "日")
            and parse_integer(original_month_values[0], 1, 12) is not None
            and all(
                inherited_months[index] == inherited_months[0]
                for index in range(first_future_index + 1)
            )
        ):
            leading_clock_minutes: list[int] = []
            for index in range(first_future_index + 1):
                options = legacy_context_options(
                    rows[index],
                    "时分",
                    parse_legacy_clock,
                    rows[index]["values"][2],
                    2,
                )
                if options:
                    minutes, _ = max(
                        options.items(), key=lambda item: item[1][0]
                    )
                    leading_clock_minutes.append(int(minutes))
            clock_rollover = any(
                following < previous
                for previous, following in zip(
                    leading_clock_minutes[:-1], leading_clock_minutes[1:]
                )
            )
            if clock_rollover:
                recovered_day = first_future_day - 1
                old_value = rows[0]["values"][1]
                rows[0]["values"][1] = str(recovered_day)
                record(
                    rows[0],
                    "日",
                    old_value,
                    str(recovered_day),
                    (
                        "站点首段含同日多时次；"
                        f"首个合法未来日期={first_future_day}，"
                        "之前存在时分回卷"
                    ),
                    "calendar-leading-clock-rollover-backcast",
                )

    # --- Day markers: legal for the inherited month and nondecreasing there.
    month_groups: dict[int, list[int]] = {}
    for index, row in enumerate(rows):
        month = inherited_months[index]
        if month is None:
            continue
        if parse_integer(row["values"][1], 1, 31) is not None:
            month_groups.setdefault(month, []).append(index)
    for month, marker_indices in month_groups.items():
        maximum_day = legacy_days_in_month(month)
        for marker_position, index in enumerate(marker_indices):
            row = rows[index]
            if locked(row, "日"):
                continue
            current = parse_integer(row["values"][1], 1, 31)
            if current is None:
                continue
            options = legacy_context_options(
                row,
                "日",
                lambda value, limit=maximum_day: parse_integer(
                    value, 1, limit
                ),
                row["values"][1],
                1,
            )
            previous = None
            following = None
            if marker_position > 0:
                previous = parse_integer(
                    rows[marker_indices[marker_position - 1]]["values"][1],
                    1,
                    maximum_day,
                )
            if marker_position + 1 < len(marker_indices):
                following = parse_integer(
                    rows[marker_indices[marker_position + 1]]["values"][1],
                    1,
                    maximum_day,
                )
            viable = {
                value: detail
                for value, detail in options.items()
                if (previous is None or value >= previous)
                and (following is None or value <= following)
            }
            current_valid = (
                1 <= current <= maximum_day
                and (previous is None or current >= previous)
                and (following is None or current <= following)
            )
            replacement: int | None = None
            source = ""
            if not current_valid and viable:
                replacement, detail = max(
                    viable.items(), key=lambda item: item[1][0]
                )
                source = detail[2]
            elif (
                not current_valid
                and previous is not None
                and following == previous + 2
                and current not in {previous, following}
            ):
                replacement = previous + 1
                source = "calendar-sandwich-inference"
            if replacement is not None and replacement != current:
                old_value = row["values"][1]
                row["values"][1] = str(replacement)
                record(
                    row,
                    "日",
                    old_value,
                    str(replacement),
                    (
                        f"月份={month}，前一日期={previous}，"
                        f"后一日期={following}"
                    ),
                    source,
                )

    # --- Month/day geometry: establish the printed marker coverage before any
    # clock-order reasoning.  Known spans are locked in day_emission(); time can
    # select among OCR clock candidates or become unresolved, but cannot move a
    # row into the next printed day group merely to repair a rollback.
    apply_legacy_date_marker_coverage(rows)

    # --- Day + time: decode the complete timestamp as one ordered sequence.
    # The old staged rule accepted ``current >= previous`` within one inherited
    # day.  It therefore turned ambiguous rows into repeated stamps such as six
    # consecutive ``23 08:00`` records.  Here equality is illegal: a clock
    # rollback selects the next date when supported by the row candidates, and
    # a clock with no chronology-compatible candidate is deliberately blank.
    decoded_day_times = decode_legacy_joint_day_time_sequence(rows, locked)
    for row, (decoded_day, decoded_time, source) in zip(
        rows, decoded_day_times
    ):
        old_day = str(row["values"][1])
        new_day = str(decoded_day)
        row["values"][1] = new_day
        if row.get("_date_span_day") is None:
            row["_date_span_resolved_day"] = int(decoded_day)
            row["_date_day_value_source"] = (
                "joint-context-for-unknown-geometry-span"
            )
        record(
            row,
            "日",
            old_day,
            new_day,
            "月、日、时分联合解码；完整时间戳必须严格递增",
            source,
        )

        old_time = str(row["values"][2])
        if (
            not decoded_time
            and row.get("_date_day_span_id") is not None
            and not locked(row, "时分")
        ):
            observed_options = legacy_context_options(
                row,
                "时分",
                parse_legacy_clock,
                old_time,
                2,
            )
            if observed_options:
                _, best_detail = max(
                    observed_options.items(),
                    key=lambda item: float(item[1][0]),
                )
                decoded_time = str(best_detail[1])
                row["_date_time_conflict"] = {
                    "date_span_day": int(decoded_day),
                    "preserved_time": decoded_time,
                    "candidate_source": str(best_detail[2]),
                    "candidate_score": round(float(best_detail[0]), 8),
                    "reason": (
                        "日期由月日标记覆盖区间确定；所有时分候选均与"
                        "区间内严格递增顺序冲突，保留最高置信度图像候选"
                        "并交由人工复核"
                    ),
                }
                source = "date-marker-span-clock-conflict"
        row["values"][2] = decoded_time
        if decoded_time:
            if row.get("_date_time_conflict"):
                reason = str(row["_date_time_conflict"]["reason"])
            else:
                reason = "月、日、时分联合解码；禁止同一站点出现重复或倒退时刻"
        else:
            reason = (
                "所有OCR时分候选均与严格递增时间轴冲突；"
                "留空并交由校正表人工确认"
            )
        record(
            row,
            "时分",
            old_time,
            decoded_time,
            reason,
            source,
        )

    correction_counts: dict[str, int] = {}
    for row in rows:
        for correction in row["_context_corrections"]:
            field = str(correction["field"])
            correction_counts[field] = correction_counts.get(field, 0) + 1
    if correction_counts:
        details = "，".join(
            f"{field}{count}处"
            for field, count in correction_counts.items()
        )
        log(f"[3/4] 月日时分上下文约束已修正：{details}")
    return rows


def split_legacy_rows_by_station(
    pdf_path: Path,
    pages: list[
        tuple[
            list[dict[str, Any]],
            dict[str, Any],
            list[dict[str, Any]],
        ]
    ],
) -> OrderedDict[str, dict[str, Any]]:
    """Assign every reconstructed row and training sample to a station.

    Station markers are compared with row-center y coordinates independently
    for every panel.  The last marker of one page is carried into a continuation
    page that does not repeat its title.
    """
    groups: OrderedDict[str, dict[str, Any]] = OrderedDict()
    current_station: str | None = None

    for page_rows, metadata, page_samples in pages:
        markers = sorted(
            metadata.get("station_markers", []),
            key=lambda item: int(item["y"]),
        )
        starting_station = current_station
        row_station_by_key: dict[tuple[int, int, int], str] = {}
        page_station_counts: OrderedDict[str, int] = OrderedDict()

        for row in page_rows:
            station = starting_station
            for marker in markers:
                if int(marker["y"]) < int(row["row_center"]):
                    station = str(marker["station"])
                else:
                    break
            if station is None:
                station = pdf_path.stem
                log(
                    f"[站点] 第{metadata['page']}页首段没有位于数据行之前的站名，"
                    f"暂用 {station}"
                )

            group = groups.setdefault(
                station,
                {
                    "raw_rows": [],
                    "pages": OrderedDict(),
                    "training_samples": [],
                    "canonical_indices": set(),
                },
            )
            group["raw_rows"].append(row)
            row_key = (
                int(row["page"]),
                int(row["panel"]),
                int(row["row"]),
            )
            row_station_by_key[row_key] = station
            page_station_counts[station] = page_station_counts.get(station, 0) + 1

        canonical_indices = {
            int(index) for index in metadata["schema"]["canonical_indices"]
        }
        for station, station_row_count in page_station_counts.items():
            group = groups[station]
            page_metadata = dict(metadata)
            page_metadata["assigned_station"] = station
            page_metadata["station_row_count"] = station_row_count
            group["pages"].setdefault(int(metadata["page"]), page_metadata)
            group["canonical_indices"].update(canonical_indices)

        for sample in page_samples:
            sample_key = (
                int(sample["page"]),
                int(sample["panel"]),
                int(sample["row"]),
            )
            station = row_station_by_key.get(sample_key)
            if station is None:
                continue
            sample["station"] = station
            groups[station]["training_samples"].append(sample)

        if markers:
            current_station = str(markers[-1]["station"])
        elif page_station_counts:
            current_station = next(reversed(page_station_counts))

    return groups


def normalize_legacy_rows(raw_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    current_month: int | None = None
    current_day: int | None = None
    current_water_integer: int | None = None
    normalized: list[dict[str, Any]] = []

    resolved_rows = resolve_legacy_calendar_time_context(raw_rows)

    # A time-only recovered row can contain perfectly usable integer/fraction
    # water sub-images while lacking an observed decimal-anchor flag.  At a
    # station boundary there is no previous integer prefix to inherit, so use
    # the preserved split candidate only when a following explicit level
    # confirms the same integer part.
    if resolved_rows:
        first_row = resolved_rows[0]
        manual_fields = set(first_row.get("_manual_corrected_fields", []))
        selected_water = str(first_row["values"][3] or "")
        selected_full = re.fullmatch(r"(\d+)\.(\d{1,2})", selected_water)
        split_text = str(first_row.get("_water_split_candidate") or "")
        split_full = re.fullmatch(r"(\d+)\.(\d{2})", split_text)
        split_score = float(first_row.get("_water_split_score") or 0.0)
        following_integers: list[int] = []
        for following_row in resolved_rows[1:9]:
            following_text = str(following_row["values"][3] or "")
            following_full = re.fullmatch(
                r"(\d+)\.(\d{1,2})", following_text
            )
            if following_full:
                following_integers.append(int(following_full.group(1)))
                if len(following_integers) >= 2:
                    break
        confirmed_integer = (
            following_integers[0]
            if following_integers
            and (
                len(following_integers) == 1
                or following_integers[0] == following_integers[1]
            )
            else None
        )
        recovered_water = ""
        integer_source = ""
        if split_full is not None and confirmed_integer is not None:
            split_integer = int(split_full.group(1))
            split_fraction = split_full.group(2)
            if split_integer == confirmed_integer:
                recovered_water = split_text
                integer_source = "拆分整数段"
            elif (
                len(following_integers) >= 2
                and following_integers[0] == following_integers[1]
                and re.fullmatch(r"\d{1,2}", selected_water)
                and int(selected_water) == int(split_fraction)
            ):
                # The narrow integer crop can still confuse a handwritten 5
                # with 3.  Two following complete levels supply stronger
                # station-local evidence for the omitted integer prefix,
                # while the first-row fraction remains image-derived.
                recovered_water = (
                    f"{confirmed_integer}.{int(split_fraction):02d}"
                )
                integer_source = "后两个完整水位的一致整数段"
        if (
            "水位" not in manual_fields
            and selected_full is None
            and split_full is not None
            and split_score >= 0.55
            and recovered_water
        ):
            first_row["values"][3] = recovered_water
            first_row.setdefault("_context_corrections", []).append(
                {
                    "field": "水位",
                    "from": selected_water,
                    "to": recovered_water,
                    "reason": (
                        "站点首行无可继承整数前缀；"
                        f"整数依据={integer_source}={confirmed_integer}"
                    ),
                    "source": "leading-water-split-forward-confirmation",
                }
            )
            log(
                "[3/4] 站点首行水位已由拆分子图恢复："
                f"{selected_water or '<空>'} -> {recovered_water}"
            )
    for row_number, raw in enumerate(resolved_rows, start=1):
        values = raw["values"]
        explicit_month = parse_integer(values[0], 1, 12)
        candidate_day = parse_integer(values[1], 1, 31)
        month_changed = False
        if explicit_month is not None:
            if current_month is None or explicit_month == current_month:
                current_month = explicit_month
            elif explicit_month > current_month or (
                current_month == 12 and explicit_month == 1
            ):
                # A genuine new month is accompanied by a day rollback.  A
                # scan speck in the sparse month column can otherwise turn a
                # mid-month row such as 7/9 into a false 8/9 and contaminate
                # every following inherited month value.
                plausible_month_boundary = bool(
                    current_day is None
                    or (
                        candidate_day is not None
                        and candidate_day < current_day
                    )
                    or (
                        candidate_day is None
                        and current_day is not None
                        and current_day >= 28
                    )
                )
                if plausible_month_boundary:
                    current_month = explicit_month
                    month_changed = True
        if current_month is None:
            raise RuntimeError(
                f"合并后的第{row_number}行没有可继承的月份，请检查分格预览。"
            )

        explicit_day = parse_integer(
            values[1], 1, legacy_days_in_month(current_month)
        )

        if explicit_day is not None:
            if current_day is None or month_changed or explicit_day >= current_day:
                current_day = explicit_day
        if current_day is None:
            raise RuntimeError(
                f"合并后的第{row_number}行没有可继承的日期，请检查分格预览。"
            )

        water_level, current_water_integer = expand_water_level(
            values[3], current_water_integer
        )
        normalized.append(
            {
                "月": current_month,
                "日": current_day,
                "时分": values[2],
                "水位(m)": water_level,
                "流量(m³/s)": parse_decimal(values[4], 2),
                "含沙量(kg/m³)": parse_decimal(values[5], 3),
                "水面比降(万分率)": parse_decimal(values[6], 3),
                "平均置信度": round(float(raw["average_confidence"]), 8),
                "_page": raw["page"],
                "_panel": raw["panel"],
                "_row": raw["row"],
                "_row_center": raw["row_center"],
                "_context_corrections": raw.get(
                    "_context_corrections", []
                ),
                "_date_month_span_id": raw.get("_date_month_span_id"),
                "_date_month_span_start_row": raw.get(
                    "_date_month_span_start_row"
                ),
                "_date_month_span_end_row": raw.get(
                    "_date_month_span_end_row"
                ),
                "_date_month_span_source": raw.get(
                    "_date_month_span_source"
                ),
                "_date_day_span_id": raw.get("_date_day_span_id"),
                "_date_day_span_start_row": raw.get(
                    "_date_day_span_start_row"
                ),
                "_date_day_span_end_row": raw.get(
                    "_date_day_span_end_row"
                ),
                "_date_day_span_source": raw.get(
                    "_date_day_span_source"
                ),
                "_date_span_resolved_day": raw.get(
                    "_date_span_resolved_day"
                ),
                "_date_day_value_source": raw.get(
                    "_date_day_value_source"
                ),
                "_date_time_conflict": raw.get("_date_time_conflict"),
            }
        )
    return normalized


def correction_label_text(value: Any, field: str) -> str:
    """Normalize one editable label cell without losing numeric zeroes."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value).strip()
    if field in ("分", "水位小数") and text.isdigit() and len(text) == 1:
        text = text.zfill(2)
    return text


def sync_context_choices_to_training_samples(
    samples: list[dict[str, Any]],
    normalized_rows: list[dict[str, Any]],
) -> int:
    """Expose contextual OCR choices in the review workbook for confirmation."""
    rows_by_key = {
        (int(row["_page"]), int(row["_panel"]), int(row["_row"])): row
        for row in normalized_rows
    }
    updated = 0
    for sample in samples:
        key = (
            int(sample["page"]),
            int(sample["panel"]),
            int(sample["row"]),
        )
        row = rows_by_key.get(key)
        if row is None:
            continue
        row_corrections = list(row.get("_context_corrections", []))
        correction_fields = {
            str(item["field"]) for item in row_corrections
        }
        field = str(sample["column"])
        date_time_conflict = bool(
            field in {"时", "分"} and row.get("_date_time_conflict")
        )
        label = ""
        if field == "月" and "月" in correction_fields:
            label = str(row["月"])
        elif field == "日" and "日" in correction_fields:
            label = str(row["日"])
        elif field in {"时", "分"} and "时分" in correction_fields:
            parsed_time = parse_legacy_clock(str(row["时分"]))
            if parsed_time is not None:
                hour, minute = parsed_time[0].split(":")
                label = str(int(hour)) if field == "时" else minute
        elif date_time_conflict:
            parsed_time = parse_legacy_clock(str(row["时分"]))
            if parsed_time is not None:
                hour, minute = parsed_time[0].split(":")
                label = str(int(hour)) if field == "时" else minute
        unresolved_time = bool(
            field in {"时", "分"}
            and "时分" in correction_fields
            and parse_legacy_clock(str(row["时分"])) is None
        )
        if not unresolved_time and not date_time_conflict and (
            not label or sample.get("correct_label") == label
        ):
            continue
        if label:
            sample["correct_label"] = label
        if sample.get("include") == "是":
            sample["include"] = "待确认"
        note = str(sample.get("note") or "").strip()
        applicable = [
            item
            for item in row_corrections
            if str(item.get("field"))
            == ("时分" if field in {"时", "分"} else field)
        ]
        audit_detail = "；".join(
            f"{item.get('reason', '')}，来源={item.get('source', '')}"
            for item in applicable
        )
        if date_time_conflict:
            conflict_detail = str(
                row.get("_date_time_conflict", {}).get("reason") or ""
            )
            audit_detail = "；".join(
                value for value in (audit_detail, conflict_detail) if value
            )
        context_note = (
            (
                "严格时间轴未找到可用时分，最终输出留空"
                if unresolved_time
                else "最终输出经月日时分上下文约束改选"
            )
            + (f"（{audit_detail}）" if audit_detail else "")
            + "；"
            "该标签请人工确认后再纳入训练"
        )
        sample["note"] = f"{note} {context_note}".strip()
        updated += 1
    return updated


def apply_training_label_corrections(
    raw_rows: list[dict[str, Any]],
    corrected_labels_path: Path,
) -> int:
    """Apply accepted labels from a review workbook to reconstructed rows."""
    workbook = load_workbook(corrected_labels_path, data_only=True, read_only=True)
    if "标签校正" not in workbook.sheetnames:
        raise RuntimeError(
            f"校正表缺少“标签校正”工作表：{corrected_labels_path}"
        )
    sheet = workbook["标签校正"]
    headers = {
        str(cell.value).strip(): index
        for index, cell in enumerate(sheet[1])
        if cell.value is not None
    }
    required = {"页", "表区", "行序号", "字段", "正确标签", "纳入训练"}
    missing = sorted(required - set(headers))
    if missing:
        raise RuntimeError(
            "校正表缺少必要列：" + "、".join(missing)
        )

    accepted: dict[tuple[int, int, int], dict[str, str]] = {}
    for cells in sheet.iter_rows(min_row=2, values_only=True):
        include = str(cells[headers["纳入训练"]] or "").strip()
        if include != "是":
            continue
        field = str(cells[headers["字段"]] or "").strip()
        label = correction_label_text(cells[headers["正确标签"]], field)
        if not field or not label:
            continue
        try:
            key = (
                int(cells[headers["页"]]),
                int(cells[headers["表区"]]),
                int(cells[headers["行序号"]]),
            )
        except (TypeError, ValueError) as exc:
            raise RuntimeError("校正表包含无效的页/表区/行序号。") from exc
        accepted.setdefault(key, {})[field] = label
    workbook.close()

    rows_by_key = {
        (int(row["page"]), int(row["panel"]), int(row["row"])): row
        for row in raw_rows
    }
    direct_columns = {
        "月": 0,
        "日": 1,
        "流量(m³/s)": 4,
        "含沙量(kg/m³)": 5,
        "水面比降(万分率)": 6,
    }
    applied = 0
    for key, fields in accepted.items():
        raw = rows_by_key.get(key)
        if raw is None:
            continue
        values = raw["values"]
        manual_fields = set(raw.get("_manual_corrected_fields", []))
        for field, column_index in direct_columns.items():
            if field in fields:
                manual_fields.add(field)
                if values[column_index] != fields[field]:
                    values[column_index] = fields[field]
                    applied += 1

        hour = fields.get("时")
        minute = fields.get("分")
        corrected_time = ""
        if hour and hour.isdigit() and 0 <= int(hour) <= 24:
            minute_value = minute if minute and minute.isdigit() else "00"
            if (
                0 <= int(minute_value) <= 59
                and (int(hour) < 24 or int(minute_value) == 0)
            ):
                corrected_time = f"{int(hour):02d}:{int(minute_value):02d}"
        if corrected_time and values[2] != corrected_time:
            values[2] = corrected_time
            applied += 1
        if corrected_time:
            manual_fields.add("时分")

        water_integer = fields.get("水位整数")
        water_fraction = fields.get("水位小数")
        corrected_water = ""
        if water_fraction and water_fraction.isdigit():
            if water_integer and water_integer.isdigit():
                corrected_water = f"{water_integer}.{water_fraction}"
            else:
                corrected_water = water_fraction
        if corrected_water and values[3] != corrected_water:
            values[3] = corrected_water
            applied += 1
        if corrected_water:
            manual_fields.add("水位")
        raw["_manual_corrected_fields"] = sorted(manual_fields)
    return applied


def merge_review_state_into_training_samples(
    samples: list[dict[str, Any]],
    corrected_labels_path: Path,
) -> int:
    """Carry reviewed labels/statuses forward into a newly exported workbook."""
    workbook = load_workbook(corrected_labels_path, data_only=True, read_only=True)
    if "标签校正" not in workbook.sheetnames:
        workbook.close()
        return 0
    sheet = workbook["标签校正"]
    headers = {
        str(cell.value).strip(): index
        for index, cell in enumerate(sheet[1])
        if cell.value is not None
    }
    required = {"样本ID", "字段", "正确标签", "纳入训练"}
    if not required.issubset(headers):
        workbook.close()
        return 0
    reviewed: dict[str, tuple[str, str]] = {}
    for cells in sheet.iter_rows(min_row=2, values_only=True):
        sample_id = str(cells[headers["样本ID"]] or "").strip()
        field = str(cells[headers["字段"]] or "").strip()
        if not sample_id:
            continue
        label = correction_label_text(cells[headers["正确标签"]], field)
        include = str(cells[headers["纳入训练"]] or "待确认").strip()
        reviewed[sample_id] = (label, include)
    workbook.close()

    merged = 0
    for sample in samples:
        state = reviewed.get(sample["sample_id"])
        if state is None:
            continue
        label, include = state
        sample["correct_label"] = label
        sample["include"] = include
        merged += 1
    return merged


def summarize_legacy_date_spans(
    rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Return auditable month/day coverage intervals for JSON and Excel."""
    summaries: list[dict[str, Any]] = []
    for kind, value_field, id_key, start_key, end_key, source_key in (
        (
            "月",
            "月",
            "_date_month_span_id",
            "_date_month_span_start_row",
            "_date_month_span_end_row",
            "_date_month_span_source",
        ),
        (
            "日",
            "日",
            "_date_day_span_id",
            "_date_day_span_start_row",
            "_date_day_span_end_row",
            "_date_day_span_source",
        ),
    ):
        seen: set[int] = set()
        for row in rows:
            span_id = row.get(id_key)
            if span_id is None or int(span_id) in seen:
                continue
            seen.add(int(span_id))
            start_row = int(row.get(start_key) or 0)
            end_row = int(row.get(end_key) or start_row)
            start = rows[start_row - 1] if start_row else row
            end = rows[end_row - 1] if end_row else row
            summaries.append(
                {
                    "类型": kind,
                    "值": row.get(value_field),
                    "覆盖起始行": start_row,
                    "覆盖结束行": end_row,
                    "覆盖行数": max(0, end_row - start_row + 1),
                    "起点": (
                        f"p{int(start['_page']):03d}/b{int(start['_panel'])}/"
                        f"r{int(start['_row']):03d}"
                    ),
                    "终点": (
                        f"p{int(end['_page']):03d}/b{int(end['_panel'])}/"
                        f"r{int(end['_row']):03d}"
                    ),
                    "来源": row.get(source_key) or "unresolved",
                    "主表状态": (
                        "留空待确认"
                        if kind == "日"
                        and (
                            str(row.get(source_key) or "").endswith(
                                "unknown-marker"
                            )
                            or row.get("_date_day_value_source")
                            == "joint-context-for-unknown-geometry-span"
                        )
                        else "已输出"
                    ),
                    "建议值": row.get(value_field),
                    "时分冲突行数": sum(
                        1
                        for candidate in rows[start_row - 1 : end_row]
                        if candidate.get("_date_time_conflict")
                    ),
                }
            )
    return summaries


def legacy_pending_output_fields(row: dict[str, Any]) -> dict[str, dict[str, Any]]:
    """Return fields whose machine candidate must not enter the main sheet."""
    pending: dict[str, dict[str, Any]] = {}
    day_source = str(row.get("_date_day_span_source") or "")
    if (
        day_source in {"unknown-marker", "covered-by-unknown-marker"}
        or row.get("_date_day_value_source")
        == "joint-context-for-unknown-geometry-span"
    ):
        pending["日"] = {
            "suggested_value": row.get("日"),
            "reason": (
                "日期覆盖边界已确认，但日期数字没有可靠图像候选；"
                "建议值仅供人工复核"
            ),
        }
    month_source = str(row.get("_date_month_span_source") or "")
    if month_source == "unknown-marker":
        pending["月"] = {
            "suggested_value": row.get("月"),
            "reason": "月份标记存在，但月份数字没有可靠图像候选",
        }
    conflict = row.get("_date_time_conflict")
    if conflict:
        pending["时分"] = {
            "suggested_value": row.get("时分"),
            "reason": str(
                conflict.get("reason")
                or "时分候选与日期覆盖段内顺序冲突"
            ),
        }
    return pending


def write_legacy_outputs(
    pdf_path: Path,
    output_dir: Path,
    station: str,
    rows: list[dict[str, Any]],
    pages: list[dict[str, Any]],
    output_columns: list[str],
    low_confidence_threshold: float,
) -> tuple[Path, Path]:
    basename = safe_filename(station)
    date_spans = summarize_legacy_date_spans(rows)
    pending_output_items: list[dict[str, Any]] = []
    pending_fields_by_row: list[dict[str, dict[str, Any]]] = []
    for output_row, row in enumerate(rows, start=1):
        pending_fields = legacy_pending_output_fields(row)
        pending_fields_by_row.append(pending_fields)
        for field, detail in pending_fields.items():
            pending_output_items.append(
                {
                    "输出行": output_row,
                    "页": row.get("_page"),
                    "表区": row.get("_panel"),
                    "原行": row.get("_row"),
                    "字段": field,
                    "建议值": detail.get("suggested_value"),
                    "主表值": None,
                    "原因": detail.get("reason"),
                    "覆盖起始行": row.get("_date_day_span_start_row"),
                    "覆盖结束行": row.get("_date_day_span_end_row"),
                }
            )
    json_path = output_dir / f"{basename}.json"
    json_path.write_text(
        json.dumps(
            {
                "input_pdf": str(pdf_path),
                "station": station,
                "row_count": len(rows),
                "output_columns": output_columns,
                "date_spans": date_spans,
                "pending_output_items": pending_output_items,
                "pages": pages,
                "rows": rows,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "合并表"
    headers = output_columns + ["平均置信度"]
    sheet.append(headers)
    for row, pending_fields in zip(rows, pending_fields_by_row):
        output_values: list[Any] = []
        for field in output_columns:
            value: Any = None if field in pending_fields else row[field]
            if (
                field == "时分"
                and isinstance(value, str)
                and re.fullmatch(r"\d{2}:\d{2}", value)
            ):
                hour = int(value[:2])
                minute = int(value[3:])
                if 0 <= hour <= 23 and 0 <= minute <= 59:
                    value = datetime.strptime(value, "%H:%M").time()
            output_values.append(value)
        output_values.append(row["平均置信度"])
        sheet.append(output_values)

    header_fill = PatternFill("solid", fgColor="EAF2F8")
    low_fill = PatternFill("solid", fgColor="FFF2CC")
    pending_fill = PatternFill("solid", fgColor="FCE8E6")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    number_formats = {
        "月": "0",
        "日": "0",
        "时分": "hh:mm",
        "水位(m)": "0.00",
        "流量(m³/s)": "0.00",
        "含沙量(kg/m³)": "0.000",
        "水面比降(万分率)": "0.000",
        "平均置信度": "0.00000000",
    }
    confidence_column = len(headers)
    for excel_row in range(2, sheet.max_row + 1):
        for column, field in enumerate(headers, start=1):
            sheet.cell(excel_row, column).number_format = number_formats[field]
            if field in {"月", "日", "时分"}:
                sheet.cell(excel_row, column).alignment = Alignment(
                    horizontal="center"
                )
        if (
            sheet.cell(excel_row, confidence_column).value or 0.0
        ) < low_confidence_threshold:
            for cell in sheet[excel_row]:
                cell.fill = low_fill
        pending_fields = pending_fields_by_row[excel_row - 2]
        for field in pending_fields:
            if field not in output_columns:
                continue
            pending_column = output_columns.index(field) + 1
            sheet.cell(excel_row, pending_column).fill = pending_fill

    column_widths = {
        "月": 8,
        "日": 8,
        "时分": 12,
        "水位(m)": 14,
        "流量(m³/s)": 16,
        "含沙量(kg/m³)": 18,
        "水面比降(万分率)": 22,
        "平均置信度": 16,
    }
    for index, field in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = (
            column_widths[field]
        )
    sheet.freeze_panes = "A2"
    last_column = get_column_letter(len(headers))
    sheet.auto_filter.ref = f"A1:{last_column}{sheet.max_row}"
    sheet.row_dimensions[1].height = 24

    audit = workbook.create_sheet("日期覆盖校核")
    audit_headers = [
        "类型",
        "值",
        "覆盖起始行",
        "覆盖结束行",
        "覆盖行数",
        "起点",
        "终点",
        "来源",
        "主表状态",
        "建议值",
        "时分冲突行数",
    ]
    audit.append(audit_headers)
    for span in date_spans:
        audit.append([span.get(field) for field in audit_headers])
    for cell in audit[1]:
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    audit.freeze_panes = "A2"
    audit.auto_filter.ref = f"A1:K{max(1, audit.max_row)}"
    for column, width in enumerate(
        (8, 8, 12, 12, 10, 18, 18, 28, 14, 10, 14), start=1
    ):
        audit.column_dimensions[get_column_letter(column)].width = width
    audit.row_dimensions[1].height = 24

    pending_sheet = workbook.create_sheet("待确认项")
    pending_headers = [
        "输出行",
        "页",
        "表区",
        "原行",
        "字段",
        "建议值",
        "主表值",
        "原因",
        "覆盖起始行",
        "覆盖结束行",
    ]
    pending_sheet.append(pending_headers)
    for item in pending_output_items:
        pending_sheet.append([item.get(field) for field in pending_headers])
    for cell in pending_sheet[1]:
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    pending_sheet.freeze_panes = "A2"
    pending_sheet.auto_filter.ref = (
        f"A1:J{max(1, pending_sheet.max_row)}"
    )
    for column, width in enumerate(
        (10, 8, 8, 8, 10, 14, 12, 72, 12, 12), start=1
    ):
        pending_sheet.column_dimensions[get_column_letter(column)].width = width
    for excel_row in range(2, pending_sheet.max_row + 1):
        pending_sheet.cell(excel_row, 8).alignment = Alignment(
            vertical="top", wrap_text=True
        )
        pending_sheet.row_dimensions[excel_row].height = 34
    pending_sheet.row_dimensions[1].height = 24

    workbook_path = output_dir / f"{basename}.xlsx"
    workbook.save(workbook_path)
    return workbook_path, json_path


def write_training_review_workbook(
    samples: list[dict[str, Any]],
    training_dir: Path,
) -> Path:
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "使用说明"
    instructions.append(["水文OCR训练样本校正说明"])
    instructions.append(["1", "打开“标签校正”工作表。"])
    instructions.append(["2", "点击图片文件名查看模型实际识别的单元格图。"])
    instructions.append(["3", "只修改“正确标签”和“纳入训练”两列。"])
    instructions.append(
        ["4", "正确标签必须与图片完全一致，例如 61、15、5、00、161。"]
    )
    instructions.append(["5", "空白格、表格线或无法判断的噪声请选择“否”。"])
    instructions.append(["6", "确认无误的样本请选择“是”；暂时无法判断可保留“待确认”。"])
    instructions.append(["7", "不要修改样本ID、图片路径、页码、表区和字段。"])
    instructions.append(
        ["8", "带冒号表按实际冒号左右边界拆成“时”和“分”；无冒号整点表把完整数字字符组作为“时”，例如2表示02:00，分钟00由版式隐含，不导出没有对应墨迹的分钟训练图。"]
    )
    instructions.append(
        ["9", "水位整格已按小数点拆成“水位整数”和“水位小数”：两个标签均不填写小数点；省略整数的行只校正实际导出的小数子样本。"]
    )
    instructions.append(
        ["10", "水位小数点与时分辅助锚点分别建立实测行序：带冒号表使用冒号中心，无冒号表使用小时数字字符组中心，再与时分、水位、流量多列墨迹联合核对。任何单一锚点都不会独自按固定行距补行，空白格不拟合坐标。"]
    )
    instructions.append(
        ["11", "分钟图不再给连写的00插入白线，也不再强制拆成十位/个位，而是整块送入中文手写模型与英文数字倾向模型；c/C/o/O/Q/)/a等零形混淆会统一转为0，例如c0、co、0)均归一为00；整对连写00若被识别为单个w/W，也按00处理。候选仍按字段合法性、数字纯度、墨迹段数、输出长度和置信度综合选择。"]
    )
    instructions.append(
        ["12", "校正完成后，可使用 --corrected-labels 指向本工作簿重新运行；程序只采用“纳入训练=是”且“正确标签”非空的记录，并将时/分、水位整数/小数合并回最终结果。"]
    )
    instructions.column_dimensions["A"].width = 18
    instructions.column_dimensions["B"].width = 100
    instructions["A1"].font = Font(bold=True, size=15, color="FFFFFF")
    instructions["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    instructions.merge_cells("A1:B1")
    instructions.row_dimensions[1].height = 28
    for row in range(2, instructions.max_row + 1):
        instructions.cell(row, 1).font = Font(bold=True, color="1F4E78")
        instructions.cell(row, 2).alignment = Alignment(wrap_text=True)
        instructions.row_dimensions[row].height = 26
    instructions.row_dimensions[9].height = 42
    instructions.row_dimensions[11].height = 52
    instructions.row_dimensions[12].height = 56
    instructions.row_dimensions[13].height = 42
    instructions.sheet_view.showGridLines = False

    sheet = workbook.create_sheet("标签校正")
    headers = [
        "样本ID",
        "页",
        "表区",
        "行序号",
        "字段",
        "OCR原文",
        "规范化标签",
        "置信度",
        "正确标签",
        "纳入训练",
        "图片文件",
        "备注",
    ]
    sheet.append(headers)
    for sample in samples:
        sheet.append(
            [
                sample["sample_id"],
                sample["page"],
                sample["panel"],
                sample["row"],
                sample["column"],
                sample["ocr_text"],
                sample["normalized_label"],
                sample["confidence"],
                sample["correct_label"],
                sample["include"],
                sample["image"],
                sample.get("note", ""),
            ]
        )
        image_cell = sheet.cell(sheet.max_row, 11)
        image_cell.hyperlink = sample["image"]
        image_cell.style = "Hyperlink"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    low_fill = PatternFill("solid", fgColor="FFF2CC")
    review_fill = PatternFill("solid", fgColor="FCE4D6")
    editable_fill = PatternFill("solid", fgColor="E2F0D9")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 8).number_format = "0.00000000"
        sheet.cell(row, 9).number_format = "@"
        sheet.cell(row, 9).fill = editable_fill
        sheet.cell(row, 10).fill = editable_fill
        if (sheet.cell(row, 8).value or 0.0) < 0.85:
            for column in range(1, 9):
                sheet.cell(row, column).fill = low_fill
        if sheet.cell(row, 10).value == "待确认":
            sheet.cell(row, 10).fill = review_fill

    validation = DataValidation(
        type="list",
        formula1='"是,否,待确认"',
        allow_blank=False,
    )
    sheet.add_data_validation(validation)
    validation.add(f"J2:J{sheet.max_row}")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:L{sheet.max_row}"
    sheet.row_dimensions[1].height = 26
    widths = [28, 8, 8, 10, 22, 18, 18, 14, 18, 14, 48, 52]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.sheet_view.showGridLines = False

    workbook_path = training_dir / "labels_review.xlsx"
    workbook.save(workbook_path)
    return workbook_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "动态识别历史扫描版重复表区数量以及每区五、六或七字段，"
            "支持同页跨站并按站点分别输出Excel。"
        ),
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("pdf", type=Path, help="输入PDF文件")
    parser.add_argument("-o", "--output", type=Path, help="输出目录")
    parser.add_argument("--device", default="gpu:0", help="Paddle推理设备")
    parser.add_argument("--batch-size", type=int, default=64)
    parser.add_argument("--low-confidence", type=float, default=0.85)
    parser.add_argument(
        "--sixth-field",
        choices=("auto", "sediment", "slope"),
        default="auto",
        help=(
            "六字段表第六列的含义；auto按打印表头识别，"
            "表头质量过差时可明确指定sediment（含沙量）或slope（水面比降）"
        ),
    )
    parser.add_argument("--reuse-structure", action="store_true")
    parser.add_argument("--save-structure-preview", action="store_true")
    parser.add_argument(
        "--numeric-model-name",
        default="en_PP-OCRv5_mobile_rec",
        help="没有本地微调模型时，用于数字候选投票的PaddleOCR模型",
    )
    parser.add_argument(
        "--numeric-model-dir",
        type=Path,
        help=(
            "本地微调数字识别模型目录；未指定时自动使用项目models目录下的"
            "hydrology_numeric_ppocrv5_mobile_v11（若存在）"
        ),
    )
    parser.add_argument(
        "--disable-numeric-model",
        action="store_true",
        help="禁用第二数字倾向模型，仅使用PP-OCRv5_server_rec",
    )
    parser.add_argument(
        "--export-training-samples",
        action="store_true",
        help="导出实际OCR单元格图片和可编辑的labels_review.xlsx",
    )
    parser.add_argument(
        "--training-samples-dir",
        type=Path,
        help="训练样本目录；默认使用输出目录下的training_samples",
    )
    parser.add_argument(
        "--corrected-labels",
        type=Path,
        help="可选的labels_review_corrected.xlsx；采用其中纳入训练=是的正确标签重建最终结果",
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()
    pdf_path = args.pdf.expanduser().resolve()
    if not pdf_path.is_file():
        raise FileNotFoundError(f"输入PDF不存在：{pdf_path}")
    output_dir = (
        args.output.expanduser().resolve()
        if args.output
        else (Path("output") / f"{pdf_path.stem}_processed_ex2").resolve()
    )
    structure_dir = output_dir / "structure"
    preview_dir = output_dir / "previews"
    training_dir = (
        args.training_samples_dir.expanduser().resolve()
        if args.training_samples_dir
        else output_dir / "training_samples"
    )
    training_images_dir = training_dir / "images"
    structure_dir.mkdir(parents=True, exist_ok=True)
    preview_dir.mkdir(parents=True, exist_ok=True)

    verify_device(args.device)
    if args.reuse_structure:
        page_results = load_saved_structure(structure_dir)
    else:
        page_results = run_structure_ocr(
            pdf_path,
            structure_dir,
            args.device,
            args.save_structure_preview,
        )

    log("[2/4] 初始化中文手写分格识别模型")
    recognizer = TextRecognition(
        model_name="PP-OCRv5_server_rec",
        device=args.device,
        engine="paddle",
    )
    numeric_recognizer: TextRecognition | None = None
    if not args.disable_numeric_model:
        numeric_model_dir: Path | None
        if args.numeric_model_dir:
            numeric_model_dir = args.numeric_model_dir.expanduser().resolve()
            if not numeric_model_dir.is_dir():
                raise FileNotFoundError(
                    f"本地数字识别模型目录不存在：{numeric_model_dir}"
                )
        elif DEFAULT_FINETUNED_NUMERIC_MODEL_DIR.is_dir():
            numeric_model_dir = DEFAULT_FINETUNED_NUMERIC_MODEL_DIR
        else:
            numeric_model_dir = None

        if numeric_model_dir is not None:
            log(f"[2/4] 初始化水文数字微调模型：{numeric_model_dir}")
            numeric_recognizer = TextRecognition(
                model_name="PP-OCRv5_mobile_rec",
                model_dir=str(numeric_model_dir),
                device=args.device,
                engine="paddle",
            )
        else:
            log(
                f"[2/4] 初始化数字倾向候选模型：{args.numeric_model_name}"
            )
            numeric_recognizer = TextRecognition(
                model_name=args.numeric_model_name,
                device=args.device,
                engine="paddle",
            )

    reconstructed_pages: list[
        tuple[
            list[dict[str, Any]],
            dict[str, Any],
            list[dict[str, Any]],
        ]
    ] = []
    raw_rows: list[dict[str, Any]] = []
    training_samples: list[dict[str, Any]] = []
    for page_number, result in enumerate(page_results, start=1):
        log(f"[2/4] 正在识别第{page_number}/{len(page_results)}页")
        rows, metadata, page_samples = reconstruct_legacy_page(
            pdf_path,
            result,
            recognizer,
            numeric_recognizer,
            preview_dir,
            args.batch_size,
            training_images_dir if args.export_training_samples else None,
            args.sixth_field,
        )
        reconstructed_pages.append((rows, metadata, page_samples))
        raw_rows.extend(rows)
        training_samples.extend(page_samples)
        counts = ", ".join(
            f"表区{item['panel']}={item['row_count']}行"
            for item in metadata["panels"]
        )
        marker_summary = "、".join(
            f"{marker['station']}@y={marker['y']}"
            for marker in metadata.get("station_markers", [])
        ) or "本页未重复站名（将继承上一页）"
        log(
            f"[2/4] 第{page_number}页完成：{counts}；"
            f"字段={'、'.join(metadata['schema']['canonical_fields'])}；"
            f"站名标记={marker_summary}"
        )
        for marker in metadata.get("station_markers", []):
            log(
                f"[站点] 第{page_number}页 y={marker['y']}："
                f"{marker['station']}（{marker['confidence']:.3f}，"
                f"{marker.get('source', 'unknown')}）"
            )

    if not raw_rows:
        raise RuntimeError("没有重建出任何数据行。")
    if args.corrected_labels:
        corrected_labels_path = args.corrected_labels.expanduser().resolve()
        if not corrected_labels_path.is_file():
            raise FileNotFoundError(
                f"标签校正表不存在：{corrected_labels_path}"
            )
        applied = apply_training_label_corrections(
            raw_rows, corrected_labels_path
        )
        merged_samples = merge_review_state_into_training_samples(
            training_samples, corrected_labels_path
        )
        log(
            f"[3/4] 已从标签校正表应用{applied}处字段修正："
            f"{corrected_labels_path}"
        )
        log(
            f"[3/4] 已继承{merged_samples}条训练样本的校正标签和审核状态"
        )
    station_groups = split_legacy_rows_by_station(pdf_path, reconstructed_pages)
    if not station_groups:
        raise RuntimeError("没有可以输出的站点数据。")
    log(
        "[3/4] 按站点独立补齐省略值并规范化动态字段数据："
        + "，".join(
            f"{station}={len(group['raw_rows'])}行"
            for station, group in station_groups.items()
        )
    )

    log("[4/4] 按站点分别写出Excel、JSON和分格预览")
    outputs: list[dict[str, Any]] = []
    for station, group in station_groups.items():
        normalized = normalize_legacy_rows(group["raw_rows"])
        context_review_count = sync_context_choices_to_training_samples(
            group["training_samples"], normalized
        )
        if context_review_count:
            log(
                f"[3/4] {station}：已将{context_review_count}条上下文改选结果"
                "标记为待人工确认"
            )
        output_columns = [
            LEGACY_COLUMN_NAMES[index]
            for index in range(len(LEGACY_COLUMN_NAMES))
            if index in group["canonical_indices"]
        ]
        workbook_path, json_path = write_legacy_outputs(
            pdf_path,
            output_dir,
            station,
            normalized,
            list(group["pages"].values()),
            output_columns,
            args.low_confidence,
        )
        outputs.append(
            {
                "station": station,
                "row_count": len(normalized),
                "output_columns": output_columns,
                "excel": str(workbook_path),
                "json": str(json_path),
            }
        )
        log(f"[完成] {station}：{len(normalized)}行")
        log(f"[完成] Excel：{workbook_path}")
        log(f"[完成] JSON：{json_path}")

    manifest_path = output_dir / f"{pdf_path.stem}_stations.json"
    manifest_path.write_text(
        json.dumps(
            {"input_pdf": str(pdf_path), "stations": outputs},
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    log(f"[完成] 站点清单：{manifest_path}")
    log(f"[完成] 分格预览：{preview_dir}")
    if args.export_training_samples:
        training_dir.mkdir(parents=True, exist_ok=True)
        review_path = write_training_review_workbook(
            training_samples,
            training_dir,
        )
        log(f"[完成] 训练样本图片：{training_images_dir}")
        log(f"[完成] 标签校正表：{review_path}")
        log(f"[完成] 训练样本数：{len(training_samples)}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        log("已取消。")
        sys.exit(130)
